"""build_template_from_parkwood.py

Build the runtime template `GGC_Blank_Underwriting_Sizer_Extended.xlsx`
directly from `Outputs/Testing/ParkwoodCorrect.xlsx` — the analyst's
hand-built Parkwood model — so the runtime workbook has the EXACT
structure (formulas, labels, layout, formatting) that an analyst
recognizes.

Why this exists:
- `Claude/CorrectOutput.xlsx` (the old build_template source) is the
  WHALESHEAD model. Its Rent Roll Input tab has Whaleshead's column
  layout (C=Unit Type canonical, D=Status, I=Lot Rent, J=Home Rent)
  with hardcoded "WHA Lot" labels. Building from there left the engine
  emitting Whaleshead-shaped workbooks for Parkwood-shaped deals.
- ParkwoodCorrect.xlsx is the analyst's MHC reference. It has the
  Parkwood layout (B=Lot#, C=Lot Type SHORT, D=Unit Type derived,
  E=Status, F=Tenant, G=Move in, H=Lot Rent, I=POH Home Rents,
  J=LTO PMT) and four-bucket Unit Mix Summary (TOH/POH/LTO/Flourish).
- Using ParkwoodCorrect as the canonical source means every run starts
  from the right structure. Deal-specific values get stripped here so
  the engine can populate them per upload.

What this script does:
1. Load Outputs/Testing/ParkwoodCorrect.xlsx
2. Map (PW) tabs → engine tab names; delete CME / combined / extras
3. Strip Parkwood-specific input values (rent values, contract price,
   per-row tenant data, etc.) while preserving formulas and labels
4. Save as GGC_Blank_Underwriting_Sizer_Extended.xlsx
5. Run fix_template.py on top for any patches the analyst's reference
   doesn't include

Run: `python3 build_template_from_parkwood.py`
"""
from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).parent
SRC  = ROOT / "Outputs" / "Testing" / "ParkwoodCorrect.xlsx"
DEST = ROOT / "GGC_Blank_Underwriting_Sizer_Extended.xlsx"


# Map ParkwoodCorrect tab names → engine-expected tab names.
# Tabs not in this map (CME, combined, comments, Property, LP, Snapshot,
# Sale Comps, Key assumptions, Waterfall variants) are dropped.
TAB_RENAME = {
    "Rent Roll Input(PW)":          "Rent Roll Input",
    "Data Consolidation(PW)":       "Data Consolidation",
    "Collections(PW)":              "Collections",
    "Unit Mix Summary(PW)":         "Unit Mix Summary",
    "Unit Mix Rent Growth(PW)":     "Unit Mix Rent Growth",
    "GGC Underwriting(PW)":         "GGC Underwriting",
    "GGC Pro Forma(PW)":            "GGC Pro Forma",
    "Sources and Uses(PW)":         "Sources and Uses",
    "Loan Scenario (PW) ":          "Loan Scenario (acquisition)",
    "Investor Return":              "Investor Return",
    "Waterfall (5-yr) ":            "Waterfall (5-yr) ",
    "Waterfall (Sparta)":           "Waterfall (10-yr)",
    "Comps":                        "Comps",
}


def _is_formula(cell_value) -> bool:
    return isinstance(cell_value, str) and cell_value.startswith("=")


def _clear_cell(cell):
    """Set a cell's value to None ONLY if it doesn't hold a formula.
    Preserves formatting, style, and column width by mutating .value."""
    if not _is_formula(cell.value):
        cell.value = None


def strip_rent_roll_input(ws):
    """Clear per-tenant data rows (3 onward) but keep header + formulas.
    Parkwood layout columns:
      A = Count (formula =A{r-1}+1)
      B = Lot #            (input)
      C = Lot Type         (input — short TOH/POH/LTO/Flourish)
      D = Unit Type        (formula =IF(C=...) deriving Type 1..4)
      E = Occupied/Vacant  (input)
      F = Tenants & lot#   (input)
      G = Move in          (input — date)
      H = Lot Rent         (input)
      I = POH Home Rents   (input)
      J = LTO PMT          (input)
      K = Combined         (formula =SUM(H:J))
    Strip B,C,E,F,G,H,I,J. Keep A,D,K (formulas).
    """
    n_cleared = 0
    for r in range(3, ws.max_row + 1):
        # Input columns (B,C,E,F,G,H,I,J): clear non-formula values only.
        for col in (2, 3, 5, 6, 7, 8):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None and not _is_formula(cell.value):
                cell.value = None
                n_cleared += 1
        # Columns I (POH Home Rents) and J (LTO PMT): wipe EVERYTHING
        # including formulas. ParkwoodCorrect carries sub-total formulas
        # at I103/I105 (=SUM(I3:I102), =I103*12) that DOUBLE-COUNT when
        # GGC Underwriting K13 sums the whole I column. Same applies to
        # any J sub-totals. The per-row writer in backend.py will
        # re-seed only the rows that have real input data.
        for col in (9, 10):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None:
                cell.value = None
                n_cleared += 1
    return n_cleared


def strip_unit_mix_summary(ws):
    """Unit Mix Summary's per-type rows (4-7) carry computed COUNTIFS
    against Rent Roll Input — those are formulas, leave them. Any
    hardcoded Parkwood-specific overrides (e.g. C8 Total Sites, or a
    hand-typed avg lot rent) need clearing. Be conservative: only
    clear cells in rows 4-12 that DON'T hold a formula."""
    n_cleared = 0
    for r in range(4, 21):
        for col in range(2, 15):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None and not _is_formula(cell.value):
                # Preserve text labels in column B (the row name).
                if col == 2 and isinstance(cell.value, str):
                    continue
                cell.value = None
                n_cleared += 1
    return n_cleared


def strip_data_consolidation(ws):
    """DC's per-line input rows 3-36 (income) and 43-102 (expense).
    Strip column A (GGC Category) too — ParkwoodCorrect has 40+ rows of
    'Gross Potential Rent' inherited from Parkwood's per-tenant data
    that aren't a useful canonical structure. The engine writes the
    category on each row when populating, so blanks are fine."""
    n_cleared = 0
    for r in list(range(3, 37)) + list(range(43, 103)):
        for col in [1, 2] + list(range(4, 22)):  # A (category), B (seller), D-U
            cell = ws.cell(row=r, column=col)
            if cell.value is not None and not _is_formula(cell.value):
                cell.value = None
                n_cleared += 1
    return n_cleared


def add_rv_reachability_row(wb):
    """The MHC-focused Parkwood layout uses GGC Underwriting row 14
    for 'LTO' (the LC payment stream). The 'RV Site Rental Income'
    enum entry — needed for RV/Whaleshead-style deals — has no
    matching SUMIFS row in this layout. Add row 18 as a parallel
    SUMIFS row so RV deals don't silently zero. Mirrors the patch
    that build_template.py applies to the legacy template."""
    if "GGC Underwriting" not in wb.sheetnames:
        return 0
    uw = wb["GGC Underwriting"]
    if (uw["A18"].value or "").strip() not in ("", "None"):
        return 0
    uw["A18"] = "RV Site Rental Income"
    rv_sumifs = {
        "B": "=SUMIFS('Data Consolidation'!$D$3:$D$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
        "C": "=SUMIFS('Data Consolidation'!$E$3:$E$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
        "D": "=SUMIFS('Data Consolidation'!$G$3:$G$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
        "E": "=SUMIFS('Data Consolidation'!$H$3:$H$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
        "F": "=SUMIFS('Data Consolidation'!$F$3:$F$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
    }
    for col, formula in rv_sumifs.items():
        uw[f"{col}18"] = formula
    uw["I18"] = 0
    uw["J18"] = "=I18/$N$7"
    uw["K18"] = 0
    # Extend EGI sum at row 19 to include row 18.
    for col in ("B", "C", "D", "E", "F", "G", "H", "I", "K"):
        cell = uw[f"{col}19"]
        val = cell.value
        if isinstance(val, str) and f"{col}17" in val and f"{col}18" not in val:
            cell.value = val.replace(f"{col}17", f"{col}17+{col}18")
    return 1


def strip_ggc_underwriting(ws):
    """GGC Underwriting input cells:
    - P4-P10: Subject property info (name, address, units, etc.)
    - R-column: Contract price, asking price
    Keep all formulas, labels, J-column $/site analyst inputs.
    Only clear non-formula values in the input block."""
    n_cleared = 0
    for addr in ["N4", "N5", "N6", "N7", "N8", "N9", "N10",
                 "P4", "P5", "P6", "P7", "P8", "P9", "P10",
                 "R4", "R9"]:
        try:
            cell = ws[addr]
            if cell.value is not None and not _is_formula(cell.value):
                cell.value = None
                n_cleared += 1
        except Exception:
            pass
    return n_cleared


def strip_sources_and_uses(ws):
    """S&U deal-specific inputs: B2 (property name), C8 (GP equity),
    C13 (contract price). C14 is a formula, preserve."""
    for addr in ["B2", "C8", "C13"]:
        try:
            cell = ws[addr]
            if cell.value is not None and not _is_formula(cell.value):
                cell.value = None
        except Exception:
            pass
    return 3


def strip_loan_scenario(ws):
    """Loan Scenario deal inputs: C6 (lender), C9 (IO mo), C14 (base
    rate), C15 (spread), C19 (price), C24 (LTV), C25 (DSCR). C19 is
    typically a formula (=Sources and Uses!C13) — preserve."""
    for addr in ["C6", "C9", "C14", "C15", "C24", "C25"]:
        try:
            cell = ws[addr]
            if cell.value is not None and not _is_formula(cell.value):
                cell.value = None
        except Exception:
            pass
    return 6


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"Source not found: {SRC}")

    print(f"[build] Source: {SRC}")
    print(f"[build] Dest:   {DEST}")

    # Backup existing template before overwriting.
    if DEST.exists():
        bak = DEST.with_suffix(".xlsx.preParkwoodTpl.bak")
        shutil.copy(DEST, bak)
        print(f"[build] Backup: {bak.name}")

    wb = load_workbook(SRC)
    src_tabs = list(wb.sheetnames)
    print(f"[build] Source tab count: {len(src_tabs)}")

    # 1. Drop every tab NOT in the rename map. Use a 2-pass to avoid
    # mutating sheetnames mid-iteration.
    keep = set(TAB_RENAME.keys())
    to_delete = [s for s in src_tabs if s not in keep]
    for s in to_delete:
        del wb[s]
    print(f"[build] Dropped {len(to_delete)} tabs not in keep-list")

    # 2. Rename remaining tabs to engine-expected names.
    for src_name, dest_name in TAB_RENAME.items():
        if src_name in wb.sheetnames and src_name != dest_name:
            ws = wb[src_name]
            ws.title = dest_name
            print(f"[build] Renamed {src_name!r} → {dest_name!r}")

    # 2a. Rewrite all formula cross-tab references to use the new
    # tab names. openpyxl renames the SHEET TITLE but leaves embedded
    # formula strings like `='Data Consolidation(PW)'!A3` untouched,
    # so every cross-sheet formula evaluates to #NAME? after the
    # rename. Walk every cell in every remaining tab and rewrite.
    n_formulas_rewritten = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                original = v
                for src_name, dest_name in TAB_RENAME.items():
                    if src_name == dest_name:
                        continue
                    # Replace both quoted ('Data Consolidation(PW)') and
                    # unquoted forms. Some formulas use 'sheetname' with
                    # quotes when the name has spaces or punctuation;
                    # tabs we're renaming all have parens so they're
                    # always quoted in formula strings.
                    v = v.replace(f"'{src_name}'", f"'{dest_name}'")
                if v != original:
                    cell.value = v
                    n_formulas_rewritten += 1
    print(f"[build] Rewrote {n_formulas_rewritten} cross-tab formula refs")

    # 3. Order tabs to match engine expectations.
    desired_order = [
        "Comps",
        "Data Consolidation",
        "Rent Roll Input",
        "Collections",
        "Unit Mix Rent Growth",
        "Unit Mix Summary",
        "GGC Underwriting",
        "GGC Pro Forma",
        "Investor Return",
        "Waterfall (10-yr)",
        "Waterfall (5-yr) ",
        "Sources and Uses",
        "Loan Scenario (acquisition)",
    ]
    # openpyxl preserves order via .move_sheet; move each to final position
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            ws = wb[name]
            current_index = wb.sheetnames.index(name)
            offset = i - current_index
            if offset != 0:
                wb.move_sheet(ws, offset=offset)

    # 4. Strip deal-specific values per tab.
    if "Rent Roll Input" in wb.sheetnames:
        n = strip_rent_roll_input(wb["Rent Roll Input"])
        print(f"[build] Rent Roll Input: cleared {n} per-tenant input cells")
    if "Unit Mix Summary" in wb.sheetnames:
        n = strip_unit_mix_summary(wb["Unit Mix Summary"])
        print(f"[build] Unit Mix Summary: cleared {n} input cells")
    if "Data Consolidation" in wb.sheetnames:
        n = strip_data_consolidation(wb["Data Consolidation"])
        print(f"[build] Data Consolidation: cleared {n} per-line input cells")
    if "GGC Underwriting" in wb.sheetnames:
        n = strip_ggc_underwriting(wb["GGC Underwriting"])
        print(f"[build] GGC Underwriting: cleared {n} property-info input cells")
    if "Sources and Uses" in wb.sheetnames:
        strip_sources_and_uses(wb["Sources and Uses"])
        print(f"[build] Sources and Uses: cleared deal-input cells")
    if "Loan Scenario (acquisition)" in wb.sheetnames:
        strip_loan_scenario(wb["Loan Scenario (acquisition)"])
        print(f"[build] Loan Scenario: cleared deal-input cells")

    # 4a. Add the RV Site Rental Income SUMIFS row so RV/Whaleshead-style
    # deals stay reachable in this MHC-focused template.
    n = add_rv_reachability_row(wb)
    if n:
        print(f"[build] Added RV Site Rental Income SUMIFS row 18")

    # 5. Save.
    wb.save(DEST)
    print(f"[build] Wrote {DEST.name}")
    print(f"[build] Final tabs ({len(wb.sheetnames)}): {wb.sheetnames}")

    # 6. Run fix_template.py for any patches the analyst reference
    # didn't include (typo fixes, S3/Scen 4 scenario blocks, etc.).
    # Skip with PARKWOOD_TPL_SKIP_FIX=1 to inspect the clean strip.
    import os
    if os.environ.get("PARKWOOD_TPL_SKIP_FIX") != "1":
        try:
            print("[build] Running fix_template.py on top...")
            import importlib
            import sys
            if "fix_template" in sys.modules:
                del sys.modules["fix_template"]
            importlib.import_module("fix_template")
            print("[build] fix_template.py completed")
        except Exception as e:
            print(f"[build] WARNING: fix_template.py raised "
                  f"{type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
