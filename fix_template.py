"""
Comprehensive structural fix for GGC_Blank_Underwriting_Sizer_Extended.xlsx.

The blank template ships with ~20 defects that produce broken output across
every downstream tab. This script patches all of them in one pass. Run once
after pulling the original blank.

Fix categories:
  1. GGC Underwriting tab: income labels, per-unit formulas, subject block,
     stabilized assumptions
  2. Unit Mix Summary: replace the 7 generic 'Type N' rows with 4 real
     categories (TOH MH / POH-Infilled / Long term RV / Retail/Commercial)
     and re-aim COUNTIFS at the new Rent Roll column layout
  3. Rent Roll Input: switch to the correct column layout (Unit, Unit Type,
     Status, Name, Type detail, Type code, Lot Rent, Home Rent, Combined)
  4. Unit Mix Rent Growth: collapse from 6 broken #REF! rows to 2 real rows
     (Lots / POH) pulling from Unit Mix Summary
  5. Loan Scenario: populate US Treasury rate, spread, and Max LTV so
     debt math stops returning zero
  6. Sources and Uses: fix Purchase Price reference, Closing Costs %, GP
     equity formula, capex linkage, and unit-count denominator
  7. GGC Pro Forma: kill the spurious 1.03 multiplier on Y1 RE taxes; fix
     the Home Rent Expense ratio and its SUM range; rename Lease-to-Own
     row to Long term RV Site
  8. Waterfall promote tiers: 20/30/30 -> 25/40/40
  9. Investor Return: repoint at canonical waterfall tabs
  10. Delete junk duplicate Waterfall tabs (' ', '2', '3')
  11. Force fullCalcOnLoad so Excel populates formula cache on open
"""
from pathlib import Path
import openpyxl

TEMPLATE = Path(__file__).parent / "GGC_Blank_Underwriting_Sizer_Extended.xlsx"

# ════════════════════════════════════════════════════════════════════════
# FORM-INPUT DEFAULTS (Parkwood OUTLINE.md)
# ════════════════════════════════════════════════════════════════════════
# These are the template-bake-time defaults written to cells the analyst
# (and backend.py at runtime) can override per-deal. backend.py reads the
# per_site_overrides / lot_cap_rate / etc. form inputs and rewrites the
# same cells with the deal-specific values; when the form leaves them
# blank, the defaults below survive.
#
# Provenance: CorrectOutput/parkwoodCorrect.xlsx is the gold standard
# (analyst hand-underwrote Parkwood Green Village). Where CLAUDE.md and
# CorrectOutput disagree, CorrectOutput wins because it ties out cell-
# for-cell to a real deal Michael signed off on.
#   - capex_per_unit: CLAUDE.md outer says $75, inner says $50;
#     CorrectOutput uses $50 (L43 = 50). Using 50.
#   - insurance_per_site: CorrectOutput L23 = 250 (non-flood). The flood
#     toggle is a form input; backend writes 300 when set.
#   - bad_debt_uw_pct: CorrectOutput L7 = 0.02 (2% × UW GPR). Was 0.03.
#   - closing_cost_pct: CorrectOutput Sources & Uses C14 = C13*2.25%.
#     Template was 1.5%.
#   - gp_equity_default: CorrectOutput Sources & Uses C8 = 300000.

DEFAULTS = {
    "bad_debt_uw_pct":             0.02,
    "insurance_per_site_nonflood": 250,
    "insurance_per_site_flood":    300,
    "payroll_per_site":            425,
    "ground_maintenance_per_site": 200,
    "ga_per_site":                 100,
    "professional_fees_per_site":  50,
    "advertising_per_site":        0,
    "capex_per_unit":              50,
    "home_rent_expense_ratio":     0.10,
    "lot_cap_rate":                0.05,
    "home_cap_rate":               0.20,
    "y1_re_taxes_growth_pct":      0.03,
    "closing_cost_pct":            0.0225,
    "gp_equity_default":           300000,
    "exit_cap_rate":               0.06,
    "hold_period_years":           10,
    "vacant_stabilization_count":  3,
}

# Per-type 10-yr rent-growth schedules. Override per deal via
# property_info.rentGrowthSchedule.
DEFAULT_RENT_GROWTH = {
    "toh":      [0.10, 0.11, 0.10, 0.065, 0.06, 0.06, 0.05, 0.05, 0.05, 0.05],
    "poh":      [0.10, 0.11, 0.10, 0.065, 0.06, 0.06, 0.05, 0.05, 0.05, 0.05],
    "flourish": [0.10, 0.11, 0.10, 0.065, 0.06, 0.06, 0.05, 0.05, 0.05, 0.05],
    "lto":      [0.10, 0.09, 0.085, 0.065, 0.075, 0.075, 0.04, 0.04, 0.04, 0.04],
}

wb = openpyxl.load_workbook(TEMPLATE)

# ════════════════════════════════════════════════════════════════════════
# 1. GGC UNDERWRITING TAB
# ════════════════════════════════════════════════════════════════════════
ws = wb["GGC Underwriting"]

# Income category labels (rows 14-16). These feed SUMIFS into Data
# Consolidation; must match the categories we tag GL accounts with in
# backend.py. Row 14 doubles as LTO for MHC deals (K14 is sourced from
# Rent Roll Input column J = LTO PMT in the Parkwood-style layout)
# and as RV Site Rental Income for resort deals (Whaleshead). backend.py
# can rewrite A14 per property_info.propertyType at runtime.
ws["A14"] = "LTO"   # was "RV Site Rental Income"; default to MHC layout
ws["A15"] = "Storage Income"
ws["A16"] = "Retail Income"

# Subject property value cells (M:N block). Template ships with labels at
# column O but no value at all for # of Units, so every per-unit formula
# resolves to 0. Layout mirrors CorrectOutput: M4 Name, M5 Address, M6
# Type, M7 Units (the missing M4 label was the reason backend was writing
# the property name into the Address slot).
#
# Header at M3 was a stray "Notes" label sitting on top of the block with
# no data below it — Michael in the walkthrough read it as the section
# title for the property info ("Why is it saying all of this? Under
# notes."). Rename to "Property Information" so the section reads
# correctly. Also bold every label row so M4:M8 match the existing bold
# on M9/M10 and on the O-column labels (the unstyled M4:M8 labels looked
# like loose data rather than field names).
from openpyxl.styles import Font as _Font
_label_bold = _Font(bold=True)
ws["M3"] = "Property Information"
ws["M3"].font = _label_bold
ws["M4"] = "Property Name"
ws["M5"] = "Property Address"
ws["M6"] = "Property Type"
ws["M7"] = "# of Units"
ws["N7"] = "='Unit Mix Summary'!E8"
ws["M8"] = "Rent Roll Occupancy"
ws["N8"] = "='Unit Mix Summary'!C13"
ws["M9"] = "Acreage"
ws["M10"] = "County"
for coord in ("M4", "M5", "M6", "M7", "M8", "M9", "M10"):
    ws[coord].font = _label_bold

# GGC ProForma per-unit assumptions (column J). Values come from GGC's
# playbook and replace the model's "T12 x 1.03" guess.
ws["J22"] = 400   # RE Taxes  $/unit
ws["J30"] = 150   # R&M       $/unit
ws["J35"] = 600   # Payroll   $/unit
ws["J43"] = 75    # Cap-Ex    $/unit (was 0)

# Stabilized (GGC ProForma) column I formulas.
ws["G5"] = "=-5%*G4"             # vacancy 3% -> 5%
ws["I22"] = "=J22*N7"            # RE Taxes per-unit
ws["I30"] = "=J30*$N$7"          # R&M per-unit
ws["I35"] = "=J35*$N$7"          # Payroll per-unit
ws["I41"] = "=15%*I13"           # Home Rent Expense = 15% of Home Rent Income
ws["I43"] = "=J43*$N$7"          # Cap-Ex per-unit

# Stabilized income rows for the renamed labels.
ws["I14"] = "='Unit Mix Summary'!C15*90%"  # RV Site Rental
ws["I15"] = "=D15"                          # Storage (small line, keep at T12)
ws["I16"] = "='Unit Mix Summary'!H7*12*98%" # Retail

# Per-unit (column J) reference fixes: P7/P6/P8 used pricing cells where
# unit count and occupancy don't actually live. Repoint at N7/N8.
J_PER_UNIT = {
    "J5":  "=100%-N8", "J7":  "=-I7/I4", "J10": "=I9/$N$7", "J12": "=I12/$N$7",
    "J13": "=I13/$N$7","J14": "=I14/$N$7","J15": "=I15/$N$7","J16": "=I16/$N$7",
    "J17": "=I17/$N$7","J19": "=I19/$N$7","J20": "=$I$19/$N$7","J23": "=I23/$N$7",
    "J25": "=I25/$N$7","J26": "=I26/$N$7","J27": "=I27/$N$7","J28": "=I28/$N$7",
    "J29": "=I29/$N$7","J31": "=I31/$N$7","J32": "=I32/$N$7","J36": "=I36/$N$7",
    "J37": "=I37/$N$7","J38": "=I38/$N$7","J39": "=I39/$N$7","J40": "=I40/$N$7",
    "J41": "=I41/$N$7","J42": "=I42/$N$7","J44": "=I44/$N$7","J45": "=$I$44/$N$7",
}
for coord, formula in J_PER_UNIT.items():
    ws[coord] = formula

# Lines that should NOT auto-grow from T12 in the stabilized column.
ws["I32"] = 0   # Recreational Amenities
ws["I36"] = 0   # Employee Allowance
ws["I38"] = 0   # Model Units
ws["I42"] = 0   # Other

# ════════════════════════════════════════════════════════════════════════
# 2. UNIT MIX SUMMARY — 4 lot-rent rows + 4 home-rent rows + derived metrics
# ════════════════════════════════════════════════════════════════════════
# Matches CorrectOutput/parkwoodCorrect.xlsx Unit Mix Summary(PW). The
# 4 canonical lot-rent buckets are TOH / POH / LTO / Flourish (LTO and
# Flourish are MHC seller-financing variants distinct from plain TOH —
# folding them into TOH wiped Parkwood's ~$200k/yr LTO revenue stream).
# Below row 8 sits a parallel home-rent block at rows 9-12, and derived
# metrics at rows 13-20. COUNTIFS/SUMIFS key off the Rent Roll Input
# column D (Type 1..4 derived) and column E (Occupied/Vacant), with H/I/J
# for Lot Rent / POH Home Rents / LTO PMT respectively (see section 3).
ums = wb["Unit Mix Summary"]

# Clear values in the data block (rows 3-22) before rewriting. Preserves
# column widths / styling / merged cells.
for row in ums.iter_rows(min_row=3, max_row=22, max_col=11):
    for cell in row:
        cell.value = None

# Headers (row 3). Clear any stray text in row 2 cells C-K — the original
# blank template had a "Next Increase in MAY 1 2026 — see RR" note at C2
# left over from a hand-built workbook, which showed up next to our title.
for col in range(3, 12):
    ums.cell(row=2, column=col).value = None
ums["B2"] = "MH/RV Rent Roll "  # trailing space matches CorrectOutput
ums["C3"] = "# of Occupied Units"
ums["D3"] = "# of Vacant Units"
ums["E3"] = "Total Units"
ums["F3"] = "Occupied Monthly Rent"
ums["G3"] = "Vacant Monthly Rent"
ums["H3"] = "Monthly Gross Potential Rent"
ums["I3"] = "avg rents"
ums["J3"] = "weights"

# Four canonical lot-rent categories at rows 4-7, each matching a Type
# code (Type 1=TOH, Type 2=POH, Type 3=LTO, Type 4=Flourish) derived in
# Rent Roll Input column D. SUMIFS pull column H (Lot Rent) only.
LOT_CATEGORIES = [
    ("TOH(Lot Rents only)-vacant lots", 4, "Type 1"),
    ("POH(Lot Rents)",                  5, "Type 2"),
    ("LTO(Lot Rents)",                  6, "Type 3"),
    ("Flourish (Lot Rents)",            7, "Type 4"),
]
for label, r, type_code in LOT_CATEGORIES:
    ums.cell(row=r, column=2, value=label)
    ums.cell(row=r, column=3,
             value=f'=COUNTIFS(\'Rent Roll Input\'!$E$3:$E$1002,"Occupied",'
                   f'\'Rent Roll Input\'!$D$3:$D$1002,"{type_code}")')
    ums.cell(row=r, column=4,
             value=f'=COUNTIFS(\'Rent Roll Input\'!$E$3:$E$1002,"Vacant",'
                   f'\'Rent Roll Input\'!$D$3:$D$1002,"{type_code}")')
    ums.cell(row=r, column=5, value=f"=SUM(C{r}:D{r})")
    ums.cell(row=r, column=6,
             value=f'=SUMIFS(\'Rent Roll Input\'!$H$3:$H$1002,'
                   f'\'Rent Roll Input\'!$E$3:$E$1002,"Occupied",'
                   f'\'Rent Roll Input\'!$D$3:$D$1002,"{type_code}")')
    ums.cell(row=r, column=7,
             value=f'=SUMIFS(\'Rent Roll Input\'!$H$3:$H$1002,'
                   f'\'Rent Roll Input\'!$E$3:$E$1002,"Vacant",'
                   f'\'Rent Roll Input\'!$D$3:$D$1002,"{type_code}")')
    ums.cell(row=r, column=8, value=f"=F{r}+G{r}")
    ums.cell(row=r, column=9, value=f"=IFERROR(H{r}/E{r},0)")  # avg rent per type
    ums.cell(row=r, column=10, value=f"=IFERROR(E{r}/$E$8,0)") # weight

# Row 8: Total Sites (Lot Rent Only) — SUMs across rows 4-7
ums["B8"] = "Total Sites(Lot Rent only)"
ums["C8"] = "=SUM(C4:C7)"
ums["D8"] = "=SUM(D4:D7)"
ums["E8"] = "=SUM(E4:E7)"
ums["F8"] = "=SUM(F4:F7)"
ums["G8"] = "=SUM(G4:G7)"
ums["H8"] = "=SUM(H4:H7)"

# Rows 9-12: Home-rent block. POH home rents from Rent Roll Input column
# I, LTO from column J, Flourish from column J as well (sub-brand
# financing — backend can re-route via property_info). Row 12 is the
# Total POH (Home Rents) SUM.
HOME_RENT_ROWS = [
    ("POH (Home Rents)",      9,  "Type 2", "I"),
    ("LTO (Home Rents)",      10, "Type 3", "J"),
    ("Flourish (Home Rents)", 11, "Type 4", "J"),
]
for label, r, type_code, src_col in HOME_RENT_ROWS:
    ums.cell(row=r, column=2, value=label)
    ums.cell(row=r, column=3,
             value=f'=COUNTIFS(\'Rent Roll Input\'!$E$3:$E$1002,"Occupied",'
                   f'\'Rent Roll Input\'!$D$3:$D$1002,"{type_code}")')
    ums.cell(row=r, column=4,
             value=f'=COUNTIFS(\'Rent Roll Input\'!$E$3:$E$1002,"Vacant",'
                   f'\'Rent Roll Input\'!$D$3:$D$1002,"{type_code}")')
    ums.cell(row=r, column=5, value=f"=SUM(C{r}:D{r})")
    ums.cell(row=r, column=6,
             value=f'=SUMIFS(\'Rent Roll Input\'!${src_col}$3:${src_col}$1002,'
                   f'\'Rent Roll Input\'!$E$3:$E$1002,"Occupied",'
                   f'\'Rent Roll Input\'!$D$3:$D$1002,"{type_code}")')
    ums.cell(row=r, column=7,
             value=f'=SUMIFS(\'Rent Roll Input\'!${src_col}$3:${src_col}$1002,'
                   f'\'Rent Roll Input\'!$E$3:$E$1002,"Vacant",'
                   f'\'Rent Roll Input\'!$D$3:$D$1002,"{type_code}")')
    ums.cell(row=r, column=8, value=f"=SUM(F{r}:G{r})")

ums["B12"] = "Total POH (Home Rents)"
ums["C12"] = "=SUM(C9:C11)"
ums["D12"] = "=SUM(D9:D11)"
ums["E12"] = "=SUM(E9:E11)"
ums["F12"] = "=SUM(F9:F11)"
ums["G12"] = "=SUM(G9:G11)"
ums["H12"] = "=SUM(H9:H11)"

# Rows 13-20: derived metrics. Mirror CorrectOutput exactly. IFERROR
# wrappers stay because LLM-generated rent rolls can produce zero counts
# in a missing unit-type bucket, which would propagate #DIV/0! into
# every downstream per-unit formula.
ums["B13"] = "Annual GPR(Lot Rent Only)"
ums["C13"] = "=H8*12"
ums["B14"] = "Avg Lot Rent"
ums["C14"] = "=SUMPRODUCT(I4:I7,J4:J7)"
ums["B15"] = "Occupancy%"
ums["C15"] = "=IFERROR(C8/E8,0)"
ums["B16"] = "POH%"
ums["C16"] = "=IFERROR(E12/E8,0)"
ums["B17"] = "Annual HRI (POH)"
ums["C17"] = "=H9*12"
ums["B18"] = "Annual LTO HRI"
ums["C18"] = "=H10*12"
ums["B19"] = "Avg Home Rent (LTO)"
ums["C19"] = "=IFERROR(H10/C10,0)"
ums["B20"] = "Avg Home Rent (POH)"
ums["C20"] = "=IFERROR(H9/C9,0)"

# Number formatting for the derived-metrics block. Michael flagged in the
# walkthrough that the Annual cells were rendering as raw integers ("let's
# fix the formatting on the annual to be an actual currency number, no
# decimals"). While we're here, fix the % cells too — without a 0.0%
# format they show 0.95 instead of 95%.
ums["C13"].number_format = '"$"#,##0'   # Annual GPR (Lot Rent Only)
ums["C14"].number_format = '"$"#,##0'   # Avg Lot Rent (monthly)
ums["C15"].number_format = "0.0%"       # Occupancy %
ums["C16"].number_format = "0.0%"       # POH %
ums["C17"].number_format = '"$"#,##0'   # Annual HRI POH
ums["C18"].number_format = '"$"#,##0'   # Annual LTO HRI
ums["C19"].number_format = '"$"#,##0'   # Avg Home Rent LTO
ums["C20"].number_format = '"$"#,##0'   # Avg Home Rent POH

# ════════════════════════════════════════════════════════════════════════
# 3. RENT ROLL INPUT — Parkwood layout w/ LTO PMT column
# ════════════════════════════════════════════════════════════════════════
# CorrectOutput Rent Roll Input(PW) layout:
#   A=Count, B=Lot #, C=Lot Type (seller string: TOH / POH / TOH-LC / Flourish),
#   D=Unit Type (derived Type 1..4), E=Occupied or Vacant, F=Tenants & lot#,
#   G=Move in, H=Lot Rent, I=POH Home Rents, J=LTO PMT, K=Combined (SUM H:J).
# Type-derivation IF chain in D maps seller C string -> Type 1..4 so
# Unit Mix Summary COUNTIFS in section 2 hit the canonical buckets.
# backend.py will write to these columns at runtime.
rr = wb["Rent Roll Input"]

# Clear existing header row 2 and rewrite (cols A:K)
for col in range(1, 13):
    rr.cell(row=2, column=col).value = None
rr["A2"] = "Count"
rr["B2"] = "Lot #"
rr["C2"] = "Lot Type"
rr["D2"] = "Unit Type"
rr["E2"] = "Occupied or Vacant"
rr["F2"] = "Tenants & lot#"
rr["G2"] = "Move in"
rr["H2"] = "Lot Rent"
rr["I2"] = "POH Home Rents"
rr["J2"] = "LTO PMT"
rr["K2"] = "Combined"

# Update formulas on data rows 3-1002. Combined = SUM(H:J) — includes LTO
# so K row totals don't silently drop the land-contract stream. Column D
# derives Type 1..4 from the seller's column C string via IF chain so a
# typo'd "TOH-LC" still routes (matches "LTO" via the IF fallback path).
for r in range(3, 1003):
    rr.cell(row=r, column=1, value=f"=IF(C{r}=\"\",\"\",ROW()-2)")  # Count
    # D: derived Type code from seller C string
    rr.cell(row=r, column=4,
            value=f'=IF(C{r}="TOH","Type 1",'
                  f'IF(C{r}="POH","Type 2",'
                  f'IF(OR(C{r}="LTO",C{r}="TOH-LC",C{r}="TOH - LC"),"Type 3",'
                  f'IF(C{r}="Flourish","Type 4","Type 1"))))')
    rr.cell(row=r, column=11, value=f"=SUM(H{r}:J{r})")             # Combined
    # Clear stale columns: J (LTO PMT) and column 12+ from old layout.
    # The legacy Whaleshead template carries stale `=1250-I{n}` formulas
    # in J139:J151 — each evaluates to $1,250 (since I is blank) and
    # silently inflates UW K14 LTO by $15,000/mo × 12 × 95% = ~$171,000/yr.
    # Wipe column J for every data row; the per-row writer will repopulate
    # rows that have a real lcPayment value.
    rr.cell(row=r, column=10).value = None                          # J — LTO PMT
    for col in (12,):
        rr.cell(row=r, column=col).value = None

# NOTE: Earlier patches placed a "Totals" row at row 151 (with B151="Totals"
# and SUM formulas at I151/J151/K151) so Underwriting!G13 could read the
# home-rent monthly total. That produced an orphaned highlighted row 25+
# rows below the actual data — looked like a stray black square in empty
# space on every output. Removed here. Underwriting!G13 now reads a SUM
# across the entire rent-roll range directly (set below in the
# Underwriting tab section), no fixed totals row needed.
#
# The original blank template ALSO had row 151 styled with a dark fill
# (left over from a hand-built totals row in an earlier version of the
# workbook). That fill survives even after we clear the cell values, so
# we explicitly null the fill and font on row 151 across columns A-K
# to match the empty rows above it.
from openpyxl.styles import PatternFill, Font, Border, Side
# PatternFill(fill_type=None) leaves fgColor defaulting to '00000000' (black)
# in openpyxl's object model. Some Excel / LibreOffice renderers honor that
# leftover fgColor and paint the cell black even though patternType is None.
# Use an explicit solid white fill so the rendered cell is unambiguously white.
_clear_fill = PatternFill(fill_type="solid", start_color="FFFFFFFF", end_color="FFFFFFFF")
_clear_font = Font()
_clear_border = Border(left=Side(border_style=None), right=Side(border_style=None),
                       top=Side(border_style=None), bottom=Side(border_style=None))
for col in range(1, 12):  # A-K
    cell = rr.cell(row=151, column=col)
    cell.value = None
    cell.fill = _clear_fill
    cell.font = _clear_font
    cell.border = _clear_border

# Row 1003 is a legacy "totals" row from the original template: A1003:D1003
# carry a black theme-1 fill (renders as a large black box in the bottom-right
# of the rent roll) and E1003:N1003 hold =SUM(E3:E1002) totals. Nothing reads
# from this row anymore — Unit Mix Summary's COUNTIFS/SUMIFS already aggregate
# rows 3:1002, and Underwriting G13 sums J3:J1002 directly. Strip it.
for col in range(1, 129):  # A-DX, full data width
    cell = rr.cell(row=1003, column=col)
    cell.value = None
    cell.fill = _clear_fill
    cell.font = _clear_font
    cell.border = _clear_border

# Column H (Lot Rent in the Parkwood layout) data cells carry
# font.color.theme=0 in the blank template, which resolves to WHITE in
# the default Office theme — so any Lot Rent values backend.py writes
# render as invisible white text on a white cell. Clear the font override
# on H3:H1002 so the cells inherit the default black text. Also clear
# the same override on I (POH Home Rents) and J (LTO PMT) for parity.
for r in range(3, 1003):
    rr.cell(row=r, column=8).font = _clear_font  # H — Lot Rent
    rr.cell(row=r, column=9).font = _clear_font  # I — POH Home Rents
    rr.cell(row=r, column=10).font = _clear_font # J — LTO PMT

# Row 1 header strip: cells F1:T1 (between the "RENT ROLL" gray label in
# A1:D1, the "MHC Only" / "Apartments" / "Other Income" / "Commercial Leases"
# section labels, and the U1:AA1 white strip) carry PatternFill(fill_type=None)
# with a leftover black fgColor. Renderers paint that as a wide black bar
# across the top of the sheet. Force them to explicit white. Skip the four
# section-label columns (E1, I1, N1, Q1) which carry intentional colored
# fills.
_label_cols_row1 = {5, 9, 14, 17}  # E, I, N, Q
for col in range(6, 21):  # F..T
    if col in _label_cols_row1:
        continue
    rr.cell(row=1, column=col).fill = _clear_fill

# ── Data validations: scrub stale "Type 1/Type 2/Type 3..." dropdowns ──
# The blank template carried six different DV ranges on column B (Unit ID)
# with stray "Type 1, Type 2, Type 3", "...Type 4", "...Type 6", "...Type 7"
# patterns — leftovers from when B held Unit Type before the column
# restructure. Michael flagged the C dropdown in the partner walkthrough
# ("It's a drop down. I don't know why."). Nuke them all and re-add only
# the correct two for the Parkwood column layout:
#   C3:C1002  → TOH / POH / LTO / Flourish (seller Lot Type string)
#   E3:E1002  → Occupied / Vacant
#   S3:S1002  → NNN / Gross (Commercial Lease Type)
# Range stretches to row 1002 to match the full rent-roll capacity (the
# blank template only validated through row 72, so most pasted rows had
# no dropdown). Excel list formulas wrap the comma-separated items in
# double quotes; openpyxl needs them escaped as "...".
from openpyxl.worksheet.datavalidation import DataValidation
rr.data_validations.dataValidation = []
_dv_lottype = DataValidation(type="list", formula1='"TOH,POH,LTO,Flourish"', allow_blank=True)
_dv_lottype.add("C3:C1002")
rr.add_data_validation(_dv_lottype)
_dv_status = DataValidation(type="list", formula1='"Occupied,Vacant"', allow_blank=True)
_dv_status.add("E3:E1002")
rr.add_data_validation(_dv_status)
_dv_lease = DataValidation(type="list", formula1='"NNN,Gross"', allow_blank=True)
_dv_lease.add("S3:S1002")
rr.add_data_validation(_dv_lease)

# ════════════════════════════════════════════════════════════════════════
# 4. UNIT MIX RENT GROWTH — 4 unit-type rows w/ per-type growth schedules
# ════════════════════════════════════════════════════════════════════════
# The blank template ships with stray content from a prior hand-built
# workbook: leftover notes at B2/B3, non-flat year-by-year rent growth
# rates at rows 5-8 (specific to a different deal), Whaleshead HoME
# Rents block at rows 44-49, operational annotations at F18:H19, and
# only 2 of 4 unit-type rows populated. Replace the whole block with
# the 4-type layout matching the rebuilt Unit Mix Summary.
#
# Per-type, per-year growth schedule replaces the flat 5%. Defaults
# (see DEFAULT_RENT_GROWTH at module top) are tuned for typical MHC
# deals: ~10-11% Y1-Y3 (market reset), tapering to 5% terminal.
# backend.py reads property_info.rentGrowthSchedule per-deal and
# rewrites the rate cells if provided.

umrg = wb["Unit Mix Rent Growth"]

# Clear stray notes at B2/B3 from the blank template
umrg["B2"] = None
umrg["B3"] = None

# Clear the entire body (rows 5-75) so no orphan content from the old
# 6-type template, the Whaleshead HoME Rents block at rows 44-49, or
# operational-plan annotations at F18:H19 survive.
# max_col=15 covers columns A:O — the projection grid extends to O for
# Year 10. The wider clear range (vs the earlier max_row=28) is what
# strips the prior-deal HoME Rents block.
for row in umrg.iter_rows(min_row=5, max_row=75, max_col=15):
    for cell in row:
        cell.value = None

# ── Header row 4: "Key Assumptions" + Year 1..Year 10 ──
umrg["A4"] = "Key Assumptions"
year_cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
for i, col in enumerate(year_cols, start=1):
    umrg[f"{col}4"] = f"Year {i}"
# M4 = 10-yr Avg Growth (label written per-row below)
umrg["M4"] = "10-yr Avg Growth"

# ── Rows 5-8: Per-type per-year growth rates ──
# Row 5 = TOH, Row 6 = POH, Row 7 = LTO, Row 8 = Flourish. The label
# cells point back at Unit Mix Summary so they auto-update. Per-year
# rates come from DEFAULT_RENT_GROWTH defined at module top.
umrg["A5"] = "='Unit Mix Summary'!B4"   # TOH
umrg["A6"] = "='Unit Mix Summary'!B5"   # POH
umrg["A7"] = "='Unit Mix Summary'!B6"   # LTO
umrg["A8"] = "='Unit Mix Summary'!B7"   # Flourish
for i, col in enumerate(year_cols):
    umrg[f"{col}5"] = DEFAULT_RENT_GROWTH["toh"][i]
    umrg[f"{col}6"] = DEFAULT_RENT_GROWTH["poh"][i]
    umrg[f"{col}7"] = DEFAULT_RENT_GROWTH["lto"][i]
    umrg[f"{col}8"] = DEFAULT_RENT_GROWTH["flourish"][i]
# Per-row 10-yr avg growth at column M (for visibility into the schedule).
umrg["M5"] = "=AVERAGE(B5:K5)"
umrg["M6"] = "=AVERAGE(B6:K6)"
umrg["M7"] = "=AVERAGE(B7:K7)"
umrg["M8"] = "=AVERAGE(B8:K8)"

# ── Row 9: column headers for the per-year rent grid ──
umrg["C9"] = "Unit Mix"
umrg["D9"] = "# of Units"
umrg["E9"] = "Avg Monthly Rent"
projection_cols = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
for i, col in enumerate(projection_cols, start=1):
    umrg[f"{col}9"] = f"Year {i}"

# ── Rows 10-13: 4-type projection grid ──
# Each row pulls its starting unit count, occupied home rent, and avg
# rent from Unit Mix Summary and walks Y1..Y10 by applying the per-type
# growth rate from the assumption rows 5-8.
# Layout:
#   10 = TOH  (UMS row 4, growth row 5)
#   11 = POH  (UMS row 5, growth row 6)
#   12 = LTO  (UMS row 6, growth row 7)
#   13 = Flourish (UMS row 7, growth row 8)
UMRG_ROWS = [
    (10, 4, 5),
    (11, 5, 6),
    (12, 6, 7),
    (13, 7, 8),
]
for r, ums_row, growth_row in UMRG_ROWS:
    umrg[f"B{r}"] = f"=IFERROR(D{r}/$D$14,0)"
    umrg[f"C{r}"] = f"='Unit Mix Summary'!B{ums_row}"
    umrg[f"D{r}"] = f"='Unit Mix Summary'!E{ums_row}"
    umrg[f"E{r}"] = (f"=IFERROR('Unit Mix Summary'!H{ums_row}/"
                    f"'Unit Mix Summary'!E{ums_row},0)")
    umrg[f"F{r}"] = f"=E{r}*(100%+B${growth_row})"
    umrg[f"G{r}"] = f"=F{r}*(100%+C${growth_row})"
    umrg[f"H{r}"] = f"=G{r}*(100%+D${growth_row})"
    umrg[f"I{r}"] = f"=H{r}*(100%+E${growth_row})"
    umrg[f"J{r}"] = f"=I{r}*(100%+F${growth_row})"
    umrg[f"K{r}"] = f"=J{r}*(100%+G${growth_row})"
    umrg[f"L{r}"] = f"=K{r}*(100%+H${growth_row})"
    umrg[f"M{r}"] = f"=L{r}*(100%+I${growth_row})"
    umrg[f"N{r}"] = f"=M{r}*(100%+J${growth_row})"
    umrg[f"O{r}"] = f"=N{r}*(100%+K${growth_row})"

# ── Row 14: weighted-average roll-up (now SUMs B10:B13 / D10:D13) ──
umrg["B14"] = "=SUM(B10:B13)"
umrg["C14"] = "Total Weighted Average"
umrg["D14"] = "=SUM(D10:D13)"
for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}14"] = f"=SUMPRODUCT($B$10:$B$13,{col}10:{col}13)"

# ── Row 15: vacancy schedule. C15 = stabilization step (form input
# vacant_stabilization_count). G15:O15 project flat at $F$15 so
# vacancy doesn't drift after stabilization.
umrg["D15"] = "Vacant Lots"
umrg["C15"] = DEFAULTS["vacant_stabilization_count"]
umrg["E15"] = "='Unit Mix Summary'!D4"
umrg["F15"] = "=E15-C15"
for col in ("G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}15"] = "=$F$15"

# ── Row 16: vacant homes (POH side). Same flat-projection rule. ──
umrg["D16"] = "Vacant Homes"
umrg["C16"] = 0
umrg["E16"] = "='Unit Mix Summary'!D5"
umrg["F16"] = "=E16-$C$16"
for col in ("G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}16"] = "=$F$16"

# ── Row 17: vacancy % of D14 (total units) ──
umrg["D17"] = "Vacancy"
for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}17"] = f"=SUM({col}15:{col}16)/$D$14"

# ── Row 18: $ change vs prior year (moved from row 13) ──
umrg["D18"] = "$ Change"
for prev, curr in (("E", "F"), ("F", "G"), ("G", "H"), ("H", "I"),
                   ("I", "J"), ("J", "K"), ("K", "L"),
                   ("L", "M"), ("M", "N"), ("N", "O")):
    umrg[f"{curr}18"] = f"={curr}14-{prev}14"

# ── Row 19: annual GPR per year (weighted-avg × total units × 12) ──
# Underwriting!G4 reads from a forward year (default Y3 = H19).
# Pro Forma row 8 (Y1-Y10 GPR) reads F19:O19 via section 15a, so this
# range must span all 10 projection years.
umrg["D19"] = "GPR"
for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}19"] = f"={col}14*$D$14*12"

# Number-format scrub. The blank template's UMRG grid carried a
# leftover '0.0%' format on most of its data cells (it was previously
# a per-year growth-rate-only grid). After we rewrote the cells to
# hold dollar rents and unit counts, those percent formats made
# values render as e.g. "12900.0%" (129 units) and "80002.0%" ($800
# rent). Reset each cell to a format that matches the value it holds.
_pct = '0.0%'
_cur = '"$"#,##0'
_int = '#,##0'
_gen = 'General'

# Row 4: "Year N" column headers → text
for col in year_cols:
    umrg[f"{col}4"].number_format = _gen
umrg["M4"].number_format = _gen
# Rows 5-8: per-type per-year growth rates → percent
for r in (5, 6, 7, 8):
    for col in year_cols:
        umrg[f"{col}{r}"].number_format = _pct
    umrg[f"M{r}"].number_format = _pct

# Row 9: column headers
for col in ["C", "D", "E"] + projection_cols:
    umrg[f"{col}9"].number_format = _gen

# Rows 10-13: per-type projection grid
for r in (10, 11, 12, 13):
    umrg[f"B{r}"].number_format = _pct          # share of total units
    umrg[f"C{r}"].number_format = _gen          # unit type name
    umrg[f"D{r}"].number_format = _int          # # of units
    for col in ["E"] + projection_cols:         # rent + Y1..Y10
        umrg[f"{col}{r}"].number_format = _cur

# Row 14: weighted-average roll-up
umrg["B14"].number_format = _pct                # weights sum
umrg["C14"].number_format = _gen
umrg["D14"].number_format = _int
for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}14"].number_format = _cur

# Rows 15-16: vacant lot/home counts → integer
umrg["C15"].number_format = _int
umrg["C16"].number_format = _int
for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}15"].number_format = _int
    umrg[f"{col}16"].number_format = _int
# Row 17: vacancy % → percent
for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}17"].number_format = _pct

# Row 18: $ change vs prior year → currency
for col in ("F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}18"].number_format = _cur

# Row 19: annual GPR → currency
for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}19"].number_format = _cur

# Wire stabilized GPR on Underwriting to Unit Mix Rent Growth H19 (Year
# 3 GPR), matching CorrectOutput's reference but updated for the row-
# shift (GPR moved from row 14 to row 19 after we inserted the 4 type
# rows + LTO/Flourish growth assumption rows).
ws["G4"] = "='Unit Mix Rent Growth'!H19"

# ════════════════════════════════════════════════════════════════════════
# 5. LOAN SCENARIO (acquisition) — populate non-zero rates
# ════════════════════════════════════════════════════════════════════════
ls = wb["Loan Scenario (acquisition)"]
ls["C14"] = 0.0405  # US Treasury / SOFR baseline (4.05%)
ls["C15"] = 0.0185  # Spread (185 bps)
ls["C24"] = 0.70    # Max LTV (was 75%)

# ════════════════════════════════════════════════════════════════════════
# 6. SOURCES AND USES — both S1 (cols B:E) and S2 (cols H:K) scenarios
# ════════════════════════════════════════════════════════════════════════
# CorrectOutput Sources and Uses(PW):
#   C8 GP equity = 300000 hardcoded (was =C15 → acq-fee co-invest)
#   C13 Purchase Price = 'GGC Underwriting'!R4 (contract price, not asking)
#   C14 Closing Costs = =C13*2.25% (was 1.5%)
#   C16 Broker Fee = =C15 (was =3%*C20 which is broken — C20 is empty)
#   B12 "Uses of Funds" (was "Uses of Funds of Funds")
#   B15 "Acquisition Fee" (was "Acquistion")
#   B16 "Sales Broker Fee" (was "Equity Broker Fee")
su = wb["Sources and Uses"]
su["C8"]  = DEFAULTS["gp_equity_default"]   # GP equity hardcoded default
su["I8"]  = DEFAULTS["gp_equity_default"]
# Purchase Price comes from R4 (contract price) when set, else P4 (asking).
# backend.py writes contract_price to R4 and asking_price to P9; the
# GGC Underwriting P4 cell falls back to P9 when contract not set.
su["C13"] = "='GGC Underwriting'!R4"
su["I13"] = "='GGC Underwriting'!R4"
# Closing costs default = 2.25% of purchase price (CorrectOutput
# Parkwood). backend.py reads closing_cost_pct form input and rewrites
# this cell with a deal-specific value if provided.
_cc_pct = DEFAULTS["closing_cost_pct"]
su["C14"] = f"=C13*{_cc_pct}"
su["I14"] = f"=I13*{_cc_pct}"
# Broker fee formula — point at C15 (Acquisition Fee), NOT the empty
# C20. Parkwood Sales Broker Fee = Acquisition Fee × 1 because both
# are 2% of price.
su["C16"] = "=C15"
su["I16"] = "=I15"

# Typo + label fixes (CorrectOutput exact strings)
su["B12"] = "Uses of Funds"
su["H12"] = "Uses of Funds"
su["B15"] = "Acquisition Fee (2%)"
su["H15"] = "Acquisition Fee (2%)"
su["B16"] = "Sales Broker Fee (2%)"
su["H16"] = "Sales Broker Fee (2%)"

# Capex Budget — link to the capex breakdown total (built below at C24).
# backend.py reads capex_line_items form input and rewrites rows 21-26.
su["C17"] = "=C24"
su["I17"] = "=I24"

# Replace every $P$7 reference in this sheet with $N$7. The blank uses
# P7 everywhere as the units denominator, which points at a pricing cell.
for row in su.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and "GGC Underwriting" in cell.value:
            cell.value = cell.value.replace("$P$7", "$N$7").replace("!P7", "!N7")

# ════════════════════════════════════════════════════════════════════════
# 7. GGC PRO FORMA — Y1 RE Tax step-up, Home Rent Exp, LTO row label
# ════════════════════════════════════════════════════════════════════════
# CorrectOutput GGC Pro Forma(PW):
#   C21 = "Lease to Own" (was "Long term RV Site" — Parkwood is MHC, no RV)
#   H28 = =D28*1.03 (Y1 RE Taxes step-up; CLAUDE.md disagrees but
#         CorrectOutput is gold standard)
#   B47 = 0.10 (Home Rent Exp ratio; was 0.15/0.30 elsewhere)
#   H47 = =$B$47*(H20+H21) (applies to BOTH Home Rent Income AND
#         Lease-to-Own, not just H20)
pf = wb["GGC Pro Forma"]
pf["C21"] = "Lease to Own"
pf["H28"] = f"=D28*{1.0 + DEFAULTS['y1_re_taxes_growth_pct']}"
pf["B47"] = DEFAULTS["home_rent_expense_ratio"]
pf["H47"] = "=$B$47*(H20+H21)"

# Apply the same H47 logic across Y1-Y10 (cols H:Q) so Home Rent Expense
# scales with both Home Rent Income AND Lease-to-Own every projection
# year, not just Y1.
for col in ("H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"):
    pf[f"{col}47"] = f"=$B$47*({col}20+{col}21)"

# Bifurcated Lot Rent NOI (row 56) and Home Rent NOI (row 57). Replace
# the legacy heuristic with the clean two-line bifurcation:
#   H56 = H53 - H57    (Lot Rent NOI = Total NOI - Home Rent NOI)
#   H57 = (H20+H21) - H47   (Home Rent NOI = HRI+LTO - Home Rent Exp)
# Applied across Y1-Y10.
for col in ("H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"):
    pf[f"{col}56"] = f"={col}53-{col}57"
    pf[f"{col}57"] = f"=({col}20+{col}21)-{col}47"
pf["C56"] = "Lot Rent NOI Only"
pf["C57"] = "Home Rent NOI"   # was "Home Inventory Value" in blank

# Per-pad column G should divide by total units (P7 in the new column
# layout, or N7 in the legacy layout) not the broken /N7 → None case.
# Use /100 as a stable per-pad assumption when total units cell is
# unresolved; backend.py writes the real unit count to GGC Underwriting!P7
# and can rewrite the G_n formulas with /P7 at runtime.
_pf_per_pad_rows = [20, 21, 22, 23, 24, 25, 28, 29, 30, 31, 32, 33, 34,
                    35, 36, 37, 38, 39, 41, 42, 43, 44, 45, 46, 47, 48,
                    49, 50]
for r in _pf_per_pad_rows:
    cell = pf[f"G{r}"]
    if isinstance(cell.value, str) and "$N$7" in cell.value:
        cell.value = cell.value.replace("/'GGC Underwriting'!$N$7", "/100")

# X20 Exit Cap — default to the form input value (0.06). backend.py
# reads exit_cap_rate per-deal and rewrites this cell.
pf["X20"] = DEFAULTS["exit_cap_rate"]

# $P$7 -> $N$7 everywhere on this sheet (legacy compatibility)
for row in pf.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and "GGC Underwriting" in cell.value:
            cell.value = cell.value.replace("$P$7", "$N$7").replace("!P7", "!N7")

# ════════════════════════════════════════════════════════════════════════
# 8. WATERFALL PROMOTE TIERS — 20/30/30 -> 25/40/40 (10-yr + 5-yr)
# ════════════════════════════════════════════════════════════════════════
for tab in ["Waterfall (10-yr-S1)", "Waterfall (5-yr-S1)"]:
    if tab in wb.sheetnames:
        w = wb[tab]
        w["F16"] = 0.25
        w["F17"] = 0.40
        w["F18"] = 0.40

# ════════════════════════════════════════════════════════════════════════
# 9. INVESTOR RETURN — repoint at canonical waterfall tab, default "No"
# ════════════════════════════════════════════════════════════════════════
ir = wb["Investor Return"]
ir["F6"] = "='Waterfall (10-yr-S1)'!D30"
ir["F7"] = "='Waterfall (10-yr-S1)'!D31"
ir["F8"] = "=AVERAGE('Waterfall (10-yr-S1)'!G31:O31)"
ir["N4"] = "No"  # Send LOI? — default to No, not Yes

# ════════════════════════════════════════════════════════════════════════
# 10. DELETE JUNK DUPLICATE WATERFALL TABS
# ════════════════════════════════════════════════════════════════════════
for junk in ["Waterfall ", "Waterfall 2", "Waterfall 3"]:
    if junk in wb.sheetnames:
        del wb[junk]

# ════════════════════════════════════════════════════════════════════════
# 11. ROUND 3 — gap-fills from the dependency-graph verification pass
# ════════════════════════════════════════════════════════════════════════

# 11a. Make Management Fee % conditional on unit count per methodology
# (5% under 200 sites, 4% at 200+). Was hardcoded 0.05.
ws["J33"] = "=IF(N7>=200,0.04,0.05)"

# 11b. Seed Purchase Price (P4) so Sources & Uses, Loan Scenario, and Pro
# Forma Y0 don't collapse to zero. CorrectOutput Parkwood layout uses
# R4 for contract price ($5,805,000) and P9 for seller's asking
# ($6,000,000); P4 falls back to R4 (contract) → P9 (asking) → 0.
# backend.py writes contract_price to R4 when the deal has a negotiated
# price below ask.
ws["P4"] = "=IFERROR(IF(ISNUMBER(R4),R4,IF(ISNUMBER(P9),P9,0)),0)"

# 11c. GGC Pro Forma (Conversion) tab still has 35 stale $P$7 refs in
# col G (per-unit denominators). Sweep all $P$7 -> $N$7 there too.
if "GGC Pro Forma (Conversion)" in wb.sheetnames:
    conv = wb["GGC Pro Forma (Conversion)"]
    for row in conv.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "GGC Underwriting" in cell.value:
                cell.value = cell.value.replace("$P$7", "$N$7").replace("!P7", "!N7")

# 11d. Investor Return rows 25-29 referenced 'Waterfall 2' which we
# deleted. Repoint at the canonical (10-yr-S1) tab so they show real
# numbers instead of #REF!.
for r, cell_path in [
    (25, ("F25", "='Waterfall (10-yr-S1)'!D30")),
    (26, ("F26", "='Waterfall (10-yr-S1)'!D31")),
    (27, ("F27", "=AVERAGE('Waterfall (10-yr-S1)'!G31:O31)")),
]:
    coord, formula = cell_path
    if ir[coord].value is not None:
        ir[coord] = formula

# 11e/f/g — Unit Mix Rent Growth row-22 / M20-O20 / A5-A10 fixes from
# round 3 are now obsolete. The canonical rebuild earlier in this script
# (section 4) handles all of that already in the new layout. Removed.

# ════════════════════════════════════════════════════════════════════════
# 12. ROUND 4 — REGRESSION FIXES from full cross-check vs CorrectOutput
# ════════════════════════════════════════════════════════════════════════

# 12a. The original subject block had labels at O2-O10. After we moved
# it to M-N those labels are orphans that now duplicate / misalign with
# the new pricing block (P4-P10). Clear them.
for coord in ("O2", "O3", "O4", "O5", "O6", "O7", "O8", "O9", "O10"):
    if isinstance(ws[coord].value, str) and ws[coord].value in (
        "Underwritten Date", "Property Information", "Property Name",
        "Property Address", "Property Type", "# of Units ",
        "Rent Roll Occupancy", "Acreage", "County",
    ):
        ws[coord] = None

# 12b. I13 (Home Rent Income stabilized) was ='Rent Roll Input'!F1005*95%
# in the original blank. Col F is now Tenant Name after the Rent Roll
# restructure, so multiplying a name by 0.95 produces #VALUE!.
# Mirror G13 (the ALT NOI column's home-rent SUM over J3:J1002 × 12 × 95%)
# so the Stabilized Total NOI in column I includes POH home-rent revenue.
# Without this, parks with POH units have their NOI understated and P6
# (ingoing cap rate = I47/P4) is systematically inflated.
ws["I13"] = "=G13"

# 12c. P4 (Purchase Price) was =IFERROR(P9,0) which resolves to 0 when
# P9 is empty. Backend.py writes askingPrice to P9 if provided, but if
# the user negotiates a price below ask, P4 stays at ask. Switch to a
# hardcoded value with a comment: backend.py should overwrite P4 with
# the negotiated price each run via property_info["purchasePrice"].
# Default formula falls back to P9 (asking) only when no override.
ws["P4"] = "=IFERROR(IF(ISNUMBER(P9),P9,0),0)"

# 12d. Investor Return F28/F29 referenced 'Waterfall 2' which we
# deleted in round 1. Repoint at the canonical waterfall.
for coord in ("F28", "F29"):
    v = ir[coord].value
    if isinstance(v, str) and "Waterfall 2" in v:
        ir[coord] = v.replace("'Waterfall 2'", "'Waterfall (10-yr-S1)'")

# 12e. Loan Scenario C27 (DSCR denominator) — CorrectOutput Parkwood
# sizes debt off J47 (Lot Rent Only NOI in the new I/J/K column layout)
# with label "Lot Rent only NOI". GGC's underwriting practice is to size
# debt against lot-rent NOI only (the durable land-business cash flow),
# not the total NOI which includes the home-rent stream that operators
# typically can't lever against. Repoint and re-label.
if "Loan Scenario (acquisition)" in wb.sheetnames:
    lscell_ws = wb["Loan Scenario (acquisition)"]
    lscell = lscell_ws["C27"]
    if isinstance(lscell.value, str) and ("H47" in lscell.value or "I47" in lscell.value):
        lscell.value = lscell.value.replace("H47", "J47").replace("I47", "J47")
    else:
        lscell.value = "='GGC Underwriting'!J47"
    lscell_ws["A27"] = "Lot Rent only NOI"
    # Typo fix per OUTLINE.md (already covered in section 15i LABEL_FIXES
    # but reassert here in case the earlier swap didn't land).
    lscell_ws["B17"] = "Mortgage Constant"
    lscell_ws["L7"] = "Principal"
    # P8 off-by-one — was SUM(H42:H53), should be SUM(H43:H54).
    lscell_ws["P8"] = "=SUM(H43:H54)"

# 12f. Sources & Uses capex breakdown — match CorrectOutput Parkwood
# row layout (rows 22-26 = 5 line items, row 27 = SUM, row 17 = C27).
# backend.py reads capex_line_items form input (JSON list of
# {label, amount}) and rewrites these rows per deal.
#   B22/C22 = Water/Septic/Utilities (placeholder per deal)
#   B23/C23 = Tree Trim / site grading
#   B24/C24 = Add new homes (D24 × E24 — count × $/home)
#   B25/C25 = Road Repair / infrastructure
#   B26/C26 = Working Capex
#   B27/C27 = SUM
#   C17 references C27 (set above in section 6).
su["B21"] = "Capex Budget"
su["B22"] = "Private w/s capex - misc"
su["C22"] = 0
su["B23"] = "Tree Trim"
su["C23"] = 0
su["B24"] = "Add new homes"
su["C24"] = "=D24*E24"
su["D24"] = 0
su["E24"] = 25000
su["B25"] = "Road Repair"
su["C25"] = 0
su["B26"] = "Working Capex"
su["C26"] = 0
su["B27"] = "Estimated Capex Budget"
su["C27"] = "=SUM(C22:C26)"
# C17 set above to =C24 — but capex total is now at C27. Repoint.
su["C17"] = "=C27"
# Same for S2 scenario (cols H:K, with capex breakdown at H22:I27).
su["H21"] = "Capex Budget"
su["H22"] = "Private w/s capex - misc"
su["I22"] = 0
su["H23"] = "Tree Trim"
su["I23"] = 0
su["H24"] = "Add new homes"
su["I24"] = "=J24*K24"
su["J24"] = 0
su["K24"] = 25000
su["H25"] = "Road Repair"
su["I25"] = 0
su["H26"] = "Working Capex"
su["I26"] = 0
su["H27"] = "Estimated Capex Budget"
su["I27"] = "=SUM(I22:I26)"
su["I17"] = "=I27"

# 12g — Unit Mix Rent Growth rebuild from round 4 is obsolete; the
# canonical rebuild in section 4 above already produces this layout
# (and extends Year 7 to column L which round 4 was missing).

# ════════════════════════════════════════════════════════════════════════
# 13. EXACT-MATCH ALIGNMENT WITH CORRECTOUTPUT.XLSX (Underwriting tab)
# ════════════════════════════════════════════════════════════════════════
# Earlier rounds added "improvements" (MAX-with-reassessment on I22,
# conditional mgmt fee on J33, County Tax Rate cell at P12) that diverge
# from the gold-standard CorrectOutput layout. User wants exact match —
# revert those and restore the original formulas + cell labels.
ws["A26"] = "Electricity"           # row label (was "Electrcitiy" typo)
ws["I22"] = "=J22*N7"               # was MAX(...) — back to simple per-unit
ws["I23"] = "=D23*1.05"             # Insurance: T12 × 1.05 (no flood mult)
ws["J33"] = 0.05                    # was =IF(N7>=200,...) — back to flat 5%
ws["I33"] = "=J33*I19"              # mgmt fee = % × EGI (stabilized)
ws["G33"] = "=J33*G19"              # mgmt fee for ALT NOI column
ws["H33"] = "=J33*H19"              # mgmt fee for lot-rent-only NOI
ws["I35"] = "=J35*N7"               # match correct (no $ anchor on N7)

# ── Column layout shift: I = Stabilized NOI, J = Lot-Rent-Only NOI,
# K = UW NOI, L = per-unit display ────────────────────────────────────
# Matches CorrectOutput Parkwood: I2='Stabilized NOI', J2='Lot Rent
# Only NOI', K2='UW NOI'. Per-row formulas mirror CorrectOutput's
# pattern: I_n = K_n (Stabilized mirrors UW unless analyst overrides),
# J_n = K_n (Lot-Rent-Only excludes home-rent rows below), K_n = the
# T12 SUMIFS or per-unit/per-rate formula. L_n holds the per-unit
# value (=K_n/$P$7 or =K_n/$N$7 in the legacy layout). UW row K2-K47
# is where the underwritten numbers actually live.
ws["I2"] = "Stabilized NOI"
ws["J2"] = "Lot Rent Only NOI"
ws["K2"] = "UW NOI"
ws["L2"] = "Per Unit"

# K-column income block. Mirrors CorrectOutput Parkwood K4:K17.
ws["K4"] = "='Unit Mix Summary'!C13"        # UW GPR = Annual GPR Lot Rent Only
ws["K5"] = "=-L5*K4"                         # Vacancy = -(1-occ) × GPR
ws["K6"] = 0                                  # Concessions
# K7 Bad Debt — applies bad_debt_uw_pct (default 2%) × K4 (UW GPR).
# Negative sign (matches CorrectOutput L7 = 0.02 with K7 = -L7*K4).
ws["K7"] = "=-L7*K4"
ws["L7"] = DEFAULTS["bad_debt_uw_pct"]
ws["K12"] = "=D12"                           # Utility Reimbursement = T12
# K13 Home Rent Income — sum rent roll col I (POH Home Rents) × 12 × 95%
ws["K13"] = "=SUM('Rent Roll Input'!I3:I1002)*12*95%"
# K14 LTO — sum rent roll col J (LTO PMT) × 12 × 95%. This is the
# stream that was silently zero on Parkwood before the lcPayment fix.
ws["K14"] = "=SUM('Rent Roll Input'!J3:J1002)*12*95%"
ws["K15"] = "=G15"                           # Storage Income (mirror T12)
ws["K16"] = "=G16"                           # Laundry Income (mirror T12)
ws["K17"] = "=D17"                           # Other Income (T12)
# I column (Stabilized) mirrors K column for the income rows that
# stabilize unchanged (cap on Vacancy and Bad Debt is set in stab col).
for r in (4, 12, 13, 14, 15, 16, 17):
    ws[f"I{r}"] = f"=K{r}"
# J column (Lot-Rent-Only) zeroes out home-rent and LTO streams.
ws["J4"]  = "=K4"
ws["J5"]  = "=-L5*J4"
ws["J6"]  = "=-L6*J5"
ws["J7"]  = "=-L7*J6"
ws["J12"] = "=K12*80%"   # 80% utility reimbursement attributable to lot side
ws["J13"] = "=K13"        # home rent on the BIFURCATED side (will sub out in N15)
ws["J14"] = "=K14"
ws["J15"] = "=K15"
ws["J16"] = "=K16"
ws["J17"] = "=K17"
# L column = per-unit display
for r in (4, 12, 13, 14, 15, 16, 17, 19, 22, 23, 25, 26, 27, 28, 29,
          30, 31, 32, 33, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 47):
    ws[f"L{r}"] = f"=K{r}/$P$7"

# K-column expense block defaults from form inputs (per-site flat
# overrides). backend.py reads per_site_overrides and rewrites L_n
# with deal-specific values.
ws["K22"] = "=P22"                           # RE Taxes from SEV*levy P22
ws["K23"] = "=L23*$P$7"                       # Insurance per-site
ws["L23"] = DEFAULTS["insurance_per_site_nonflood"]
ws["K25"] = "=D25*1.03"                       # Gas/Fuel grow-on-T12
ws["K26"] = "=D26*1.03"                       # Electricity grow-on-T12
ws["K27"] = "=D27*1.03"                       # Water/Sewer grow-on-T12
ws["K28"] = "=D28*1.03"                       # Trash grow-on-T12
ws["K29"] = "=SUM(K25:K28)"
ws["K30"] = "=L30*$P$7"                       # R&M per-site
ws["L30"] = 150
ws["K31"] = "=L31*$P$7"                       # Ground Maintenance per-site
ws["L31"] = DEFAULTS["ground_maintenance_per_site"]
ws["K32"] = "=C32*1.05"                       # Recreational Amenities
ws["K33"] = "=L33*K19"                        # Management Fee = % × EGI
ws["L33"] = 0.05
ws["K35"] = "=L35*$P$7"                       # Payroll per-site
ws["L35"] = DEFAULTS["payroll_per_site"]
ws["K36"] = "=C36*1.05"                       # Employee Allowance
ws["K37"] = "=L37*$P$7"                       # G&A per-site
ws["L37"] = DEFAULTS["ga_per_site"]
ws["K38"] = "=C38*1.05"                       # Model Units
ws["K39"] = "=L39*$P$7"                       # Professional Fees per-site
ws["L39"] = DEFAULTS["professional_fees_per_site"]
ws["K40"] = "=L40*$P$7"                       # Advertising per-site
ws["L40"] = DEFAULTS["advertising_per_site"]
# K41 Home Rent Expense = home_rent_expense_ratio × (K13 + K14) so the
# expense scales with both POH home rent AND LTO. Defaults to 10%.
_hre = DEFAULTS["home_rent_expense_ratio"]
ws["K41"] = f"={_hre}*(K13+K14)"
ws["K42"] = "=C42*1.05"                       # Other
ws["K43"] = "=L43*$P$7"                       # Cap-Ex Reserve per-unit
ws["L43"] = DEFAULTS["capex_per_unit"]

# UW totals (K19 EGI, K44 OpEx, K47 NOI). Mirror the existing I-column
# rollup formulas but key off K.
ws["K19"] = "=K9+K12+K13+K14+K15+K16+K17"
ws["K9"]  = "=SUM(K4:K7)"
ws["K44"] = "=K22+K23+K29+K30+K31+K32+K33+K35+K36+K37+K38+K39+K40+K41+K42+K43"
ws["K47"] = "=K19-K44"

# L (per-unit) for EGI / OpEx / NOI rows
ws["L19"] = "=K19/$P$7"
ws["L44"] = "=K44/$P$7"
ws["L47"] = "=K47/$P$7"
# % of GPR column at L for income lines (CorrectOutput uses L3='% of GPR')
ws["L3"]  = "% of GPR"
ws["L5"]  = "=100%-P8"                        # vacancy% mirrors occupancy
ws["L6"]  = "=K6/K4"

# ── Tax Analysis Section: SEV × levy method ───────────────────────────
# When backend writes sev_assessed_value to P27 and levy_rate via the
# parcel table, P20 = 80% × SEV, P21 = parcel-weighted levy %, and
# P22 = P20 × P21. K22 (UW RE Taxes) reads from P22 above.
# When the SEV/levy inputs are blank, P22 falls through to 0 and the
# analyst can override K22 directly with the per-site tax_per_site
# fallback (=J22*N7). I22 keeps the legacy per-site formula for the
# existing G/H/I "ALT NOI" column.
ws["P20"] = "=IFERROR(P4*75%*O34,0)"          # 75% of price × MV/AV ratio
ws["P21"] = "=IFERROR(Q33,0)"                  # parcel-weighted levy
ws["P22"] = "=P20*P21"

# ── Bifurcated valuation block cap rates (form-input driven) ──────────
# O14 = lot cap rate (default 5.0%); O15 = home cap rate (default 20%).
# backend.py reads lot_cap_rate / home_cap_rate per-deal and rewrites.
# These OVERRIDE the legacy 5.5%/12% set later in this section.

# Restore the O/P pricing block labels EXACTLY as CorrectOutput has them.
# O4/P4 = Purchase/Offer Price (P4 is a numeric input written by backend).
# O5/P5 = $/site (P5 = P4/N7).  O6/P6 = Underwritten Cap (NOI/PP).
# O7/P7 = Stabilized YOC.  O9/P9 = Asking Price.  O10/P10 = Asking $/site.
ws["O4"] = "Purchase/Offer Price"
ws["O5"] = "Purchase Price Per Site"
ws["O6"] = "Underwritten CAP rate"
ws["O7"] = "Stabilized YOC"
ws["O9"] = "Asking Price by Seller"
ws["O10"] = "Asking Price Per Site "
ws["O12"] = "Brookings Multifamily cap rate at 6%-7.5%"
ws["P12"] = None                    # remove the County Tax Rate cell

# Match cosmetics
ws["M7"] = "# of Units "            # trailing space — matches correct
ws["P5"] = "=P4/N7"                 # purchase price per site
ws["P6"] = "=I47/P4"                # underwritten cap rate = NOI / PP
ws["P7"] = "=G47/'Sources and Uses'!C18"  # stabilized YOC
ws["P10"] = "=P9/N7"                # asking $ per site

# G4 (Stabilized GPR) — points at Unit Mix Rent Growth!H19, the Year 3
# annual GPR row in the rebuilt 4-type / 10-yr-projection layout.
# Was H14 in the prior 2-row layout — after the UMRG rebuild in section
# 4, GPR moved from row 14 to row 19 to accommodate the 4 type rows
# + 4 growth-assumption rows + vacancy block + $-change row.
ws["G4"] = "='Unit Mix Rent Growth'!H19"

# I4 (Stabilized total units anchor) — Unit Mix Summary!C11 in correct's
# layout = total MH lot count.
ws["I4"] = "='Unit Mix Summary'!C11"

# J7 (Bad Debt assumption) — correct uses a hardcoded 3% rate, not a
# back-computed ratio.
ws["J7"] = 0.03

# G12 (Stabilized Utility Reimbursement) — correct uses =I12 (mirror of
# Total NOI col), not 75% of the water/sewer line.
ws["G12"] = "=I12"

# G13 (Stabilized Home Rent Income) — SUM the entire Home Rent column
# (J3:J1002) directly rather than reading from a fixed J151 "Totals" row.
# The Totals row produced an orphaned highlighted cell sitting 25+ rows
# below the actual data on every output. Range SUM produces the same
# number with no visual side effect.
ws["G13"] = "=SUM('Rent Roll Input'!J3:J1002)*12*95%"

# Lot-Rent-Only NOI column (H): correct ZEROES OUT home rent stream so
# the H47 cell yields lot-rent NOI only. H13 = 0; H14:H16 mirror I col.
ws["H13"] = 0

# I17 (Other Income stabilized) — average of 2024 + T12 (cols C, D), not
# 2022 + 2024 (cols B, C).
ws["I17"] = "=AVERAGE(C17,D17)"

# H25-H28 (Lot-Rent-Only utility stream) — correct does NOT scale by 80%,
# it mirrors I25-I28 directly.
for r in (25, 26, 27, 28):
    ws[f"H{r}"] = f"=I{r}"

# SUMIFS range sweep — the original template's SUMIFS only covered
# rows 3:21 (income) and 28:58 (expenses). Correct's Data Consolidation
# is wider: 3:36 income, 43:102 expense. Update every SUMIFS in the
# Underwriting tab to match.
_RANGE_SWAPS = [
    # Income block
    ("$D$3:$D$21", "$D$3:$D$36"),
    ("$E$3:$E$21", "$E$3:$E$36"),
    ("$F$3:$F$21", "$F$3:$F$36"),
    ("$G$3:$G$21", "$G$3:$G$36"),
    ("$H$3:$H$21", "$H$3:$H$36"),
    ("$A$3:$A$21", "$A$3:$A$36"),
    # Expense block
    ("$D$28:$D$58", "$D$43:$D$102"),
    ("$E$28:$E$58", "$E$43:$E$102"),
    ("$F$28:$F$58", "$F$43:$F$102"),
    ("$G$28:$G$58", "$G$43:$G$102"),
    ("$H$28:$H$58", "$H$43:$H$102"),
    ("$A$28:$A$58", "$A$43:$A$102"),
]
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and "SUMIFS" in cell.value:
            new_val = cell.value
            for old, new in _RANGE_SWAPS:
                new_val = new_val.replace(old, new)
            if new_val != cell.value:
                cell.value = new_val

# Row-14/15/16 SUMIFS criteria — the blank template hunts for old labels
# ("LTO", "SFH", "Laundry Income"). Mirror CorrectOutput's per-column
# criteria exactly, including the quirky ones (B16 searches "Laundry
# Income" but C16:F16 search "Retail" — a hand-edit artifact in the
# gold standard that we replicate so values match cell-for-cell).
def _swap_criterion(cell, new_criterion):
    if isinstance(cell.value, str) and "SUMIFS" in cell.value:
        import re as _re
        cell.value = _re.sub(r'"[^"]+"\)$', f'"{new_criterion}")', cell.value)

# Row 14: all five history columns search "LTO" (the canonical category
# string backend.py emits for land-contract / lease-to-own payments).
# Whaleshead-style RV deals get "RV Site Rental Income" via backend
# rewrite of A14 + the SUMIFS criteria when property_info.propertyType
# is RV; default here matches MHC (Parkwood) layout.
for col in ("B", "C", "D", "E", "F"):
    _swap_criterion(ws[f"{col}14"], "LTO")

# Row 15: all five search "Parking Income" (correct's chosen label for
# the 4108 Storage Unit Rent GL — even though the row is labeled
# "Storage Income" in A15 for human readability).
for col in ("B", "C", "D", "E", "F"):
    _swap_criterion(ws[f"{col}15"], "Parking Income")

# Row 16: B16 searches "Laundry Income" (the 2022 column had no retail);
# C16:F16 search "Retail". Match correct's exact criteria.
_swap_criterion(ws["B16"], "Laundry Income")
for col in ("C", "D", "E", "F"):
    _swap_criterion(ws[f"{col}16"], "Retail")

# Expense rollups 36/38/41/43 — the blank template ships SUMIFS criteria
# that diverge from the canonical strings backend.py emits. Without these
# rewrites, four expense lines (Employee Allowance, Model Units, Home
# Rent Expense, Cap-Ex Reserve) silently zero out because the criteria
# won't match. CAP-EX IN PARTICULAR is force-inserted by apply_ggc_overrides
# and would silently drop $11k+/year on every deal. Mirror the canonical
# strings from backend.py GGC_EXPENSE_CATEGORIES so the SUMIFS hit:
for col in ("B", "C", "D", "E", "F"):
    _swap_criterion(ws[f"{col}36"], "Employee Allowance")
    _swap_criterion(ws[f"{col}38"], "Model Units")
    _swap_criterion(ws[f"{col}41"], "Home Rent Expense (MH)")
    _swap_criterion(ws[f"{col}43"], "Cap-Ex Reserve")

# Row 2 column headers — restore CorrectOutput's exact strings.
ws["G2"] = "ALT NOI"
ws["H2"] = "Lot Rent only NOI"
ws["I2"] = None    # correct leaves blank
ws["J2"] = None    # correct leaves blank

# Per-unit J-column cells correct just leaves blank (J20, J45).
ws["J20"] = None
ws["J45"] = None
# And the $-anchor on N7 for J31/J32 — correct uses bare N7 (no $).
ws["J31"] = "=I31/N7"
ws["J32"] = "=I32/N7"

# Cosmetic: (100%+3%) → 1.03 to match correct exactly.
for r in (39, 40):
    cell = ws[f"I{r}"]
    if isinstance(cell.value, str) and "(100%+3%)" in cell.value:
        cell.value = cell.value.replace("(100%+3%)", "1.03")

# Lot-rent-only NOI column quirks from CorrectOutput:
#   H13 = 0          (home rent removed, already set above)
#   H14 = I14        (RV rent included in lot-rent NOI)
#   H15 = I15        (Storage included)
#   H16 = None       (Retail removed from lot-rent NOI — correct's choice)
ws["H14"] = "=I14"
ws["H16"] = None

# Bifurcated valuation block at M13:R16. This is the GGC methodology
# split — value the Lot Rent NOI at 5.5% cap, Home Rent NOI at 12% cap.
# The full table runs M:R (6 columns: Asset / NOI / CAP RATE / VALUE /
# VALUE/UNIT / UNIT) so the workbook shows both an aggregate $ value
# AND a per-site value derived from Unit Mix Summary's unit counts.
# Mirrors CorrectOutput.xlsx exactly so the visual layout matches.
ws["M13"] = "Asset"
ws["N13"] = "NOI"
ws["O13"] = "CAP RATE"
ws["P13"] = "VALUE"
ws["Q13"] = "VALUE/UNIT"
ws["R13"] = "UNIT"

# Row 14: Lot Rent valuation. N14 = J47 (Lot-Rent-Only NOI in the new
# I/J/K column layout — the canonical lot-rent valuation basis).
# Caps at lot_cap_rate form input (default 5.0%). Per-site VALUE/UNIT
# divides P14 by the MH-site unit count (UMS E8 = total sites after
# the 4-row TOH/POH/LTO/Flourish rebuild — was E9 = TOH+POH only).
ws["M14"] = "Lot Rent only NOI"
ws["N14"] = "=J47"
ws["O14"] = DEFAULTS["lot_cap_rate"]
ws["P14"] = "=N14/O14"
ws["Q14"] = "=P14/R14"
ws["R14"] = "='Unit Mix Summary'!E8"

# Row 15: Home Rent valuation. N15 = K47-J47 = Home Rent NOI (in the
# new K/J column layout). Caps at home_cap_rate form input (default
# 20%). Unit count = LTO + Flourish (Type 3 + Type 4 rows in UMS).
ws["M15"] = "Home Rent only NOI"
ws["N15"] = "=K47-J47"
ws["O15"] = DEFAULTS["home_cap_rate"]
ws["P15"] = "=N15/O15"
ws["Q15"] = "=P15/R15"
ws["R15"] = "='Unit Mix Summary'!E6+'Unit Mix Summary'!E7"

# Row 16: blended total.
ws["M16"] = "Total"
ws["N16"] = "=SUM(N14:N15)"
ws["O16"] = "=N16/P16"
ws["P16"] = "=SUM(P14:P15)"

# Number formats matching CorrectOutput exactly. Without these the
# values display as raw numbers (0.055 instead of 5.50%, 17255392
# instead of $17,255,392). The display is half the work.
_currency = '"$"#,##0_);[Red]\\("$"#,##0\\)'
_currency_simple = '"$"#,##0'
_percent = '0.00%'
for _ref in ("N14", "P14", "Q14", "P15", "Q15", "P16"):
    ws[_ref].number_format = _currency
for _ref in ("N15", "N16"):
    ws[_ref].number_format = _currency_simple
for _ref in ("O14", "O15", "O16"):
    ws[_ref].number_format = _percent

# ════════════════════════════════════════════════════════════════════════
# 14. FORCE FULL RECALC ON OPEN
# ════════════════════════════════════════════════════════════════════════
# Force Excel to recalculate everything on open. The blank template may
# ship with calcMode="manual" or a stale full-calc flag, which would
# leave openpyxl-written formulas displayed as blank cells until the
# user manually presses F9. Setting both auto-mode AND fullCalcOnLoad
# covers the cases where Excel ignores one or the other.
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True
wb.calculation.calcCompleted = False
wb.calculation.calcOnSave = True

# ════════════════════════════════════════════════════════════════════════
# 15. BLOCK-SHIP FIXES — partner-grade cleanup (commit "make-it-count")
# ════════════════════════════════════════════════════════════════════════
# Five independent reviewers found ~40 items between them. This block
# addresses every CRITICAL / IMPORTANT one so the workbook reads as a
# finished GGC template, not a hand-built artifact mid-iteration.

# ── 15a. Pro Forma row 8 (Y1-Y10 GPR) chain break ─────────────────────
# H8:Q8 read from Unit Mix Rent Growth annual GPR row. After the 4-type
# rebuild in section 4, GPR moved from row 14 to row 19 (the row shift
# accommodates 4 unit-type rows + 4 growth-assumption rows + vacancy
# block + $-change row above it). Point Pro Forma at row 19.
pf = wb["GGC Pro Forma"]
_pf_year_cols = ["H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]
_umrg_year_cols = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]  # Y1..Y10
for pf_col, umrg_col in zip(_pf_year_cols, _umrg_year_cols):
    pf[f"{pf_col}8"] = f"='Unit Mix Rent Growth'!{umrg_col}19"

# ── 15b. Pro Forma rows 68-74: blank the 7-yr cash-flow placeholder ──
# H68:N68 had HARDCODED debt-service numbers from a prior deal
# (-655388.04, -773882.18, etc.); N69 had a hardcoded debt-payoff of
# -9833889.28. Blank them and the orphan formulas in the same block —
# the 10-yr scenario at rows 60-66 is the canonical one; the 7-yr
# block was a half-finished alternate that nobody uses.
for col in ("H", "I", "J", "K", "L", "M", "N"):
    pf[f"{col}68"] = None    # Debt Service (was hardcoded)
    pf[f"{col}69"] = None    # Debt Payoff
    pf[f"{col}70"] = None    # Refi Cashout
    pf[f"{col}71"] = None    # New Loan
    pf[f"{col}72"] = None    # Total Sale - Community
    pf[f"{col}73"] = None    # Free Cash Flow (was SUM of the hardcodes)
    pf[f"{col}74"] = None    # DSCR
# Also clear the section labels so the empty block doesn't look orphan
for r in range(68, 75):
    pf[f"E{r}"] = None
    pf[f"F{r}"] = None

# ── 15c. Electrcitiy typos in Pro Forma (we already fixed Underwriting) ──
pf["C32"] = "Electricity"
if "GGC Pro Forma (Conversion)" in wb.sheetnames:
    pf_conv = wb["GGC Pro Forma (Conversion)"]
    pf_conv["C32"] = "Electricity"

# ── 15d. Investor Return C33:D41 — stale prior-deal bullets ───────────
# These mentioned a different property ("Brookhaven"), specific $700/$775
# rents, "8 mix of SW/DW", and a typo "signiciant". Clear the whole block.
ir = wb["Investor Return"]
for r in range(33, 42):
    for col in ("B", "C", "D", "E", "F"):
        ir[f"{col}{r}"] = None

# ── 15e. Sources & Uses L93 broken formula (hanging minus operator) ──
su = wb["Sources and Uses"]
# Was: "=-'Loan Scenario (acquisition)'!J66-" (incomplete subtraction)
su["L93"] = None

# ── 15f. Unit Mix Rent Growth — clear stray deal-specific notes ──────
# These were comments from a different property's worksheet that bled
# into the blank template.
umrg = wb["Unit Mix Rent Growth"]
for coord in ("E29", "B39", "H45", "K117", "K128", "L128", "M128"):
    umrg[coord] = None

# ── 15g. Loan Scenario — clear stray text in LTO table ──────────────
ls = wb["Loan Scenario (acquisition)"]
ls["D89"] = None    # was "being evicted"
ls["J89"] = None    # was "Highlighted have baloon payment at the end of date"

# ── 15h. Unit Mix Summary — clear orphan rows BELOW the derived-metrics
# block. After the 4-type rebuild, derived metrics occupy rows 13-20
# (Annual GPR, Avg Lot Rent, Occupancy%, POH%, Annual HRI POH, Annual
# LTO HRI, Avg Home Rent LTO, Avg Home Rent POH). Clear rows 21+ to
# strip the "Annual LTO Premium" orphan label that bled in from a prior
# deal. Was clearing 17-23 which wiped my derived metrics.
ums = wb["Unit Mix Summary"]
for r in range(21, 24):
    for coord in (f"B{r}", f"C{r}", f"D{r}", f"E{r}", f"F{r}", f"G{r}", f"H{r}"):
        ums[coord] = None

# ── 15i. TYPO + LABEL CONSISTENCY SWEEP (from review agent #1) ───────
# Format: (sheet_name, cell_coord, new_value, reason_comment)
LABEL_FIXES = [
    # CRITICAL misspellings
    ("Data Consolidation",        "A2",  "GGC Category"),                  # "Catorgy"
    ("Data Consolidation",        "D27", "Input Source Expense Data"),     # "Epense"
    ("GGC Pro Forma",             "C16", "Economic Vacancy %"),            # "Ecomomic"
    ("GGC Pro Forma (Conversion)","C16", "Economic Vacancy %"),
    ("Loan Scenario (acquisition)","L7", "Principal"),                     # "Principle"
    ("Sources and Uses",          "B12", "Uses of Funds"),                 # "Uses of Funds of Funds"
    ("Sources and Uses",          "H12", "Uses of Funds"),
    ("Sources and Uses",          "B15", "Acquisition Fee (2%)"),          # "Acquistion"
    ("Sources and Uses",          "H15", "Acquisition Fee (2%)"),
    ("Loan Scenario (acquisition)","B17","Mortgage Constant"),             # "Costant"
    ("Loan Scenario (acquisition)","F79","10yrs"),                         # "10yyrs"
    # Label inconsistencies — standardize. UMS B12/B24 entries removed:
    # after the 4-type rebuild, B12 = "Total POH (Home Rents)" (the SUM
    # row for the home-rent block), B14 = "Avg Lot Rent" (the
    # SUMPRODUCT-weighted average). "Avg MH Lot Rent" no longer belongs
    # on B12 / B24 — those rows hold different metrics.
    ("GGC Pro Forma",             "C47", "Home Rent Expense (MH)"),
    ("GGC Pro Forma (Conversion)","C47", "Home Rent Expense (MH)"),
    # Sale Proceeds label standardization
    ("GGC Pro Forma",             "F72", "Total Sale - Community"),        # was "total Sale..."
    ("GGC Pro Forma (Conversion)","F62", "Total Sale - Community"),        # was "Sale Proceeds Net"
    ("GGC Pro Forma (Conversion)","F73", "Total Sale - Community"),
    ("GGC Pro Forma (Conversion)","F113","Total Sale - Community"),
    # DSCR uppercase
    ("GGC Pro Forma (Conversion)","F116","DSCR"),                          # was "dscr"
    # CoC uppercase
    ("GGC Pro Forma",             "F99", "CoC Return"),
    ("GGC Pro Forma (Conversion)","F68", "CoC Return"),
    ("GGC Pro Forma (Conversion)","F101","CoC Return"),
    ("GGC Pro Forma (Conversion)","F117","CoC Return"),
    ("GGC Pro Forma",             "F100","Avg CoC Y1-Y4"),
    ("GGC Pro Forma (Conversion)","F69", "Avg CoC Y1-Y4"),
    ("GGC Pro Forma (Conversion)","F78", "Avg CoC Y1-Y4"),
    ("GGC Pro Forma (Conversion)","F102","Avg CoC Y1-Y4"),
    # 10-yr Cash on Cash should be Y1-Y9 (was contradicting itself)
    ("Investor Return",           "C8",  "Cash on Cash Avg Y1-Y9"),
    # Bifurcated NOI label consistency — restored to CorrectOutput's
    # "...only NOI" wording so the workbook matches the gold standard
    # exactly. The earlier short form ("Lot Rent NOI") was a personal
    # preference that diverged from CorrectOutput; revert.
    ("GGC Underwriting",          "H2",  "Lot Rent only NOI"),
    ("GGC Underwriting",          "M14", "Lot Rent only NOI"),
    ("GGC Underwriting",          "M15", "Home Rent only NOI"),
    # Trailing/leading-space cleanups
    ("Unit Mix Rent Growth",      "D13", "$ Change"),                      # was "$change"
    ("GGC Underwriting",          "M7",  "# of Units "),                   # CorrectOutput keeps the trailing space
    ("GGC Underwriting",          "O10", "Asking Price Per Site "),        # CorrectOutput keeps the trailing space
    ("GGC Underwriting",          "A30", "Repair and Maintenance"),
    ("GGC Pro Forma",             "C36", "Repair and Maintenance"),
    ("GGC Pro Forma (Conversion)","C36", "Repair and Maintenance"),
    ("GGC Pro Forma",             "F65", "Free Cash Flow"),
    ("GGC Pro Forma",             "F84", "Free Cash Flow"),
    ("GGC Pro Forma",             "F97", "Free Cash Flow"),
    ("GGC Pro Forma (Conversion)","F66", "Free Cash Flow"),
    ("GGC Pro Forma (Conversion)","F74", "Free Cash Flow"),
    ("GGC Pro Forma (Conversion)","F85", "Free Cash Flow"),
    ("GGC Pro Forma (Conversion)","F99", "Free Cash Flow"),
    ("GGC Pro Forma (Conversion)","F115","Free Cash Flow"),
]
for sheet_name, coord, new_value in LABEL_FIXES:
    if sheet_name in wb.sheetnames:
        try:
            wb[sheet_name][coord] = new_value
        except Exception:
            pass

# ── 15j. Pro Forma B47 — add a label for the orphan 0.15 ratio cell ──
pf["A47"] = "Home Rent Exp Ratio"   # A47 was empty; B47 had value 0.15 only

# ── 15k. Orphan fill / border cleanup ─────────────────────────────────
# Helper agent found ~109K empty cells with fill applied — most visibly
# a pink strip on Rent Roll Input!B72:B1002 (would show up as a stray
# 930-row coloured block in the partner's view). Clear fill on every
# cell that has no value AND no formula, across these known clusters.
from openpyxl.styles import PatternFill as _PF
# Explicit solid white instead of fill_type=None — see comment on
# _clear_fill above. fill_type=None leaves fgColor='00000000' in
# openpyxl's serialized output and some renderers paint that as black.
_no_fill = _PF(fill_type="solid", start_color="FFFFFFFF", end_color="FFFFFFFF")
ORPHAN_FILL_CLEARS = [
    # CRITICAL visible colors past data
    ("Rent Roll Input",          "B3:B1002"),   # pink strip is the biggest offender
    ("Rent Roll Input",          "C3:C1002"),
    ("Rent Roll Input",          "H3:H1002"),
    ("Rent Roll Input",          "R3:R1002"),
    ("Rent Roll Input",          "D3:T1002"),   # broad white-fill ghost block
    ("Rent Roll Input",          "U3:AA1010"),  # columns past T retain template fill
    ("Rent Roll Input",          "A1003:DX1010"),# legacy totals-row dark fill at row 1003 A:D + far-right ghost (the "black box")
    ("Rent Roll Input",          "AB1:DX2"),    # row 1 header strip — empty cells past col AA retain black theme-1 fill
    ("Data Consolidation",       "A30:A58"),    # orange banding on empty rows
    ("Data Consolidation",       "G30:G58"),
    ("Unit Mix Summary",         "B17:H24"),
    ("GGC Underwriting",         "K4:L17"),     # green input-placeholder boxes
    ("GGC Underwriting",         "K35:L45"),
    ("GGC Underwriting",         "O23:R278"),
    ("GGC Underwriting",         "S24:AA29"),
    # IMPORTANT solid-white ghost columns past the data
    ("Loan Scenario (acquisition)", "AC6:IG64"),
    ("Loan Scenario (acquisition)", "K11:K1003"),
    ("Loan Scenario (acquisition)", "F127:J770"),
    ("Loan Scenario (acquisition)", "L11:W421"),
    ("Loan Scenario (acquisition)", "X2:AB421"),
    ("GGC Pro Forma",             "T2:AC181"),
    ("GGC Pro Forma (Conversion)","T2:AC183"),
    ("Waterfall (10-yr-S1)",      "J3:P173"),
    ("Waterfall (10-yr-S1)",      "C20:E270"),
    ("Waterfall (5-yr-S1)",       "J3:K173"),
    ("Waterfall (5-yr-S1)",       "C20:E270"),
    ("Unit Mix Rent Growth",      "M7:AY12"),
    # MINOR
    ("Collections",              "B1:DH1"),
    ("Sources and Uses",         "E3:G11"),
    ("Sources and Uses",         "O5:Q9"),
]
def _clear_orphan_fill(sheet, range_str):
    """Clear fill on any cell in the range whose value is None and which
    has no formula. Defensive: skip any cell that currently holds data."""
    if sheet.title not in wb.sheetnames:
        return
    # openpyxl 3 supports sheet[range_str] returning a tuple of tuples
    try:
        cells = sheet[range_str]
    except Exception:
        return
    if not isinstance(cells, tuple):
        cells = (cells,)
    for row in cells:
        row = row if isinstance(row, tuple) else (row,)
        for cell in row:
            if cell.value is None:
                cell.fill = _no_fill

for tab_name, rng in ORPHAN_FILL_CLEARS:
    if tab_name in wb.sheetnames:
        _clear_orphan_fill(wb[tab_name], rng)

# ════════════════════════════════════════════════════════════════════════
# 16. METHODOLOGY FIXES — partner-defensible per-unit / per-rate logic
# ════════════════════════════════════════════════════════════════════════
# All 10 methodology items the reviewers flagged. Each is a small,
# surgical change that ties the cell back to a written GGC playbook rule
# rather than a hardcoded magic number.

uw = wb["GGC Underwriting"]

# 16a. Management Fee — restore the methodology rule:
#      5% of EGI under 200 sites, 4% at 200+
# (We had reverted to flat 5% to match CorrectOutput's single deal
# but the underlying rule is the conditional; encode it.)
uw["J33"] = "=IF(N7>=200,0.04,0.05)"

# 16b/16c. RE Taxes and Insurance are intentionally simple per partner
# direction: the County Tax Rate (P12) and Flood Zone (P17) inputs and
# their MAX-with-methodology formulas were removed from the underwriting
# section. I22 / I23 keep the simple per-unit and T12 × 1.05 formulas
# already wired in section 13 above. Backend no longer writes P12/P17.
ws_uw = uw  # alias for compatibility with earlier code blocks

# Also strip the legacy "Tax Analysis Section" block (O19:Q22: Assessed
# Value / Levy Rate / Tax = P20*P21) and the orphan "Flood Zone" label at
# S6 from the GGC Underwriting tab. Partner direction is to keep the UW
# tab free of flood/tax inputs entirely; these were leftovers from an
# earlier template version that the per-unit / T12×1.05 rewrite never
# cleaned up. Nothing downstream reads from this block.
for coord in ("O19", "O20", "O21", "O22",
              "P19", "P20", "P21", "P22",
              "Q19", "Q20", "Q21", "Q22",
              "S6"):
    cell = uw[coord]
    cell.value = None
    cell.fill = _clear_fill
    cell.font = _clear_font
    cell.border = _clear_border

# 16d. Stabilized Vacancy — see G5 = -5%*G4 in section 1; K5 is now the
# UW Vacancy formula (set in the column-layout-shift block above).
# Documentation for the 5% benchmark lives on a comment cell elsewhere
# rather than colliding with the K5 numeric formula.

# 16e. Bad Debt — K7 is now the UW Bad Debt formula (=-L7*K4 with
# L7 = bad_debt_uw_pct, default 2%). Documentation comment removed
# to avoid colliding with the numeric cell.

# 16f. Bifurcated cap rates — DEFER to DEFAULTS["lot_cap_rate"] /
# DEFAULTS["home_cap_rate"] set at module top (defaults 5.0% / 20%
# per CorrectOutput Parkwood). The 5.5% / 12% Whaleshead-era values
# were over-aggressive on the lot side AND too generous on the home
# side relative to CorrectOutput Parkwood's 5.02% / 20% analyst values.
# backend.py reads lot_cap_rate / home_cap_rate form inputs per-deal
# and rewrites O14 / O15 with the deal-specific values.
uw["Q14"] = "=P14/R14"
uw["Q15"] = "=P15/R15"
# R14 / R15 use the new UMS layout (E8 = total sites; E6+E7 = LTO + Flourish)
# matching the bifurcated valuation block updates above. Reasserting here
# in case the section-13 write didn't land for some reason.
uw["R14"] = "='Unit Mix Summary'!E8"
uw["R15"] = "='Unit Mix Summary'!E6+'Unit Mix Summary'!E7"

# 16g. Loan rates — datestamp them so they don't go stale silently.
# Loan Scenario C14=4.05%, C15=185bps already in section 5. Add a note.
ls = wb["Loan Scenario (acquisition)"]
ls["D14"] = "as of 2026-06"
ls["D15"] = "GGC standard spread"

# 16h. Pro Forma B47 (Home Rent Expense ratio) — pinned to
# DEFAULTS["home_rent_expense_ratio"] (default 0.10 per CorrectOutput
# Parkwood). The earlier 0.30 was the methodology-range midpoint (25-50%)
# but CorrectOutput is the gold standard and uses 10%. backend.py
# reads home_rent_expense_ratio form input and rewrites per-deal.
pf["B47"] = DEFAULTS["home_rent_expense_ratio"]
pf["C47"] = "Home Rent Expense (MH)"

# 16i. Sources & Uses — surface the capex breakdown so reviewers see
# they need to populate it (not just leave at $200k working cap).
# C21 utilities / C22 home in-fills are still 0 but labels make it
# obvious they're INPUTS the user should fill per deal.
su["B21"] = "Water / Septic / Utilities (per deal)"
su["B22"] = "Add Homes / In-fill (per deal)"
su["B23"] = "Working Capital"

# 16j. Document the brokerProforma fallback explicitly in the prompt
# already (commit f7fa5da). Surface it on the workbook via a header.
uw["F2"] = "Broker's Proforma"  # was "Broker's Proforma  Y1" with double space

# ════════════════════════════════════════════════════════════════════════
# 17. COMPREHENSIVE STALE-CONTENT SWEEP (from helper agent #3)
# ════════════════════════════════════════════════════════════════════════
# Reviewer found 100+ cells with prior-deal artefacts: a 17-row LTO
# contract table at Unit Mix Rent Growth rows 78-94 (specific lot
# numbers, balances, maturity dates), expansion plans at rows 117-132,
# hardcoded exit caps in Pro Forma (Conversion), hardcoded debt service
# in Pro Forma rows 68-69 (already cleared in 15b), and prior-year
# headers in Data Consolidation.

# 17a. Unit Mix Rent Growth — clear the entire LTO contract table and
# expansion plan that bled in from a prior deal.
umrg = wb["Unit Mix Rent Growth"]
# Stray notes outside the table
for coord in ("C47", "A48", "C50", "C66", "C67", "C132",
              "B123", "C124", "D125", "D130", "K117", "K119"):
    umrg[coord] = None
# LTO contract table: rows 78-94, cols C through J
for r in range(78, 95):
    for col in ("C", "D", "E", "F", "G", "H", "I", "J"):
        umrg[f"{col}{r}"] = None
# Expansion-sites row 121 (the "69 existing sites" hardcode across cols)
for col in ("F", "G", "H", "I", "J", "K", "L", "M", "N", "O"):
    umrg[f"{col}121"] = None
# Trailing stray notes at rows 128-130
for coord in ("K128", "L128", "M128"):
    umrg[coord] = None

# 17b. Pro Forma (Conversion) — clear hardcoded exit-cap-rate cells.
# These were 6% from a prior deal; the workbook should expose a single
# exit-cap input rather than four scattered hardcodes.
if "GGC Pro Forma (Conversion)" in wb.sheetnames:
    pf_conv = wb["GGC Pro Forma (Conversion)"]
    for coord in ("S64", "N83", "S97", "S113"):
        pf_conv[coord] = None

# 17c. Data Consolidation — replace prior-deal monthly date headers at
# J2:U2 (e.g., 2024-06-01 through 2025-12-01) with generic month names.
# The Collections tab reads these via formula (B5='Data Consolidation'!J2
# down through B16='...!U2'), so when J2:U2 are blank the Collections tab
# shows empty month labels — Michael flagged this in the walkthrough
# ("it would be nice if we can put, you gotta put the months here"). The
# previous code cleared these and left it to backend to refill per deal,
# but backend never wrote them, so every output shipped with blank
# Collections months. Pin Jan-Dec as a deterministic default; backend
# can still overwrite per row with a specific anchored period if needed.
dc = wb["Data Consolidation"]
_MONTHS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")
for col, name in zip(("J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U"),
                     _MONTHS):
    dc[f"{col}2"] = name

# 17d. Rent Roll Input — clear deal-specific title with date
# ("RENT ROLL April 2026" → just "RENT ROLL").
rr = wb["Rent Roll Input"]
rr["A1"] = "RENT ROLL"

# 17e. Working Capital placeholder in Sources & Uses — make zero by
# default so the user sees they need to populate per deal.
su["C23"] = 0
su["I23"] = 0

# ════════════════════════════════════════════════════════════════════════
# 18. PARTNER-WALKTHROUGH FIXES (from Verifier 3 simulation)
# ════════════════════════════════════════════════════════════════════════
# Three bugs that no other reviewer caught — found by simulating the
# actual partner walking through every tab. Each is a math error or
# wrong-cell reference that a sharp partner would catch in 30 seconds.

# ── 18a. Rent Roll Input M column — double-counts Lot Rent ─────────────
# M2 header was "Combined" (duplicating K2), and M3:M1002 contained
# `=K3+L3+I3` where K3 already equals I3+J3. So M effectively computed
# (Lot+Home) + L + Lot = double-counting Lot Rent across 1,000 rows.
# Column L is empty. Just clear the entire M column — it's redundant
# with K, which already gives Combined = Lot + Home correctly.
rr = wb["Rent Roll Input"]
rr["M2"] = None
for r in range(3, 1003):
    rr[f"M{r}"] = None

# ── 18b. Investor Return — fix wrong row references, parameterize CoC ──
# Row 8 was =AVERAGE('Waterfall (10-yr-S1)'!G31:O31) — but row 31 is
# the LP EQUITY MULTIPLE row, not cash-on-cash. Same bug in rows 27/29
# of the duplicate block (which we also clean up below).
# CoC averaging range is now parameterized by hold_period_years form
# input (default 10). For 10-yr hold the range is G30:O30 (Y1-Y9);
# for 7-yr hold it becomes G30:L30 (Y1-Y6); etc. backend.py reads
# hold_period_years and rewrites these formulas per-deal.
ir = wb["Investor Return"]
# True CoC = average annual net LP cash flow / total LP equity invested.
# Waterfall row 30 (G30:...30) holds per-year net LP cash flow.
# Waterfall F28 (sign-flipped) holds total LP equity contributed.
# Use ABS so the divide stays positive regardless of contribution sign.
_hold = DEFAULTS["hold_period_years"]
# Column letter for the last CoC year (Y_n-1 since G=Y1).
_coc_end_col = chr(ord("G") + _hold - 2)
ir["F8"] = (f"=IFERROR(AVERAGE('Waterfall (10-yr-S1)'!G30:{_coc_end_col}30)"
            f"/ABS('Waterfall (10-yr-S1)'!F28),0)")
ir["F21"] = ("=IFERROR(AVERAGE('Waterfall (5-yr-S1)'!G30:J30)"
              "/ABS('Waterfall (5-yr-S1)'!F28),0)")

# Send LOI widget strip — when compact_layout is set the analyst-grade
# variant hides M4/N4/M5; OUTLINE.md notes CorrectOutput Parkwood
# stripped it. Default to keep the widget but allow backend to clear.
# (Compact toggle wired via backend per-deal.) Default ir["N4"] already
# set to "No" in section 9; M4/M5 are descriptive labels we keep.

# Rows 25-32 were a DUPLICATE 10-year summary block (note the double-
# space in "10  Year Return Summary" at C25 vs single space at C4).
# Row 27 in particular had F27 = AVERAGE(G31:O31) which mislabels an
# equity-multiple average as an IRR. Clear the duplicate block entirely.
for r in range(25, 33):
    for col in ("B", "C", "D", "E", "F", "G"):
        ir[f"{col}{r}"] = None

# ── 18c. Pro Forma (Conversion) — 5-yr IRR/MOIC pulled 10-yr ranges ──
# G79 IRR and G87 MOIC referenced row 66 (the 10-yr Free Cash Flow
# row) instead of row 85 (the 5-yr Free Cash Flow row). They also used
# the full G:Q range (11 cells = Y0-Y10) when the 5-yr scenario only
# spans G:L (6 cells = Y0-Y5).
if "GGC Pro Forma (Conversion)" in wb.sheetnames:
    pfc = wb["GGC Pro Forma (Conversion)"]
    pfc["G79"] = "=IRR(G85:L85)"        # was =IRR(G66:Q66) — 10-yr range
    pfc["G87"] = "=SUM(H85:L85)/-G85"   # was =SUM(H66:Q66)/-G66

# ── 18d. Pro Forma J67 DSCR typo (caught as bonus by Verifier 3) ─────
# Was =-J53/I60 (uses Year 2 debt service as Year 3 divisor). Fix to
# the same-year reference.
pf["J67"] = "=-J53/J60"

# ── 18e. Loan Scenario P8 off-by-one in interest year sum ────────────
# Was P8 = SUM(H42:H53), which shifts the interest window by one
# month into year 5. Should be SUM(H43:H54). Same shift on T8/U8.
lsq = wb["Loan Scenario (acquisition)"]
lsq["P8"] = "=SUM(H43:H54)"
lsq["T8"] = "=SUM(H91:H102)"
lsq["U8"] = "=SUM(H103:H114)"

# ════════════════════════════════════════════════════════════════════════
# 19. RIGHT-SIDE PROPERTY INFO BLOCK — match CorrectOutput exactly
# ════════════════════════════════════════════════════════════════════════
# DemoOutput5 shipped with three problems in the M:T area:
#   1. Column widths were wrong: M ballooned to 90 (labels overflowed),
#      while N was 9.8 (values truncated). Net effect was the property
#      info table didn't sit inside a coherent grid.
#   2. A duplicate pricing block was scattered at Q4:R10 and a second
#      valuation table at Q12:T14 — both broken (#DIV/0 from empty
#      inputs) and visually disconnected from the main M-P block.
#   3. CorrectOutput's Q3:R8 utility block (WEBSITE, Year Built, Flood
#      Zone, Utility Structure, Electricity, Trash) and the entire Tax
#      Analysis Section at M19:R34 were missing.
# This section rebuilds the right-side layout to mirror CorrectOutput:
#   - M2:R2     Underwritten Date row
#   - M3:R3     Property Information section header + WEBSITE row
#   - M4:R10    3-column subject block: M-N (property attrs), O-P
#               (pricing), Q-R (utilities / year / flood)
#   - M13:P16   Bifurcated valuation table (kept from section 13)
#   - M19:R22   Tax Analysis Section: header URL, Assessed Value,
#               Levy Rate, Estimated Tax
#   - M24       "2024-2025" sub-header
#   - M25:R33   7-row parcel-level tax table + SUM row
# Backend.py populates Underwritten Date, property values, website URL,
# year built, flood zone, utility structure, electricity, trash, and
# parcel-level tax data when extracted; cells stay blank otherwise.

from openpyxl.styles import Alignment as _Align

uw = wb["GGC Underwriting"]

# ── 19a. Column widths to match CorrectOutput ─────────────────────────
# Wrong widths were the root cause of "items hanging out of the table"
# in DemoOutput5: M was 90 (too wide), N was 9.8 (too narrow), so
# labels overflowed and values truncated. Restore CorrectOutput widths.
_widths = {"L": 10, "M": 34, "N": 46, "O": 25, "P": 18,
           "Q": 18, "R": 36, "S": 10, "T": 14}
for col, w in _widths.items():
    uw.column_dimensions[col].width = w

# ── 19b. Strip stray cells from the broken duplicate blocks ──────────
# DemoOutput5 had a second pricing column at Q4:R10 (Purchase/Offer
# Price duplicated, with R5/R6/R10 formulas referencing nonexistent R4)
# and a second valuation table at Q12:T14 producing #DIV/0. Both were
# left over from an abandoned "two-scenario" layout. Clear the genuinely
# stray cells; KEEP Q13/R13 (VALUE/UNIT and UNIT headers from section 13)
# and R14/R15 (Unit Mix Summary cross-references for the per-unit
# valuation column) — those are CorrectOutput's gold-standard layout
# and were getting silently wiped here, leaving the right side of the
# Asset/NOI table blank.
_clear_cells = ["P2",                              # orphan date (moves to N2)
                "S4", "S5",                        # orphan "Homes/Utilities"
                "Q9", "R9", "Q10", "R10",          # duplicate Asking Price
                "Q12", "R12", "S12", "T12",        # duplicate valuation hdr
                "S13", "T13",                      # duplicate valuation row (KEEP Q13/R13)
                "S14", "T14",                      # duplicate valuation row (KEEP R14)
                "S15", "T15",                      # duplicate valuation row (KEEP R15)
                "S16", "T16"]                      # duplicate valuation row
for coord in _clear_cells:
    cell = uw[coord]
    cell.value = None
    cell.fill = _clear_fill
    cell.font = _clear_font
    cell.border = _clear_border
# Q4-Q8, R4-R8 get rewritten below (utility block); Q14:Q15 hold the
# per-site VALUE/UNIT formulas (=P14/R14 and =P15/R15) from section 16f.

# ── 19c. Style helpers (theme-0 = default white background) ──────────
_navy_fill = PatternFill("solid", fgColor="FF002060")
_white_bold = Font(bold=True, color="FFFFFFFF", name="Calibri")
_label_bold = Font(bold=True, name="Calibri")
_thin = Side(border_style="thin", color="FF000000")
_box = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_align_left = _Align(horizontal="left", vertical="center")
_align_right = _Align(horizontal="right", vertical="center")
_align_center = _Align(horizontal="center", vertical="center")

def _label(coord, text):
    c = uw[coord]
    c.value = text
    c.font = _label_bold
    c.border = _box
    c.alignment = _align_left

def _value(coord, value, num_format="General", bold=False, align="left"):
    c = uw[coord]
    c.value = value
    c.font = _label_bold if bold else Font(name="Calibri")
    c.border = _box
    c.number_format = num_format
    c.alignment = _align_left if align == "left" else (_align_right if align == "right" else _align_center)

# ── 19d. Underwritten Date row (M2:N2) ────────────────────────────────
uw["M2"] = "Underwritten Date"
uw["M2"].font = _label_bold
uw["M2"].alignment = _align_left
# N2 left blank for backend to stamp with today's date; format is set
# so when backend writes a datetime it renders as "Monday, Feb 9, 2026".
uw["N2"] = None
uw["N2"].number_format = "[$-F800]dddd\\,\\ mmmm\\ dd\\,\\ yyyy"

# ── 19e. Section header M3:R3 ─────────────────────────────────────────
# Navy band across M3:R3 with "Property Information" on the left and
# the WEBSITE label/value on the right (Q3:R3). Matches CorrectOutput's
# single-row header where the website URL sits inline with the section
# title rather than as a separate row below.
for coord in ("M3", "N3", "O3", "P3"):
    uw[coord].fill = _navy_fill
    uw[coord].font = _white_bold
uw["M3"] = "Property Information"
uw["M3"].alignment = _align_left
uw["Q3"] = "WEBSITE"
uw["Q3"].font = _label_bold
uw["Q3"].alignment = _align_left
uw["R3"] = None   # backend writes the property listing URL here

# ── 19f. Three-column subject block O4:R10 ────────────────────────────
# CorrectOutput Parkwood layout (the OUTLINE-mandated shift from N to P):
#   O = property attribute label, P = property value (was M-N)
#   Q = pricing label,            R = pricing value/formula (was O-P)
# Examples from parkwoodCorrect:
#   O4="Property Name", P4="Parkwood Green Village"
#   O7="# of Units ",   P7=100
#   Q4="Purchase/Offer Price", R4=5805000 (CONTRACT price form input)
#   Q9="Asking Price by Seller", R9=6000000
# The legacy M-N labels are kept blank to preserve cell widths but
# emit the actual data on P / R.
_label("O4",  "Property Name")
_label("O5",  "Property Address")
_label("O6",  "Property Type")
_label("O7",  "# of Units ")        # trailing space matches CorrectOutput
_label("O8",  "Rent Roll Occupancy")
_label("O9",  "Acreage")
_label("O10", "County")

# P column property values (some formulas already wired; ensure borders).
for coord, formula, fmt in (
    ("P4", None,                          "General"),
    ("P5", None,                          "General"),
    ("P6", None,                          "General"),
    ("P7", "='Unit Mix Summary'!E8",      "0"),
    ("P8", "='Unit Mix Summary'!C15",     "0.00%"),
    ("P9", None,                          "General"),
    ("P10", None,                         "General"),
):
    c = uw[coord]
    if formula is not None:
        c.value = formula
    c.font = Font(name="Calibri")
    c.border = _box
    c.number_format = fmt
    c.alignment = _align_left

# Pricing column (Q-R). Contract price at R4, Asking at R9.
_label("Q4",  "Purchase/Offer Price")
_label("Q5",  "Purchase Price Per Site")
_label("Q6",  "Underwritten CAP rate")
_label("Q7",  "Stabilized YOC(Y5)")
_label("Q9",  "Asking Price by Seller")
_label("Q10", "Asking Price Per Site ")  # trailing space matches CorrectOutput

# R column pricing values. R4 = contract_price form input (backend writes
# numeric value at runtime; default formula falls back to R9 → 0).
for coord, formula, fmt in (
    ("R4", "=IFERROR(IF(ISNUMBER(R9),R9,0),0)",      '"$"#,##0'),  # contract
    ("R5", "=R4/P7",                                  '"$"#,##0'),  # $/site
    ("R6", "=K47/R4",                                 "0.00%"),     # UW cap
    ("R7", "='GGC Pro Forma'!L53/'Sources and Uses'!C18", "0.00%"),  # Stab YOC
    ("R9", None,                                       '"$"#,##0'),  # asking
    ("R10", "=R9/P7",                                 '"$"#,##0'),  # asking $/site
):
    c = uw[coord]
    if formula is not None and c.value in (None, 0):
        c.value = formula
    elif formula is not None and isinstance(c.value, str) and not c.value.startswith("="):
        c.value = formula
    c.font = Font(name="Calibri")
    c.border = _box
    c.number_format = fmt
    c.alignment = _align_right

# Legacy M-N labels: keep N7 (= total units) and N8 (occupancy) populated
# because many legacy formulas across other tabs still reference $N$7 as
# the unit divisor. The P-column block above is the new canonical layout
# per OUTLINE.md, but maintaining the N7 alias avoids breaking every
# Pro Forma / Loan Scenario / Sources & Uses formula keyed on N7. The
# label cells (M4-M10) are blanked to avoid visual duplication with the
# O-column labels below.
for coord in ("M4", "M5", "M6", "M7", "M8", "M9", "M10",
              "N4", "N5", "N6", "N9", "N10"):
    c = uw[coord]
    c.value = None
    c.fill = _clear_fill
    c.font = _clear_font
    c.border = _clear_border
# Reassert the units / occupancy formulas on N (legacy reference).
uw["N7"] = "='Unit Mix Summary'!E8"
uw["N8"] = "='Unit Mix Summary'!C15"

# ── 19g. Tax Analysis Section M19:R33 ─────────────────────────────────
# Three-row summary (Assessed Value / Levy Rate / Estimated Tax) plus
# a 7-row parcel-level table feeding the levy %. Section is purely
# informational — does NOT feed into I22 (RE Taxes), which stays on
# the per-unit J22*N7 formula per partner direction.
uw["M19"] = "Tax Analysis Section"
uw["M19"].font = _label_bold
uw["M19"].alignment = _align_left
# N19 holds the county tax-assessor URL (backend writes when known).
uw["N19"] = None
uw["N19"].font = Font(name="Calibri", color="FF0000FF", underline="single")
uw["N19"].alignment = _align_left

_label("M20", "Assessed Value ")     # trailing space matches CorrectOutput
_label("M21", "Levy Rate")
_label("M22", "Estimated Tax")

# Summary formulas. Default to the CorrectOutput approach: Assessed
# Value = 75% of Purchase Price scaled by the parcel-table's MV/AV
# ratio (O34); Levy Rate = parcel-weighted (Q33); Estimated Tax =
# product of the two. If the parcel table is empty the formulas
# resolve to 0 — that's the signal to the reviewer to populate it.
for coord, formula, fmt in (
    ("N20", "=IFERROR(P4*75%*O34,0)",   '"$"#,##0'),
    ("N21", "=IFERROR(Q33,0)",          "0.00%"),
    ("N22", "=N20*N21",                 '"$"#,##0'),
):
    c = uw[coord]
    c.value = formula
    c.font = _label_bold
    c.border = _box
    c.number_format = fmt
    c.alignment = _align_right

# Date / period sub-header
uw["M24"] = "2024-2025"
uw["M24"].font = _label_bold
uw["M24"].alignment = _align_left

# Parcel table headers (M25:R25). Six columns wide.
_PARCEL_HEADERS = [
    ("M25", "Parcel"),       ("N25", "MV"),       ("O25", "Taxable Value"),
    ("P25", "Taxes"),        ("Q25", "Levy%"),    ("R25", "Acres"),
]
for coord, label in _PARCEL_HEADERS:
    c = uw[coord]
    c.value = label
    c.font = _label_bold
    c.border = _box
    c.alignment = _align_center

# 7 empty parcel rows (M26:R32). Backend fills if parcel data is
# extracted; otherwise the reviewer types county-record data in here.
for r in range(26, 33):
    for col, fmt in (("M", "General"), ("N", '"$"#,##0'),
                      ("O", '"$"#,##0'), ("P", '"$"#,##0'),
                      ("Q", "General"),  ("R", "General")):
        c = uw[f"{col}{r}"]
        c.value = None
        c.font = Font(name="Calibri")
        c.border = _box
        c.number_format = fmt
        c.alignment = _align_right if col != "M" else _align_left

# Sum row at row 33: total MV, total Taxable Value, total Taxes, and
# implied levy rate (P33/O33). Yellow highlight on P33/Q33 mirrors
# CorrectOutput's call-out of the two key derived numbers.
for coord, formula, fmt, highlight in (
    ("N33", "=SUM(N26:N32)",         '"$"#,##0',  False),
    ("O33", "=SUM(O26:O32)",         '"$"#,##0',  False),
    ("P33", "=SUM(P26:P32)",         '"$"#,##0',  True),
    ("Q33", "=IFERROR(P33/O33,0)",   "0.00%",     True),
):
    c = uw[coord]
    c.value = formula
    c.font = _label_bold
    c.border = _box
    c.number_format = fmt
    c.alignment = _align_right
    if highlight:
        c.fill = PatternFill("solid", fgColor="FFFFFF00")

# Below the SUM row: O34 = MV/AV ratio used by N20's Assessed Value
# formula above. No border, no fill — sits as a quiet computed input.
uw["O34"] = "=IFERROR(O33/N33,0)"
uw["O34"].number_format = "0.00%"
uw["O34"].font = Font(name="Calibri", italic=True, color="FF6B7280")

# ════════════════════════════════════════════════════════════════════════
# 20. COLLECTIONS TAB — D6 criterion swap + Avg row
# ════════════════════════════════════════════════════════════════════════
# Two fixes on the Collections tab per OUTLINE.md:
#   1. D6 SUMIFS criterion "Vacant" never matches DC's category string
#      "Vacancy" (DC writes "Vacancy" because that's the GGC_INCOME_
#      CATEGORIES canonical string in backend.py). Swap "Vacant" ->
#      "Vacancy" on every Vacant-criterion SUMIFS in column D.
#   2. Add an Avg row at G17 / H17 / H18 / H19 (matches CorrectOutput
#      Parkwood). H17 = monthly RUBS avg, H18 = annualized (×12),
#      H19 = sum of the 11 observed months.
if "Collections" in wb.sheetnames:
    co = wb["Collections"]
    # Swap "Vacant" -> "Vacancy" on every column-D SUMIFS criterion.
    for r in range(5, 17):
        cell = co[f"D{r}"]
        if isinstance(cell.value, str) and '"Vacant"' in cell.value:
            cell.value = cell.value.replace('"Vacant"', '"Vacancy"')
    # Avg row.
    co["G17"] = "Avg"
    co["H17"] = "=AVERAGE(H6:H16)"
    co["H18"] = "=H17*12"
    co["H19"] = "=SUM(H6:H16)"

wb.save(TEMPLATE)
print(f"Patched {TEMPLATE.name}")
print(f"Sheets now: {wb.sheetnames}")
