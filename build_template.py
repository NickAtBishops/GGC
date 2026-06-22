"""
build_template.py — generate the blank GGC underwriting template from
the CorrectOutput.xlsx gold standard.

Why: the prior blank template (GGC_Blank_Underwriting_Sizer_Extended.xlsx)
was hand-patched over time via fix_template.py and drifted from CorrectOutput's
layout in ways that didn't reproduce the gold-standard look. Starting from
CorrectOutput as the literal base — and only stripping the Whaleshead-specific
input values — guarantees the workbook produced by the engine matches the
analyst-built reference structurally and visually.

How: this script
  1. Loads `Claude/CorrectOutput.xlsx` (the analyst's hand-built model)
  2. Drops engine-managed tabs (Comps, questions, miscellaneous — the
     engine creates fresh `Comps Analysis`, `Extraction Check`, `Miscellaneous`
     tabs at fill_template time)
  3. Strips deal-specific input cells while preserving every formula,
     label, header, structural cell, number format, font, fill, border,
     column width, row height, and freeze pane
  4. Restores P4 (Purchase/Offer Price) as a formula reading from P9
     so the engine's single write to P9 cascades correctly
  5. Applies a sweep of structural patches that were previously living in
     fix_template.py but were NOT making it into the runtime template:
       - label typo fixes ("Uses of Funds of Funds", "Acquistion Fee",
         "Mortgage Costant", "Principle")
       - Pro Forma S3 (rows 81-87, BRIDGE LOAN -6 MONTHS) + Scen 4
         (rows 91-97, Seller Carry) scenario blocks + IRR rows H89/H100
       - Collections Avg row at G17/H17/H18/H19
       - Loan Scenario P8/T8/U8 amortization sum ranges fixed off-by-one
  6. Saves as `GGC_Blank_Underwriting_Sizer_Extended.xlsx` (the deployed
     template that backend.py loads at fill_template time)
  7. Optionally re-applies fix_template.py on top so any later patch in
     fix_template lands too. This guarantees fix_template's defects are
     never silently dropped by a fresh build.

Run: `python3 build_template.py`. After it finishes the template is
guaranteed to be in the post-fix state — no separate `python3 fix_template.py`
step is needed for these patches to land.

INVOCATION CHAIN (the canonical, post-fix state):
    Claude/CorrectOutput.xlsx
        ↓ build_template.main()
        ↓ strips deal-specific cells, applies patches in step (5) above
        ↓ optionally invokes fix_template.py for label/typo/layout sweeps
    GGC_Blank_Underwriting_Sizer_Extended.xlsx   (the runtime template)
        ↓ backend.fill_template(...) per deal
    populated 16-tab workbook                    (download artifact)

KNOWN GAP (CLAUDE.md §5.2) — NOT folded in here:
The GGC Underwriting column shift from G/H/I → I/J/K (Stabilized /
Lot-Rent-Only / UW NOI columns) and the Property Information block move
from N → P-R are a multi-tab cascade that also touches every cross-tab
reference in Pro Forma F8:F53, Loan Scenario C27, Investor Return F6:F8,
Sources & Uses C13. Doing it from build_template.py alone risks decoupling
the layout from backend.py's writers (which still target N4/P9 etc.) and
producing a workbook that's structurally right but functionally empty.
The shift needs to land in lockstep across build_template.py + fix_template.py
+ backend.py writers — see analysis_dumps/parkwood_compare/OUTLINE.md §2
"GGC Underwriting" for the exact cell map and §3 cross-cutting items
#13/#14 for the full coordination.
"""
from __future__ import annotations
import re
from pathlib import Path
from openpyxl import load_workbook

SRC  = Path(__file__).parent / "Claude" / "CorrectOutput.xlsx"
DEST = Path(__file__).parent / "GGC_Blank_Underwriting_Sizer_Extended.xlsx"

# Rent Roll Input capacity. CorrectOutput was hand-tailored to Whaleshead's
# 148 units (scan range `Rent Roll Input!$X$3:$X$150`), but the deployed
# template needs to cover any deal. backend.py truncates at this same cap
# and emits a hard-fail extraction check when a real deal exceeds it.
RENT_ROLL_CAPACITY = 2000   # 2000 tenant rows → scan range 3:2002
LAST_RR_ROW = 2 + RENT_ROLL_CAPACITY


# ─────────────────────────────────────────────────────────────────────────────
# Per-sheet strip lists. Each entry is either a single coordinate ("N4") or a
# range ("M26:R32"). Only the VALUE on each cell is cleared — number format,
# font, fill, border, alignment, merged-cell state, etc. are preserved.
# ─────────────────────────────────────────────────────────────────────────────

# GGC Underwriting: the right-side property info block and tax-parcel section
# are deal-specific. Everything in the left-side P&L block (rows 3-50) is
# already formula-driven from Data Consolidation.
STRIP_UNDERWRITING = [
    # Property Information block (M3:R10)
    "N4",   # Property Name        — backend writes
    "N5",   # Property Address     — backend writes
    "N6",   # Property Type        — backend writes (when methodology provides)
    "N9",   # Acreage              — backend writes (when methodology provides)
    "N10",  # County               — backend writes
    "R3",   # Website URL          — backend writes
    "R4",   # Year Built           — backend writes
    "R5",   # Flood Zone           — backend writes
    "R6",   # Utility Structure    — backend writes
    "R7",   # Electricity notes    — backend writes
    "R8",   # Trash notes          — backend writes
    # Pricing block
    "P4",   # Purchase Price — REPLACED below with formula reading from P9
    "P9",   # Asking Price by Seller — backend writes
    "N2",   # Underwritten Date — backend writes
    # Tax Analysis Section (M19:R32)
    "N19",  # Tax assessor URL — backend writes
    # Parcel table M26:R32 — all parcel-specific data
    "M26", "N26", "O26", "P26", "Q26", "R26",
    "M27", "N27", "O27", "P27", "Q27", "R27",
    "M28", "N28", "O28", "P28", "Q28", "R28",
    "M29", "N29", "O29", "P29", "Q29", "R29",
    "M30", "N30", "O30", "P30", "Q30", "R30",
    "M31", "N31", "O31", "P31", "Q31", "R31",
    "M32", "N32", "O32", "P32", "Q32", "R32",
]

# Data Consolidation: all rows 3-36 (income leaves) and 43-102 (expense leaves)
# carry deal-specific values in columns A/B/D-G/J-U. Strip those; preserve
# the section-sum rows (38, 104), the input-source-check rows (39-40, 105),
# and the section markers (row 42 "Choose Expense Category" stays as a
# template separator — CorrectOutput keeps it).
DC_INCOME_LEAF_ROWS  = list(range(3, 37))   # rows 3-36
DC_EXPENSE_LEAF_ROWS = list(range(43, 103)) # rows 43-102
DC_STRIP_COLS = ["A", "B", "D", "E", "F", "G"]    # cat, name, fyPrior, fyCurrent, brokerProforma, t12
DC_STRIP_MONTHLY_COLS = list("JKLMNOPQRSTU")      # 12 monthly cols (J-U)

# Rent Roll Input: data rows 3 to LAST_RR_ROW (RENT_ROLL_CAPACITY rows
# after the 2 header rows). Strip columns B-J (unit id, type, status,
# name, type detail, type code, lot rent, home rent) — keep column A
# (Count formula = previous + 1) and column K (Combined formula = lot
# + home).
RR_DATA_ROWS = list(range(3, LAST_RR_ROW + 1))
RR_STRIP_COLS = list("BCDEFGHIJ")

# Comps: strip rows 3+ (comp data). Keep row 2 header. The engine adds a
# dedicated "Comps Analysis" tab at runtime, but keeping the gold-standard
# `Comps` tab with a clean structure lets analysts reference it as a hand-fill.
COMPS_STRIP_ROWS = list(range(3, 12))


def strip_cell(ws, coord):
    """Clear the value AND any attached hyperlink on a cell. Formatting
    (number format, font, fill, border, alignment, merged-cell state) is
    preserved — openpyxl's .value = None / .hyperlink = None mutations
    don't touch styling. CorrectOutput attaches hyperlinks to R3 (website)
    and N19 (county assessor URL) that survive a .value strip; the
    hyperlink object holds the URL even when the displayed text is blank,
    so it has to be explicitly nulled."""
    cell = ws[coord]
    cell.value = None
    cell.hyperlink = None


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"Source gold standard missing: {SRC}")
    print(f"[build] Loading {SRC.name}")
    wb = load_workbook(SRC)
    print(f"[build] Source tabs: {wb.sheetnames}")

    # ── 1. Remove engine-managed tabs ───────────────────────────────────
    # `questions` and `miscellaneous` are analyst scratch tabs — the engine
    # doesn't use them and they shouldn't ship in the blank template.
    # The hand-built `Comps` tab from CorrectOutput is kept (analysts may
    # populate it manually); the engine's runtime `Comps Analysis` tab is
    # a different sheet and gets added separately at fill_template time.
    for name in ("questions", "miscellaneous"):
        if name in wb.sheetnames:
            del wb[name]
            print(f"[build] Removed '{name}' tab")

    # ── 1b. Strip external workbook links ───────────────────────────────
    # CorrectOutput references two external workbooks (probably the
    # seller's files or earlier GGC templates) that don't exist on the
    # deploy target. Excel pops "We found a problem with some content"
    # on open when these dangle. Clear them.
    if getattr(wb, "_external_links", None):
        n = len(wb._external_links)
        wb._external_links = []
        print(f"[build] Removed {n} external workbook link(s)")

    # ── 1c. Drop defined names with invalid characters ──────────────────
    # CorrectOutput has a defined name "CU?" that probably originally
    # had a non-ASCII char. Excel can't resolve it and warns on open.
    if hasattr(wb.defined_names, "delete"):
        bad = [n for n in list(wb.defined_names) if "?" in n or not n.isascii()]
        for n in bad:
            wb.defined_names.delete(n)
        if bad:
            print(f"[build] Removed {len(bad)} invalid defined name(s): {bad}")

    # ── 2. GGC Underwriting — strip property + parcel inputs ─────────────
    uw = wb["GGC Underwriting"]
    for coord in STRIP_UNDERWRITING:
        strip_cell(uw, coord)
    # P4 needs to remain functional. Restore as formula reading from P9 so
    # backend.py's single write to P9 cascades through to P4 (Purchase
    # Price), I47/P4 (cap rate), and the rest of the pricing chain.
    uw["P4"] = "=IFERROR(IF(ISNUMBER(P9),P9,0),0)"
    # Preserve P4's number format from CorrectOutput (currency).
    uw["P4"].number_format = '"$"#,##0_);[Red]\\("$"#,##0\\)'
    print(f"[build] GGC Underwriting: stripped {len(STRIP_UNDERWRITING)} input cells, "
          f"restored P4 formula")

    # ── 3. Data Consolidation — strip per-line input data ────────────────
    # Each leaf row carries the seller's account info (col B), category
    # tag (col A), FY columns (D/E/F), T12 (G), and 12 monthly values
    # (J-U). The engine writes all of these per-run; strip CorrectOutput's
    # Whaleshead values so the blank template carries no signal.
    dc = wb["Data Consolidation"]
    n_cleared = 0
    for r in DC_INCOME_LEAF_ROWS + DC_EXPENSE_LEAF_ROWS:
        for col in DC_STRIP_COLS + DC_STRIP_MONTHLY_COLS:
            cell = dc[f"{col}{r}"]
            if cell.value is not None and not (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                # Skip cells that hold a formula — preserve structural math.
                cell.value = None
                n_cleared += 1
    print(f"[build] Data Consolidation: cleared {n_cleared} per-line input cells "
          f"(94 leaf rows × ~18 cols)")

    # ── 4. Rent Roll Input — strip tenant rows ───────────────────────────
    # Per-tenant data lives in rows 3-1002 (the template scans 3:1002 in
    # its COUNTIFS/SUMIFS). Strip the input columns; keep column A
    # (=A_prev+1 count formula) and column K (=I+J combined rent).
    rr = wb["Rent Roll Input"]
    n_cleared = 0
    for r in RR_DATA_ROWS:
        for col in RR_STRIP_COLS:
            cell = rr[f"{col}{r}"]
            if cell.value is not None and not (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                cell.value = None
                n_cleared += 1
    print(f"[build] Rent Roll Input: cleared {n_cleared} per-tenant input cells "
          f"({len(RR_DATA_ROWS)} rows × {len(RR_STRIP_COLS)} cols)")

    # ── 3b. Strip "Input Source Data" reference rows ─────────────────────
    # Data Consolidation rows 39 (income) and 105 (expense) are "Input
    # Source Data" check rows that the analyst hand-typed with Whaleshead
    # historical totals ($1.4M income, $494K expense) for comparison
    # against the SUM aggregates above them. These hardcoded Whaleshead
    # values must NOT ship in a generic blank — they'd compare every new
    # deal's income against Whaleshead's $1.4M and the check rows would
    # always look wrong. Clear the value cells; the row labels in column B
    # ("Input Source Data") stay for the analyst to optionally re-fill.
    # Includes column V (the SUM aggregation column that F39/F105 forward
    # from via =V39 / =V105) — without this V holds the Whaleshead total.
    for r in (39, 105):
        for col in DC_STRIP_COLS + DC_STRIP_MONTHLY_COLS + ["V"]:
            cell = dc[f"{col}{r}"]
            if cell.value is not None and not (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                cell.value = None
    print(f"[build] Data Consolidation: cleared 'Input Source Data' check rows 39 + 105")

    # ── 3c. Strip embedded images ────────────────────────────────────────
    # CorrectOutput has an Oregon Department of Revenue / FEMA flood map
    # image on Unit Mix Rent Growth and 3 deal-specific images on GGC
    # Underwriting (right-side Whaleshead photos, etc.). These are tied
    # to the Whaleshead deal and shouldn't ship in a generic blank.
    # The engine's Miscellaneous tab is where Street View / Static Maps
    # for the actual deal get embedded at fill_template time.
    n_images = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws._images:
            n_images += len(ws._images)
            ws._images = []
    print(f"[build] Stripped {n_images} embedded image(s) from the workbook")

    # ── 4b. Extend Rent Roll scan ranges + per-row formulas ──────────────
    # CorrectOutput's Unit Mix Summary COUNTIFS/SUMIFS scan only rows 3:150
    # of Rent Roll Input (Whaleshead-specific 148-unit fit). Bump every
    # such reference to row LAST_RR_ROW (2002 with the default cap) so the
    # template works on any deal up to RENT_ROLL_CAPACITY tenants.
    print(f"[build] Extending Rent Roll scan ranges to row {LAST_RR_ROW}...")
    _range_pat = re.compile(
        r"(Rent Roll Input'?!\$?[A-Z]+\$?)(\d+):(\$?[A-Z]+\$?)(\d+)"
    )
    n_extended = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or "Rent Roll Input" not in v:
                    continue
                def _bump(m):
                    return f"{m.group(1)}{m.group(2)}:{m.group(3)}{LAST_RR_ROW}"
                new = _range_pat.sub(_bump, v)
                if new != v:
                    cell.value = new
                    n_extended += 1
    print(f"[build] Rewrote {n_extended} formulas referencing Rent Roll Input ranges")

    # Extend the per-row formulas in column A (=A_prev+1 row count) and
    # column K (=I+J combined rent) so the new rows participate in the
    # count chain and combined-rent display. Without this, rows past 150
    # have no row count and combined-rent column stays blank.
    rr_ws = wb["Rent Roll Input"]
    n_seeded = 0
    for r in range(3, LAST_RR_ROW + 1):
        a_cell = rr_ws.cell(row=r, column=1)
        if a_cell.value is None or (isinstance(a_cell.value, int) and a_cell.value == r - 2):
            a_cell.value = f"=A{r-1}+1" if r > 3 else 1
            n_seeded += 1
        k_cell = rr_ws.cell(row=r, column=11)
        if k_cell.value is None:
            k_cell.value = f"=I{r}+J{r}"
    print(f"[build] Seeded {n_seeded} new Rent Roll Input row-formula pairs")

    # ── 4c. Rewrite Unit Mix Summary COUNTIFS to match engine output ─────
    # CorrectOutput's Unit Mix Summary hunts column C of Rent Roll Input
    # for the analyst's INTERNAL CODES "Type 1" / "Type 2" / "Type 3" /
    # "Type 4". But backend.py's _canonicalize_unit_type writes the
    # canonical strings "TOH MH Site" / "POH-Infilled units" /
    # "Long term RV Site" / "Retail/Commercial" directly to column C.
    # Result on Run3: every COUNTIFS returns 0, Unit Mix Summary stays
    # blank, every Underwriting cell that references unit counts (Total
    # Units N7, Occupancy N8, GPR via Unit Mix Rent Growth, etc.) breaks.
    # Rewrite the criterion strings so the SUMIFS hit the canonical names.
    TYPE_CODE_MAP = {
        '"Type 1"': '"TOH MH Site"',
        '"Type 2"': '"POH-Infilled units"',
        '"Type 3"': '"Long term RV Site"',
        '"Type 4"': '"Retail/Commercial"',
    }
    ums = wb["Unit Mix Summary"]
    n_rewritten = 0
    for row in ums.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or "Type " not in v:
                continue
            new = v
            for old, repl in TYPE_CODE_MAP.items():
                new = new.replace(old, repl)
            if new != v:
                cell.value = new
                n_rewritten += 1
    print(f"[build] Unit Mix Summary: rewrote {n_rewritten} COUNTIFS/SUMIFS "
          f"criteria from 'Type N' → canonical unit-type strings")

    # Also fix the typo in the display label B7 ("Retail/Comemrcial" →
    # "Retail/Commercial") so the visible row label matches the actual
    # COUNTIFS criterion the engine emits.
    if ums["B7"].value == "Retail/Comemrcial":
        ums["B7"].value = "Retail/Commercial"
        print(f"[build] Unit Mix Summary: fixed 'Retail/Comemrcial' typo at B7")

    # ── 5. Comps — strip comp data rows ──────────────────────────────────
    if "Comps" in wb.sheetnames:
        cp = wb["Comps"]
        n_cleared = 0
        for r in COMPS_STRIP_ROWS:
            for col_idx in range(1, cp.max_column + 1):
                cell = cp.cell(row=r, column=col_idx)
                if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    cell.value = None
                    n_cleared += 1
        print(f"[build] Comps: cleared {n_cleared} cells in data rows")

    # ── 5. Direct label / typo fixes ─────────────────────────────────────
    # These were historically applied by fix_template.py but did not always
    # reach the runtime template (fix_template ran against a different
    # snapshot, or build_template overwrote it). Apply them here so they
    # are part of every fresh build by construction.
    LABEL_FIXES = [
        # (sheet_name, coord, new_value)
        # CRITICAL: Sources and Uses "Uses of Funds of Funds" typo
        ("Sources and Uses",            "B12", "Uses of Funds"),
        ("Sources and Uses",            "H12", "Uses of Funds"),
        # "Acquistion Fee" → "Acquisition Fee"
        ("Sources and Uses",            "B15", "Acquisition Fee (2%)"),
        ("Sources and Uses",            "H15", "Acquisition Fee (2%)"),
        # Loan Scenario "Mortgage Costant" → "Mortgage Constant"
        ("Loan Scenario (acquisition)", "B17", "Mortgage Constant"),
        # Loan Scenario "Principle" → "Principal"
        ("Loan Scenario (acquisition)", "L7",  "Principal"),
    ]
    n_label_fixes = 0
    for sheet_name, coord, new_value in LABEL_FIXES:
        if sheet_name in wb.sheetnames:
            wb[sheet_name][coord] = new_value
            n_label_fixes += 1
    print(f"[build] Applied {n_label_fixes} label / typo fixes")

    # NOTE: Pro Forma S3 / Scen 4 scenario blocks are applied AFTER
    # fix_template.py runs (see "_apply_post_fix_patches" below). fix_template
    # writes "Free Cash Flow" labels to F84 / F97 / F100 from a stale layout
    # that pre-dated the scenario blocks moving to rows 81-87 / 91-97.
    # Doing those writes here would let fix_template silently overwrite them.

    # ── 5c. Collections — Avg row at G17/H17/H18/H19 ─────────────────────
    # CorrectOutput has an Avg row immediately below the 11 monthly rows
    # (rows 6-16). Templates that don't carry this row force the analyst
    # to compute T11 monthly averages by hand. Add the row so every deal
    # gets the avg / annualized / sum block.
    if "Collections" in wb.sheetnames:
        coll = wb["Collections"]
        coll["G17"] = "Avg"
        coll["H17"] = "=AVERAGE(H6:H16)"
        coll["H18"] = "=H17*12"
        coll["H19"] = "=SUM(H6:H16)"
        print(f"[build] Collections: added Avg row at G17/H17/H18/H19")

    # ── 5d. Loan Scenario — fix off-by-one sum ranges at P8 / T8 / U8 ────
    # The acquisition amort table runs rows 7-126 (Y1-Y120). P7/T7/U7 sum
    # column I rows 43:54 / 91:102 / 103:114; the immediately-below P8/T8/U8
    # sum column H over an off-by-one range that drops a row. Repoint to
    # the matching 12-month windows used by P7/T7/U7.
    ls_name = (
        "Loan Scenario (acquisition)"
        if "Loan Scenario (acquisition)" in wb.sheetnames
        else "Loan Scenario"
    )
    if ls_name in wb.sheetnames:
        ls = wb[ls_name]
        ls["P8"] = "=SUM(H43:H54)"
        ls["T8"] = "=SUM(H91:H102)"
        ls["U8"] = "=SUM(H103:H114)"
        print(f"[build] {ls_name}: rewrote P8/T8/U8 amort sum ranges")

    # ── 6. Force Excel to fully recalculate on open ──────────────────────
    # Without this, openpyxl-written formulas show blank in Excel until
    # the user presses F9. Setting both calcMode=auto AND fullCalcOnLoad
    # covers the cases where Excel ignores one or the other.
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcCompleted = False
    wb.calculation.calcOnSave = True

    # ── 7. Save ──────────────────────────────────────────────────────────
    wb.save(DEST)
    print(f"[build] Wrote {DEST.name}")
    print(f"[build] Final tabs: {load_workbook(DEST).sheetnames}")

    # ── 8. Apply fix_template.py on top ──────────────────────────────────
    # fix_template.py contains a large body of label / formula / formatting
    # patches accumulated over the project's history. Historically it was
    # invoked separately and could silently miss the runtime template if
    # the order was wrong. Run it here as a guaranteed post-step so any
    # patch it carries lands on the file we just wrote.
    #
    # Skip with BUILD_TEMPLATE_SKIP_FIX=1 if you need a clean
    # CorrectOutput-strip-only build (e.g. to diff against fix_template's
    # output and see what it changed).
    import os
    if os.environ.get("BUILD_TEMPLATE_SKIP_FIX") != "1":
        try:
            print("[build] Running fix_template.py on top of the fresh build...")
            import importlib
            import sys
            # fix_template.py is a top-level script that mutates and saves
            # the template on import. Re-import every run so it executes.
            if "fix_template" in sys.modules:
                del sys.modules["fix_template"]
            importlib.import_module("fix_template")
            print("[build] fix_template.py completed")
        except Exception as e:
            # Don't fail the build if fix_template trips — log and continue.
            # The label/typo/scenario-block fixes above are the minimum
            # guaranteed post-fix state.
            print(f"[build] WARNING: fix_template.py raised {type(e).__name__}: {e}")
            print("[build] Continuing with build_template-only patches.")

    # ── 9. Post-fix patches that fix_template.py would otherwise overwrite ──
    # fix_template.py's label sweep at line ~1154/1175/1176 writes
    # "Free Cash Flow" and "Avg CoC Y1-Y4" labels to F84, F97, and F100 on
    # GGC Pro Forma — values that pre-date the scenario blocks moving to
    # rows 81-87 (S3 bridge) and 91-97 (Scen 4 seller carry). Re-apply the
    # correct CorrectOutput layout here after fix_template has run.
    _apply_post_fix_patches()
    print("[build] Post-fix patches applied (Pro Forma S3 + Scen 4 scenario blocks)")


def _apply_post_fix_patches():
    """Patches that must land AFTER fix_template.py runs because fix_template
    writes stale labels to the same cells. Loads the template that
    build_template + fix_template just produced and re-saves it with the
    canonical Pro Forma scenario block layout from CorrectOutput.
    """
    wb = load_workbook(DEST)
    pf = wb["GGC Pro Forma"]

    # S3 block: rows 81-87 (BRIDGE LOAN -6 MONTHS) — see CorrectOutput
    # GGC Pro Forma(PW) rows 80-89 for the source layout.
    pf["H80"] = "BRIDGE LOAN -6 MONTHS"
    pf["F81"] = "Debt Service"
    pf["H81"] = "=-'Loan Scenario (acquisition)'!M10"
    pf["F82"] = "Debt Payoff"
    pf["F83"] = "Refi Cashout"
    pf["F84"] = "New Loan"
    pf["F85"] = "Total Sale - Community"
    pf["F86"] = "Free Cash Flow "
    pf["H86"] = "=H53+SUM(H81:H85)"
    pf["F87"] = "DSCR"
    pf["H87"] = "=-H53/H81"
    pf["F89"] = "IRR"
    pf["H89"] = "=IRR(G86:Q86)"

    # Scen 4 block: rows 91-97 (SELLER CARRY) — H91 derives a 5%
    # debt-service-only carry off Sources and Uses!J5 (the seller note).
    pf["F91"] = "Debt Service"
    pf["H91"] = "=-'Sources and Uses'!$J$5*5%"
    pf["F92"] = "Debt Payoff"
    pf["F93"] = "Refi Cashout"
    pf["F94"] = "New Loan"
    pf["F95"] = "Total Sale - Community"
    pf["F96"] = "Free Cash Flow "
    pf["H96"] = "=H53+SUM(H91:H95)"
    pf["F97"] = "DSCR"
    pf["H97"] = "=-H53/H91"
    pf["F100"] = "IRR"
    pf["H100"] = "=IRR(G96:Q96)"

    # ── RV Site Rental Income reachability ─────────────────────────────
    # fix_template.py repurposes GGC Underwriting row 14 (originally
    # "RV Site Rental Income") to "LTO" so MHC deals like Parkwood get
    # their land-contract revenue in the right bucket. Without that swap,
    # the SUMIFS at B/C/D/E/F14 searched for the string "RV Site Rental
    # Income" and matched it; afterward, they search for "LTO" and the
    # RV income string has no row at all — any RV deal (Whaleshead-style)
    # would silently zero that line.
    #
    # Add a parallel SUMIFS row at row 18 (currently blank between Other
    # Income at row 17 and Total EGI at row 19) so both enums stay
    # reachable. Update the EGI sum at row 19 to include row 18.
    uw = wb["GGC Underwriting"]
    if (uw["A18"].value or "").strip() in ("", "None"):
        uw["A18"] = "RV Site Rental Income"
        rv_sumifs = {
            "B": "=SUMIFS('Data Consolidation'!$D$3:$D$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
            "C": "=SUMIFS('Data Consolidation'!$E$3:$E$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
            "D": "=SUMIFS('Data Consolidation'!$G$3:$G$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
            "E": "=SUMIFS('Data Consolidation'!$H$3:$H$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
            "F": "=SUMIFS('Data Consolidation'!$F$3:$F$36,'Data Consolidation'!$A$3:$A$36,\"RV Site Rental Income\")",
        }
        for col, formula in rv_sumifs.items():
            uw[f"{col}18"] = formula
        uw["G18"] = "=I18"
        uw["H18"] = "=I18"
        uw["I18"] = 0
        uw["J18"] = "=I18/$N$7"
        # K18 holds a placeholder for the RV income value in the K (UW)
        # column. Use 0 (not a text note) so it stays numeric and the EGI
        # SUMs at row 19 don't trip #VALUE! when they include K18. The
        # analyst can overwrite with a real RV income figure for RV deals.
        uw["K18"] = 0
        # Update Total EGI sum at row 19 to include row 18 across every
        # column where the EGI sum lives. The CorrectOutput-style layout
        # uses K as the UW column with its own EGI formula at K19, so
        # patch K too (in addition to the legacy B-I columns).
        for col in ("B", "C", "D", "E", "F", "G", "H", "I", "K"):
            cell = uw[f"{col}19"]
            val = cell.value
            if isinstance(val, str) and f"{col}17" in val and f"{col}18" not in val:
                cell.value = val.replace(f"{col}17", f"{col}17+{col}18")

    # ── Storage / Retail Income circular-formula fix ──────────────────
    # The legacy template wires K15=G15, G15=I15, I15=K15 (three cells
    # each referencing the next). Same chain at row 16. The result is
    # #VALUE! → poisons I19 EGI → I44 Total OpEx → I47 NOI Total. CorrectOutput
    # short-circuits the cycle by setting G15/G16 to literal 0; do the
    # same here so MHC deals (which have no Storage or Retail income)
    # don't trip the cascade. The analyst can type a value into G15/G16
    # for a deal that actually has storage / retail revenue (Whaleshead).
    if isinstance(uw["G15"].value, str) and "=" in str(uw["G15"].value):
        uw["G15"] = 0
    if isinstance(uw["G16"].value, str) and "=" in str(uw["G16"].value):
        uw["G16"] = 0

    wb.save(DEST)


if __name__ == "__main__":
    main()
