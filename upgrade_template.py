"""
upgrade_template.py — One-time script to extend GGC's blank template
from 70 rent roll rows to 1000.

What it does:
1. Extends rows 72 → 1002 (1000 unit slots: rows 3-1002)
2. Adds the count formula =A_prev+1 to new rows 73-1002 in column A
3. Adds the combined-rent formula =E+F to new rows 73-1002 in column H
4. Moves totals row 73 → 1003 (with updated SUM ranges)
5. Moves averages row 74 → 1004
6. Moves annual row 75 → 1005
7. Updates all 8 cross-sheet references that pointed to old rows 73-75
   (Unit Mix Rent Growth, Unit Mix Summary, GGC Underwriting)

Run once: python upgrade_template.py
Output: GGC_Blank_Underwriting_Sizer_Extended.xlsx
"""

from openpyxl import load_workbook
from pathlib import Path
from copy import copy

INPUT_PATH  = Path("GGC_Blank_Underwriting_Sizer.xlsx")
OUTPUT_PATH = Path("GGC_Blank_Underwriting_Sizer_Extended.xlsx")

# Old → new row mapping for the totals block
OLD_LAST_DATA_ROW = 72
NEW_LAST_DATA_ROW = 1002   # rows 3-1002 = 1000 unit slots
ROW_SHIFT = NEW_LAST_DATA_ROW - OLD_LAST_DATA_ROW   # +930

print(f"Loading {INPUT_PATH}...")
wb = load_workbook(INPUT_PATH)
ws = wb["Rent Roll Input"]

# ── Step 1: Capture the existing formulas/styles from the last data row (72) ──
# Row 72 has: A72=A71+1 (count formula), H72=E72+F72 (combined formula)
# We want to copy these patterns down to row 1002.

reference_row = OLD_LAST_DATA_ROW  # row 72
print(f"Capturing styles/formulas from row {reference_row}...")

# Capture styles for each column we care about
style_cache = {}
for c in range(1, 12):
    cell = ws.cell(row=reference_row, column=c)
    style_cache[c] = {
        "font":      copy(cell.font) if cell.has_style else None,
        "fill":      copy(cell.fill) if cell.has_style else None,
        "border":    copy(cell.border) if cell.has_style else None,
        "alignment": copy(cell.alignment) if cell.has_style else None,
        "number_format": cell.number_format,
    }

# ── Step 2: Capture the totals/avg/annual rows (73, 74, 75) ───────────────────
totals_row_data = {}
for src_row, label in [(73, "totals"), (74, "averages"), (75, "annual")]:
    totals_row_data[src_row] = {}
    for c in range(1, 15):
        cell = ws.cell(row=src_row, column=c)
        totals_row_data[src_row][c] = {
            "value": cell.value,
            "font":      copy(cell.font) if cell.has_style else None,
            "fill":      copy(cell.fill) if cell.has_style else None,
            "border":    copy(cell.border) if cell.has_style else None,
            "alignment": copy(cell.alignment) if cell.has_style else None,
            "number_format": cell.number_format,
        }

# ── Step 3: Clear old totals rows 73-75 (they're moving) ──────────────────────
print("Clearing old totals rows 73-75...")
for r in [73, 74, 75]:
    for c in range(1, 15):
        ws.cell(row=r, column=c).value = None

# ── Step 4: Extend data rows 73 to 1002 ───────────────────────────────────────
print(f"Extending data rows from 73 to {NEW_LAST_DATA_ROW}...")
for r in range(73, NEW_LAST_DATA_ROW + 1):
    # Column A: count formula =A_prev+1
    cell_a = ws.cell(row=r, column=1, value=f"=A{r-1}+1")
    if style_cache[1]["font"]:
        cell_a.font = copy(style_cache[1]["font"])
        cell_a.fill = copy(style_cache[1]["fill"])
        cell_a.border = copy(style_cache[1]["border"])
        cell_a.alignment = copy(style_cache[1]["alignment"])
        cell_a.number_format = style_cache[1]["number_format"]

    # Column H: combined rent formula =E+F
    cell_h = ws.cell(row=r, column=8, value=f"=E{r}+F{r}")
    if style_cache[8]["font"]:
        cell_h.font = copy(style_cache[8]["font"])
        cell_h.fill = copy(style_cache[8]["fill"])
        cell_h.border = copy(style_cache[8]["border"])
        cell_h.alignment = copy(style_cache[8]["alignment"])
        cell_h.number_format = style_cache[8]["number_format"]

    # Apply styles to other columns (B-G, I-K) to keep formatting consistent
    for c in [2, 3, 4, 5, 6, 7, 9, 10, 11]:
        if style_cache[c]["font"]:
            cell = ws.cell(row=r, column=c)
            cell.font = copy(style_cache[c]["font"])
            cell.fill = copy(style_cache[c]["fill"])
            cell.border = copy(style_cache[c]["border"])
            cell.alignment = copy(style_cache[c]["alignment"])
            cell.number_format = style_cache[c]["number_format"]

# ── Step 5: Place totals/avg/annual rows at new positions ─────────────────────
# New positions: 1003 (totals), 1004 (averages), 1005 (annual)
new_totals_row  = NEW_LAST_DATA_ROW + 1  # 1003
new_avg_row     = NEW_LAST_DATA_ROW + 2  # 1004
new_annual_row  = NEW_LAST_DATA_ROW + 3  # 1005

print(f"Writing totals row at {new_totals_row}...")
# Row 73 was: =SUM(E3:E72) etc. New: =SUM(E3:E1002)
for c, info in totals_row_data[73].items():
    val = info["value"]
    if isinstance(val, str) and val.startswith("="):
        # Replace E3:E72 → E3:E1002 etc.
        col_letter = chr(64 + c)
        new_val = val.replace(f":{col_letter}72", f":{col_letter}{NEW_LAST_DATA_ROW}")
        ws.cell(row=new_totals_row, column=c, value=new_val)
    else:
        ws.cell(row=new_totals_row, column=c, value=val)
    if info["font"]:
        cell = ws.cell(row=new_totals_row, column=c)
        cell.font = info["font"]
        cell.fill = info["fill"]
        cell.border = info["border"]
        cell.alignment = info["alignment"]
        cell.number_format = info["number_format"]

print(f"Writing averages row at {new_avg_row}...")
for c, info in totals_row_data[74].items():
    val = info["value"]
    if isinstance(val, str) and val.startswith("="):
        col_letter = chr(64 + c)
        new_val = val.replace(f":{col_letter}72", f":{col_letter}{NEW_LAST_DATA_ROW}")
        ws.cell(row=new_avg_row, column=c, value=new_val)
    else:
        ws.cell(row=new_avg_row, column=c, value=val)
    if info["font"]:
        cell = ws.cell(row=new_avg_row, column=c)
        cell.font = info["font"]
        cell.fill = info["fill"]
        cell.border = info["border"]
        cell.alignment = info["alignment"]
        cell.number_format = info["number_format"]

print(f"Writing annual row at {new_annual_row}...")
# Row 75 had =E73*12 etc. New: =E1003*12
for c, info in totals_row_data[75].items():
    val = info["value"]
    if isinstance(val, str) and val.startswith("="):
        col_letter = chr(64 + c)
        # E73 → E1003 (the new totals row)
        new_val = val.replace(f"{col_letter}73", f"{col_letter}{new_totals_row}")
        ws.cell(row=new_annual_row, column=c, value=new_val)
    else:
        ws.cell(row=new_annual_row, column=c, value=val)
    if info["font"]:
        cell = ws.cell(row=new_annual_row, column=c)
        cell.font = info["font"]
        cell.fill = info["fill"]
        cell.border = info["border"]
        cell.alignment = info["alignment"]
        cell.number_format = info["number_format"]

# ── Step 6: Update cross-sheet references ────────────────────────────────────
# These need updating across other sheets:
#   F73 → F1003 (totals)
#   G73 → G1003
#   F74 → F1004 (averages)
#   G74 → G1004
#   F75 → F1005 (annual)
#   G75 → G1005

reference_updates = {
    "F73": f"F{new_totals_row}",
    "G73": f"G{new_totals_row}",
    "E73": f"E{new_totals_row}",
    "H73": f"H{new_totals_row}",
    "F74": f"F{new_avg_row}",
    "G74": f"G{new_avg_row}",
    "F75": f"F{new_annual_row}",
    "G75": f"G{new_annual_row}",
    "E75": f"E{new_annual_row}",
    "H75": f"H{new_annual_row}",
}

print(f"\nUpdating cross-sheet references in {len(wb.sheetnames)} sheets...")
update_count = 0
for sheet_name in wb.sheetnames:
    sht = wb[sheet_name]
    for row in sht.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            if "Rent Roll Input" not in v:
                continue
            new_v = v
            # Replace references in order from longest match to shortest
            # to avoid partial replacements (e.g. F75 inside F750)
            for old_ref, new_ref in reference_updates.items():
                # Look for !F73 or 'Rent Roll Input'!F73 (with possible $ for absolute)
                # The pattern: !{column}{row} where row is one of our targets
                # We need to be careful: F73 in the formula is preceded by ! and not by another digit
                # Use a careful string replace: only at boundaries
                col_letter = old_ref[0]
                row_num = old_ref[1:]
                # Patterns to replace (handle both $-prefixed and non-prefixed)
                patterns = [
                    (f"!{col_letter}{row_num}",  f"!{new_ref[0]}{new_ref[1:]}"),
                    (f"!${col_letter}${row_num}", f"!${new_ref[0]}${new_ref[1:]}"),
                    (f"!{col_letter}${row_num}", f"!{new_ref[0]}${new_ref[1:]}"),
                    (f"!${col_letter}{row_num}", f"!${new_ref[0]}{new_ref[1:]}"),
                ]
                for old_pat, new_pat in patterns:
                    if old_pat in new_v:
                        # Make sure we're not catching e.g. "!F730" — check the next char
                        # Since these are exact references, full match is safest
                        # We need to ensure the row number isn't followed by another digit
                        idx = 0
                        result_parts = []
                        while idx < len(new_v):
                            found_idx = new_v.find(old_pat, idx)
                            if found_idx == -1:
                                result_parts.append(new_v[idx:])
                                break
                            # Check char after the match — must not be a digit
                            after_idx = found_idx + len(old_pat)
                            if after_idx < len(new_v) and new_v[after_idx].isdigit():
                                # It's a longer reference like !F730 — skip
                                result_parts.append(new_v[idx:after_idx])
                                idx = after_idx
                            else:
                                result_parts.append(new_v[idx:found_idx])
                                result_parts.append(new_pat)
                                idx = after_idx
                        new_v = "".join(result_parts)

            if new_v != v:
                cell.value = new_v
                update_count += 1
                print(f"  {sheet_name}!{cell.coordinate}: {v[:60]} → {new_v[:60]}")

print(f"\nTotal cross-sheet references updated: {update_count}")

# ── Step 7: Save ──────────────────────────────────────────────────────────────
print(f"\nSaving to {OUTPUT_PATH}...")
wb.save(OUTPUT_PATH)
print("Done.")

# ── Step 8: Verify ────────────────────────────────────────────────────────────
print("\nVerification:")
wb2 = load_workbook(OUTPUT_PATH)
ws2 = wb2["Rent Roll Input"]
print(f"  Row 1002 col A: {ws2['A1002'].value} (should be =A1001+1)")
print(f"  Row 1002 col H: {ws2['H1002'].value} (should be =E1002+F1002)")
print(f"  Row 1003 col E: {ws2['E1003'].value} (should be =SUM(E3:E1002))")
print(f"  Row 1005 col E: {ws2['E1005'].value} (should be =E1003*12)")

# Check downstream references
ws_uw = wb2["GGC Underwriting"]
print(f"\n  GGC Underwriting!I13: {ws_uw['I13'].value} (should reference F1005)")
print(f"  GGC Underwriting!I14: {ws_uw['I14'].value} (should reference G1003)")

ws_um = wb2["Unit Mix Summary"]
print(f"  Unit Mix Summary!F13: {ws_um['F13'].value} (should reference F1003)")
print(f"  Unit Mix Summary!C22: {ws_um['C22'].value} (should reference F1004)")
