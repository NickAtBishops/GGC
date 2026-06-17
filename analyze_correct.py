"""Deep analysis of the Whaleshead Correct Output workbook."""
import os
from openpyxl import load_workbook

PATH = "/Users/nicholasrevencomacbookair/Desktop/GGC/V5/ggc-deal-engine/Outputs/CorrectOutput copy.xlsx"
OUT_DIR = "/Users/nicholasrevencomacbookair/Desktop/GGC/V5/ggc-deal-engine/analysis_dumps"
os.makedirs(OUT_DIR, exist_ok=True)

wb_f = load_workbook(PATH, data_only=False)
wb_v = load_workbook(PATH, data_only=True)

# 1. Overview of every sheet
overview_path = os.path.join(OUT_DIR, "00_overview.txt")
with open(overview_path, "w") as f:
    f.write("=" * 100 + "\n")
    f.write("WORKBOOK SHEET OVERVIEW\n")
    f.write("=" * 100 + "\n\n")
    for i, name in enumerate(wb_f.sheetnames):
        ws = wb_f[name]
        # count non-empty cells
        nonempty = 0
        max_r, max_c = ws.max_row, ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c, values_only=True):
            for v in row:
                if v not in (None, ""):
                    nonempty += 1
        f.write(f"[{i}] '{name}'\n")
        f.write(f"    dimensions: {max_r} rows x {max_c} cols\n")
        f.write(f"    non-empty cells: {nonempty}\n\n")
print(f"Wrote {overview_path}")

# Helper to dump full sheet
def dump_sheet(name, prefix):
    out = os.path.join(OUT_DIR, f"{prefix}_{name.replace(' ', '_').replace('/', '_')}.txt")
    ws_f = wb_f[name]
    ws_v = wb_v[name]
    with open(out, "w") as f:
        f.write(f"SHEET: '{name}'\n")
        f.write(f"Dimensions: {ws_f.max_row} rows x {ws_f.max_column} cols\n")
        f.write("=" * 100 + "\n")
        f.write(f"{'CELL':<8} {'VALUE (resolved)':<60} {'FORMULA / RAW':<60}\n")
        f.write("=" * 100 + "\n")
        for row in ws_f.iter_rows():
            for cell in row:
                raw = cell.value
                if raw is None or raw == "":
                    continue
                resolved = ws_v[cell.coordinate].value
                raw_s = str(raw)[:58]
                res_s = str(resolved)[:58] if resolved is not None else "None"
                # only write formula in raw col if it differs from resolved
                if isinstance(raw, str) and raw.startswith("="):
                    f.write(f"{cell.coordinate:<8} {res_s:<60} {raw_s:<60}\n")
                else:
                    f.write(f"{cell.coordinate:<8} {res_s:<60} {'':<60}\n")
    print(f"Wrote {out}")
    return out

for name in wb_f.sheetnames:
    dump_sheet(name, f"sheet_{wb_f.sheetnames.index(name):02d}")

print("DONE")
