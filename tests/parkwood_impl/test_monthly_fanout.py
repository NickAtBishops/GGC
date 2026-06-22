"""Deterministic tests for the monthly fan-out fallback in `fill_template`.

CONTEXT (OUTLINE root cause #4 / CLAUDE.md §0)
-----------------------------------------------
Before the Parkwood fix, `fill_template` flat-spread `annual / 12` across
the Data Consolidation J:U monthly columns whenever the methodology
returned a line with a `t12Total` but no monthly array. Twelve identical
values look fine on inspection but lie about the trend, which silently
corrupts:

  - the Collections build (§5.1) — T3 / T6 / T12 trend bands collapse to
    a single number,
  - the bad-debt goal-seek that relies on those bands,
  - any monthly anomaly detection (§7) — every month becomes its own
    average so nothing ever looks anomalous.

The fix in `backend.py` lines 6910-6927 (`_write_item` inside
`fill_template`):

  monthly = item.get("monthly") or []
  if len(monthly) == 12:
      for m_i, val in enumerate(monthly):
          ws.cell(row=r, column=10 + m_i, value=val)
  elif item.get("t12Total"):
      # Annual-only P&L: do NOT fan out annual/12.
      annual_only_lines.append(item.get("sellerName") or
                               item.get("ggcCategory") or "?")

…and the WARN entry it appends to `financials["_extractionChecks"]` after
the write loop.

WHAT THESE TESTS PIN
--------------------
1. The DC monthly cells (columns J-U = openpyxl columns 10-21) are left
   EMPTY (cell.value is None) for an annual-only line. No "annual/12"
   number lands in any of the 12 monthly cells.
2. An Extraction Check entry exists with status="warn", an item label of
   "Annual-only P&L coverage", and a detail string naming the affected
   sellerName.
3. For comparison: a line WITH a full 12-month series gets each monthly
   value written into J-U (the happy path still works).

DESIGN COMPROMISES (documented per task instructions)
-----------------------------------------------------
- `_write_item` is a nested closure inside `fill_template`, so it cannot
  be imported and called in isolation without re-creating its enclosing
  scope (ws, annual_only_lines list, _structural_rows skip set, etc.).
  Source-slicing the def out — the approach `test_canonicalize_unit_type`
  takes — would still need a stub worksheet and an annual_only_lines
  list, AND it would not exercise the real WARN-append code path that
  lives in fill_template proper. So the primary test drives the full
  `fill_template` against a tempfile workbook, matching the style of
  `tests/test_template_contract.py::test_end_to_end_write_preserves_structural_rows`.
  Runtime is a few seconds (one openpyxl load + save of the real
  template) — acceptable for the safety this pins.

- The synthetic `financials` payload is intentionally MINIMAL: a single
  income line with t12Total but no monthly, one income line with a full
  monthly array (the happy-path control), and a single expense line with
  t12Total but no monthly. We do NOT exercise the Underwriting tab here
  — only the Data Consolidation J:U write behavior and the
  _extractionChecks list. Other Underwriting numerics are validated by
  `test_pipeline.py` and `test_template_contract.py`.

- The `market` dict is a near-empty stub. `fill_template` accesses every
  market field defensively with `market.get(..., default) or default`,
  so the stub is enough to let the write-back complete.

- `propertyInfo.totalUnits` is set so the rent-roll write loop doesn't
  blow up downstream. The actual rent-roll content is empty — we don't
  care about it for this test.

Run:  pytest tests/parkwood_impl/test_monthly_fanout.py -v
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

import backend  # noqa: E402


# ─────────────────────────────────────────────────────────────────────────── #
# Constants — the cell range under test                                       #
# ─────────────────────────────────────────────────────────────────────────── #
# Data Consolidation monthly columns J..U map to openpyxl numeric columns
# 10..21 (12 columns), per `_write_item` (`ws.cell(row=r, column=10 + m_i)`).
DC_MONTHLY_COL_START = 10  # J
DC_MONTHLY_COL_END   = 21  # U  (inclusive)
DC_MONTHLY_COL_COUNT = DC_MONTHLY_COL_END - DC_MONTHLY_COL_START + 1
assert DC_MONTHLY_COL_COUNT == 12, "J:U is twelve columns by definition"


# Sentinel labels — easy to grep for in the WARN detail string and easy
# to spot in the workbook column B if the assertion ever fires.
ANNUAL_ONLY_INCOME_LABEL  = "4999 Annual-Only Lot Rent Aggregate"
ANNUAL_ONLY_EXPENSE_LABEL = "5999 Annual-Only Insurance Aggregate"
HAPPY_PATH_INCOME_LABEL   = "4101 Monthly Detail Lot Rent"


# ─────────────────────────────────────────────────────────────────────────── #
# Fixtures                                                                    #
# ─────────────────────────────────────────────────────────────────────────── #
@pytest.fixture
def synthetic_financials() -> dict:
    """A minimal financials payload exercising both code paths.

    - One income line with monthly=None  (the annual-only case under test)
    - One income line with a full 12-month array (the happy-path control)
    - One expense line with monthly=None  (a second annual-only case)

    All numerics are deliberately non-round so a flat-spread regression
    (annual/12 = 1000) would be visually obvious in the workbook output.
    """
    return {
        "income": [
            {
                "ggcCategory":     "Gross Potential Rent",
                "sellerName":      ANNUAL_ONLY_INCOME_LABEL,
                "t12Total":        144_376,   # not divisible by 12 to make
                                              # any accidental fan-out
                                              # numerically obvious
                "monthly":         None,      # the case under test
                "ggcUnderwritten": 144_376,
                "confidence":      "medium",
            },
            {
                "ggcCategory":     "Other Income",
                "sellerName":      HAPPY_PATH_INCOME_LABEL,
                "t12Total":        24_000,
                "monthly":         [2_000.0] * 12,  # full series — happy path
                "ggcUnderwritten": 24_000,
                "confidence":      "high",
            },
        ],
        "expenses": [
            {
                "ggcCategory":     "Insurance",
                "sellerName":      ANNUAL_ONLY_EXPENSE_LABEL,
                "t12Total":        18_500,
                "monthly":         None,      # second annual-only case
                "ggcUnderwritten": 18_500,
                "confidence":      "medium",
            },
        ],
        "rentRoll":     {"totalUnits": 100, "unitGroups": []},
        "propertyInfo": {"name": "Annual-Only Test Park",
                         "askingPrice": 5_000_000,
                         "totalUnits": 100},
    }


@pytest.fixture
def stub_market() -> dict:
    """`fill_template` reads market fields defensively; empty containers
    are enough to let the write-back run without raising."""
    return {
        "rentComps":               [],
        "saleComps":               [],
        "landmarks":               [],
        "demographics":            {},
        "altHousing":              {},
        "visuals":                 {},
        "marketRentConclusion":    "",
        "marketCapRateConclusion": "",
        "demandSignal":            "MODERATE",
        "demandRationale":         "",
    }


@pytest.fixture
def filled_workbook_path(synthetic_financials, stub_market):
    """Run fill_template once and hand the resulting .xlsx path to the
    test. The same workbook is reused across the assertions in a single
    test invocation to keep the runtime down — every test still gets its
    own fresh workbook because the fixture is function-scoped."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out_path = Path(f.name)
    try:
        backend.fill_template(synthetic_financials, stub_market, out_path)
        yield out_path, synthetic_financials
    finally:
        out_path.unlink(missing_ok=True)


# ─────────────────────────────────────────────────────────────────────────── #
# Assertion helpers                                                           #
# ─────────────────────────────────────────────────────────────────────────── #
def _find_dc_row_by_seller_name(ws, seller_name: str) -> int | None:
    """Return the row index in Data Consolidation whose column B holds
    `seller_name`, or None if no such row exists.

    Column B is where `_write_item` writes the sellerName (see
    `ws.cell(row=r, column=2, value=item.get("sellerName", ""))` in
    backend.py around line 6912)."""
    # The income band runs rows 3-36, expense band rows 43-102. Scan
    # both — the seller name pins the row regardless of which slot it
    # landed in.
    for r in list(range(3, 37)) + list(range(43, 103)):
        if ws.cell(row=r, column=2).value == seller_name:
            return r
    return None


def _read_monthly_cells(ws, row: int) -> list:
    """Return the 12 raw cell values from J..U on the given row."""
    return [ws.cell(row=row, column=c).value
            for c in range(DC_MONTHLY_COL_START, DC_MONTHLY_COL_END + 1)]


# ─────────────────────────────────────────────────────────────────────────── #
# Tests                                                                       #
# ─────────────────────────────────────────────────────────────────────────── #
def test_annual_only_income_line_leaves_monthly_cells_blank(filled_workbook_path):
    """An income line with t12Total but `monthly=None` must NOT have any
    annual/12 value written into J..U. Every monthly cell must be None.

    REGRESSION CAUGHT: prior behavior populated all 12 cells with
    t12Total/12 (= 12_031.33 for the sentinel value 144_376), which
    silently flattened the Collections trend. If the regression returns,
    every cell here would hold ~12_031 and this assertion would scream.
    """
    out_path, _ = filled_workbook_path
    wb = load_workbook(out_path, data_only=False)
    ws = wb["Data Consolidation"]

    row = _find_dc_row_by_seller_name(ws, ANNUAL_ONLY_INCOME_LABEL)
    assert row is not None, (
        f"Could not locate the annual-only income line "
        f"{ANNUAL_ONLY_INCOME_LABEL!r} in Data Consolidation column B. "
        f"Either the _keep filter dropped it, the income slot list "
        f"shifted, or fill_template's write contract changed."
    )

    monthly_cells = _read_monthly_cells(ws, row)
    non_empty = [(DC_MONTHLY_COL_START + i, v)
                 for i, v in enumerate(monthly_cells) if v is not None]
    assert non_empty == [], (
        f"Annual-only income line {ANNUAL_ONLY_INCOME_LABEL!r} at row "
        f"{row} should have BLANK monthly cells (J..U). Found values: "
        f"{non_empty}. This is the annual/12 flat-spread regression — "
        f"see CLAUDE.md §0 and the OUTLINE root cause #4 note."
    )


def test_annual_only_expense_line_leaves_monthly_cells_blank(filled_workbook_path):
    """Same guarantee on the expense band. Insurance is a common
    annual-only field (sellers often quote a single annual premium with
    no monthly breakdown), so this case is the realistic one we'll hit.
    """
    out_path, _ = filled_workbook_path
    wb = load_workbook(out_path, data_only=False)
    ws = wb["Data Consolidation"]

    row = _find_dc_row_by_seller_name(ws, ANNUAL_ONLY_EXPENSE_LABEL)
    assert row is not None, (
        f"Could not locate the annual-only expense line "
        f"{ANNUAL_ONLY_EXPENSE_LABEL!r} in Data Consolidation column B."
    )

    monthly_cells = _read_monthly_cells(ws, row)
    non_empty = [(DC_MONTHLY_COL_START + i, v)
                 for i, v in enumerate(monthly_cells) if v is not None]
    assert non_empty == [], (
        f"Annual-only expense line {ANNUAL_ONLY_EXPENSE_LABEL!r} at row "
        f"{row} should have BLANK monthly cells (J..U). Found values: "
        f"{non_empty}."
    )


def test_happy_path_line_still_writes_all_twelve_months(filled_workbook_path):
    """The control case: a line that DOES supply a 12-element monthly
    array must still get every monthly value written into J..U. This
    guards against an overcorrection where the fix accidentally blanks
    monthly cells across the board.
    """
    out_path, _ = filled_workbook_path
    wb = load_workbook(out_path, data_only=False)
    ws = wb["Data Consolidation"]

    row = _find_dc_row_by_seller_name(ws, HAPPY_PATH_INCOME_LABEL)
    assert row is not None, (
        f"Could not locate the happy-path income line "
        f"{HAPPY_PATH_INCOME_LABEL!r} in Data Consolidation column B."
    )

    monthly_cells = _read_monthly_cells(ws, row)
    # Each cell should hold 2000 (the synthetic monthly value). openpyxl
    # may return int 2000 or float 2000.0 depending on how the formula
    # protector wrote it — accept either.
    for i, v in enumerate(monthly_cells):
        col_letter = chr(ord("J") + i)
        assert v == 2_000 or v == 2_000.0, (
            f"Happy-path monthly cell {col_letter}{row} expected 2000, "
            f"got {v!r}. The fix may have overshot and blanked legitimate "
            f"monthly arrays."
        )


def test_extraction_check_warn_row_names_the_annual_only_lines(filled_workbook_path):
    """A WARN-status entry must be appended to `_extractionChecks` whose
    detail string names BOTH annual-only sellerNames. This is the
    reviewer's signal — without it the blank cells just look like a bug
    rather than a deliberate refusal to invent a trend.
    """
    _, financials = filled_workbook_path
    checks = financials.get("_extractionChecks") or []
    warn_entries = [
        c for c in checks
        if (c.get("status") == "warn"
            and c.get("item") == "Annual-only P&L coverage")
    ]
    assert warn_entries, (
        f"Expected one 'Annual-only P&L coverage' WARN entry in "
        f"_extractionChecks, found none. All entries: "
        f"{[(c.get('item'), c.get('status')) for c in checks]}"
    )
    assert len(warn_entries) == 1, (
        f"Expected exactly one annual-only WARN entry, found "
        f"{len(warn_entries)}. fill_template should aggregate all "
        f"annual-only lines into a single WARN, not one per line."
    )

    detail = warn_entries[0].get("detail") or ""
    # Both sellerNames (income + expense) must appear in the detail so
    # the reviewer can act on them. The detail format from backend.py
    # joins names with ", " and includes "(+N more)" when there are
    # more than 5 — we have 2, so both should appear verbatim.
    assert ANNUAL_ONLY_INCOME_LABEL in detail, (
        f"WARN detail does not name the annual-only income line "
        f"{ANNUAL_ONLY_INCOME_LABEL!r}. Detail was: {detail!r}"
    )
    assert ANNUAL_ONLY_EXPENSE_LABEL in detail, (
        f"WARN detail does not name the annual-only expense line "
        f"{ANNUAL_ONLY_EXPENSE_LABEL!r}. Detail was: {detail!r}"
    )
    # The count in the detail should be 2 (one income + one expense).
    assert "2 line(s)" in detail, (
        f"WARN detail should report '2 line(s)' annual-only; got: "
        f"{detail!r}. Count drift here usually means the _keep filter "
        f"silently dropped one of the synthetic lines before write."
    )


def test_extraction_check_warn_absent_when_all_lines_have_monthly(
    stub_market, tmp_path
):
    """Negative control: when every line carries a full 12-month series
    there is NO 'Annual-only P&L coverage' WARN. Without this assertion a
    bug that mis-classifies happy-path lines as annual-only would slip
    past test_annual_only_*: those tests would still see blank cells if
    the misclassification dropped the monthly values too.

    This test uses its own minimal financials (one fully-monthly line)
    rather than reusing the shared fixture, because the shared one
    deliberately includes annual-only cases.
    """
    financials = {
        "income": [
            {
                "ggcCategory":     "Gross Potential Rent",
                "sellerName":      "4101 All-Monthly Lot Rent",
                "t12Total":        24_000,
                "monthly":         [2_000.0] * 12,
                "ggcUnderwritten": 24_000,
                "confidence":      "high",
            },
        ],
        "expenses":     [],
        "rentRoll":     {"totalUnits": 100, "unitGroups": []},
        "propertyInfo": {"name": "All-Monthly Park", "totalUnits": 100,
                         "askingPrice": 5_000_000},
    }
    out_path = tmp_path / "all_monthly.xlsx"
    backend.fill_template(financials, stub_market, out_path)

    checks = financials.get("_extractionChecks") or []
    warn_entries = [
        c for c in checks
        if (c.get("status") == "warn"
            and c.get("item") == "Annual-only P&L coverage")
    ]
    assert warn_entries == [], (
        f"No annual-only line is present, so the 'Annual-only P&L "
        f"coverage' WARN should NOT have been raised. Got: {warn_entries}"
    )
