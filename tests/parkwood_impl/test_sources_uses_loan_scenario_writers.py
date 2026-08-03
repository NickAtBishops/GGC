"""Deterministic tests for `fill_sources_and_uses` and `fill_loan_scenario`
(backend.py) — the two Sources & Uses / Loan Scenario writer functions.

Both bugs pinned here were silent-dollar-error template-wiring bugs (an
ultrareview finding) that violated CLAUDE.md's "never underwrite the wrong
numbers" mandate: the workbook rendered fine and every formula evaluated
cleanly, but the numbers were wrong because the writers targeted the wrong
cells relative to the live template layout.

1. `fill_sources_and_uses` wrote analyst-supplied capex_line_items starting
   at row 21, but the template's total formula is C27=SUM(C22:C26) — row
   21 (labeled "Water / Septic / Utilities (per deal)") sits OUTSIDE the
   SUM window regardless of what's written there. The first capex line
   item was silently dropped from the total.

2. `fill_loan_scenario` wrote amort_months/term_months to BOTH the
   task-spec cells (C20/C21) AND the live-template cells (C10/C8). But
   C20/C21 in the live template are NOT loan-term cells — they're
   "Closing Costs" (0) and "Capital Expenditure" (250,000), both literals
   feeding Total Cost Basis via C22=C19+C20+C21. Writing loan terms there
   silently clobbered the $250K capex budget.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

import backend  # noqa: E402


@pytest.fixture
def wb():
    return load_workbook(backend.TEMPLATE_PATH)


# --------------------------------------------------------------------------- #
# fill_sources_and_uses — capex line items land in the SUM(C22:C26) window   #
# --------------------------------------------------------------------------- #
def test_capex_line_items_land_in_sum_window_not_before_it(wb):
    """All analyst-supplied capex line items must land in rows 22-26 (the
    SUM(C22:C26) window feeding C27). None should land in row 21 — its
    label ("Water / Septic / Utilities (per deal)") sits outside the SUM
    window, so a value written there is silently excluded from the total."""
    ws = wb["Sources and Uses"]
    label_before = ws["B21"].value
    assert label_before == "Water / Septic / Utilities (per deal)", (
        f"Fixture assumption broken — B21 should hold the template's "
        f"stock row-21 label, got {label_before!r}"
    )

    property_info = {
        "capex_line_items": [
            {"label": "Water/Sewer", "amount": 100_000},
            {"label": "Homes",       "amount":  50_000},
        ],
    }
    backend.fill_sources_and_uses(wb, {}, property_info)

    # Row 21's label must survive untouched — not overwritten by the
    # first item — and its C-value must stay unwritten (outside the SUM).
    assert ws["B21"].value == "Water / Septic / Utilities (per deal)", ws["B21"].value
    assert ws["C21"].value in (None, ""), (
        f"C21 must stay outside the line-item write range (it's outside "
        f"SUM(C22:C26)); got {ws['C21'].value!r}"
    )

    # Both line items must land inside rows 22-26.
    assert ws["B22"].value == "Water/Sewer"
    assert ws["C22"].value == 100_000
    assert ws["B23"].value == "Homes"
    assert ws["C23"].value == 50_000

    # C27 formula must still cover the window the items were written into.
    assert ws["C27"].value == "=SUM(C22:C26)", ws["C27"].value


def test_capex_line_items_caps_at_five_to_fit_sum_window(wb):
    """SUM(C22:C26) is a 5-row window. A 6th line item must not silently
    fall outside it (rows 22-26 = 5 rows, not 6)."""
    property_info = {
        "capex_line_items": [
            {"label": f"Item {i}", "amount": 1_000 * (i + 1)}
            for i in range(6)
        ],
    }
    backend.fill_sources_and_uses(wb, {}, property_info)

    ws = wb["Sources and Uses"]
    written_rows = [r for r in range(22, 27) if ws.cell(row=r, column=2).value]
    assert len(written_rows) == 5, (
        f"Expected exactly 5 line items written (rows 22-26), got "
        f"{len(written_rows)}: {written_rows}"
    )
    # The 6th item must not have landed on row 27 (that's the SUM formula
    # row — clobbering it would break the total entirely).
    assert ws["C27"].value == "=SUM(C22:C26)", ws["C27"].value


# --------------------------------------------------------------------------- #
# fill_loan_scenario — loan-term writes must not touch Closing Costs/CapEx   #
# --------------------------------------------------------------------------- #
def test_amort_and_term_overrides_do_not_clobber_closing_costs_or_capex(wb):
    """Setting amort_months/term_months must only touch C10/C8 (the live
    template's actual amort/term cells) and must leave C20 (Closing Costs)
    and C21 (Capital Expenditure) at their template defaults."""
    ws = wb["Loan Scenario (acquisition)"]
    closing_costs_before = ws["C20"].value
    capex_before = ws["C21"].value
    assert closing_costs_before == 0, closing_costs_before
    assert capex_before == 250_000, capex_before

    property_info = {"amort_months": 360, "term_months": 60}
    backend.fill_loan_scenario(wb, {}, property_info)

    assert ws["C10"].value == 360, "C10 (Amortization) should be overwritten"
    assert ws["C8"].value == 60, "C8 (Term) should be overwritten"
    assert ws["C20"].value == closing_costs_before, (
        f"C20 (Closing Costs) must NOT be touched by the loan-term "
        f"override; got {ws['C20'].value!r}, expected {closing_costs_before!r}"
    )
    assert ws["C21"].value == capex_before, (
        f"C21 (Capital Expenditure) must NOT be touched by the loan-term "
        f"override; got {ws['C21'].value!r}, expected {capex_before!r}"
    )
    # Total Cost Basis formula must stay intact so it keeps summing the
    # untouched capex value.
    assert ws["C22"].value == "=C19+C20+C21", ws["C22"].value
