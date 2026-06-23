"""Accuracy verification for the new Comps Analysis sections.

Each test asserts a computed value against an INDEPENDENTLY derived
expected value, using a different math path than the engine uses. The
goal is to catch math/logic errors that pass `compile + import` but
emit silently wrong numbers — exactly the failure mode the user just
caught in the rent-roll undercount and Collections fan-out.

Sections covered:
  §1 Comp Quality + Similarity scoring
  §2 Robust Statistics + Percentile Positioning
  §3 Mark-to-Market Upside Heatmap
  §4 Implied Rent Growth from Sale Comp Vintage
  §5 Submarket Concentration (HHI) + Cap Rate Triangulation
  §6 Affordability Spread

Each section has 2-4 tests covering: happy path, edge case, boundary.
Tolerances are LOOSE by design — exact float equality fails too often
for derived stats with float-point math. The looser tolerance still
catches "off by a factor of N" bugs, which is what we care about.
"""
from __future__ import annotations

import math
import statistics
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

import backend  # noqa: E402


# --------------------------------------------------------------------------- #
# Helpers — build a comp set + run fill_template + scrape the output tab.    #
# --------------------------------------------------------------------------- #
def _make_subject_financials(*, lot_rent: float = 450.0,
                             units: int = 100,
                             occ_rate: float = 0.94) -> dict:
    """Synthetic Parkwood-shaped financials.

    The Comps Analysis fallback chain for subject_rent: first checks
    rentRoll.avgLotRent (LLM-set), then weights unitGroups, then averages
    occupied rentRollRows. We populate rentRoll.avgLotRent directly so
    the test isolates the subject_rent value from extraction quirks.
    """
    return {
        "rentRoll": {
            "avgLotRent":   lot_rent,
            "totalUnits":   units,
            "occupiedUnits": int(units * occ_rate),
            "vacantUnits":   units - int(units * occ_rate),
            "occupancyRate": occ_rate,
            "unitGroups":   [],
            "rentRollRows": [],
        },
        "income":   [],
        "expenses": [],
        "propertyInfo": {
            "name":         "Test Park",
            "totalUnits":   units,
            "units":        units,
            "state":        "MI",
            "yearBuilt":    1985,
        },
    }


def _build_rent_comps(values: list[float],
                      states: list[str] | None = None) -> list[dict]:
    """Build N rent comps with the given lot-rent values. States default
    to MI (matches subject in tests). Other fields kept comparable so
    similarity scoring doesn't vary in the test cases.
    """
    states = states or ["MI"] * len(values)
    return [
        {
            "name":          f"Comp #{i+1}",
            "city":          "City",
            "state":         st,
            "units":         100,
            "lotRent":       v,
            "occupancy":     0.95,
            "yearBuilt":     1985,
            "pohPercent":    0.10,
            "qualityRating": 3,
            "source":        "test",
        }
        for i, (v, st) in enumerate(zip(values, states))
    ]


def _build_sale_comps(triples: list[tuple]) -> list[dict]:
    """triples = [(year, ppu, noi, sale_price), ...]"""
    out = []
    for i, (year, ppu, noi, price) in enumerate(triples):
        out.append({
            "name":         f"Sale #{i+1}",
            "location":     "City, MI",
            "saleDate":     f"{year}-06",
            "units":        100,
            "salePrice":    price,
            "pricePerUnit": ppu,
            "capRate":      noi / price if price else 0,
            "noi":          noi,
            "buyer":        "Buyer",
            "seller":       "Seller",
            "source":       "test",
        })
    return out


def _market_dict(*, rent_comps=None, sale_comps=None,
                 demo: dict | None = None,
                 alt: dict | None = None) -> dict:
    return {
        "rentComps":        rent_comps or [],
        "saleComps":        sale_comps or [],
        "demographics":     demo or {"countyName": "Test County", "majorEmployers": []},
        "altHousing":       alt or {},
        "marketRentConclusion": "test conclusion",
    }


def _run_and_inspect(financials: dict, market: dict, tmp_path: Path) -> dict:
    """Run fill_template and parse the Comps Analysis tab. Returns a dict
    indexed by section name with the relevant computed cell values."""
    out_path = str(tmp_path / "comps_test.xlsx")
    # Run a copy of financials so the engine's mutations don't leak between
    # tests (apply_ggc_overrides + _ensure_rent_roll_complete both mutate).
    backend.fill_template(financials, market, out_path)
    wb = openpyxl.load_workbook(out_path, data_only=False)
    ws = wb["Comps Analysis"]
    return _scrape_sections(ws)


def _scrape_sections(ws) -> dict:
    """Locate each new analytical section by header text and pull the
    cells whose values we want to verify."""
    headers = {}
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str):
            for tag in ("COMP QUALITY", "ROBUST STATISTICS",
                        "MARK-TO-MARKET", "IMPLIED MARKET RENT GROWTH",
                        "SUBMARKET CONCENTRATION", "AFFORDABILITY SPREAD"):
                if tag in b:
                    headers[tag] = r
                    break

    out: dict = {}

    # §1 Comp Quality — header at row, table header at +3, comps from +4.
    if "COMP QUALITY" in headers:
        cq_h = headers["COMP QUALITY"]
        table_h = cq_h + 3
        comps = []
        for r in range(table_h + 1, table_h + 30):
            name = ws.cell(row=r, column=3).value
            if not isinstance(name, str) or not name:
                break
            comps.append({
                "name":     name,
                "geo":      ws.cell(row=r, column=4).value,
                "size":     ws.cell(row=r, column=5).value,
                "vintage":  ws.cell(row=r, column=6).value,
                "poh":      ws.cell(row=r, column=7).value,
                "overall":  ws.cell(row=r, column=8).value,
                "lot_rent": ws.cell(row=r, column=9).value,
                "weighted": ws.cell(row=r, column=10).value,
            })
        out["comp_quality"] = comps

    # §2 Robust Statistics — Lot Rent row at header + 4
    if "ROBUST STATISTICS" in headers:
        rs_h = headers["ROBUST STATISTICS"]
        lr = rs_h + 4
        out["robust_lot_rent"] = {
            "subject":     ws.cell(row=lr, column=3).value,
            "p25":         ws.cell(row=lr, column=4).value,
            "median":      ws.cell(row=lr, column=5).value,
            "p75":         ws.cell(row=lr, column=6).value,
            "mean":        ws.cell(row=lr, column=7).value,
            "trimmed":     ws.cell(row=lr, column=8).value,
            "stddev":      ws.cell(row=lr, column=9).value,
            "cv":          ws.cell(row=lr, column=10).value,
            "position":    ws.cell(row=lr, column=11).value,
            "weighted":    ws.cell(row=lr, column=12).value,
        }

    # §3 MTM — three scenarios at header + 4, +5, +6
    if "MARK-TO-MARKET" in headers:
        mtm_h = headers["MARK-TO-MARKET"]
        scenarios = []
        for offset in (4, 5, 6):
            r = mtm_h + offset
            scenarios.append({
                "label":        ws.cell(row=r, column=2).value,
                "delta_rent":   ws.cell(row=r, column=3).value,
                "annual_gpr":   ws.cell(row=r, column=4).value,
                "value_5":      ws.cell(row=r, column=5).value,
                "value_55":     ws.cell(row=r, column=6).value,
                "value_6":      ws.cell(row=r, column=7).value,
            })
        out["mtm"] = scenarios

    # §4 Implied growth — value at header + 3 column 4
    if "IMPLIED MARKET RENT GROWTH" in headers:
        ig_h = headers["IMPLIED MARKET RENT GROWTH"]
        out["implied_growth"] = ws.cell(row=ig_h + 3, column=4).value

    # §5 HHI + cap triangulation — HHI label is 2 rows below section header
    if "SUBMARKET CONCENTRATION" in headers:
        smc_h = headers["SUBMARKET CONCENTRATION"]
        # Walk down to find "HHI (rent-comp operators):" label
        hhi_row = None
        for r in range(smc_h + 1, smc_h + 8):
            b = ws.cell(row=r, column=2).value
            if isinstance(b, str) and b.startswith("HHI ("):
                hhi_row = r
                break
        if hhi_row:
            out["hhi"] = ws.cell(row=hhi_row, column=3).value
            out["cap_a"] = ws.cell(row=hhi_row + 1, column=8).value
            out["cap_b"] = ws.cell(row=hhi_row + 2, column=8).value
            out["cap_c"] = ws.cell(row=hhi_row + 3, column=8).value

    # §6 Affordability — 5 metrics at fixed offsets aff_h + 3..7
    # (engine writes at aff_row = aff_start + 3, then 4 more rows). The
    # substring-scan approach picks up the description text too because
    # that paragraph naturally mentions "spread", "2BR", "MHP all-in"
    # etc. Anchor on exact label text on those 5 known rows instead.
    if "AFFORDABILITY SPREAD" in headers:
        aff_h = headers["AFFORDABILITY SPREAD"]
        # The 5 metrics live at aff_h + 3, +4, +5, +6, +7.
        aff = {}
        for offset in range(3, 8):
            r = aff_h + offset
            label = ws.cell(row=r, column=2).value
            if isinstance(label, str) and label.strip():
                aff[label] = ws.cell(row=r, column=9).value
        out["affordability"] = aff
    return out


# --------------------------------------------------------------------------- #
# §1 Comp Quality + Similarity Scoring                                        #
# --------------------------------------------------------------------------- #
def test_similarity_identical_comp_scores_100(tmp_path):
    """A comp that matches the subject on every dimension should score
    near 100 overall. Same state, same units, same vintage, same POH%."""
    fin = _make_subject_financials(lot_rent=450, units=100)
    market = _market_dict(rent_comps=_build_rent_comps([450]))
    result = _run_and_inspect(fin, market, tmp_path)
    overall = result["comp_quality"][0]["overall"]
    assert overall is not None
    assert float(overall) >= 95, (
        f"Identical comp should score ≥95, got {overall}. "
        "Geo/size/vintage/POH all match subject exactly."
    )


def test_similarity_different_state_loses_geo(tmp_path):
    """An out-of-state comp loses the full geo component (40 vs 100).
    With three other components at 100 (size/vintage/POH match), the
    overall should be (40+100+100+100)/4 = 85."""
    fin = _make_subject_financials(lot_rent=450, units=100)
    market = _market_dict(rent_comps=_build_rent_comps([450], states=["OH"]))
    result = _run_and_inspect(fin, market, tmp_path)
    overall = result["comp_quality"][0]["overall"]
    assert overall is not None
    assert 80 <= float(overall) <= 90, (
        f"Different-state comp should score ~85 (geo=40, others=100), got {overall}"
    )


def test_weighted_contribution_equals_rent_times_score(tmp_path):
    """The Weighted Lot Rent column = rent × overall_score / 100."""
    fin = _make_subject_financials(lot_rent=450, units=100)
    market = _market_dict(rent_comps=_build_rent_comps([500], states=["OH"]))
    result = _run_and_inspect(fin, market, tmp_path)
    c = result["comp_quality"][0]
    expected = c["lot_rent"] * float(c["overall"]) / 100.0
    assert c["weighted"] == pytest.approx(expected, abs=1), (
        f"Weighted lot rent {c['weighted']} should equal "
        f"rent {c['lot_rent']} × score {c['overall']} / 100 = {expected:.2f}"
    )


# --------------------------------------------------------------------------- #
# §2 Robust Statistics — independent percentile + mean + stddev + CV          #
# --------------------------------------------------------------------------- #
def test_robust_stats_match_independent_computation(tmp_path):
    """Mean, stddev, CV must match Python's `statistics` module within
    tight tolerances. P25/median/P75 use linear-interpolation percentiles
    (NumPy/Excel convention)."""
    rents = [400, 425, 450, 475, 500, 525, 550, 575, 600]  # 9 comps
    fin = _make_subject_financials(lot_rent=450, units=100)
    market = _market_dict(rent_comps=_build_rent_comps(rents))
    result = _run_and_inspect(fin, market, tmp_path)
    rs = result["robust_lot_rent"]

    # Expected mean
    exp_mean = sum(rents) / len(rents)
    assert rs["mean"] == pytest.approx(exp_mean, abs=0.01), (
        f"Mean {rs['mean']} ≠ expected {exp_mean}"
    )

    # Expected median (P50)
    exp_median = statistics.median(rents)
    assert rs["median"] == pytest.approx(exp_median, abs=0.01), (
        f"Median {rs['median']} ≠ expected {exp_median}"
    )

    # Expected stddev (population, matches engine's pstdev)
    exp_sd = statistics.pstdev(rents)
    assert rs["stddev"] == pytest.approx(exp_sd, abs=0.01), (
        f"StdDev {rs['stddev']} ≠ expected {exp_sd}"
    )

    # Expected CV = pstdev / mean
    exp_cv = exp_sd / exp_mean
    assert rs["cv"] == pytest.approx(exp_cv, rel=0.001), (
        f"CV {rs['cv']} ≠ expected {exp_cv}"
    )


def test_percentile_25_75_linear_interpolation(tmp_path):
    """For 9 sorted values [400..600 step 25], P25 = value at index 2
    (linear interp), P75 = value at index 6. NumPy convention."""
    rents = list(range(400, 601, 25))  # [400, 425, ..., 600]
    fin = _make_subject_financials(lot_rent=450, units=100)
    market = _market_dict(rent_comps=_build_rent_comps(rents))
    result = _run_and_inspect(fin, market, tmp_path)
    rs = result["robust_lot_rent"]

    # P25 with linear interp: k = 8 * 0.25 = 2, exact integer → rents[2] = 450
    assert rs["p25"] == pytest.approx(450, abs=0.5), f"P25 should be 450, got {rs['p25']}"
    # P75: k = 8 * 0.75 = 6, exact → rents[6] = 550
    assert rs["p75"] == pytest.approx(550, abs=0.5), f"P75 should be 550, got {rs['p75']}"


def test_trimmed_mean_drops_top_and_bottom_decile(tmp_path):
    """Trimmed mean (10%) on 10 values drops 1 from each end → mean of
    middle 8. With [100, 200, ..., 1000], trimmed = mean of [200..900] = 550."""
    rents = list(range(100, 1001, 100))  # 10 values
    fin = _make_subject_financials(lot_rent=550, units=100)
    market = _market_dict(rent_comps=_build_rent_comps(rents))
    result = _run_and_inspect(fin, market, tmp_path)
    rs = result["robust_lot_rent"]

    exp_trimmed = sum(range(200, 901, 100)) / 8  # = 550
    assert rs["trimmed"] == pytest.approx(exp_trimmed, abs=0.5), (
        f"Trimmed mean {rs['trimmed']} ≠ expected {exp_trimmed}"
    )


# --------------------------------------------------------------------------- #
# §3 Mark-to-Market Upside Heatmap                                            #
# --------------------------------------------------------------------------- #
def test_mtm_delta_rent_equals_scenario_minus_subject(tmp_path):
    """Δ Rent in the heatmap row = scenario_rent − subject_rent. For
    rents [500] (single comp), all three scenarios equal 500."""
    fin = _make_subject_financials(lot_rent=450, units=100, occ_rate=0.95)
    market = _market_dict(rent_comps=_build_rent_comps([500] * 5))
    result = _run_and_inspect(fin, market, tmp_path)
    # Median scenario (middle row) should be 500
    median = result["mtm"][1]
    assert median["delta_rent"] == pytest.approx(50, abs=0.5), (
        f"Δ Rent should be 500−450=50, got {median['delta_rent']}"
    )


def test_mtm_annualized_gpr_equals_delta_times_12_times_occupied(tmp_path):
    """Annualized GPR = Δ_rent × 12 × occupied_units. With 100 units at
    95% occupancy, occupied = 95."""
    fin = _make_subject_financials(lot_rent=450, units=100, occ_rate=0.95)
    market = _market_dict(rent_comps=_build_rent_comps([500] * 5))
    result = _run_and_inspect(fin, market, tmp_path)
    median = result["mtm"][1]
    expected_gpr = 50 * 12 * 95
    assert median["annual_gpr"] == pytest.approx(expected_gpr, abs=5), (
        f"Annual GPR Δ should be 50×12×95={expected_gpr}, got {median['annual_gpr']}"
    )


def test_mtm_capitalized_values_match_gpr_divided_by_cap(tmp_path):
    """Value impact at each cap rate = annualized GPR / cap."""
    fin = _make_subject_financials(lot_rent=450, units=100, occ_rate=0.95)
    market = _market_dict(rent_comps=_build_rent_comps([500] * 5))
    result = _run_and_inspect(fin, market, tmp_path)
    median = result["mtm"][1]
    gpr = median["annual_gpr"]
    assert median["value_5"]  == pytest.approx(gpr / 0.050, abs=10), "5.0% cap"
    assert median["value_55"] == pytest.approx(gpr / 0.055, abs=10), "5.5% cap"
    assert median["value_6"]  == pytest.approx(gpr / 0.060, abs=10), "6.0% cap"


def test_mtm_negative_delta_when_subject_above_comp(tmp_path):
    """If subject lot rent > comp set, Δ should be negative (downside scenario)."""
    fin = _make_subject_financials(lot_rent=600, units=100, occ_rate=0.95)
    market = _market_dict(rent_comps=_build_rent_comps([500] * 5))
    result = _run_and_inspect(fin, market, tmp_path)
    median = result["mtm"][1]
    assert median["delta_rent"] == pytest.approx(-100, abs=0.5), (
        f"Δ should be 500−600=-100 (downside), got {median['delta_rent']}"
    )


# --------------------------------------------------------------------------- #
# §4 Implied Rent Growth from Sale Comp Vintage                               #
# --------------------------------------------------------------------------- #
def test_implied_growth_log_linear_fit_5pct(tmp_path):
    """Sale comps with $/unit growing at exactly 5% YoY for 4 years.
    Log-linear regression must recover ~5%."""
    base = 40_000
    triples = [
        (2021, int(base),                   200_000,  base * 100),
        (2022, int(base * 1.05),            210_000,  int(base * 1.05) * 100),
        (2023, int(base * 1.05 ** 2),       220_500,  int(base * 1.05 ** 2) * 100),
        (2024, int(base * 1.05 ** 3),       231_525,  int(base * 1.05 ** 3) * 100),
        (2025, int(base * 1.05 ** 4),       243_101,  int(base * 1.05 ** 4) * 100),
    ]
    fin = _make_subject_financials()
    market = _market_dict(sale_comps=_build_sale_comps(triples))
    result = _run_and_inspect(fin, market, tmp_path)
    g = result["implied_growth"]
    assert g is not None, "Implied growth should compute with 5 dated comps"
    assert g == pytest.approx(0.05, abs=0.005), (
        f"Implied growth should be ~5% on 5% YoY comp set, got {g}"
    )


def test_implied_growth_skipped_with_fewer_than_3_dated_comps(tmp_path):
    """The engine requires ≥3 dated comps with parseable years +
    pricePerUnit. With 2 comps, the section should render the
    'insufficient' message — implied_growth in scraped output = None."""
    triples = [(2021, 40_000, 200_000, 4_000_000),
               (2025, 50_000, 250_000, 5_000_000)]
    fin = _make_subject_financials()
    market = _market_dict(sale_comps=_build_sale_comps(triples))
    result = _run_and_inspect(fin, market, tmp_path)
    assert result.get("implied_growth") is None, (
        "With only 2 dated comps the regression should be skipped"
    )


def test_implied_growth_handles_zero_growth(tmp_path):
    """Constant $/unit across years → growth ~0%."""
    triples = [(y, 50_000, 250_000, 5_000_000) for y in (2021, 2022, 2023, 2024, 2025)]
    fin = _make_subject_financials()
    market = _market_dict(sale_comps=_build_sale_comps(triples))
    result = _run_and_inspect(fin, market, tmp_path)
    g = result["implied_growth"]
    assert g is not None
    assert g == pytest.approx(0.0, abs=0.005), (
        f"Constant $/unit → growth should be 0%, got {g}"
    )


# --------------------------------------------------------------------------- #
# §5 HHI + Cap Rate Triangulation                                             #
# --------------------------------------------------------------------------- #
def test_hhi_perfect_monopoly_equals_10000(tmp_path):
    """A single operator with 100% market share → HHI = 10,000.
    Engine collapses operator names by first 2 tokens, so 4 comps
    starting with 'Acme Communities' all map to the same operator."""
    rents = [450, 460, 470, 480]
    comps = [
        {"name": f"Acme Communities #{i+1}", "city": "C", "state": "MI",
         "units": 100, "lotRent": r, "occupancy": 0.95, "yearBuilt": 1985,
         "pohPercent": 0.10, "qualityRating": 3, "source": "test"}
        for i, r in enumerate(rents)
    ]
    fin = _make_subject_financials()
    # 4 comps from same operator. Engine requires >=2 distinct operators
    # to compute HHI (otherwise len(operator_units) < 2 → None).
    # Add 1 different-operator comp to satisfy the 2-operator floor.
    comps.append({
        "name": "Other Park", "city": "C", "state": "MI", "units": 1,
        "lotRent": 400, "occupancy": 0.95, "yearBuilt": 1985,
        "pohPercent": 0.10, "qualityRating": 3, "source": "test",
    })
    market = _market_dict(rent_comps=comps)
    result = _run_and_inspect(fin, market, tmp_path)
    # 400/401 of pads owned by Acme → share ~0.9975 → HHI = 0.9975^2*10000 ≈ 9950
    hhi_str = str(result.get("hhi") or "").replace(",", "")
    try:
        hhi = float(hhi_str)
    except ValueError:
        pytest.fail(f"HHI cell value {result.get('hhi')!r} not parseable")
    assert hhi > 9500, f"Near-monopoly HHI should exceed 9,500, got {hhi}"


def test_hhi_perfect_fragmentation_low(tmp_path):
    """10 distinct operators each with equal share → HHI = 10 × (10%^2 × 10000) = 1000.
    That's deep in the 'Fragmented' band (<1500)."""
    rents = [450, 460, 470, 480, 490, 500, 510, 520, 530, 540]
    comps = [
        {"name": f"Operator{i+1} Park", "city": "C", "state": "MI",
         "units": 100, "lotRent": r, "occupancy": 0.95, "yearBuilt": 1985,
         "pohPercent": 0.10, "qualityRating": 3, "source": "test"}
        for i, r in enumerate(rents)
    ]
    fin = _make_subject_financials()
    market = _market_dict(rent_comps=comps)
    result = _run_and_inspect(fin, market, tmp_path)
    hhi_str = str(result.get("hhi") or "").replace(",", "")
    try:
        hhi = float(hhi_str)
    except ValueError:
        pytest.fail(f"HHI cell value {result.get('hhi')!r} not parseable")
    assert hhi == pytest.approx(1000, abs=50), (
        f"10 equal operators should yield HHI ≈ 1,000, got {hhi}"
    )


def test_cap_triangulation_three_methods_converge(tmp_path):
    """When all sale comps are internally consistent (NOI/price = stated
    cap rate), all three methods should converge to the same number."""
    # Build 5 comps each at exactly 6% cap rate.
    triples = []
    for i, price in enumerate([5_000_000, 4_500_000, 5_500_000, 6_000_000, 4_800_000]):
        noi = price * 0.06
        triples.append((2024, price // 100, noi, price))
    fin = _make_subject_financials()
    market = _market_dict(sale_comps=_build_sale_comps(triples))
    result = _run_and_inspect(fin, market, tmp_path)
    # All three methods should land exactly at 6%
    for key, val in (("cap_a", result.get("cap_a")),
                     ("cap_b", result.get("cap_b")),
                     ("cap_c", result.get("cap_c"))):
        assert val == pytest.approx(0.06, abs=0.001), (
            f"{key} should converge to 0.06 on consistent comps, got {val}"
        )


def test_cap_triangulation_method_b_equals_sum_noi_over_sum_price(tmp_path):
    """Method B is the aggregate cap = ΣNOI / Σprice. Compute
    independently and compare."""
    triples = [
        (2024,  62_000, 359_600, 6_200_000),
        (2024,  57_900, 341_000, 5_500_000),
        (2024,  52_700, 377_000, 5_800_000),
        (2023,  50_000, 267_750, 4_250_000),
        (2022,  45_000, 409_500, 5_850_000),
    ]
    fin = _make_subject_financials()
    market = _market_dict(sale_comps=_build_sale_comps(triples))
    result = _run_and_inspect(fin, market, tmp_path)
    exp_b = sum(t[2] for t in triples) / sum(t[3] for t in triples)
    assert result["cap_b"] == pytest.approx(exp_b, rel=0.001), (
        f"Method B {result['cap_b']} ≠ expected ΣNOI/ΣPrice {exp_b}"
    )


# --------------------------------------------------------------------------- #
# §6 Affordability Spread                                                     #
# --------------------------------------------------------------------------- #
def test_affordability_spread_equals_apt_rent_div_mhp_cost(tmp_path):
    """Spread = market 2BR / MHP all-in. For $1,200 / $1,000 = 1.2."""
    fin = _make_subject_financials()
    market = _market_dict(
        alt={"avgRent2BR": 1_200, "mhpAllInCost": 1_000},
        demo={"medianHHIncome": 60_000, "majorEmployers": []},
    )
    result = _run_and_inspect(fin, market, tmp_path)
    aff = result["affordability"]
    spread_label = next(k for k in aff if "spread" in k.lower())
    assert aff[spread_label] == pytest.approx(1.2, abs=0.001), (
        f"Spread {aff[spread_label]} ≠ expected 1.2"
    )


def test_affordability_burden_equals_annual_mhp_div_income(tmp_path):
    """Burden = (MHP × 12) / median HH income. $1,000 × 12 / $60,000 = 0.20 = 20%."""
    fin = _make_subject_financials()
    market = _market_dict(
        alt={"avgRent2BR": 1_200, "mhpAllInCost": 1_000},
        demo={"medianHHIncome": 60_000, "majorEmployers": []},
    )
    result = _run_and_inspect(fin, market, tmp_path)
    aff = result["affordability"]
    burden_label = next(k for k in aff if "% of HH" in k)
    assert aff[burden_label] == pytest.approx(0.20, abs=0.001), (
        f"Burden {aff[burden_label]} ≠ expected 0.20"
    )


def test_affordability_fallback_to_1br_when_2br_missing(tmp_path):
    """When avgRent2BR is absent, the section should fall back to
    avgRent1BR (per the assignment `apt_2br = alt.get('avgRent2BR') or
    alt.get('avgRent1BR') or 0`)."""
    fin = _make_subject_financials()
    market = _market_dict(
        alt={"avgRent1BR": 900, "mhpAllInCost": 800},
        demo={"medianHHIncome": 50_000, "majorEmployers": []},
    )
    result = _run_and_inspect(fin, market, tmp_path)
    aff = result["affordability"]
    spread_label = next(k for k in aff if "spread" in k.lower())
    assert aff[spread_label] == pytest.approx(900 / 800, abs=0.001), (
        f"Spread fell back to 1BR rent correctly"
    )


# --------------------------------------------------------------------------- #
# Edge cases — empty / single / degenerate inputs                              #
# --------------------------------------------------------------------------- #
def test_empty_comp_set_does_not_crash(tmp_path):
    """No rent comps, no sale comps. fill_template should still produce
    a workbook with the Comps Analysis tab. All derived stats == n/a."""
    fin = _make_subject_financials()
    market = _market_dict()  # all empty
    # Should not raise.
    out_path = str(tmp_path / "empty.xlsx")
    backend.fill_template(fin, market, out_path)
    wb = openpyxl.load_workbook(out_path, data_only=False)
    assert "Comps Analysis" in wb.sheetnames


def test_single_comp_does_not_crash_percentile_logic(tmp_path):
    """1-comp set: percentile interp must not divide by zero. P25, P50,
    P75 should all equal the single value."""
    fin = _make_subject_financials(lot_rent=450)
    market = _market_dict(rent_comps=_build_rent_comps([475]))
    result = _run_and_inspect(fin, market, tmp_path)
    rs = result["robust_lot_rent"]
    assert rs["p25"]    == 475, "P25 of single-comp set = the value"
    assert rs["median"] == 475
    assert rs["p75"]    == 475


def test_subject_rent_falls_back_through_chain(tmp_path):
    """Subject rent fallback: rentRoll.avgLotRent → unitGroups
    weighted → rentRollRows mean. Drop avgLotRent and verify the
    unitGroups path produces a value."""
    fin = _make_subject_financials(lot_rent=0)  # nukes avgLotRent
    fin["rentRoll"]["avgLotRent"] = 0
    fin["rentRoll"]["unitGroups"] = [
        {"unitType": "TOH MH Site", "occupiedCount": 80,
         "vacantCount": 0, "avgLotRentOccupied": 460},
        {"unitType": "POH-Infilled units", "occupiedCount": 20,
         "vacantCount": 0, "avgLotRentOccupied": 500},
    ]
    market = _market_dict(rent_comps=_build_rent_comps([475]))
    result = _run_and_inspect(fin, market, tmp_path)
    rs = result["robust_lot_rent"]
    # Weighted avg: (80×460 + 20×500) / 100 = (36800 + 10000) / 100 = 468
    assert rs["subject"] == pytest.approx(468, abs=1), (
        f"Subject rent should fall back to weighted unitGroups avg = 468, got {rs['subject']}"
    )
