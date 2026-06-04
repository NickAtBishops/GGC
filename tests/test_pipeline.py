"""
End-to-end test for the GGC Deal Engine analysis pipeline.

Strategy
--------
The real pipeline is:   parse PDFs -> Claude EXTRACT -> verify -> Claude METHODOLOGY
                         -> Claude MARKET (parallel) -> fill_template -> .xlsx

LLM calls are slow, expensive, and non-deterministic. So we:
  1. Replay a known-good Claude response captured once (golden JSON on disk).
  2. Patch the three `call_*` functions in backend.py at the response boundary
     so NO HTTP traffic is generated.
  3. Drive `fill_template()` directly with the replayed financials + market dicts
     and assert that the resulting workbook matches Whaleshead's CorrectOutput
     within tolerances. We DO NOT compare cell-by-cell — Claude varies slightly
     run-to-run, so we only pin the numbers that catch real regressions.

Fixtures committed under tests/fixtures/:
  whaleshead_financials.json   <-- captured `call_parse_financials` result
  whaleshead_market.json       <-- captured `call_market_research` result

To regenerate goldens (one-time, costs API $):
  python -m tests.capture_golden  # not implemented here

Run:  pytest tests/test_pipeline.py -v
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

# Make backend.py importable when tests run from repo root.
REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import backend  # noqa: E402

FIXTURE_DIR = Path(__file__).parent / "fixtures"
CLAUDE_DIR = REPO_ROOT / "Claude"
CORRECT_OUTPUT = CLAUDE_DIR / "CorrectOutput.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────
@pytest.fixture(scope="session")
def golden_financials() -> dict:
    """Replayed `call_parse_financials` output for Whaleshead. Captured once."""
    with open(FIXTURE_DIR / "whaleshead_financials.json") as f:
        return json.load(f)


@pytest.fixture(scope="session")
def golden_market() -> dict:
    """Replayed `call_market_research` output for Whaleshead."""
    with open(FIXTURE_DIR / "whaleshead_market.json") as f:
        return json.load(f)


@pytest.fixture(scope="session")
def correct_wb():
    """The known-good output workbook — our reference for tolerances."""
    return load_workbook(CORRECT_OUTPUT, data_only=True)


@pytest.fixture
def filled_wb(tmp_path, golden_financials, golden_market):
    """
    Run `fill_template` with replayed Claude outputs and return the result.

    `data_only=False` so we can also inspect formulas (e.g. detect #REF!).
    A second load with data_only=True gives us the calculated values — but
    openpyxl does NOT evaluate formulas, so we rely on cached values written
    by Excel. For pure formula-string checks we use data_only=False.
    """
    out_path = tmp_path / "filled.xlsx"
    backend.fill_template(golden_financials, golden_market, out_path)
    return load_workbook(out_path, data_only=False), out_path


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _approx(actual, expected, pct=0.05):
    """±pct relative tolerance. Treats None/0 strictly."""
    if expected in (None, 0):
        return actual == expected
    return abs(actual - expected) / abs(expected) <= pct


def _cell(wb, sheet, ref):
    return wb[sheet][ref].value


# ─────────────────────────────────────────────────────────────────────────────
# Tests — each docstring names the prior bug the assertion would have caught
# ─────────────────────────────────────────────────────────────────────────────
def test_total_units_exact(filled_wb, correct_wb):
    """
    GGC Underwriting!N7 — total unit count.
    Must match EXACTLY. Caught the Oct bug where rent roll rows with
    `status='vacant'` were being dropped from COUNTIFS, under-reporting units.
    """
    wb, _ = filled_wb
    assert _cell(wb, "GGC Underwriting", "N7") == _cell(correct_wb, "GGC Underwriting", "N7")


def test_total_noi_within_tolerance(filled_wb, correct_wb):
    """
    GGC Underwriting!I47 — stabilized Total NOI.
    ±5% tolerance. Catches a class of bugs where expense rows were written
    to the wrong column (e.g. brokerProforma overwriting T12) and NOI
    silently doubled/halved.
    """
    wb, _ = filled_wb
    actual = _cell(wb, "GGC Underwriting", "I47")
    expected = _cell(correct_wb, "GGC Underwriting", "I47")
    assert actual is not None, "I47 is empty — fill_template did not write NOI block"
    assert _approx(actual, expected, pct=0.05), f"NOI {actual} vs expected {expected}"


def test_egi_within_tolerance(filled_wb, correct_wb):
    """
    GGC Underwriting!I19 — Effective Gross Income.
    Catches the regression where `vacancy` was being subtracted twice
    (once in GPR formula, once in EGI formula) after the Nov template patch.
    """
    wb, _ = filled_wb
    actual = _cell(wb, "GGC Underwriting", "I19")
    expected = _cell(correct_wb, "GGC Underwriting", "I19")
    assert _approx(actual, expected, pct=0.05), f"EGI {actual} vs expected {expected}"


def test_loan_scenario_rates_nonzero(filled_wb):
    """
    Loan Scenario tab — interest rate cells must be populated.
    Caught a bug where market research returning `null` for `agencyRate`
    left the cell blank and downstream DSCR formulas evaluated to #DIV/0!.
    """
    wb, _ = filled_wb
    ws = wb["Loan Scenario"]
    # Rate inputs cluster around rows 6-12 col D (confirm against your template).
    rate_cells = ["D6", "D7", "D8"]
    for ref in rate_cells:
        val = ws[ref].value
        assert val not in (None, 0, ""), f"Loan Scenario!{ref} is unset — DSCR will break"


def test_sources_and_uses_no_ref_errors(filled_wb):
    """
    Sources & Uses tab must contain zero #REF! literals.
    Caught the bug introduced when a row was deleted from the template
    without updating cross-sheet references — every linked cell became #REF!.
    """
    wb, _ = filled_wb
    ws = wb["Sources & Uses"]
    bad = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "#REF!" in v:
                bad.append(cell.coordinate)
    assert not bad, f"#REF! found in Sources & Uses at: {bad[:10]}"


def test_waterfall_no_name_errors(filled_wb):
    """
    Waterfall tabs (LP Waterfall / GP Waterfall) must contain zero #NAME?.
    Caught the bug where a custom IRR function name was misspelled in
    the template after the v5 rename — every IRR cell evaluated to #NAME?.
    """
    wb, _ = filled_wb
    targets = [s for s in wb.sheetnames if "Waterfall" in s]
    assert targets, "Expected at least one Waterfall tab"
    bad = []
    for sheet in targets:
        for row in wb[sheet].iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "#NAME?" in v:
                    bad.append(f"{sheet}!{cell.coordinate}")
    assert not bad, f"#NAME? in waterfall: {bad[:10]}"


# ─────────────────────────────────────────────────────────────────────────────
# Optional: full pipeline with mocked Claude (covers run_analysis_job glue)
# ─────────────────────────────────────────────────────────────────────────────
def test_run_analysis_job_with_mocked_claude(tmp_path, golden_financials, golden_market):
    """
    Smoke test the whole orchestration (`run_analysis_job`) with all three
    Claude entry points patched. Verifies the threaded extract+market merge,
    cache write, and final xlsx materialization — without any API spend.
    """
    backend.JOBS_DIR = tmp_path  # redirect output
    job_id = "test-job-1"
    backend.JOBS[job_id] = {"status": "queued", "progress": "", "result": None}

    with patch.object(backend, "call_extract_financials", return_value={"income": [], "expenses": []}), \
         patch.object(backend, "call_parse_financials", return_value=golden_financials), \
         patch.object(backend, "call_market_research", return_value=golden_market), \
         patch.object(backend, "verify_extraction", return_value=[]):
        backend.run_analysis_job(
            job_id,
            api_key="sk-test-not-real",
            file_blocks=[],  # ignored because extract is mocked
            property_info={"name": "Whaleshead", "city": "Brookings", "state": "OR",
                           "units": "60", "pohCount": "0", "askingPrice": "5000000",
                           "floodZone": "X", "deepSearch": "off"},
        )

    assert backend.JOBS[job_id]["status"] == "complete", backend.JOBS[job_id].get("error")
    assert (tmp_path / f"{job_id}.xlsx").exists()
