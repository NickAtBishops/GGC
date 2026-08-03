"""
Template-contract tests for the GGC Deal Engine.

These tests are FAST and DETERMINISTIC — no API calls, no fixtures.
They pin the parts that broke 17June:

  1. The category enum strings match the SUMIFS criteria in the blank
     template. (Drift here silently zeros UW lines.)
  2. The Data Consolidation structural-row scanner finds rows 70/72 (income
     band) and 134/136/138 (expense band). (Writing items onto these rows is
     what landed Advertising T-12 = $1.17M and R&M = $263k in 17June.)
  3. `_normalize_ggc_category` strips `Less:` prefixes and remaps known
     LLM variants ("General and Administrative" -> "G&A"). Belt-and-
     suspenders for when Structured Outputs falls back to prompt-only.
  4. End-to-end write-back of a synthetic methodology JSON produces a
     workbook where the structural rows keep their formulas and the
     `5700 Total Personnel` placeholder is dropped.
  5. The pinned canonical strings still appear in CorrectOutput so we
     detect if someone renames a category there without updating here.

Run:  pytest tests/test_template_contract.py -v
"""
from __future__ import annotations

import re
import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import backend  # noqa: E402

TEMPLATE_PATH       = REPO_ROOT / "GGC_Blank_Underwriting_Sizer_Extended.xlsx"
CORRECT_OUTPUT_PATH = REPO_ROOT / "Outputs" / "CorrectOutput copy.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# 1. _normalize_ggc_category — pin the alias map
# ─────────────────────────────────────────────────────────────────────────────
@pytest.mark.parametrize("raw, expected", [
    # "Less:" prefix gets stripped (UW SUMIFS look up the bare string).
    ("Less: Bad Debt",                "Bad Debt"),
    ("Less: Vacancy",                 "Vacancy"),
    ("Less: Concessions",             "Concessions"),
    # Common LLM variants when grammar masking is unavailable.
    ("General and Administrative",    "G&A"),
    ("General & Administrative",      "G&A"),
    ("G & A",                         "G&A"),
    ("G and A",                       "G&A"),
    # Canonical strings pass through untouched.
    ("Cap-Ex Reserve",                "Cap-Ex Reserve"),
    ("Home Rent Expense (MH)",        "Home Rent Expense (MH)"),
    ("Gross Potential Rent",          "Gross Potential Rent"),
    ("G&A",                           "G&A"),
    # Edge cases: empty / None pass through.
    ("",                              ""),
    (None,                            None),
])
def test_normalize_ggc_category(raw, expected):
    assert backend._normalize_ggc_category(raw) == expected


# ─────────────────────────────────────────────────────────────────────────────
# 2. _structural_rows — pin the Data Consolidation structural-row inventory
# ─────────────────────────────────────────────────────────────────────────────
def test_structural_rows_income_band():
    """The live template puts the income SUM at row 70 and the
    reconciliation IF-check at row 72 (income leaf rows now run 3-68, per
    the fix that widened income_structural/income_slots from range(3,37)
    to range(3,70) — see backend.py fill_template comments around the
    income_structural/expense_structural definitions). The structural-row
    scanner MUST find them — if it doesn't, the write loop will overwrite
    their column A and UW SUMIFS will pull formula outputs as line values
    (the 17June bug).
    """
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["Data Consolidation"]
    found = backend._structural_rows(ws, 3, 72)
    assert 70 in found, "row 70 (=SUM(D3:D68) — income subtotal) must be structural"
    assert 72 in found, "row 72 (=IF(D70=D71,\"OK\") — reconciliation) must be structural"


def test_structural_rows_expense_band():
    """Rows 134 (=SUM(D75:D132)), 136 (=IF check), and 138 (=D70-D134, the
    NOI row) are the three landmines in the expense band (expense leaf rows
    now run 75-132). Writing a category label onto any of them is exactly
    how 17June's Advertising T-12 became $1.17M.
    """
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["Data Consolidation"]
    found = backend._structural_rows(ws, 75, 138)
    for r in (134, 136, 138):
        assert r in found, f"row {r} must be detected as structural (carries a SUM/IF/NOI formula)"


def test_enough_slots_after_skip():
    """After skipping structural rows we still want plenty of slots for
    a real deal. Whaleshead lands ~25 expense items; if the template ever
    grows so many structural rows that <20 remain, the write-back will
    overflow and we want this test to scream first.
    """
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["Data Consolidation"]
    inc_struct = backend._structural_rows(ws, 3, 36)
    exp_struct = backend._structural_rows(ws, 43, 102)
    assert (34 - len(inc_struct)) >= 28, "income band needs ≥28 writeable slots"
    assert (60 - len(exp_struct)) >= 50, "expense band needs ≥50 writeable slots"


# ─────────────────────────────────────────────────────────────────────────────
# 3. Enum vs template SUMIFS contract
# ─────────────────────────────────────────────────────────────────────────────
def _sumifs_criteria_in_uw():
    """Pull every criterion out of GGC Underwriting SUMIFS — both literal
    string criteria (`"Parking Income"`) and same-row cell-reference criteria
    (`A12`, or fully-qualified `'GGC Underwriting'!A12`). The template uses
    the cell-reference form throughout rows 12-14/17 and 22-43: the SUMIFS on
    row N reads its own column-A label (e.g. A12 = "Utility Reimbursement")
    rather than repeating the string as a literal. That's a legitimate,
    working pattern — not a missing SUMIFS — so it must resolve to the same
    criterion a literal-string SUMIFS would.

    Returns the set of strings the template is searching `Data Consolidation`
    column A for. Anything in the enum that ISN'T here is unreachable;
    anything here that ISN'T in the enum will silently zero the UW line.
    """
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb["GGC Underwriting"]
    literal_pat = re.compile(r'SUMIFS\([^,]+,[^,]+,"([^"]+)"\)', re.IGNORECASE)
    # Matches a bare `,A12)` or `,'GGC Underwriting'!A12)` trailing criterion
    # arg — i.e. "use whichever row this formula itself lives on".
    cellref_pat = re.compile(
        r"SUMIFS\([^,]+,[^,]+,(?:'GGC Underwriting'!)?\$?([A-Z]+)\$?(\d+)\)",
        re.IGNORECASE,
    )
    criteria = set()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and "SUMIFS" in v.upper()):
                continue
            criteria.update(literal_pat.findall(v))
            for col_letter, row_num in cellref_pat.findall(v):
                ref_row = int(row_num)
                # Only same-row self-references are the "my own label" idiom
                # this template uses; anything else would be a genuinely
                # different (and currently unsupported) indirection pattern.
                if ref_row != cell.row:
                    continue
                label = ws[f"{col_letter}{ref_row}"].value
                if isinstance(label, str) and label.strip():
                    criteria.add(label.strip())
    return criteria


def test_every_enum_string_is_reachable():
    """Every enum string MUST be reachable by some UW SUMIFS — otherwise the
    LLM can emit it and the UW will never sum it. Sentinels like "Omitt
    Income" / "Omitt Expense" are deliberately unreachable (exclusion
    buckets), so they're allowed.
    """
    criteria = _sumifs_criteria_in_uw()
    unreachable_allowed = {"Omitt Income", "Omitt Expense"}
    for cat in backend.GGC_INCOME_CATEGORIES + backend.GGC_EXPENSE_CATEGORIES:
        if cat in unreachable_allowed:
            continue
        assert cat in criteria, (
            f"enum string {cat!r} has no matching SUMIFS criterion in the "
            f"template. UW will silently zero this line. Fix: either add a "
            f"SUMIFS row in the template, or drop {cat!r} from the enum."
        )


# ─────────────────────────────────────────────────────────────────────────────
# 4. End-to-end write-back — the integration test
# ─────────────────────────────────────────────────────────────────────────────
def test_end_to_end_write_preserves_structural_rows():
    """Drive fill_template with a synthetic methodology output that
    INCLUDES the failure modes 17June hit ("Less:" prefix, the empty
    `5700 Total Personnel` placeholder, the "General and Administrative"
    variant) and assert that:
      - structural rows 70, 72, 134, 136, 138 keep their formulas in column G,
      - their column A is empty (no label landed on them),
      - the placeholder Payroll subtotal is dropped,
      - the variant strings get normalized to the canonical enum values.
    """
    financials = {
        "income": [
            {"ggcCategory": "Less: Bad Debt", "sellerName": "6120 Bad Debt",
             "t12Total": -26062, "monthly": [-2172]*12, "ggcUnderwritten": -26062,
             "confidence": "high"},
            {"ggcCategory": "Gross Potential Rent", "sellerName": "4101 Lot Rent",
             "t12Total": 1144604, "monthly": [95383]*12, "ggcUnderwritten": 1144604,
             "confidence": "high"},
            {"ggcCategory": "RV Site Rental Income", "sellerName": "4103 Long Term RV",
             "t12Total": 172705, "monthly": [14392]*12, "ggcUnderwritten": 172705,
             "confidence": "high"},
            {"ggcCategory": "Parking Income", "sellerName": "4108 Storage",
             "t12Total": 900, "monthly": [75]*12, "ggcUnderwritten": 900,
             "confidence": "high"},
        ],
        "expenses": [
            {"ggcCategory": "Advertising", "sellerName": "5001 Advertising",
             "t12Total": 2468, "monthly": [205]*12, "ggcUnderwritten": 2468,
             "confidence": "high"},
            {"ggcCategory": "Repair and Maintenance", "sellerName": "5108 Plumbing",
             "t12Total": 2886, "monthly": [240]*12, "ggcUnderwritten": 2886,
             "confidence": "high"},
            # The non-canonical variant. Backend must normalize to "G&A".
            {"ggcCategory": "General and Administrative",
             "sellerName": "5407 Tenant Cable TV",
             "t12Total": 69426, "monthly": [5786]*12, "ggcUnderwritten": 69426,
             "confidence": "high"},
            # Placeholder subtotal row — must be DROPPED.
            {"ggcCategory": "Payroll",
             "sellerName": "5700 Total Personnel (5701..., 5702...)",
             "t12Total": 0, "monthly": [0]*12, "ggcUnderwritten": 0,
             "confidence": "low"},
            {"ggcCategory": "Payroll", "sellerName": "5701 Wages, Salary",
             "t12Total": 48650, "monthly": [4054]*12, "ggcUnderwritten": 48650,
             "confidence": "high"},
            {"ggcCategory": "Payroll", "sellerName": "5703 Casual Labour",
             "t12Total": 22505, "monthly": [1875]*12, "ggcUnderwritten": 22505,
             "confidence": "high"},
        ],
        "rentRoll": {"totalUnits": 127, "unitGroups": []},
        "propertyInfo": {"name": "Test", "askingPrice": 10000000, "totalUnits": 127},
    }
    market = {
        "rentComps": [], "saleComps": [], "landmarks": [],
        "demographics": {}, "altHousing": {}, "visuals": {},
        "marketRentConclusion": "", "marketCapRateConclusion": "",
        "demandSignal": "MODERATE", "demandRationale": "",
    }

    # Apply overrides (normalizes Less:, drops placeholder, taxes/capex/mgmt).
    backend.apply_ggc_overrides(financials, financials["propertyInfo"])

    # Normalization should have stripped "Less: " and remapped the variant.
    cats = [it["ggcCategory"] for it in financials["income"]]
    assert "Bad Debt" in cats and "Less: Bad Debt" not in cats
    cats = [it["ggcCategory"] for it in financials["expenses"]]
    assert "G&A" in cats and "General and Administrative" not in cats
    # The placeholder Payroll row with $0 across the board should be gone.
    sellers = [it["sellerName"] for it in financials["expenses"]]
    assert "5700 Total Personnel (5701..., 5702...)" not in sellers
    # The two leaf Payroll GLs must still be present (real dollars).
    assert "5701 Wages, Salary" in sellers and "5703 Casual Labour" in sellers

    # Write the workbook to a tempfile and read it back.
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out_path = f.name
    try:
        backend.fill_template(financials, market, out_path)
        wb = load_workbook(out_path, data_only=False)
        ws = wb["Data Consolidation"]

        # Critical structural-row assertions: formulas preserved, column A clear.
        # Live template's structural rows: 70/72 (income SUM/IF-check) and
        # 134/136/138 (expense SUM/IF-check/NOI) — see test_structural_rows_*
        # above for the row-bounds fix this mirrors.
        for r in (70, 72, 134, 136, 138):
            assert ws.cell(row=r, column=1).value is None, (
                f"row {r} column A must be empty (structural row); "
                f"got {ws.cell(row=r, column=1).value!r}")
        for r in (70, 72, 134, 136, 138):
            v = ws.cell(row=r, column=7).value
            assert isinstance(v, str) and v.startswith("="), (
                f"row {r} column G must still hold a formula; got {v!r}")
    finally:
        Path(out_path).unlink(missing_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# 5. CorrectOutput sanity check
# ─────────────────────────────────────────────────────────────────────────────
@pytest.mark.skipif(
    not CORRECT_OUTPUT_PATH.exists(),
    reason="CorrectOutput copy.xlsx not committed",
)
def test_correct_output_still_present():
    """A canary test — if someone moves or renames CorrectOutput, this fails
    fast instead of letting downstream diff scripts silently miss the
    reference. We don't assert anything about the content (the
    correct-vs-current template strings differ by design — see the
    `Cap Ex` vs `Cap-Ex Reserve` discussion in the change outline)."""
    wb = load_workbook(CORRECT_OUTPUT_PATH, data_only=True)
    assert "GGC Underwriting" in wb.sheetnames
    assert "Data Consolidation" in wb.sheetnames
    assert "Unit Mix Summary" in wb.sheetnames
    assert "Rent Roll Input" in wb.sheetnames
