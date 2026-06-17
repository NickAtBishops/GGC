"""
GGC Deal Engine — Backend Server v3

CHANGES vs. v4:
- Uses GGC's official blank template directly (GGC_Blank_Underwriting_Sizer.xlsx)
- Maps to GGC's exact category names (matching what they use in SUMIFS, including
  the "Electrcitiy" typo which is in their model)
- Adds a NEW "Miscellaneous" tab on top of GGC's template (since their model
  doesn't have one) — preserves Google Maps imagery, landmarks, demographics
- Adds Demographics section per Michael's feedback
- Lists individual rent roll units (one row per unit) instead of flattening counts

Run: python backend.py
Then open http://localhost:5001
"""

import os
import json
import time
import base64
import hashlib
import statistics
import traceback
import re
from pathlib import Path
import threading
from threading import Thread
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter
from io import BytesIO
from urllib.parse import quote_plus
from google.cloud import documentai
from google.api_core.client_options import ClientOptions

import requests
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from dotenv import load_dotenv
from pydantic import BaseModel, Field, ValidationError, field_validator, model_validator

# ═══════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════
ANTHROPIC_API_URL = "https://api.anthropic.com/v1/messages"
# Two-stage financial pipeline — every stage now runs on Claude Fable 5
# (claude-fable-5), Anthropic's most capable tier (above Opus). The north star
# is accuracy and cost is explicitly not the constraint (CLAUDE.md §0/§11).
#   EXTRACTION  → Fable 5, no thinking. Faithful transcription only. Fable 5
#                 removed temperature/top_p/top_k (400 if sent) and rejects an
#                 explicit thinking:"disabled" — call_claude simply omits both
#                 keys. temperature=0 never made hosted LLMs deterministic
#                 anyway (CLAUDE.md §0); run-to-run consistency comes from
#                 structured outputs + self-consistency voting + the
#                 deterministic verifier + the versioned extraction cache.
#   METHODOLOGY → Fable 5, adaptive thinking (effort via THINKING_EFFORT).
#                 Judgment-heavy: GGC categorization, collections, POH
#                 bifurcation, taxes.
#   MARKET      → Fable 5, adaptive thinking + web_search.
# Default to Opus 4.8 — Fable 5 is not yet generally available on most
# Anthropic orgs (Anthropic returns "Claude Fable 5 is not available. Please
# use Opus 4.8."). Opus 4.8 is the most capable Claude model and supports
# Structured Outputs + adaptive thinking + web_search, so it's a clean
# drop-in. Override per-stage via env when Fable 5 access lands:
#   MODEL_EXTRACTION=claude-fable-5 etc. call_claude re-attaches temperature
# only for models that accept it (Opus does; Fable 5 does not).
MODEL_EXTRACTION  = os.environ.get("MODEL_EXTRACTION",  "claude-opus-4-8")
MODEL_METHODOLOGY = os.environ.get("MODEL_METHODOLOGY", "claude-opus-4-8")
MODEL_MARKET      = os.environ.get("MODEL_MARKET",      "claude-opus-4-8")

# Cost-mode model presets. Selected on the upload form per-run and read
# from property_info inside the call sites — beats hard-coded constants
# so users can dial cost down for demos without redeploying.
COST_MODE_MODELS = {
    "economy":  {"extraction": "claude-haiku-4-5",  "methodology": "claude-haiku-4-5",  "market": "claude-haiku-4-5"},
    "balanced": {"extraction": "claude-haiku-4-5",  "methodology": "claude-opus-4-8",   "market": "claude-opus-4-8"},
    "max":      {"extraction": MODEL_EXTRACTION,    "methodology": MODEL_METHODOLOGY,   "market": MODEL_MARKET},
}

def _model_for_stage(property_info, stage):
    """Resolve the model id for a stage. Per-run cost mode wins over the
    module-level default so a single demo can use Haiku while production
    runs use Opus, no redeploy required. stage ∈ {extraction, methodology, market}."""
    mode = (property_info.get("_costMode") or "max").lower()
    preset = COST_MODE_MODELS.get(mode) or COST_MODE_MODELS["max"]
    return preset.get(stage) or COST_MODE_MODELS["max"][stage]
# Thinking depth for adaptive-thinking calls (methodology + market). "high" is
# Anthropic's recommended default for intelligence-sensitive work; "max" trades
# tokens/latency for ceiling accuracy on the hardest deals.
THINKING_EFFORT   = os.environ.get("THINKING_EFFORT", "high")
API_VERSION       = "2023-06-01"
MAX_TOKENS             = 32000  # default; safe for non-thinking calls (Fable 5 caps at 128k, Sonnet 4.6 at 64k)
# Fable 5 with adaptive thinking + effort=high spends most of the budget on
# thinking — a complex deal can burn 30-50k thinking tokens before emitting
# any visible JSON. max_tokens is the COMBINED ceiling for thinking + output,
# so the methodology + market stages need much more headroom than extraction.
# Fable 5 supports up to 128k output via streaming (which we already use).
MAX_TOKENS_METHODOLOGY = 96000  # ~80k thinking headroom + ~16k for JSON
MAX_TOKENS_MARKET      = 64000  # thinking + web_search results + comp tables
MAX_RETRIES       = 6
BASE_BACKOFF_SEC  = 2

# Anthropic Structured Outputs is now GA (verified against
# platform.claude.com/docs/en/build-with-claude/structured-outputs as of
# 2026-06). When a JSON Schema is passed to call_claude(), it's compiled into
# a token grammar and invalid tokens are masked at inference — guarantees
# schema compliance and eliminates the class of variance where Claude emits
# the wrong field types / out-of-enum categories / malformed JSON.
# Confirmed supported (GA list): Fable 5, Sonnet 4.5/4.6, Opus 4.5/4.6/4.7/4.8,
# Haiku 4.5. No beta header required.
# STRUCTURED_OUTPUTS_BETA is kept for backward compatibility with deployments
# pinned to a model snapshot where only the beta path is recognized.
STRUCTURED_OUTPUTS_BETA = "structured-outputs-2025-11-13"
USE_STRUCTURED_OUTPUTS  = os.environ.get("USE_STRUCTURED_OUTPUTS", "1") == "1"
# Default to off — GA path doesn't need it. Flip on if your pinned snapshot
# rejects output_config.format without the legacy beta header.
SO_LEGACY_BETA_HEADER   = os.environ.get("SO_LEGACY_BETA_HEADER", "0") == "1"

# Retries when verify_extraction returns any status=fail. Each retry feeds the
# failure messages back to Claude (Instructor pattern).
MAX_PARSE_RETRIES = int(os.environ.get("MAX_PARSE_RETRIES", "2"))

# N parallel runs for self-consistency voting (Wang et al.). Per CLAUDE.md
# §0 the bar is zero acknowledged accuracy gap, so voting is the DEFAULT —
# not an opt-in deep_search feature. Numeric fields merge by confidence-
# weighted median; categorical (ggcCategory in particular) by confidence-
# weighted mode. Cost is N× tokens per stage, which the §0 policy explicitly
# permits.
FINANCIAL_PARSE_RUNS      = int(os.environ.get("FINANCIAL_PARSE_RUNS",      "3"))
FINANCIAL_PARSE_RUNS_DEEP = int(os.environ.get("FINANCIAL_PARSE_RUNS_DEEP", "5"))
METHODOLOGY_RUNS          = int(os.environ.get("METHODOLOGY_RUNS",          "3"))
METHODOLOGY_RUNS_DEEP     = int(os.environ.get("METHODOLOGY_RUNS_DEEP",     "5"))

# ── Per-job token & cost accounting ──────────────────────────────────────
# Each run_analysis_job spawns its own Thread, so a thread-local accumulator
# attributes every call_claude usage block to the right job without
# threading job_id through every call site. Reset at the start of a job
# (reset_usage_tracking) and snapshot at the end (get_usage_summary).
#
# Prices are per 1M tokens, sourced from anthropic.com/pricing. Cached-read
# input is billed at ~10% of the normal input rate. Cache-write costs the
# same as a normal input token (no premium).
MODEL_PRICING = {
    # model_id_prefix : (input, output, cached_read) per 1M tokens.
    # First prefix match wins (dicts preserve insertion order), so the more
    # specific pins MUST stay above the generic family fallbacks.
    # Verified against anthropic.com/pricing 2026-06: Fable 5 = $10/$50;
    # Opus 4.5-4.8 = $5/$25; the $15/$75 rate only applies to legacy
    # Opus 4.0/4.1, kept as the "claude-opus-4" fallback.
    "claude-fable-5":  (10.00, 50.00, 1.00),
    "claude-opus-4-8": (5.00,  25.00, 0.50),
    "claude-opus-4-7": (5.00,  25.00, 0.50),
    "claude-opus-4-6": (5.00,  25.00, 0.50),
    "claude-opus-4-5": (5.00,  25.00, 0.50),
    "claude-sonnet-4": (3.00, 15.00, 0.30),
    "claude-opus-4":   (15.00, 75.00, 1.50),
    "claude-haiku-4":  (1.00,  5.00,  0.10),
}
_USAGE_LOCAL = threading.local()

def _pricing_for(model_id):
    for prefix, prices in MODEL_PRICING.items():
        if model_id and model_id.startswith(prefix):
            return prices
    return (15.00, 75.00, 1.50)  # default to Opus pricing — overstate, not under

def reset_usage_tracking():
    _USAGE_LOCAL.calls = []

def record_usage(model_id, usage):
    """Append one call's usage block to the current thread's accumulator."""
    calls = getattr(_USAGE_LOCAL, "calls", None)
    if calls is None:
        return  # tracking not enabled for this thread; ignore
    calls.append({
        "model":             model_id or "",
        "input_tokens":      int(usage.get("input_tokens", 0) or 0),
        "output_tokens":     int(usage.get("output_tokens", 0) or 0),
        "cache_read_input_tokens":
            int(usage.get("cache_read_input_tokens", 0) or 0),
        "cache_creation_input_tokens":
            int(usage.get("cache_creation_input_tokens", 0) or 0),
    })

def get_usage_summary():
    """Return aggregated tokens + estimated $ cost for the current thread."""
    calls = getattr(_USAGE_LOCAL, "calls", None) or []
    totals = {"input_tokens": 0, "output_tokens": 0,
              "cache_read_input_tokens": 0,
              "cache_creation_input_tokens": 0, "cost_usd": 0.0}
    per_model = {}
    for c in calls:
        in_price, out_price, cached_price = _pricing_for(c["model"])
        cost = (
            (c["input_tokens"]                  / 1_000_000) * in_price +
            (c["cache_creation_input_tokens"]   / 1_000_000) * in_price +
            (c["cache_read_input_tokens"]       / 1_000_000) * cached_price +
            (c["output_tokens"]                 / 1_000_000) * out_price
        )
        totals["input_tokens"]                += c["input_tokens"]
        totals["output_tokens"]               += c["output_tokens"]
        totals["cache_read_input_tokens"]     += c["cache_read_input_tokens"]
        totals["cache_creation_input_tokens"] += c["cache_creation_input_tokens"]
        totals["cost_usd"]                    += cost
        pm = per_model.setdefault(c["model"], {"calls": 0, "cost_usd": 0.0})
        pm["calls"] += 1
        pm["cost_usd"] += cost
    return {"calls": len(calls), "totals": totals, "per_model": per_model}

# Versioned extraction cache: (file content + property_info + config) hash →
# the full financial_pipeline output. A hit returns the exact same numbers
# every time for previously-seen inputs. Lives in EXTRACTION_CACHE_DIR.
EXTRACTION_CACHE_ENABLED = os.environ.get("EXTRACTION_CACHE_ENABLED", "1") == "1"

# Parser backend: which OCR/layout engine converts PDFs to text for Claude.
# Options: "docai" (Google, current), "azure" (Document Intelligence Layout,
# deterministic), "tensorlake" (vision-first, 91.7% F1 per vendor),
# "reducto" (vision-first, 90.2% on RD-TableBench per vendor).
# Per the 2026 playbook, Doc AI scores 64.6% on complex tables — swap once a
# 50-doc benchmark confirms a better fit. Backend name + version go into the
# extraction cache key, so swapping invalidates the cache.
PARSER_BACKEND = os.environ.get("PARSER_BACKEND", "docai").lower()
PARSER_VERSION = "v1"  # bump when changing parser configuration

GOOGLE_STATIC_MAPS_URL       = "https://maps.googleapis.com/maps/api/staticmap"
GOOGLE_STATIC_STREETVIEW_URL = "https://maps.googleapis.com/maps/api/streetview"

# Secrets only ever come from the environment (or .env in local dev). No
# hardcoded fallbacks — a key committed to source ends up in git history
# forever, which is why CLAUDE.md §12.4 calls out historical leaks as a
# rotate-immediately item.
load_dotenv()

DEFAULT_ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
GOOGLE_MAPS_API_KEY   = os.environ.get("GOOGLE_MAPS_API_KEY", "")

GCP_PROJECT_ID         = os.environ.get("GCP_PROJECT_ID", "")
GCP_LOCATION           = os.environ.get("GCP_LOCATION", "us")
GCP_LAYOUT_PROCESSOR_ID = os.environ.get("GCP_LAYOUT_PROCESSOR_ID", "")
DOC_AI_ENABLED = bool(GCP_PROJECT_ID and GCP_LAYOUT_PROCESSOR_ID)

# Alternate parser backends (Phase 2 of the 2026 playbook). Set the env vars
# for whichever backend you select in PARSER_BACKEND.
AZURE_DOC_INTEL_ENDPOINT = os.environ.get("AZURE_DOC_INTEL_ENDPOINT", "")
AZURE_DOC_INTEL_KEY      = os.environ.get("AZURE_DOC_INTEL_KEY", "")
TENSORLAKE_API_KEY       = os.environ.get("TENSORLAKE_API_KEY", "")
REDUCTO_API_KEY          = os.environ.get("REDUCTO_API_KEY", "")

# IMPORTANT: GGC's official blank template, extended to 1000 rent roll rows
TEMPLATE_PATH        = Path(__file__).parent / "GGC_Blank_Underwriting_Sizer_Extended.xlsx"
# Defaults preserve the historical V5-level dirs for local dev; containers
# (Cloud Run) point these at a writable path like /tmp/ggc/* via env. Note
# Cloud Run's filesystem is in-memory and per-instance — durable copies of
# finished models go to Firebase Storage (see _fb_store_output).
JOBS_DIR             = Path(os.environ.get("JOBS_DIR")             or Path(__file__).parent.parent / "jobs")
IMG_CACHE_DIR        = Path(os.environ.get("IMG_CACHE_DIR")        or Path(__file__).parent.parent / "img_cache")
EXTRACTION_CACHE_DIR = Path(os.environ.get("EXTRACTION_CACHE_DIR") or Path(__file__).parent.parent / "extraction_cache")
JOBS_DIR.mkdir(parents=True, exist_ok=True)
IMG_CACHE_DIR.mkdir(parents=True, exist_ok=True)
EXTRACTION_CACHE_DIR.mkdir(parents=True, exist_ok=True)

# GGC's exact category strings — these MUST match column A in Data
# Consolidation in `CorrectOutput.xlsx` (the gold-standard analyst output),
# because the GGC Underwriting tab's SUMIFS criteria key off these strings
# verbatim. Any drift silently zeros the line. Strings verified against
# CorrectOutput rows 5-32 (income) and 44-101 (expenses).
GGC_INCOME_CATEGORIES = [
    # Core MHC/RV revenue
    "Gross Potential Rent",       # 4101 Lot Rent
    "RV Site Rental Income",      # 4103 Long Term RV Lot Rent
    "Parking Income",             # 4108 Storage Unit Rent
    "Retail",                     # 4110 Retail Unit Rent
    # Recoveries / fees
    "Utility Reimbursement",      # 4403/4404 utility tenant pass-through
    "Other Income",               # 4304/4905/4908-4915 fees, rev-share
    # Adjustments
    "Bad Debt",                   # 6120 (CorrectOutput uses "Bad Debt", NOT "Less: Bad Debt")
    "Omitt Income",               # Non-recurring / Discontinued / Seller-Specific exclusions
    # POH-related (when present)
    "Home Rent Income",           # POH home-rent component
    # Less-common buckets retained for completeness
    "Employee Allowance",
    "Model Units",
]

GGC_EXPENSE_CATEGORIES = [
    # Big-five operating
    "RE Taxes",                   # 5301
    "Insurance",                  # 5053 Liability Insurance (vehicle ins → Omitt Expense)
    "Water and Sewer",            # 5402 + 5403
    "Electricity",                # 5404 (spelling matches gold; template Underwriting label has "Electrcitiy" typo but SUMIFS criterion is "Electricity")
    "Gas/Fuel",                   # 5406 Gas & Propane (5401 Vehicle Fuel → Omitt Expense)
    "Trash Removal",              # 5405
    # Maintenance buckets
    "Ground Maintenance",         # 5102, 5103, 5104
    "Repair and Maintenance",     # 5107-5111, 5200, 5409 (Coin Laundry rentals)
    "Recreational Amenities",
    # People / overhead
    "Management Fee",             # 5000 (GGC override: 5% under 200 sites, 4% at 200+)
    "Payroll",                    # 5700 series — emit ONE ROW PER GL, never the "Total Personnel" subtotal
    "G&A",                        # 5070, 5072, 5407, 5601, 5602, 5603, 5606, 5650 (CorrectOutput uses "G&A", not "General and Administrative")
    "Professional Fees",          # 5061, 5062, 5066
    "Advertising",                # 5001
    # POH-related
    "Home Rent Expense (MH)",
    # Exclusions / reserves
    "Omitt Expense",              # Vehicle (5051, 5401) / Seller-Specific (5605 Postage) / Discontinued
    "Other",
    "Cap-Ex Reserve",             # GGC override: $75/unit/year (gold standard, per CorrectOutput I43)
]

# Canonical unit-type taxonomy. Unit Mix Summary COUNTIFS/SUMIFS in the
# template key on these exact strings — any drift breaks unit counts,
# which cascades to # of Units (Underwriting!N7) and every per-unit
# expense formula downstream.
CANONICAL_UNIT_TYPES = [
    "TOH MH Site",         # Tenant-owned manufactured-home lots
    "POH-Infilled units",  # Park-owned home rentals
    "Long term RV Site",   # Annual RV lots
    "Retail/Commercial",   # Storefronts, commercial space, storage
]

import secrets
from collections import OrderedDict

# Job state. Mutated from the request thread (insert) and the worker
# thread (status / progress / result writes); read from /api/status.
# CPython dict ops are atomic, but read-modify-write on nested values is
# not, so all mutations go through JOBS_LOCK. OrderedDict + LRU
# eviction keeps memory and disk bounded.
JOBS = OrderedDict()
JOBS_LOCK = threading.Lock()
JOBS_MAX = 50  # cap concurrent + recent finished jobs in memory

def _evict_old_jobs():
    """Remove the oldest completed/errored jobs (and their .xlsx files)
    once we exceed JOBS_MAX. Running/queued jobs are skipped — we never
    evict work in flight. Caller MUST already hold JOBS_LOCK."""
    if len(JOBS) <= JOBS_MAX:
        return
    overflow = len(JOBS) - JOBS_MAX
    victims = []
    for jid, job in JOBS.items():
        if job.get("status") in ("complete", "error"):
            victims.append(jid)
            if len(victims) >= overflow:
                break
    for jid in victims:
        JOBS.pop(jid, None)
        try:
            (JOBS_DIR / f"{jid}.xlsx").unlink(missing_ok=True)
        except Exception:
            pass

# Upload guard rails. Catches obvious abuse before parsing.
MAX_UPLOAD_BYTES = 150 * 1024 * 1024  # 150 MB total per request
ALLOWED_UPLOAD_EXTS = {
    ".pdf", ".xlsx", ".xls", ".csv",
    ".png", ".jpg", ".jpeg",
    ".txt", ".md",
}

# Job IDs are 256-bit random tokens — not enumerable by timestamp.
# The validation pattern below is also used as a path-traversal guard.
JOB_ID_RE = re.compile(r"^[A-Za-z0-9_-]{16,64}$")

def _new_job_id():
    return secrets.token_urlsafe(24)

def _valid_job_id(job_id):
    return isinstance(job_id, str) and bool(JOB_ID_RE.match(job_id))

app = Flask(__name__, static_folder=".", static_url_path="")
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES
# Local default stays wide-open for the localhost UI. In hosted mode, set
# ALLOWED_ORIGINS to the Vercel app origin(s), comma-separated, so other
# sites can't script against the engine with a victim's bearer token.
_cors_origins = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "*").split(",") if o.strip()]
CORS(app, origins=_cors_origins if _cors_origins != ["*"] else "*")

# ═══════════════════════════════════════════════════════════════════════════
# FIREBASE — hosted mode: auth + run persistence. Entirely optional:
#   REQUIRE_AUTH=1            → /api/analyze|status|download demand a Firebase
#                               ID token (Authorization: Bearer <token>)
#   ALLOWED_EMAILS=a@x,b@y    → optional allowlist on top of sign-in
#   FIREBASE_STORAGE_BUCKET   → finished models upload to Cloud Storage and
#                               run status mirrors to Firestore `deal_runs`
# Credentials: FIREBASE_SERVICE_ACCOUNT_JSON (one-line JSON) if set, else
# Application Default Credentials (Cloud Run's service account locally via
# GOOGLE_APPLICATION_CREDENTIALS — the file Document AI already uses).
# With neither env var set this whole section is inert and the app behaves
# exactly as it always has on localhost.
# ═══════════════════════════════════════════════════════════════════════════
from functools import wraps
from flask import g

REQUIRE_AUTH            = os.environ.get("REQUIRE_AUTH", "0") == "1"
FIREBASE_STORAGE_BUCKET = os.environ.get("FIREBASE_STORAGE_BUCKET", "")
ALLOWED_EMAILS = {e.strip().lower()
                  for e in os.environ.get("ALLOWED_EMAILS", "").split(",") if e.strip()}

_FB_APP = _FB_DB = _FB_BUCKET = None
if REQUIRE_AUTH or FIREBASE_STORAGE_BUCKET:
    # Hard import on purpose: if hosted mode is requested, fail at boot —
    # never fall open to unauthenticated serving because a dep is missing.
    import firebase_admin
    from firebase_admin import auth as fb_auth
    from firebase_admin import credentials as fb_credentials
    from firebase_admin import firestore as fb_firestore
    from firebase_admin import storage as fb_storage

    _svc = os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON", "").strip()
    _fb_cred = fb_credentials.Certificate(json.loads(_svc)) if _svc else None
    _fb_opts = {"storageBucket": FIREBASE_STORAGE_BUCKET} if FIREBASE_STORAGE_BUCKET else None
    _FB_APP = firebase_admin.initialize_app(_fb_cred, _fb_opts)
    _FB_DB = fb_firestore.client()
    if FIREBASE_STORAGE_BUCKET:
        _FB_BUCKET = fb_storage.bucket()
    print(f"[Firebase] initialized — auth={'ON' if REQUIRE_AUTH else 'off'}, "
          f"firestore=ON, storage={'ON' if _FB_BUCKET else 'off'}, "
          f"allowlist={sorted(ALLOWED_EMAILS) if ALLOWED_EMAILS else 'any signed-in user'}")


def require_auth(fn):
    """Verify the Firebase ID token on a route. No-op unless REQUIRE_AUTH=1,
    so local development needs no Firebase project at all."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not REQUIRE_AUTH:
            g.user_uid, g.user_email = None, None
            return fn(*args, **kwargs)
        hdr = request.headers.get("Authorization", "")
        token = hdr[7:].strip() if hdr.startswith("Bearer ") else ""
        if not token:
            return jsonify({"error": "Sign-in required"}), 401
        try:
            decoded = fb_auth.verify_id_token(token)
        except Exception:
            return jsonify({"error": "Invalid or expired session — sign in again"}), 401
        email = (decoded.get("email") or "").lower()
        if ALLOWED_EMAILS and email not in ALLOWED_EMAILS:
            return jsonify({"error": "This account is not authorized for the GGC Deal Engine"}), 403
        g.user_uid, g.user_email = decoded.get("uid"), email
        return fn(*args, **kwargs)
    return wrapper


def _fb_run_create(job_id, property_info):
    """Create the Firestore run doc when a job is accepted. Best-effort —
    the analysis must never die because telemetry hiccuped."""
    if _FB_DB is None:
        return
    try:
        _FB_DB.collection("deal_runs").document(job_id).set({
            "uid":       getattr(g, "user_uid", None),
            "email":     getattr(g, "user_email", None),
            "name":      property_info.get("name") or "Untitled deal",
            "city":      property_info.get("city", ""),
            "state":     property_info.get("state", ""),
            "status":    "queued",
            "progress":  "Queued",
            "createdAt": fb_firestore.SERVER_TIMESTAMP,
            "updatedAt": fb_firestore.SERVER_TIMESTAMP,
        })
    except Exception as e:
        print(f"[Firebase] Firestore create failed for {job_id}: {e}")


def _fb_run_upsert(job_id, fields):
    """Merge light job-state fields into deal_runs/{job_id}. Best-effort."""
    if _FB_DB is None or not fields:
        return
    try:
        doc = dict(fields)
        doc["updatedAt"] = fb_firestore.SERVER_TIMESTAMP
        _FB_DB.collection("deal_runs").document(job_id).set(doc, merge=True)
    except Exception as e:
        print(f"[Firebase] Firestore upsert failed for {job_id}: {e}")


def _fb_store_output(job_id, output_path):
    """Upload the finished 16-tab model to Storage at runs/{uid}/{job_id}.xlsx
    (the path shape storage.rules grants per-user read on) and record the
    path on the run doc. Best-effort; the local file remains the source the
    engine itself serves via /api/download."""
    if _FB_BUCKET is None:
        return None
    try:
        with JOBS_LOCK:
            uid = (JOBS.get(job_id) or {}).get("uid") or "anon"
        blob_path = f"runs/{uid}/{job_id}.xlsx"
        _FB_BUCKET.blob(blob_path).upload_from_filename(str(output_path))
        _fb_run_upsert(job_id, {"storagePath": blob_path})
        return blob_path
    except Exception as e:
        print(f"[Firebase] Storage upload failed for {job_id}: {e}")
        return None

# ═══════════════════════════════════════════════════════════════════════════
# GOOGLE MAPS HELPERS
# ═══════════════════════════════════════════════════════════════════════════
def fetch_google_static_map(address, map_type="satellite", zoom=17, size="600x400"):
    if not GOOGLE_MAPS_API_KEY:
        return None

    geo = geocode_address(address)
    if not geo:
        # Fallback to address-based
        center_param = address
        cache_key = f"map_{abs(hash(address))}_{map_type}_{zoom}_{size}"
    else:
        center_param = f"{geo['lat']},{geo['lng']}"
        cache_key = f"map_{geo['lat']:.6f}_{geo['lng']:.6f}_{map_type}_{zoom}_{size}"

    cache_path = IMG_CACHE_DIR / f"{cache_key}.png"
    if cache_path.exists():
        return str(cache_path)

    params = {
        "center": center_param,
        "zoom": zoom,
        "size": size,
        "maptype": map_type,
        "markers": f"color:red|{center_param}",
        "key": GOOGLE_MAPS_API_KEY,
    }
    try:
        r = requests.get(GOOGLE_STATIC_MAPS_URL, params=params, timeout=15)
        if r.status_code == 200 and r.headers.get("content-type", "").startswith("image/"):
            cache_path.write_bytes(r.content)
            return str(cache_path)
    except Exception as e:
        print(f"[GoogleMaps] Static map error: {e}")
    return None


GOOGLE_STREETVIEW_METADATA_URL = "https://maps.googleapis.com/maps/api/streetview/metadata"

def fetch_google_streetview(address, heading=0, pitch=0, fov=90, size="600x400",
                             radius=200):
    """
    Fetch a Street View image, intelligently handling cases where the address
    doesn't sit on a road with panorama coverage.

    Strategy:
    1. Geocode the address to lat/lng (with centroid offset for MHCs)
    2. Query the Street View Metadata API to find the nearest panorama
       within `radius` meters
    3. If found, request the image using the panorama's actual lat/lng
       and compute a heading that points BACK TOWARD the property
    4. If no panorama within radius, fall back to address-based request
    """
    if not GOOGLE_MAPS_API_KEY:
        return None

    geo = geocode_address(address)
    if not geo:
        return _fetch_streetview_by_address(address, heading, pitch, fov, size)

    cache_key = f"sv_{geo['lat']:.6f}_{geo['lng']:.6f}_{heading}_{pitch}_{fov}_{size}"
    cache_path = IMG_CACHE_DIR / f"{cache_key}.png"
    if cache_path.exists():
        return str(cache_path)

    # Find the closest panorama within `radius` meters of the geocoded point
    try:
        meta = requests.get(GOOGLE_STREETVIEW_METADATA_URL,
                             params={
                                 "location": f"{geo['lat']},{geo['lng']}",
                                 "radius": radius,
                                 "source": "outdoor",
                                 "key": GOOGLE_MAPS_API_KEY,
                             }, timeout=10).json()
    except Exception as e:
        print(f"[StreetView] Metadata fetch failed: {e}")
        meta = {"status": "ERROR"}

    if meta.get("status") == "OK":
        # The panorama exists. Get the panorama's actual location, then
        # compute a heading that points from the panorama AT the property
        pano_lat = meta["location"]["lat"]
        pano_lng = meta["location"]["lng"]
        bearing_to_property = _bearing(pano_lat, pano_lng, geo["lat"], geo["lng"])

        # The `heading` arg lets the caller request "looking left", "looking
        # right" etc. Treat it as an offset from the property-facing bearing.
        effective_heading = (bearing_to_property + heading) % 360

        params = {
            "location": f"{pano_lat},{pano_lng}",
            "size": size,
            "heading": effective_heading,
            "pitch": pitch,
            "fov": fov,
            "key": GOOGLE_MAPS_API_KEY,
            "source": "outdoor",
        }
        try:
            r = requests.get(GOOGLE_STATIC_STREETVIEW_URL, params=params, timeout=15)
            if r.status_code == 200 and r.headers.get("content-type", "").startswith("image/") and len(r.content) > 5000:
                cache_path.write_bytes(r.content)
                return str(cache_path)
        except Exception as e:
            print(f"[StreetView] Image fetch failed: {e}")

    # Fallback: address-based, original behavior — better than nothing
    return _fetch_streetview_by_address(address, heading, pitch, fov, size)


def _fetch_streetview_by_address(address, heading, pitch, fov, size):
    """Original fallback path — used when geocoding or metadata fails."""
    cache_key = f"sv_addr_{abs(hash(address))}_{heading}_{pitch}_{fov}_{size}"
    cache_path = IMG_CACHE_DIR / f"{cache_key}.png"
    if cache_path.exists():
        return str(cache_path)
    params = {"location": address, "size": size, "heading": heading,
              "pitch": pitch, "fov": fov, "key": GOOGLE_MAPS_API_KEY,
              "source": "outdoor"}
    try:
        r = requests.get(GOOGLE_STATIC_STREETVIEW_URL, params=params, timeout=15)
        if r.status_code == 200 and r.headers.get("content-type", "").startswith("image/"):
            if len(r.content) < 5000:
                return None
            cache_path.write_bytes(r.content)
            return str(cache_path)
    except Exception as e:
        print(f"[StreetView] Error: {e}")
    return None


def _bearing(lat1, lng1, lat2, lng2):
    """Calculate compass bearing in degrees from point 1 to point 2."""
    import math
    lat1_r, lat2_r = math.radians(lat1), math.radians(lat2)
    diff = math.radians(lng2 - lng1)
    y = math.sin(diff) * math.cos(lat2_r)
    x = math.cos(lat1_r) * math.sin(lat2_r) - math.sin(lat1_r) * math.cos(lat2_r) * math.cos(diff)
    return (math.degrees(math.atan2(y, x)) + 360) % 360


def embed_image_in_cell(ws, image_path, anchor_cell, width_px=400, height_px=280):
    if not image_path or not Path(image_path).exists():
        return False
    try:
        img = XLImage(image_path)
        img.width = width_px
        img.height = height_px
        img.anchor = anchor_cell
        ws.add_image(img)
        return True
    except Exception as e:
        print(f"[Excel] Image embed failed: {e}")
        return False


def to_decimal_pct(value):
    """Coerce a percentage-ish value to its DECIMAL form for Excel cells
    formatted as ``0.00%``. The LLM is inconsistent: it may return 7.5 to
    mean 7.5% (percent form) or 0.075 (decimal form). Excel's percent format
    multiplies by 100, so a percent-form 7.5 would render as 750%.

    Heuristic: anything with abs(v) < 1.5 is already decimal (0.075 → 0.075).
    Anything larger is presumed percent form and divided by 100 (7.5 → 0.075).
    Returns None if the value can't be parsed.
    """
    if value is None or value == "":
        return None
    if isinstance(value, str):
        s = value.strip().rstrip("%").strip()
        try:
            value = float(s)
        except (TypeError, ValueError):
            return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    return v if abs(v) < 1.5 else v / 100


def safe_pct(value):
    """
    Format a value as a percentage string, regardless of whether Claude returned:
      - None              → ""
      - a number 0.045    → "4.5%"
      - a number 4.5      → "4.5%"  (already in percent form)
      - a string "4.5%"   → "4.5%"  (passthrough)
      - a string "4.5"    → "4.5%"

    Routes the decimal-vs-percent disambiguation through to_decimal_pct
    so this helper and to_decimal_pct share a single threshold. Previously
    safe_pct used abs(v) < 1 while to_decimal_pct used abs(v) < 1.5, which
    meant the two helpers disagreed for values in [1.0, 1.5) — the same
    raw LLM output could render as 1.2% in one Excel cell and 120% in
    another. Now both helpers route through the same heuristic.
    """
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        s = value.strip()
        if s.endswith("%"):
            return s
    decimal = to_decimal_pct(value)
    if decimal is None:
        return str(value) if not isinstance(value, str) else value
    return f"{decimal * 100:.1f}%"


def safe_money(value, suffix=""):
    """Format a value as $X,XXX. Handles strings like '$78,000' or numbers or None."""
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        s = value.strip().replace("$", "").replace(",", "").replace("/mo", "").strip()
        if s.endswith("%"):
            return value  # weird input, return as-is
        try:
            value = float(s)
        except ValueError:
            return value  # not numeric, return as-is
    try:
        return f"${float(value):,.0f}{suffix}"
    except (TypeError, ValueError):
        return str(value)

# ═══════════════════════════════════════════════════════════════════════════
# PDF PARSER ABSTRACTION
# Single entry point `parse_pdf(bytes, filename) → markdown | None`.
# Dispatches by PARSER_BACKEND. Parser name + version are baked into the
# cache filename so swapping backends or bumping config invalidates stale
# parses (the file SHA stays stable so identical bytes hit the cache).
#
# Backends:
#   docai      — Google Document AI Layout Parser (current default).
#                Per RD-TableBench Nov 2024, scored 64.6% on complex tables
#                — lowest of the major cloud providers.
#   azure      — Azure Document Intelligence Layout. ~82.7% on the same
#                benchmark; deterministic given identical input.
#   tensorlake — Vision-first parser. 91.7% F1 per vendor (Nov 2025).
#   reducto    — Vision-first parser. 90.2% on RD-TableBench per vendor.
# ═══════════════════════════════════════════════════════════════════════════

def _stable_pdf_hash(pdf_bytes):
    return hashlib.sha256(pdf_bytes).hexdigest()[:16]


def parse_pdf(pdf_bytes, filename):
    """
    Parse a PDF using whichever backend PARSER_BACKEND names. Returns the
    markdown string on success, None on failure (caller falls back to
    Anthropic's native PDF handling).
    """
    backend = PARSER_BACKEND
    cache_key  = f"{backend}_{PARSER_VERSION}_{_stable_pdf_hash(pdf_bytes)}"
    cache_path = IMG_CACHE_DIR / f"{cache_key}.md"
    if cache_path.exists():
        print(f"[Parser:{backend}] Cache hit for {filename}")
        return cache_path.read_text()

    try:
        if backend == "docai":
            markdown = _parse_via_docai(pdf_bytes, filename)
        elif backend == "azure":
            markdown = _parse_via_azure(pdf_bytes, filename)
        elif backend == "tensorlake":
            markdown = _parse_via_tensorlake(pdf_bytes, filename)
        elif backend == "reducto":
            markdown = _parse_via_reducto(pdf_bytes, filename)
        else:
            print(f"[Parser] Unknown PARSER_BACKEND='{backend}' — "
                  "supported: docai, azure, tensorlake, reducto")
            return None
    except Exception as e:
        print(f"[Parser:{backend}] Failed for {filename}: {e}")
        traceback.print_exc()
        return None

    if markdown:
        # Diagnostic hash — if this changes across runs of the same file,
        # the parser itself is non-deterministic (and Claude variance has
        # nothing to do with run-to-run drift)
        md_hash = hashlib.md5(markdown.encode()).hexdigest()[:12]
        print(f"[Parser:{backend}] Markdown hash for {filename}: "
              f"{md_hash} ({len(markdown)} chars)")
        cache_path.write_text(markdown)
    return markdown


# Backward-compat alias — encode_file_for_claude callers and any external
# wrappers expecting the old name still work.
parse_pdf_with_document_ai = parse_pdf


def _parse_via_docai(pdf_bytes, filename):
    """Google Document AI Layout Parser — markdown with tables/headers."""
    if not DOC_AI_ENABLED:
        print("[Parser:docai] GCP_PROJECT_ID / GCP_LAYOUT_PROCESSOR_ID not "
              "set — skipping Doc AI")
        return None

    opts = ClientOptions(api_endpoint=f"{GCP_LOCATION}-documentai.googleapis.com")
    client = documentai.DocumentProcessorServiceClient(client_options=opts)
    processor_name = (
        f"projects/{GCP_PROJECT_ID}/locations/{GCP_LOCATION}"
        f"/processors/{GCP_LAYOUT_PROCESSOR_ID}"
    )
    raw_document = documentai.RawDocument(content=pdf_bytes, mime_type="application/pdf")
    process_options = documentai.ProcessOptions(
        layout_config=documentai.ProcessOptions.LayoutConfig(
            chunking_config=documentai.ProcessOptions.LayoutConfig.ChunkingConfig(
                chunk_size=1000,
                include_ancestor_headings=True,
            )
        )
    )
    request = documentai.ProcessRequest(
        name=processor_name,
        raw_document=raw_document,
        process_options=process_options,
    )

    print(f"[Parser:docai] Parsing {filename} ({len(pdf_bytes)//1024} KB)...")
    t0 = time.time()
    result = client.process_document(request=request)
    print(f"[Parser:docai] Parsed in {time.time() - t0:.1f}s")

    doc = result.document
    parts = [f"# Document: {filename}\n"]
    if doc.chunked_document and doc.chunked_document.chunks:
        for chunk in doc.chunked_document.chunks:
            if chunk.page_headers:
                for hdr in chunk.page_headers:
                    parts.append(f"\n## {hdr.text}\n")
            parts.append(chunk.content)
            parts.append("\n")
    else:
        parts.append(doc.text)
    return "\n".join(parts)


def _parse_via_azure(pdf_bytes, filename):
    """
    Azure Document Intelligence Layout (prebuilt-layout model) — returns
    markdown when outputContentFormat=markdown is requested. Async: POST
    returns 202 with an operation-location header; poll until done.

    Verified 2026-06 against
      learn.microsoft.com/en-us/azure/ai-services/document-intelligence/
    REST API v4.0 (api-version=2024-11-30) is the current GA. v4.0 changed
    table representation to HTML inside the markdown stream to support
    merged cells + multirow headers, which is desirable for T12s.
    """
    if not AZURE_DOC_INTEL_ENDPOINT or not AZURE_DOC_INTEL_KEY:
        print("[Parser:azure] AZURE_DOC_INTEL_ENDPOINT / AZURE_DOC_INTEL_KEY "
              "not set — skipping")
        return None

    base = AZURE_DOC_INTEL_ENDPOINT.rstrip("/")
    api_version = os.environ.get("AZURE_DOC_INTEL_API_VERSION", "2024-11-30")
    start_url = (
        f"{base}/documentintelligence/documentModels/"
        f"prebuilt-layout:analyze"
        f"?api-version={api_version}&outputContentFormat=markdown"
    )
    headers = {
        "Ocp-Apim-Subscription-Key": AZURE_DOC_INTEL_KEY,
        "Content-Type": "application/pdf",
    }

    print(f"[Parser:azure] Parsing {filename} ({len(pdf_bytes)//1024} KB)...")
    t0 = time.time()
    r = requests.post(start_url, headers=headers, data=pdf_bytes, timeout=120)
    if r.status_code not in (200, 202):
        print(f"[Parser:azure] start failed: {r.status_code} {r.text[:300]}")
        return None
    op_url = r.headers.get("operation-location") or r.headers.get("Operation-Location")
    if not op_url:
        print(f"[Parser:azure] no operation-location header: {dict(r.headers)}")
        return None

    poll_headers = {"Ocp-Apim-Subscription-Key": AZURE_DOC_INTEL_KEY}
    deadline = time.time() + 300
    while time.time() < deadline:
        time.sleep(2)
        pr = requests.get(op_url, headers=poll_headers, timeout=30)
        if pr.status_code != 200:
            print(f"[Parser:azure] poll {pr.status_code}: {pr.text[:200]}")
            return None
        data = pr.json()
        status = data.get("status")
        if status == "succeeded":
            print(f"[Parser:azure] Parsed in {time.time() - t0:.1f}s")
            content = data.get("analyzeResult", {}).get("content", "")
            return f"# Document: {filename}\n\n{content}"
        if status in ("failed", "canceled"):
            print(f"[Parser:azure] terminal status {status}: {data}")
            return None
    print("[Parser:azure] polling timed out (5 min)")
    return None


def _parse_via_tensorlake(pdf_bytes, filename):
    """
    Tensorlake document parsing (v2 async) per docs.tensorlake.ai/api-reference
    (verified 2026-06):
      1) POST /documents/v2/files (multipart) → {"file_id": "..."}
         NOTE: the upload endpoint URL is the one piece NOT explicitly shown
         in the public docs preview — verify against your account's actual
         endpoint or use the official Tensorlake Python SDK for upload.
         Override via TENSORLAKE_UPLOAD_PATH if needed.
      2) POST /documents/v2/parse with {"file_id", "mime_type",
         "parsing_options": {"table_output_mode": "markdown"}} → {"parse_id"}
      3) GET /documents/v2/parse/{parse_id} until status="successful"
      4) Markdown is the joined chunks[].content
    Pricing: $0.01/page flat ($10/1k pages).
    """
    if not TENSORLAKE_API_KEY:
        print("[Parser:tensorlake] TENSORLAKE_API_KEY not set — skipping")
        return None

    base = os.environ.get("TENSORLAKE_BASE_URL", "https://api.tensorlake.ai").rstrip("/")
    upload_path = os.environ.get("TENSORLAKE_UPLOAD_PATH", "/documents/v2/files")
    auth = {"Authorization": f"Bearer {TENSORLAKE_API_KEY}"}

    # Step 1 — upload, get file_id
    print(f"[Parser:tensorlake] Uploading {filename} "
          f"({len(pdf_bytes)//1024} KB)...")
    t0 = time.time()
    up = requests.post(f"{base}{upload_path}", headers=auth,
                       files={"file": (filename, pdf_bytes, "application/pdf")},
                       timeout=300)
    if up.status_code not in (200, 201):
        print(f"[Parser:tensorlake] upload {up.status_code}: {up.text[:300]} "
              f"— if 404/405 the upload endpoint differs; set "
              "TENSORLAKE_UPLOAD_PATH or use the Tensorlake SDK")
        return None
    try:
        file_id = (up.json().get("file_id") or up.json().get("id"))
    except json.JSONDecodeError:
        print(f"[Parser:tensorlake] upload non-JSON: {up.text[:200]}")
        return None
    if not file_id:
        print(f"[Parser:tensorlake] upload missing file_id: {up.text[:200]}")
        return None

    # Step 2 — submit parse job
    body = {
        "file_id":   file_id,
        "mime_type": "application/pdf",
        "parsing_options": {"table_output_mode": "markdown"},
    }
    sub = requests.post(f"{base}/documents/v2/parse",
                        headers={**auth, "Content-Type": "application/json"},
                        json=body, timeout=120)
    if sub.status_code not in (200, 201, 202):
        print(f"[Parser:tensorlake] parse submit {sub.status_code}: {sub.text[:300]}")
        return None
    try:
        parse_id = sub.json().get("parse_id") or sub.json().get("id")
    except json.JSONDecodeError:
        print(f"[Parser:tensorlake] submit non-JSON: {sub.text[:200]}")
        return None
    if not parse_id:
        print(f"[Parser:tensorlake] submit missing parse_id: {sub.text[:200]}")
        return None

    # Step 3 — poll until successful (or timeout at 5 min)
    poll_url = f"{base}/documents/v2/parse/{parse_id}"
    deadline = time.time() + 300
    while time.time() < deadline:
        time.sleep(3)
        pr = requests.get(poll_url, headers=auth, timeout=30)
        if pr.status_code != 200:
            print(f"[Parser:tensorlake] poll {pr.status_code}: {pr.text[:200]}")
            return None
        data = pr.json()
        status = (data.get("status") or "").lower()
        if status in ("successful", "succeeded", "completed", "complete"):
            chunks = (data.get("chunks") or data.get("result", {}).get("chunks")
                      or [])
            parts = [ch.get("content", "") for ch in chunks if ch.get("content")]
            if not parts:
                print(f"[Parser:tensorlake] no chunks in result; keys="
                      f"{list(data.keys())}")
                return None
            print(f"[Parser:tensorlake] Parsed in {time.time() - t0:.1f}s "
                  f"({len(parts)} chunks)")
            return f"# Document: {filename}\n\n" + "\n\n".join(parts)
        if status in ("failed", "error", "errored", "canceled"):
            print(f"[Parser:tensorlake] terminal status {status}: {data}")
            return None
    print("[Parser:tensorlake] polling timed out (5 min)")
    return None


def _parse_via_reducto(pdf_bytes, filename):
    """
    Reducto vision-first parser. Two-step synchronous flow per
    docs.reducto.ai/quickstart (verified 2026-06):
      1) POST /upload (multipart) → {"file_id": "reducto://..."}
      2) POST /parse (JSON, input=file_id) → {"result": {"chunks":[{"content":...}]}}
    Markdown is the concatenated chunks[].content. Published pricing
    starts at $0.015/page.
    """
    if not REDUCTO_API_KEY:
        print("[Parser:reducto] REDUCTO_API_KEY not set — skipping")
        return None

    base = os.environ.get("REDUCTO_BASE_URL", "https://platform.reducto.ai").rstrip("/")
    auth = {"Authorization": f"Bearer {REDUCTO_API_KEY}"}

    # Step 1 — upload bytes, get file_id
    print(f"[Parser:reducto] Uploading {filename} "
          f"({len(pdf_bytes)//1024} KB)...")
    t0 = time.time()
    up = requests.post(f"{base}/upload", headers=auth,
                       files={"file": (filename, pdf_bytes, "application/pdf")},
                       timeout=300)
    if up.status_code not in (200, 201):
        print(f"[Parser:reducto] upload {up.status_code}: {up.text[:300]}")
        return None
    try:
        file_id = up.json().get("file_id")
    except json.JSONDecodeError:
        print(f"[Parser:reducto] upload returned non-JSON: {up.text[:200]}")
        return None
    if not file_id:
        print(f"[Parser:reducto] upload response missing file_id: {up.text[:200]}")
        return None

    # Step 2 — parse with markdown table formatting
    body = {"input": file_id, "formatting": {"table_output_format": "md"}}
    pr = requests.post(f"{base}/parse",
                       headers={**auth, "Content-Type": "application/json"},
                       json=body, timeout=600)
    if pr.status_code not in (200, 201):
        print(f"[Parser:reducto] parse {pr.status_code}: {pr.text[:300]}")
        return None

    try:
        payload = pr.json()
    except json.JSONDecodeError:
        print(f"[Parser:reducto] parse non-JSON: {pr.text[:200]}")
        return None

    chunks = (payload.get("result") or {}).get("chunks") or payload.get("chunks") or []
    parts = []
    for ch in chunks:
        # Per docs, ch.content is the full text; ch.blocks[] is per-element.
        # Prefer content; fall back to joined block contents.
        txt = ch.get("content")
        if not txt and ch.get("blocks"):
            txt = "\n\n".join(b.get("content", "") for b in ch["blocks"]
                              if b.get("content"))
        if txt:
            parts.append(txt)

    if not parts:
        print(f"[Parser:reducto] parse response had no chunks: "
              f"keys={list(payload.keys())}")
        return None

    print(f"[Parser:reducto] Parsed in {time.time() - t0:.1f}s "
          f"({payload.get('usage', {}).get('num_pages', '?')} pages)")
    return f"# Document: {filename}\n\n" + "\n\n".join(parts)

# ═══════════════════════════════════════════════════════════════════════════
# VERSIONED EXTRACTION CACHE
# Key = SHA-256 of (file_blocks + property_info + model pins + parser config +
# prompt hashes + schema hashes + n_runs). On a hit, returns the exact same
# financial_pipeline output that was produced last time the same inputs were
# seen. This is the operational fix for the "same PDF produces different
# output" complaint — fingerprinted memoization.
# Bumping any prompt, schema, or model invalidates all entries automatically
# (their hash changes). To force-clear, delete EXTRACTION_CACHE_DIR.
# ═══════════════════════════════════════════════════════════════════════════

def _hash_obj(obj):
    """Stable SHA-256 hex digest of any JSON-serializable object."""
    return hashlib.sha256(
        json.dumps(obj, sort_keys=True, default=str).encode()
    ).hexdigest()


def extraction_cache_key(file_blocks, property_info, n_extraction_runs,
                         extraction_prompt, methodology_prompt,
                         n_methodology_runs=1):
    """
    Build the cache key tuple, then hash it. Anything that changes the output
    must be in here. PARSER_BACKEND is in via PARSER_VERSION + the fact that
    PARSER_BACKEND affects what's *in* file_blocks (the parser's markdown is
    embedded directly).
    """
    key_obj = {
        "files":        _hash_obj(file_blocks),
        "property":     _hash_obj(property_info),
        "model_x":      MODEL_EXTRACTION,
        "model_m":      MODEL_METHODOLOGY,
        "use_so":       USE_STRUCTURED_OUTPUTS,
        "parser":       PARSER_BACKEND,
        "parser_v":     PARSER_VERSION,
        "ext_prompt":   _hash_obj(extraction_prompt)[:16],
        "meth_prompt":  _hash_obj(methodology_prompt)[:16],
        "ext_schema":   _hash_obj(EXTRACTION_OUTPUT_SCHEMA)[:16],
        "meth_schema":  _hash_obj(METHODOLOGY_OUTPUT_SCHEMA)[:16],
        "n_runs":       n_extraction_runs,
        "n_meth_runs":  n_methodology_runs,
    }
    return _hash_obj(key_obj)[:32]


def extraction_cache_get(key):
    if not EXTRACTION_CACHE_ENABLED:
        return None
    path = EXTRACTION_CACHE_DIR / f"{key}.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text())
    except Exception as e:
        print(f"[ExtractionCache] read failed for {key[:8]}: {e}")
        return None


def extraction_cache_put(key, payload):
    if not EXTRACTION_CACHE_ENABLED:
        return
    path = EXTRACTION_CACHE_DIR / f"{key}.json"
    try:
        path.write_text(json.dumps(payload))
    except Exception as e:
        print(f"[ExtractionCache] write failed for {key[:8]}: {e}")


GOOGLE_GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"

def geocode_address(address):
    """
    Resolve an address to its best lat/lng, using the property centroid
    when the geocoder returns a bounding box (common for MHCs and complexes).
    Returns dict with 'lat', 'lng', 'precision', and 'place_id'.
    """
    if not GOOGLE_MAPS_API_KEY or not address:
        return None

    cache_key = f"geo_{abs(hash(address))}"
    cache_path = IMG_CACHE_DIR / f"{cache_key}.json"
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text())
        except Exception:
            pass

    try:
        r = requests.get(GOOGLE_GEOCODE_URL,
                          params={"address": address, "key": GOOGLE_MAPS_API_KEY},
                          timeout=15)
        if r.status_code != 200:
            return None
        data = r.json()
        if data.get("status") != "OK" or not data.get("results"):
            print(f"[Geocode] No result for {address}: {data.get('status')}")
            return None

        result = data["results"][0]
        loc = result["geometry"]["location"]
        precision = result["geometry"].get("location_type", "UNKNOWN")

        # If the geocoder returns a bounding box (common for MHCs, apartment
        # complexes, business parks), prefer the box centroid over the snap
        # point — it's typically further inside the property
        bounds = result["geometry"].get("bounds") or result["geometry"].get("viewport")
        if bounds:
            ne = bounds["northeast"]
            sw = bounds["southwest"]
            centroid_lat = (ne["lat"] + sw["lat"]) / 2
            centroid_lng = (ne["lng"] + sw["lng"]) / 2
            # Only use centroid if it's reasonably close to the geocoder pin
            # (sanity check — sometimes viewports are huge city-level boxes)
            lat_diff = abs(centroid_lat - loc["lat"])
            lng_diff = abs(centroid_lng - loc["lng"])
            if lat_diff < 0.005 and lng_diff < 0.005:  # ~500m
                loc = {"lat": centroid_lat, "lng": centroid_lng}

        out = {
            "lat": loc["lat"],
            "lng": loc["lng"],
            "precision": precision,
            "place_id": result.get("place_id"),
            "formatted_address": result.get("formatted_address"),
        }
        cache_path.write_text(json.dumps(out))
        return out
    except Exception as e:
        print(f"[Geocode] Error for {address}: {e}")
        return None

# ═══════════════════════════════════════════════════════════════════════════
# CLAUDE API CLIENT
# ═══════════════════════════════════════════════════════════════════════════
def _accepts_sampling(model_id):
    """True if the model still accepts temperature/top_p/top_k. Fable 5 and
    Opus 4.7+ removed sampling params entirely (the API returns 400 if any is
    sent). Allowlist the families that DO accept them so an env override back
    to e.g. Sonnet 4.6 automatically regains its temperature=0 config."""
    mid = (model_id or "").lower()
    if mid.startswith(("claude-sonnet-", "claude-haiku-", "claude-3")):
        return True
    # Opus 4.6 and earlier accept temperature; Opus 4.7+ and Fable 5 do not.
    return mid.startswith(("claude-opus-4-0", "claude-opus-4-1",
                           "claude-opus-4-5", "claude-opus-4-6",
                           "claude-opus-4-2"))  # ...-4-2 = claude-opus-4-20250514


def _supports_adaptive_thinking(model_id):
    """True if the model accepts `thinking: {type: 'adaptive'}`. Haiku models
    don't expose extended thinking (Anthropic returns 400 with
    `adaptive thinking is not supported on this model`). Opus 4.x and Fable 5
    do. Used to silently disable thinking when an Economy-mode run lands on
    Haiku, instead of crashing the analysis."""
    mid = (model_id or "").lower()
    if mid.startswith("claude-haiku-"):
        return False
    if mid.startswith(("claude-opus-", "claude-fable-", "claude-sonnet-")):
        return True
    return True  # Default to true so new model families don't silently lose thinking.


# Per-model max-output-tokens cap. Anthropic rejects requests above each
# model's hard ceiling (400 with `max_tokens: N > LIMIT`). The methodology
# stage requests 96K to leave headroom for thinking on Opus/Fable 5, which
# is fine for them but blows past Haiku 4.5's 64K cap. Clamp at the call
# site so a single Economy-mode run doesn't crash.
_MODEL_MAX_OUTPUT_TOKENS = {
    "claude-haiku-":  64_000,
    "claude-sonnet-": 64_000,
    "claude-opus-":  128_000,
    "claude-fable-": 128_000,
}

def _clamp_max_tokens(model_id, requested):
    mid = (model_id or "").lower()
    for prefix, cap in _MODEL_MAX_OUTPUT_TOKENS.items():
        if mid.startswith(prefix):
            return min(int(requested), cap)
    return int(requested)


def call_claude(api_key, system_prompt, user_content, tools=None,
                use_thinking=True, temperature=None, model=None,
                output_schema=None, max_tokens=None):
    """
    Call Claude with streaming enabled. Streaming keeps the connection alive
    during long-running requests (which can hit 3+ minutes when Claude is doing
    heavy thinking + web search) instead of timing out at the request level.

    Model routing:
    - Defaults to MODEL_MARKET (Fable 5) for market research with adaptive thinking
    - Pass model=MODEL_METHODOLOGY (Fable 5) for GGC categorization + methodology
    - Pass model=MODEL_EXTRACTION with use_thinking=False for document
      extraction. temperature is only attached when the target model still
      accepts sampling params (_accepts_sampling) — Fable 5 and Opus 4.7+
      reject temperature/top_p/top_k AND an explicit thinking:"disabled",
      so for them a no-thinking call simply omits both keys (omission is the
      documented way to run Fable 5 without thinking).
    """
    headers = {"x-api-key": api_key, "anthropic-version": API_VERSION,
               "content-type": "application/json"}
    # Wrap the system prompt in Anthropic's prompt-cache marker. The
    # FINANCIAL_PARSE_PROMPT is ~50K tokens of static methodology that
    # changes only when the codebase is redeployed; cached input is billed
    # at ~10% of the normal rate after the first hit (5-minute TTL). This
    # cuts the dominant Opus call's input cost ~80-90% on cache hits.
    # Strings shorter than the cache-eligibility floor are passed through
    # in their normal form to avoid wasting a cache slot.
    CACHE_MIN_TOKENS = 1024  # Anthropic's ephemeral cache floor
    if isinstance(system_prompt, str) and len(system_prompt) >= CACHE_MIN_TOKENS * 4:
        # ~4 chars/token heuristic — only cache prompts long enough to amortize.
        system_field = [{"type": "text", "text": system_prompt,
                         "cache_control": {"type": "ephemeral"}}]
    else:
        system_field = system_prompt

    chosen_model = model or MODEL_MARKET
    body = {"model": chosen_model,
            "max_tokens": _clamp_max_tokens(chosen_model, max_tokens or MAX_TOKENS),
            "system": system_field,
            "messages": [{"role": "user", "content": user_content}],
            "stream": True,}
    # Silently disable thinking when the chosen model doesn't expose
    # adaptive thinking (e.g. Haiku 4.5). Without this guard, an Economy-
    # mode run that lands on Haiku 400s with "adaptive thinking is not
    # supported on this model" and the whole analysis crashes — even
    # though the methodology call would have run fine without thinking.
    if use_thinking and not _supports_adaptive_thinking(body["model"]):
        print(f"[Claude] {body['model']} does not support adaptive thinking — "
              f"running without it.")
        use_thinking = False
    if use_thinking:
        body["thinking"] = {"type": "adaptive"}
        body["output_config"] = {"effort": THINKING_EFFORT}
    elif temperature is not None and _accepts_sampling(body["model"]):
        # Pre-4.7 models keep their deterministic temperature=0 config.
        # Fable 5 / Opus 4.7+ get neither key: sampling params 400, and the
        # only valid "no thinking" on Fable 5 is omitting the field entirely.
        body["temperature"] = temperature

    # Structured outputs (GA). The schema is compiled into a token-level
    # grammar — Claude literally cannot emit a property outside it.
    # Composes with thinking by sharing output_config.
    use_so = bool(output_schema) and USE_STRUCTURED_OUTPUTS
    if use_so:
        # The GA path doesn't need a beta header; SO_LEGACY_BETA_HEADER opts
        # back in if a pinned snapshot still requires the transition header.
        if SO_LEGACY_BETA_HEADER:
            headers["anthropic-beta"] = STRUCTURED_OUTPUTS_BETA
        body.setdefault("output_config", {})["format"] = {
            "type": "json_schema",
            "schema": output_schema,
        }
    if tools:
        body["tools"] = tools

    last_err = None
    for attempt in range(MAX_RETRIES):
        try:
            # connect_timeout = 30s, read_timeout = 600s (10 min) per chunk
            # Streaming keeps each chunk's idle time low, so this is safe
            resp = requests.post(ANTHROPIC_API_URL, headers=headers,
                                 json=body, timeout=(30, 600), stream=True)
            if resp.status_code != 200:
                # If structured outputs is rejected for the pinned snapshot,
                # drop it and retry — prompt + Pydantic still enforce the
                # shape, just without grammar-level masking.
                if resp.status_code == 400 and use_so:
                    err_text = resp.text[:500].lower()
                    if any(t in err_text for t in ("structured", "beta", "output_config", "schema")):
                        print(f"[Claude] Structured outputs rejected for "
                              f"{body['model']} ({resp.status_code}: "
                              f"{resp.text[:160]}); falling back to "
                              f"prompt-only schema")
                        headers.pop("anthropic-beta", None)
                        if isinstance(body.get("output_config"), dict):
                            body["output_config"].pop("format", None)
                            if not body["output_config"]:
                                body.pop("output_config", None)
                        use_so = False
                        continue
                if resp.status_code in (429, 500, 502, 503, 504, 529):
                    retry_after = resp.headers.get("retry-after")
                    delay = int(retry_after) if retry_after else BASE_BACKOFF_SEC * (2 ** attempt)
                    print(f"[Claude] {resp.status_code} — retrying in {delay}s")
                    time.sleep(delay)
                    last_err = f"HTTP {resp.status_code}"
                    continue
                raise RuntimeError(f"Claude API error {resp.status_code}: {resp.text[:500]}")

            # Parse the SSE stream and reassemble the message
            parsed = _parse_stream(resp)
            record_usage(body.get("model"), parsed.get("usage") or {})
            return parsed

        except requests.exceptions.Timeout:
            print(f"[Claude] Request timeout — retrying...")
            time.sleep(BASE_BACKOFF_SEC * (2 ** attempt))
            last_err = "Timeout"
        except requests.exceptions.ChunkedEncodingError as e:
            # Mid-stream connection drop — retry
            print(f"[Claude] Stream interrupted — retrying...")
            time.sleep(BASE_BACKOFF_SEC * (2 ** attempt))
            last_err = f"Stream interrupted: {e}"
        except requests.exceptions.ConnectionError as e:
            print(f"[Claude] Connection error (likely DNS/network) — retrying...")
            time.sleep(BASE_BACKOFF_SEC * (2 ** attempt))
            last_err = f"Connection error: {e}"
        except Exception as e:
            last_err = str(e)
            if attempt == MAX_RETRIES - 1:
                raise
            time.sleep(BASE_BACKOFF_SEC * (2 ** attempt))
    raise RuntimeError(f"Claude API failed after {MAX_RETRIES} retries: {last_err}")

def _parse_stream(resp):
    """
    Parse Anthropic's Server-Sent Events stream into the same dict shape
    that the non-streaming endpoint returns.
    """
    content_blocks = []   # accumulated content blocks
    current_block = None  # block currently being streamed
    stop_reason = None
    usage = {}

    for line in resp.iter_lines(decode_unicode=True):
        if not line or not line.startswith("data: "):
            continue
        try:
            event = json.loads(line[6:])
        except json.JSONDecodeError:
            continue

        etype = event.get("type")
        if etype == "content_block_start":
            current_block = dict(event.get("content_block", {}))
            content_blocks.append(current_block)
        elif etype == "content_block_delta":
            delta = event.get("delta", {})
            dtype = delta.get("type")
            if dtype == "text_delta" and current_block is not None:
                current_block["text"] = current_block.get("text", "") + delta.get("text", "")
            elif dtype == "input_json_delta" and current_block is not None:
                current_block["partial_json"] = current_block.get("partial_json", "") + delta.get("partial_json", "")
            elif dtype == "thinking_delta" and current_block is not None:
                current_block["thinking"] = current_block.get("thinking", "") + delta.get("thinking", "")
        elif etype == "content_block_stop":
            current_block = None
        elif etype == "message_delta":
            md = event.get("delta", {})
            if "stop_reason" in md:
                stop_reason = md["stop_reason"]
            if "usage" in event:
                usage.update(event["usage"])
        elif etype == "message_stop":
            break
        elif etype == "error":
            err = event.get("error", {})
            raise RuntimeError(f"Claude stream error: {err.get('type', '?')}: {err.get('message', '')}")

    return {"content": content_blocks, "stop_reason": stop_reason, "usage": usage}


def extract_text(response):
    return "\n".join(b.get("text", "") for b in response.get("content", []) if b.get("type") == "text")


def extract_json(response):
    text = extract_text(response)
    stop_reason = response.get("stop_reason", "")

    if stop_reason == "max_tokens":
        raise RuntimeError(
            "Claude hit max_tokens before finishing. The response was truncated. "
            "This usually means the deal has a very large rent roll. "
            "Try analyzing without an OM PDF, or split the analysis across two runs."
        )

    # Strip code fences if present
    text = re.sub(r"```(?:json)?\s*", "", text).replace("```", "")
    start, end = text.find("{"), text.rfind("}")
    if start == -1 or end == -1:
        raise ValueError(f"No JSON object found in response:\n{text[:500]}")

    candidate = text[start:end + 1]

    # Strategy 1: Try as-is
    try:
        return json.loads(candidate)
    except json.JSONDecodeError as e1:
        first_err = e1

    # Strategy 2: Escape unescaped control characters inside string values.
    # Claude sometimes inserts literal newlines/tabs inside strings, which is invalid JSON.
    # We walk the text character-by-character, tracking whether we're inside a quoted
    # string, and escape any raw control characters we encounter.
    def escape_control_chars(s):
        out = []
        in_string = False
        prev_was_backslash = False
        for ch in s:
            if in_string:
                if ch == '"' and not prev_was_backslash:
                    in_string = False
                    out.append(ch)
                elif ch == '\n':
                    out.append('\\n')
                elif ch == '\r':
                    out.append('\\r')
                elif ch == '\t':
                    out.append('\\t')
                elif ord(ch) < 0x20:
                    out.append(f'\\u{ord(ch):04x}')
                else:
                    out.append(ch)
            else:
                if ch == '"' and not prev_was_backslash:
                    in_string = True
                out.append(ch)
            prev_was_backslash = (ch == '\\') and not prev_was_backslash
        return ''.join(out)

    try:
        escaped = escape_control_chars(candidate)
        return json.loads(escaped)
    except json.JSONDecodeError:
        pass

    # Strategy 3: Strip trailing commas before } or ]
    cleaned = re.sub(r",(\s*[}\]])", r"\1", candidate)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # Strategy 4: Combine — escape control chars AND strip trailing commas
    try:
        combo = re.sub(r",(\s*[}\]])", r"\1", escape_control_chars(candidate))
        return json.loads(combo)
    except json.JSONDecodeError:
        pass

    # Strategy 5: Truncate at the failure point and auto-close any open brackets/braces
    # This recovers a partial-but-valid JSON when Claude went off the rails mid-output
    try:
        truncate_at = first_err.pos
        partial = candidate[:truncate_at]
        depth_obj = partial.count("{") - partial.count("}")
        depth_arr = partial.count("[") - partial.count("]")
        partial = re.sub(r',\s*"[^"]*"?\s*:?\s*$', '', partial)
        partial = re.sub(r',\s*$', '', partial)
        partial = partial.rstrip()
        partial += "]" * max(0, depth_arr) + "}" * max(0, depth_obj)
        return json.loads(escape_control_chars(partial))
    except (json.JSONDecodeError, Exception):
        pass

    # All strategies failed — give a useful error with context around the failure point
    err_pos = first_err.pos
    context_start = max(0, err_pos - 200)
    context_end = min(len(candidate), err_pos + 200)
    context = candidate[context_start:err_pos] + " >>>HERE>>> " + candidate[err_pos:context_end]
    raise RuntimeError(
        f"Failed to parse JSON from Claude. Stop reason: {stop_reason}. "
        f"JSON error at position {err_pos}: {first_err.msg}.\n"
        f"Context around error:\n...{context}..."
    )


# ═══════════════════════════════════════════════════════════════════════════
# FILE ENCODING
# ═══════════════════════════════════════════════════════════════════════════
def encode_file_for_claude(file_storage):
    filename = file_storage.filename
    ext = Path(filename).suffix.lower()
    data = file_storage.read()
    b64 = base64.standard_b64encode(data).decode("utf-8")

    if ext == ".pdf":
        # Try the configured PARSER_BACKEND first for deterministic,
        # high-accuracy table extraction. Falls back to Anthropic's native
        # PDF handling if the parser is disabled, misconfigured, or fails.
        markdown = parse_pdf(data, filename)
        if markdown:
            return {"type": "text",
                    "text": f"[PDF parsed by {PARSER_BACKEND}: {filename}]\n\n{markdown[:200000]}"}
        print(f"[Encode] Parser {PARSER_BACKEND} unavailable; "
              f"falling back to Anthropic native PDF for {filename}")
        return {"type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
                "title": filename}
    elif ext in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
        media_type = f"image/{'jpeg' if ext == '.jpg' else ext[1:]}"
        return {"type": "image",
                "source": {"type": "base64", "media_type": media_type, "data": b64}}
    elif ext in (".xlsx", ".xls"):
        # Preserve column letters, row numbers, and merged-cell ranges so
        # the LLM can reason about the geometry of the seller's workbook
        # ("the monthly columns are C-N, the total is O, the broker
        # proforma is Q"). The old flatten-to-tabs approach lost all of
        # that and was the root cause of the categorization mismatch on
        # multi-column P&Ls.
        try:
            from openpyxl.utils import get_column_letter
            wb = load_workbook(BytesIO(data), data_only=True)
            # Per-sheet cap. Default 200K chars is enough for typical
            # P&Ls (12 columns × 60 rows × ~30 chars/cell ≈ 22K). But
            # rent-roll-shaped sheets can be 1000 rows × 12 cols × 30
            # chars ≈ 360K and were silently getting cut to ~266 rows,
            # leaving the LLM unable to count vacant units past row 266.
            # Detect rent-roll sheets by name and raise to 1M, which is
            # enough for any realistic MH/RV park (largest GGC asset ≈
            # 500 units; even a 2,000-row commercial rent roll fits).
            DEFAULT_SHEET_CAP = 200_000
            RENT_ROLL_SHEET_CAP = 1_500_000
            sheet_blocks = [f"[Spreadsheet: {filename}]"]
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Heuristic: any sheet whose name hints "rent roll" gets
                # the larger budget. False positives are harmless.
                _name_lower = (sheet_name or "").lower()
                is_rent_roll = any(k in _name_lower for k in
                                    ("rent roll", "rentroll", "rent_roll",
                                     "rr ", "rr_", "tenant"))
                PER_SHEET_CAP = RENT_ROLL_SHEET_CAP if is_rent_roll else DEFAULT_SHEET_CAP
                max_row, max_col = ws.max_row or 0, ws.max_column or 0
                max_col_letter = get_column_letter(max_col) if max_col else "A"
                header = (
                    f"\n## Sheet: {sheet_name}  "
                    f"(rows 1-{max_row}, cols A-{max_col_letter})\n"
                    f"Format: <ColLetter><RowNum>: <value>  |  blanks shown as '·'  "
                    f"|  merged ranges expanded (same value repeated in every cell)"
                )
                # Map every cell inside a merged range to the top-left value.
                merged_map = {}
                merged_anchors = set()
                for mr in ws.merged_cells.ranges:
                    top_left = ws.cell(mr.min_row, mr.min_col).value
                    merged_anchors.add((mr.min_row, mr.min_col, mr.coord))
                    for r in range(mr.min_row, mr.max_row + 1):
                        for c in range(mr.min_col, mr.max_col + 1):
                            merged_map[(r, c)] = top_left
                anchor_lookup = {(r, c): coord for (r, c, coord) in merged_anchors}
                lines, used = [header], len(header)
                truncated = False
                for r in range(1, max_row + 1):
                    row_cells, row_has_value = [], False
                    for c in range(1, max_col + 1):
                        if (r, c) in merged_map:
                            v = merged_map[(r, c)]
                            suffix = (f"[merged {anchor_lookup[(r, c)]}]"
                                      if (r, c) in anchor_lookup else "")
                        else:
                            v, suffix = ws.cell(r, c).value, ""
                        if v is None:
                            token = "·"
                        else:
                            token = str(v).replace("\n", " ").replace("\t", " ").strip() or "·"
                            row_has_value = True
                        row_cells.append(f"{get_column_letter(c)}{r}: {token}{suffix}")
                    if row_has_value:
                        line = "  |  ".join(row_cells)
                        if used + len(line) + 1 > PER_SHEET_CAP:
                            truncated = True
                            break
                        lines.append(line)
                        used += len(line) + 1
                if truncated:
                    lines.append(f"... [TRUNCATED at {PER_SHEET_CAP} chars; "
                                 f"sheet has {max_row} rows total]")
                sheet_blocks.append("\n".join(lines))
            return {"type": "text", "text": "\n".join(sheet_blocks)}
        except Exception as e:
            return {"type": "text", "text": f"[Could not parse {filename}: {e}]"}
    elif ext in (".txt", ".csv", ".md"):
        return {"type": "text",
                "text": f"[File: {filename}]\n{data.decode('utf-8', errors='replace')[:200000]}"}
    else:
        return {"type": "text",
                "text": f"[File: {filename}]\n{data.decode('utf-8', errors='replace')[:100000]}"}


# ═══════════════════════════════════════════════════════════════════════════
# JSON SCHEMAS + PYDANTIC MODELS
# Two correctness layers:
#   (1) JSON Schema sent to Anthropic Structured Outputs → token-grammar
#       compliance. Eliminates malformed JSON + category-mapping drift.
#   (2) Pydantic models run after extract_json → mirror the schema for the
#       Python side and give the validator something to type-check against.
# verify_extraction (the pure-Python tie-out function) runs in addition and
# catches numeric/cross-field problems the schema layer can't see.
# ═══════════════════════════════════════════════════════════════════════════

CONFIDENCE_LEVELS = ["high", "medium", "low"]
SEVERITY_LEVELS   = ["high", "medium", "low"]
PROPERTY_TYPES    = ["MHC", "RV", "Hybrid"]
RECONCILE_UNIT    = ["match", "rent_roll_short", "rent_roll_long"]
RECONCILE_POH     = ["match", "stated_high", "stated_low", "stated_zero_but_found"]
SECTION_INCOME    = ["income"]
SECTION_EXPENSE   = ["expense"]


# ── EXTRACTION SCHEMA (Stage-1 / MODEL_EXTRACTION output) ─────────────────

def _extracted_line_schema(section_values):
    """Income/expense row from the extraction stage."""
    return {
        "type": "object",
        "additionalProperties": False,
        "required": ["sellerLabel", "annualTotal", "monthly", "section",
                     "isSubtotal", "sellerNotes", "proFormaTotal"],
        "properties": {
            "sellerLabel": {"type": "string"},
            "annualTotal": {"type": ["number", "null"]},
            # The seller's "PRO FORMA" column (when present) — their own
            # adjusted forward-looking estimate. GGC sometimes uses this
            # verbatim when the source P&L marks the line as adjusted
            # (e.g., insurance "Last quarter annualized"). Null when the
            # source has no pro-forma column.
            "proFormaTotal": {"type": ["number", "null"]},
            # The seller's "NOTES TO PRO FORMA" column verbatim. Drives the
            # Omitt routing in the methodology stage (see the FINANCIAL_PARSE
            # _PROMPT hard rule). Common values: "T12", "Non-recurring",
            # "Discontinued", "Seller Specific", "Last quarter annualized".
            "sellerNotes":  {"type": "string"},
            # monthly: 12-element array OR null. Anthropic schema accepts the
            # type-array form for nullables.
            "monthly": {
                # Anthropic structured outputs only accept minItems 0 or 1;
                # 12-element constraint is enforced by the Pydantic validator
                # downstream (monthly_ties_to_total skips items where the
                # array isn't exactly 12, so a sloppy LLM response gets
                # caught there rather than refused by the API).
                "type": ["array", "null"],
                "items": {"type": "number"},
            },
            "section":    {"type": "string", "enum": section_values},
            "isSubtotal": {"type": "boolean"},
        },
    }


EXTRACTION_OUTPUT_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "required": ["reportingPeriod", "income", "expenses", "rentRoll",
                 "documentsSeen", "extractionNotes"],
    "properties": {
        "reportingPeriod": {
            "type": "object",
            "additionalProperties": False,
            "required": ["periodUsed", "dateRange", "monthsCovered",
                         "candidatePeriodsSeen", "notes"],
            "properties": {
                "periodUsed":           {"type": "string"},
                "dateRange":            {"type": "string"},
                "monthsCovered":        {"type": ["integer", "null"]},
                "candidatePeriodsSeen": {"type": "array", "items": {"type": "string"}},
                "notes":                {"type": "string"},
            },
        },
        "income":   {"type": "array", "items": _extracted_line_schema(SECTION_INCOME)},
        "expenses": {"type": "array", "items": _extracted_line_schema(SECTION_EXPENSE)},
        "rentRoll": {
            "type": "object",
            "additionalProperties": False,
            "required": ["totalRowsInRentRoll", "statedTotalRentMonthly",
                         "statedTotalIsMonthly", "occupiedCount", "vacantCount",
                         "unitTypes", "rentRollRows"],
            "properties": {
                "totalRowsInRentRoll":    {"type": ["integer", "null"]},
                "statedTotalRentMonthly": {"type": ["number", "null"]},
                "statedTotalIsMonthly":   {"type": "boolean"},
                "occupiedCount":          {"type": ["integer", "null"]},
                "vacantCount":            {"type": ["integer", "null"]},
                "unitTypes": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "required": ["unitType", "count", "occupiedCount",
                                     "vacantCount", "avgLotRentOccupied",
                                     "hasHomeRentEntries", "avgHomeRent"],
                        "properties": {
                            "unitType":           {"type": "string"},
                            "count":              {"type": "integer"},
                            "occupiedCount":      {"type": "integer"},
                            "vacantCount":        {"type": "integer"},
                            "avgLotRentOccupied": {"type": ["number", "null"]},
                            "hasHomeRentEntries": {"type": "boolean"},
                            "avgHomeRent":        {"type": ["number", "null"]},
                        },
                    },
                },
                # Per-tenant data. CRITICAL: emit ONE ROW per data row in the
                # source rent roll. Aggregates (unitTypes above) alone are
                # not enough — Unit Mix Summary's COUNTIFS scans the per-row
                # data to populate Total Units, Occupancy, GPR, and the
                # bifurcated lot/home rent NOI. When this array is empty,
                # the template falls back to synthesizing rows with average
                # rents which zeros out every downstream metric.
                "rentRollRows": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "required": ["tenantName", "unitId", "unitType",
                                     "status", "lotRent", "homeRent",
                                     "marketRent"],
                        "properties": {
                            "tenantName": {"type": "string"},
                            "unitId":     {"type": "string"},
                            "unitType":   {"type": "string"},
                            "status":     {"type": "string"},  # "Occupied" / "Vacant"
                            "lotRent":    {"type": ["number", "null"]},
                            "homeRent":   {"type": ["number", "null"]},
                            "marketRent": {"type": ["number", "null"]},
                        },
                    },
                },
            },
        },
        "documentsSeen":   {"type": "array", "items": {"type": "string"}},
        "extractionNotes": {"type": "string"},
    },
}


# ── METHODOLOGY SCHEMA (Stage-2 / MODEL_METHODOLOGY output) ────────────────

def _methodology_line_schema(category_enum):
    return {
        "type": "object",
        "additionalProperties": False,
        "required": ["ggcCategory", "sellerName", "fyPrior", "fyCurrent",
                     "brokerProforma", "t12Total", "monthly",
                     "ggcUnderwritten", "confidence", "notes"],
        "properties": {
            "ggcCategory":     {"type": "string", "enum": category_enum},
            "sellerName":      {"type": "string"},
            "fyPrior":         {"type": "number"},
            "fyCurrent":       {"type": "number"},
            "brokerProforma":  {"type": "number"},
            "t12Total":        {"type": "number"},
            "monthly":         {"type": "array", "items": {"type": "number"}},
            "ggcUnderwritten": {"type": "number"},
            "confidence":      {"type": "string", "enum": CONFIDENCE_LEVELS},
            "notes":           {"type": "string"},
        },
    }


METHODOLOGY_OUTPUT_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "required": ["propertyInfo", "income", "expenses", "rentRoll",
                 "flags", "dataQualityChecks", "questions", "dataQuality"],
    "properties": {
        "propertyInfo": {
            "type": "object",
            "additionalProperties": False,
            "required": ["name", "address", "city", "state", "county",
                         "totalUnits", "askingPrice", "propertyType",
                         "ingoingCapRate", "stabilizedYieldOnCost",
                         "spreadBps", "meetsInvestmentCriteria"],
            "properties": {
                "name":        {"type": "string"},
                "address":     {"type": "string"},
                "city":        {"type": "string"},
                "state":       {"type": "string"},
                "county":      {"type": "string"},
                "totalUnits":  {"type": "integer"},
                "askingPrice": {"type": "number"},
                "propertyType":            {"type": "string", "enum": PROPERTY_TYPES},
                "ingoingCapRate":          {"type": "number"},
                "stabilizedYieldOnCost":   {"type": "number"},
                "spreadBps":               {"type": "integer"},
                "meetsInvestmentCriteria": {"type": "boolean"},
                # Optional metadata fields used to populate the Property
                # Information block on the Underwriting tab. All optional
                # — the LLM emits them only when the OM/T12/listing
                # surfaces the value; the cells stay blank otherwise so
                # the reviewer can fill them in by hand.
                "yearBuilt":         {"type": ["integer", "null"]},
                "websiteUrl":        {"type": ["string", "null"]},
                "acreage":           {"type": ["number", "null"]},
                "utilityStructure":  {"type": ["string", "null"]},
                "electricityNotes":  {"type": ["string", "null"]},
                "trashNotes":        {"type": ["string", "null"]},
                "taxAssessorUrl":    {"type": ["string", "null"]},
                "taxParcels": {
                    "type": ["array", "null"],
                    "items": {
                        "type": "object",
                        "properties": {
                            "parcelId":     {"type": ["string", "null"]},
                            "marketValue":  {"type": ["number", "null"]},
                            "taxableValue": {"type": ["number", "null"]},
                            "taxes":        {"type": ["number", "null"]},
                            "acres":        {"type": ["number", "null"]},
                        },
                    },
                },
            },
        },
        "income":   {"type": "array", "items": _methodology_line_schema(GGC_INCOME_CATEGORIES)},
        "expenses": {"type": "array", "items": _methodology_line_schema(GGC_EXPENSE_CATEGORIES)},
        "rentRoll": {
            # Simplified schema: dropped rentRollRows + unitMixSummary
            # (both optional and redundant with unitGroups), dropped enum
            # constraint on unitGroups.unitType (we map post-hoc in Python),
            # dropped additionalProperties:false on nested objects. The
            # methodology schema otherwise compiled into a grammar too
            # large for Anthropic's structured outputs to accept, which
            # silently downgraded every call to prompt-only enforcement.
            "type": "object",
            "required": ["totalUnits", "occupiedUnits", "vacantUnits",
                         "occupancyRate", "avgLotRent", "parkOwnedHomes",
                         "pohPercent", "unitGroups"],
            "properties": {
                "totalUnits":     {"type": "integer"},
                "occupiedUnits":  {"type": "integer"},
                "vacantUnits":    {"type": "integer"},
                "occupancyRate":  {"type": "number"},
                "avgLotRent":     {"type": "number"},
                "parkOwnedHomes": {"type": "integer"},
                "pohPercent":     {"type": "number"},
                "unitGroups": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "required": ["unitType", "occupiedCount", "vacantCount",
                                     "lotRent", "pohRent"],
                        "properties": {
                            "unitType":      {"type": "string"},
                            "occupiedCount": {"type": "integer"},
                            "vacantCount":   {"type": "integer"},
                            "lotRent":       {"type": "number"},
                            "pohRent":       {"type": "number"},
                            "tenantNamePattern": {"type": "string"},
                            "sellerUnitLabel":   {"type": "string"},
                        },
                    },
                },
            },
        },
        "flags": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "required": ["item", "issue", "severity", "recommendation"],
                "properties": {
                    "item":           {"type": "string"},
                    "issue":          {"type": "string"},
                    "severity":       {"type": "string", "enum": SEVERITY_LEVELS},
                    "recommendation": {"type": "string"},
                },
            },
        },
        "dataQualityChecks": {
            "type": "object",
            "additionalProperties": False,
            "required": ["totalUnitsStated", "rentRollRowsFound",
                         "unitCountReconciliation", "pohCountStated",
                         "pohRowsFound", "pohReconciliation",
                         "finalPohCount", "finalPohPercent"],
            "properties": {
                "totalUnitsStated":         {"type": "integer"},
                "rentRollRowsFound":        {"type": "integer"},
                "unitCountReconciliation":  {"type": "string", "enum": RECONCILE_UNIT},
                "pohCountStated":           {"type": "integer"},
                "pohRowsFound":             {"type": "integer"},
                "pohReconciliation":        {"type": "string", "enum": RECONCILE_POH},
                "finalPohCount":            {"type": "integer"},
                "finalPohPercent":          {"type": "number"},
            },
        },
        "questions": {"type": "array", "items": {"type": "string"}},
        "dataQuality": {
            "type": "object",
            "additionalProperties": False,
            "required": ["hasT12", "hasT3", "hasRentRoll",
                         "hasMonthlyBreakdown", "t12Period", "missingData"],
            "properties": {
                "hasT12":              {"type": "boolean"},
                "hasT3":               {"type": "boolean"},
                "hasRentRoll":         {"type": "boolean"},
                "hasMonthlyBreakdown": {"type": "boolean"},
                "t12Period":           {"type": "string"},
                "missingData":         {"type": "array", "items": {"type": "string"}},
            },
        },
    },
}


# ── Pydantic mirrors (extraction stage) ───────────────────────────────────

class ExtractedLineItem(BaseModel):
    sellerLabel:   str
    annualTotal:   float | None = None
    monthly:       list[float] | None = None
    section:       str
    isSubtotal:    bool = False
    proFormaTotal: float | None = None
    sellerNotes:   str = ""


class ExtractedRentRoll(BaseModel):
    totalRowsInRentRoll:    int | None = None
    statedTotalRentMonthly: float | None = None
    statedTotalIsMonthly:   bool = True
    occupiedCount:          int | None = None
    vacantCount:            int | None = None
    unitTypes:              list[dict] = Field(default_factory=list)
    rentRollRows:           list[dict] = Field(default_factory=list)


class ExtractedReportingPeriod(BaseModel):
    periodUsed:           str = ""
    dateRange:            str = ""
    monthsCovered:        int | None = None
    candidatePeriodsSeen: list[str] = Field(default_factory=list)
    notes:                str = ""


class ExtractedFinancials(BaseModel):
    reportingPeriod: ExtractedReportingPeriod
    income:          list[ExtractedLineItem]
    expenses:        list[ExtractedLineItem]
    rentRoll:        ExtractedRentRoll
    documentsSeen:   list[str] = Field(default_factory=list)
    extractionNotes: str = ""


# ── Pydantic mirrors (methodology stage) ──────────────────────────────────

class MethodologyLineItem(BaseModel):
    # All numeric fields accept None — the LLM frequently emits null when
    # a particular FY column doesn't exist in the source P&L (e.g. fyPrior
    # null for a deal that only has 2024 + T12). Treating null as 0 via a
    # pre-validator keeps the downstream math clean without flooding the
    # Extraction Check tab with "Input should be a valid number" noise.
    ggcCategory: str
    sellerName: str = ""
    fyPrior: float | None = 0
    fyCurrent: float | None = 0
    brokerProforma: float | None = 0
    t12Total: float | None = 0
    monthly: list[float] | None = Field(default_factory=lambda: [0.0] * 12)
    ggcUnderwritten: float | None = 0
    confidence: str = "medium"
    notes: str = ""

    @field_validator("fyPrior", "fyCurrent", "brokerProforma",
                      "t12Total", "ggcUnderwritten", mode="before")
    @classmethod
    def _none_to_zero(cls, v):
        return 0 if v is None else v

    @model_validator(mode="after")
    def monthly_ties_to_total(self):
        # Skip when there's nothing to tie out OR when the LLM declined
        # to emit a monthly series (some derived/synthetic lines like
        # GGC's Cap-Ex Reserve don't have monthly history).
        if not self.monthly or len(self.monthly) != 12 or abs(self.t12Total) < 1:
            return self
        # 5% tolerance matches the methodology's other thresholds (spike
        # detection, run-to-run disagreement). 1% was too tight — every
        # methodology-adjusted line item failed on rounding noise and the
        # Extraction Check tab drowned in false positives. Above 5% the
        # mismatch is real (LLM confused monthly with adjusted total, or
        # arithmetic error worth surfacing).
        total = sum(self.monthly)
        tolerance = max(50.0, abs(self.t12Total) * 0.05)
        if abs(total - self.t12Total) > tolerance:
            raise ValueError(
                f"{self.ggcCategory} ({self.sellerName}): monthly sum "
                f"${total:,.2f} != t12Total ${self.t12Total:,.2f} "
                f"(diff ${abs(total - self.t12Total):,.2f}, "
                f"{abs(total - self.t12Total) / abs(self.t12Total):.1%})"
            )
        return self


# Markers in sellerName that indicate the line is a subtotal/aggregation, NOT
# a leaf GL row. The methodology must emit one row per leaf GL — never the
# rolled-up "5700 Total Personnel" or "4100 Total Rental Income (non-posting)"
# rows from the seller's chart of accounts, because those carry no t12Total
# and silently zero out NOI when written to the workbook (the original
# payroll-aggregation bug in 17June).
_SUBTOTAL_MARKERS = (
    "non-posting", "(non-posting)", "non posting", "(non posting)",
    "total personnel", "total income", "total expense",
    "total rental", "total maintenance", "total utility",
    "total other", "total insurance",
    "total repairs", "total taxes", "total office",
)

def _looks_like_subtotal(seller_name):
    s = (seller_name or "").lower()
    return any(m in s for m in _SUBTOTAL_MARKERS)


class MethodologyIncomeItem(MethodologyLineItem):
    @field_validator("ggcCategory")
    @classmethod
    def category_in_enum(cls, v):
        if v not in GGC_INCOME_CATEGORIES:
            raise ValueError(
                f"income.ggcCategory='{v}' not in GGC_INCOME_CATEGORIES"
            )
        return v

    @model_validator(mode="after")
    def not_a_subtotal_row(self):
        if _looks_like_subtotal(self.sellerName):
            raise ValueError(
                f"income.sellerName='{self.sellerName}' looks like a subtotal/"
                f"aggregation row. Emit the underlying leaf GL accounts "
                f"individually instead."
            )
        return self


class MethodologyExpenseItem(MethodologyLineItem):
    @field_validator("ggcCategory")
    @classmethod
    def category_in_enum(cls, v):
        if v not in GGC_EXPENSE_CATEGORIES:
            raise ValueError(
                f"expense.ggcCategory='{v}' not in GGC_EXPENSE_CATEGORIES"
            )
        return v

    @model_validator(mode="after")
    def not_a_subtotal_row(self):
        if _looks_like_subtotal(self.sellerName):
            raise ValueError(
                f"expense.sellerName='{self.sellerName}' looks like a "
                f"subtotal/aggregation row (e.g. '5700 Total Personnel'). "
                f"Emit the leaf GL accounts individually instead."
            )
        return self


# ═══════════════════════════════════════════════════════════════════════════
# CLAUDE CALL #1 — EXTRACTION (faithful transcription, Fable 5, no thinking)
# ═══════════════════════════════════════════════════════════════════════════
# This call does ONE job: read the documents and pull out clean numbers.
# No GGC categorization, no underwriting methodology, no judgment. Just faithful
# transcription of what's actually in the source, with the correct reporting
# period identified. Keeping this separate from the methodology call is what
# makes the numbers reliable — the model isn't juggling extraction AND analysis
# at the same time, which is what was causing wrong periods and wrong values.
EXTRACTION_PROMPT = """You are a financial data extraction engine for a real estate underwriting team. Your ONLY job is to faithfully transcribe the numbers from the attached seller documents into clean structured JSON. You do NOT categorize, analyze, or apply any methodology. You transcribe exactly what is in the documents.

CRITICAL OUTPUT RULES:
- Your response MUST start with `{` and end with `}`. NOTHING before or after.
- DO NOT add preamble, explanation, or commentary.
- DO NOT wrap the JSON in markdown code fences.
- Transcribe numbers EXACTLY as they appear. Do not round, adjust, or "correct" them.
- If a value is genuinely not present in the documents, use null. Never invent a number.

## STEP 1 — IDENTIFY THE CORRECT REPORTING PERIOD (most important)

Seller financials frequently contain MULTIPLE time periods side by side. You MUST identify which one is the operative trailing-12-month (T12) period before extracting anything.

- Look for column headers like "T-12 Ended 5/23", "T-12 Ended 9/22", "Oct 2022 - May 2023", "FY2023", individual month columns, etc.
- The operative T12 is the MOST RECENT complete trailing-12-month column. If there are several "T-12 Ended X" columns, use the one with the latest ending date.
- A column covering fewer than 12 months (e.g. "Oct 2022 - May 2023" = 8 months) is a PARTIAL period. NEVER treat a partial period as the annual figure.
- If individual monthly columns are present, the 12 most recent consecutive months ARE the T12. Sum them to cross-check against any stated T12 total.
- Report exactly which period you used and its date range in the "reportingPeriod" field.

## STEP 2 — EXTRACT THE INCOME STATEMENT (P&L / T12)

For EVERY line item in the seller's operating statement, transcribe:
- "sellerLabel": the exact label as written (e.g. "6950 · UTILITIES", "4010 · RENTAL INCOME"). Include account numbers verbatim.
- "annualTotal": the value from the operative T12 column you identified in Step 1
- "monthly": an array of the 12 monthly values for that line, IN CHRONOLOGICAL ORDER, if monthly detail exists. If no monthly detail exists, use null.
- "section": your best read of whether this is "income" or "expense" based on where it sits in the statement (this is structural, NOT GGC categorization — just income vs expense)
- "proFormaTotal": if the source has a "PRO FORMA" column (or any equivalent forward-looking adjusted column), transcribe its value here. Use null when no such column exists or the cell is blank.
- "sellerNotes": if the source has a "NOTES TO PRO FORMA" or any notes column adjacent to the pro-forma column, copy its contents verbatim. Common values are "T12", "Non-recurring", "Discontinued", "Seller Specific", "Last quarter annualized". Empty string when no notes column exists. These notes drive the downstream Omitt-routing decisions — do not omit them.

Rules:
- If monthly values exist, they should sum to (or very close to) the annualTotal. If they don't, still transcribe both faithfully — the verification step will flag the discrepancy.
- Include EVERY line, even ones that look like subtotals or totals. Mark subtotals/totals with "isSubtotal": true so they can be excluded from sums later. Common subtotal labels: "Total", "non-posting", "Total Personnel", "Total Income", "TOTAL EXPENSE".
- Preserve the seller's account numbers in the label if present.

## STEP 3 — EXTRACT THE RENT ROLL

Transcribe the rent roll into BOTH structured-aggregate AND per-tenant-row form. The per-tenant rows are NOT optional — the downstream Unit Mix Summary counts each row, and missing per-row data zeros out every per-unit metric.

Aggregates:
- "totalRowsInRentRoll": the number of unit/space rows you found (integer)
- "statedTotalRentMonthly": if the rent roll shows a "Total Possible Rent" or "Totals" figure, transcribe it here. Note whether it is monthly or annual.
- "occupiedCount": number of occupied units
- "vacantCount": number of vacant units
- "unitTypes": array of distinct unit types found, each with:
    - "unitType": the label (e.g. "WHA Lot", "WHA RV", "Commercial Space", "Storage")
    - "count": how many of this type
    - "occupiedCount": occupied of this type
    - "vacantCount": vacant of this type
    - "avgLotRentOccupied": average lot rent of the OCCUPIED units of this type
    - "hasHomeRentEntries": true if any unit of this type shows a home rent / POH rent value
    - "avgHomeRent": average home rent among units of this type that have one (null if none)

Per-tenant rows (REQUIRED — one entry per data row in the source rent roll):
- "rentRollRows": array of every tenant/space row, each with:
    - "tenantName": tenant name as listed (use "" for vacant rows)
    - "unitId": the unit/pad/space identifier (e.g. "A05", "B12", "EL02A"). Use "" when not present.
    - "unitType": the seller's unit-type label for this row VERBATIM (do NOT map to GGC's canonical types here — the methodology stage does that).
    - "status": "Occupied" or "Vacant". When the source uses other words ("Occ", "OCC", "Y", "X") translate to the canonical Occupied/Vacant.
    - "lotRent": the lot/site rent for this row (number; 0 for vacant or for non-MH types)
    - "homeRent": the home/POH rent for this row (number; 0 when not applicable)
    - "marketRent": the market rent for this row when separately listed (null when the seller doesn't break it out)

## OUTPUT SCHEMA (JSON only)

{
  "reportingPeriod": {
    "periodUsed": "string — exact label of the column you used, e.g. 'T-12 Ended 5/23'",
    "dateRange": "string — e.g. 'Jun 2022 - May 2023'",
    "monthsCovered": 12,
    "candidatePeriodsSeen": ["list every period column header you saw in the document"],
    "notes": "string — explain any ambiguity in choosing the period"
  },
  "income": [
    {"sellerLabel": "string", "annualTotal": number|null, "monthly": [12 numbers]|null, "section": "income", "isSubtotal": false, "proFormaTotal": number|null, "sellerNotes": "string"}
  ],
  "expenses": [
    {"sellerLabel": "string", "annualTotal": number|null, "monthly": [12 numbers]|null, "section": "expense", "isSubtotal": false, "proFormaTotal": number|null, "sellerNotes": "string"}
  ],
  "rentRoll": {
    "totalRowsInRentRoll": integer|null,
    "statedTotalRentMonthly": number|null,
    "statedTotalIsMonthly": true,
    "occupiedCount": integer|null,
    "vacantCount": integer|null,
    "unitTypes": [
      {"unitType": "string", "count": integer, "occupiedCount": integer, "vacantCount": integer,
       "avgLotRentOccupied": number|null, "hasHomeRentEntries": false, "avgHomeRent": number|null}
    ],
    "rentRollRows": [
      {"tenantName": "string", "unitId": "string", "unitType": "string", "status": "Occupied|Vacant",
       "lotRent": number|null, "homeRent": number|null, "marketRent": number|null}
    ]
  },
  "documentsSeen": ["list each document by what it appears to be, e.g. 'T12 operating statement', 'rent roll', 'offering memorandum'"],
  "extractionNotes": "string — anything that was hard to read, ambiguous, or that the downstream analyst should know"
}"""


def call_extract_financials(api_key, file_blocks, property_info):
    """
    Call 1 of the financial pipeline: faithful extraction.
    MODEL_EXTRACTION (Fable 5, no thinking) reads the documents and returns
    clean numbers with the correct reporting period identified. No GGC
    methodology applied. (temperature=0 is still passed but only attaches on
    models that accept sampling params, e.g. a Sonnet env override.)

    Two correctness layers:
      (a) Structured outputs grammar (Anthropic beta) — guarantees schema
          compliance; the model literally cannot emit a property outside
          EXTRACTION_OUTPUT_SCHEMA.
      (b) Pydantic structural validation + verify_extraction tie-outs. On
          either kind of failure, the errors are appended to the prompt and
          extraction is re-run (up to MAX_PARSE_RETRIES). Failures past the
          last retry are surfaced by verify_extraction downstream — not
          silently dropped.
    """
    base_context = {
        "type": "text",
        "text": f"""Documents are attached above. Property context for reference only:
- Name: {property_info.get('name', 'N/A')}
- Total Units (user-stated): {property_info.get('units', 'N/A')}
- Park-Owned Home Count (user-stated): {property_info.get('pohCount', '0')}

Extract the income statement, rent roll, and reporting period into the structured JSON. Transcribe faithfully — do not categorize or analyze."""
    }
    base_user_blocks = file_blocks + [base_context]

    user_blocks    = base_user_blocks
    last_extracted = None

    model_id = _model_for_stage(property_info, "extraction")
    for attempt in range(MAX_PARSE_RETRIES + 1):
        print(f"[Claude] Stage 1/2 — EXTRACTION attempt "
              f"{attempt+1}/{MAX_PARSE_RETRIES+1} ({model_id})...")
        t0 = time.time()
        response = call_claude(api_key, EXTRACTION_PROMPT, user_blocks,
                               use_thinking=False, temperature=0,
                               model=model_id,
                               output_schema=EXTRACTION_OUTPUT_SCHEMA)
        elapsed = time.time() - t0
        print(f"[Claude] Extraction returned in {elapsed:.1f}s "
              f"(stop_reason: {response.get('stop_reason', '?')})")
        extracted = extract_json(response)
        last_extracted = extracted

        # Layer (b1): Pydantic structural check
        pydantic_errors = []
        try:
            ExtractedFinancials(**extracted)
        except ValidationError as e:
            for err in e.errors():
                loc = ".".join(str(x) for x in err.get("loc", []))
                pydantic_errors.append(f"{loc}: {err.get('msg', 'invalid')}")

        # Layer (b2): pure-Python tie-out checks
        checks   = verify_extraction(extracted, property_info)
        failures = [c for c in checks if c.get("status") == "fail"]

        if not pydantic_errors and not failures:
            print(f"[Extract] Validation passed on attempt {attempt+1}")
            return extracted

        print(f"[Extract] Attempt {attempt+1} issues: "
              f"{len(pydantic_errors)} schema, {len(failures)} tie-out")
        for e in pydantic_errors[:5]:
            print(f"  - schema: {e}")
        for c in failures[:5]:
            print(f"  - tie-out: {c.get('item', '')}: {c.get('detail', '')}")

        if attempt < MAX_PARSE_RETRIES:
            err_lines = ([f"- schema: {e}" for e in pydantic_errors[:8]]
                         + [f"- {c.get('item','?')}: {c.get('detail','?')}"
                            for c in failures[:10]])
            user_blocks = base_user_blocks + [{
                "type": "text",
                "text": (
                    "Your previous extraction failed these checks. "
                    "Re-read the documents and produce a CORRECTED "
                    "extraction:\n\n"
                    + "\n".join(err_lines)
                    + "\n\nPay attention to: (a) the monthly array of each "
                    "income/expense line summing to its annualTotal, "
                    "(b) the rent roll occupiedCount + vacantCount equaling "
                    "totalRowsInRentRoll, (c) picking the MOST RECENT "
                    "12-month period (not a partial period), (d) section "
                    "labels exactly matching 'income' or 'expense'."
                ),
            }]

    # Retries exhausted. Return the last extraction. Downstream
    # verify_extraction surfaces the failures on the Extraction Check tab —
    # they are visible to the reviewer, not silently passed.
    print(f"[Extract] {MAX_PARSE_RETRIES+1} attempts exhausted; "
          "downstream verify_extraction will surface residual failures")
    return last_extracted


def verify_extraction(extracted, property_info):
    """
    Pure-Python verification of the extracted data — NO AI involved, so this is
    fully deterministic. Catches the failure modes that showed up in testing:
    monthly values not summing to the annual total, rent roll row count not
    matching the stated unit count, partial periods used as annual, etc.

    Returns a list of check dicts: {item, check, status, detail}
    status is "ok" | "warn" | "fail". These get surfaced on the Extraction
    Check tab so the reviewer can confirm the numbers tie before trusting
    anything downstream.
    """
    checks = []

    # ── Reporting period sanity ──────────────────────────────────────────
    rp = extracted.get("reportingPeriod", {}) or {}
    months = rp.get("monthsCovered")
    if months == 12:
        checks.append({"item": "Reporting period", "check": "T12 (12 months)",
                       "status": "ok",
                       "detail": f"{rp.get('periodUsed', '?')} ({rp.get('dateRange', '?')})"})
    elif months is not None:
        checks.append({"item": "Reporting period", "check": "Should be 12 months",
                       "status": "fail",
                       "detail": f"Period used covers {months} months — this is a PARTIAL period, not a T12. {rp.get('periodUsed', '?')}"})
    else:
        checks.append({"item": "Reporting period", "check": "12 months",
                       "status": "warn",
                       "detail": "Could not confirm months covered. Verify the period manually."})
    candidates = rp.get("candidatePeriodsSeen") or []
    if len(candidates) > 1:
        checks.append({"item": "Multiple periods in doc", "check": "Picked most recent T12",
                       "status": "warn",
                       "detail": f"Document had {len(candidates)} period columns: {', '.join(str(c) for c in candidates[:6])}. Confirm the right one was used."})

    # ── Monthly vs annual reconciliation for each line ───────────────────
    def _check_lines(lines, label):
        for ln in lines or []:
            if ln.get("isSubtotal"):
                continue
            monthly = ln.get("monthly")
            annual = ln.get("annualTotal")
            name = ln.get("sellerLabel", "(unlabeled)")
            if isinstance(monthly, list) and len(monthly) == 12 and isinstance(annual, (int, float)):
                msum = sum(v for v in monthly if isinstance(v, (int, float)))
                if annual == 0:
                    pct_off = 0 if msum == 0 else 100
                else:
                    pct_off = abs(msum - annual) / abs(annual) * 100
                if pct_off <= 1:
                    checks.append({"item": f"{label}: {name}", "check": "Monthlies sum to annual",
                                   "status": "ok", "detail": f"Σmonthly={msum:,.0f} vs annual={annual:,.0f}"})
                elif pct_off <= 5:
                    checks.append({"item": f"{label}: {name}", "check": "Monthlies ≈ annual",
                                   "status": "warn",
                                   "detail": f"Σmonthly={msum:,.0f} vs annual={annual:,.0f} ({pct_off:.1f}% off)"})
                else:
                    checks.append({"item": f"{label}: {name}", "check": "Monthlies vs annual MISMATCH",
                                   "status": "fail",
                                   "detail": f"Σmonthly={msum:,.0f} vs annual={annual:,.0f} ({pct_off:.1f}% off) — check extraction"})

    _check_lines(extracted.get("income"), "Income")
    _check_lines(extracted.get("expenses"), "Expense")

    # ── Rent roll row count vs user-stated unit count ────────────────────
    rr = extracted.get("rentRoll", {}) or {}
    rows = rr.get("totalRowsInRentRoll")
    try:
        stated_units = int(str(property_info.get("units", "")).strip() or 0)
    except (ValueError, TypeError):
        stated_units = 0
    if rows is not None and stated_units:
        if rows == stated_units:
            checks.append({"item": "Rent roll rows vs unit count", "check": "Match",
                           "status": "ok", "detail": f"{rows} rows = {stated_units} units"})
        elif rows < stated_units:
            checks.append({"item": "Rent roll rows vs unit count", "check": "Rent roll short",
                           "status": "warn",
                           "detail": f"{rows} rows but {stated_units} units stated — {stated_units - rows} likely vacant lots omitted by seller. Should be imputed at market rent."})
        else:
            checks.append({"item": "Rent roll rows vs unit count", "check": "Rent roll long",
                           "status": "warn",
                           "detail": f"{rows} rows but only {stated_units} units stated — likely the form unit count is low or the rent roll includes non-unit rows (model homes, manager-comp lots, storage). Workbook still produced; reviewer should confirm."})

    # ── Rent roll stated total vs sum of unit-type rents (rough) ─────────
    stated_total = rr.get("statedTotalRentMonthly")
    annual_from_monthly = None
    if isinstance(stated_total, (int, float)) and stated_total:
        annual_from_monthly = stated_total * 12 if rr.get("statedTotalIsMonthly", True) else stated_total
        checks.append({"item": "Rent roll total", "check": "Monthly → annual",
                       "status": "ok",
                       "detail": f"Stated rent roll total {stated_total:,.0f}/mo → {annual_from_monthly:,.0f}/yr (this should approximate GPR)"})

    # ── Rent anomaly across unit types (hybrid 2σ + ratio-to-median) ─────
    # With only 3 unit types and one extreme outlier, the outlier itself
    # inflates the stdev and z stays under 2. So we also check ratio to
    # the median (robust statistic): flag any rent ≥3× or ≤1/3 the median.
    # Catches both subtle and obvious mispricings.
    rents = [ut.get("avgLotRentOccupied") for ut in (rr.get("unitTypes") or [])
             if isinstance(ut.get("avgLotRentOccupied"), (int, float))
             and ut.get("avgLotRentOccupied") > 0]
    if len(rents) >= 3:
        mean = sum(rents) / len(rents)
        median = statistics.median(rents)
        try:
            sd = statistics.stdev(rents)
        except statistics.StatisticsError:
            sd = 0
        for ut in rr.get("unitTypes") or []:
            rent = ut.get("avgLotRentOccupied")
            if not isinstance(rent, (int, float)) or rent <= 0:
                continue
            z = (rent - mean) / sd if sd > 0 else 0
            ratio = rent / median if median > 0 else 1
            if abs(z) >= 2 or ratio >= 3 or ratio <= 1/3:
                checks.append({
                    "item":  f"Rent anomaly: {ut.get('unitType', '?')}",
                    "check": "Lot rent within property range",
                    "status": "warn",
                    "detail": (
                        f"avgLotRent ${rent:,.0f} vs median ${median:,.0f} "
                        f"(ratio {ratio:.2f}×, z={z:+.2f}σ). "
                        "Verify this isn't a data entry error or a "
                        "mispriced unit type."
                    ),
                })

    # ── POH cross-check vs user-stated count ─────────────────────────────
    try:
        stated_poh = int(str(property_info.get("pohCount", "")).strip() or 0)
    except (ValueError, TypeError):
        stated_poh = 0
    any_home_entries = any(ut.get("hasHomeRentEntries")
                           for ut in (rr.get("unitTypes") or []))
    if stated_poh > 0 and not any_home_entries:
        checks.append({
            "item": "POH count vs rent roll",
            "check": f"User stated {stated_poh} POH",
            "status": "warn",
            "detail": (
                f"User stated {stated_poh} park-owned homes but no unit "
                "type in the rent roll has home rent entries. The seller "
                "may be hiding home rent income (or POH lots are comped "
                "for on-site staff). Verify with seller."
            ),
        })
    elif stated_poh == 0 and any_home_entries:
        checks.append({
            "item": "POH count vs rent roll",
            "check": "User stated 0 POH",
            "status": "fail",
            "detail": (
                "User stated 0 park-owned homes but the rent roll has "
                "home rent entries. Use the rent-roll count as "
                "authoritative — the user input was wrong."
            ),
        })

    # ── Cross-document parity: rent-roll annual GPR vs sum of income ─────
    docs = extracted.get("documentsSeen") or []
    if len(docs) >= 2 and annual_from_monthly:
        income_sum = sum(
            ln.get("annualTotal") or 0
            for ln in (extracted.get("income") or [])
            if not ln.get("isSubtotal")
            and isinstance(ln.get("annualTotal"), (int, float))
        )
        if income_sum > 0:
            spread = abs(annual_from_monthly - income_sum) / max(income_sum, 1)
            if spread > 0.20:
                checks.append({
                    "item": "Cross-doc revenue parity",
                    "check": "Rent roll annual ≈ Σ income lines",
                    "status": "warn",
                    "detail": (
                        f"Rent roll implies ${annual_from_monthly:,.0f}/yr "
                        f"GPR but income lines sum to ${income_sum:,.0f}/yr "
                        f"({spread*100:.0f}% gap). Either one source is "
                        "stale (different period) or income lines are missing."
                    ),
                })

    # ── Run-to-run disagreement, if N=3 merge ran ────────────────────────
    note = (extracted.get("extractionNotes") or "")
    if "Run-to-run disagreement" in note:
        checks.append({
            "item": "Run-to-run disagreement",
            "check": "N=3 merge: line-item annualTotal spread",
            "status": "warn",
            "detail": note.split("Run-to-run disagreement")[-1].strip()[:300],
        })

    # ── Section subtotal sums (§7 check #2): line items must sum to the
    # extracted subtotal row within tolerance. Catches the case where one
    # line item was dropped or duplicated by the parser.
    def _check_section_subtotal(lines, label):
        if not lines:
            return
        items   = [ln for ln in lines if not ln.get("isSubtotal")
                                    and isinstance(ln.get("annualTotal"), (int, float))]
        totals  = [ln for ln in lines if ln.get("isSubtotal")
                                    and isinstance(ln.get("annualTotal"), (int, float))]
        if not items or not totals:
            return
        items_sum    = sum(ln["annualTotal"] for ln in items)
        subtotal_val = max((t["annualTotal"] for t in totals), key=abs)
        if subtotal_val == 0:
            return
        pct_off = abs(items_sum - subtotal_val) / abs(subtotal_val) * 100
        # Section subtotals can legitimately differ from line-item sums by a
        # few percent on seller P&Ls (rounding, omitted small lines, mid-
        # period adjustments). Only catastrophic mismatches indicate a real
        # parsing failure that would corrupt the workbook.
        if pct_off <= 1:
            status, prefix = "ok",   "Lines sum to subtotal"
        elif pct_off <= 20:
            status, prefix = "warn", "Lines ≈ subtotal"
        else:
            status, prefix = "fail", "Lines vs subtotal MISMATCH"
        checks.append({
            "item": f"{label} subtotal sum",
            "check": prefix,
            "status": status,
            "detail": (f"Σ{label.lower()} items ${items_sum:,.0f} vs subtotal "
                       f"${subtotal_val:,.0f} ({pct_off:.1f}% off)"),
        })

    _check_section_subtotal(extracted.get("income"),   "Income")
    _check_section_subtotal(extracted.get("expenses"), "Expense")

    # ── Rent-roll cross-check: occupied × avg lot rent ≈ scheduled rent
    # (§7 rent-roll check #1). Within 5% is OK; outside is a real signal.
    types = rr.get("unitTypes") or []
    implied_monthly = 0.0
    for ut in types:
        oc  = ut.get("occupiedCount") or 0
        rent = ut.get("avgLotRentOccupied") or 0
        if isinstance(oc, (int, float)) and isinstance(rent, (int, float)):
            implied_monthly += oc * rent
    if implied_monthly > 0 and isinstance(stated_total, (int, float)) and stated_total > 0:
        stated_monthly = stated_total if rr.get("statedTotalIsMonthly", True) \
                                       else stated_total / 12
        spread = abs(implied_monthly - stated_monthly) / max(stated_monthly, 1)
        pct = spread * 100
        if pct <= 5:
            status, prefix = "ok",   "Occupied × avg = stated rent"
        elif pct <= 15:
            status, prefix = "warn", "Occupied × avg ≈ stated rent"
        else:
            status, prefix = "fail", "Occupied × avg ≠ stated rent"
        checks.append({
            "item": "Rent-roll cross-check",
            "check": prefix,
            "status": status,
            "detail": (f"Σ(occupied × avg lot rent) ${implied_monthly:,.0f}/mo "
                       f"vs stated total ${stated_monthly:,.0f}/mo "
                       f"({pct:.1f}% off)"),
        })

    return checks


# ═══════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════
# verify_methodology — post-categorization parity checks
# ═══════════════════════════════════════════════════════════════════════════
# Runs after call_parse_financials. Compares the rent roll's canonical
# unit types against the income categories the methodology assigned —
# if the rent roll has Long-term RV sites but RV Site Rental Income is
# missing or zero, the LLM almost certainly collapsed it into Gross
# Potential Rent (the bug that prompted this whole investigation).
def verify_methodology(financials):
    checks = []
    income = financials.get("income") or []
    expenses = financials.get("expenses") or []
    rr = financials.get("rentRoll") or {}

    # Index income by category (multiple rows per category aggregate).
    by_cat = {}
    for it in income:
        c = (it.get("ggcCategory") or "").strip()
        by_cat.setdefault(c, []).append(it)

    def _cat_total(cat):
        return sum(float(i.get("t12Total") or i.get("ggcUnderwritten") or 0)
                   for i in by_cat.get(cat, []))

    # Map canonical unit types → required income categories.
    unit_to_cat = {
        "TOH MH Site":        "Gross Potential Rent",
        "POH-Infilled units": "Home Rent Income",
        "Long term RV Site":  "RV Site Rental Income",
        # Note: enum was renamed "Retail Income" -> "Retail" to match
        # CorrectOutput's SUMIFS criterion. The parity check has to use
        # the new name or it fails every deal with retail units.
        "Retail/Commercial":  "Retail",
    }
    unit_groups = rr.get("unitGroups") or []
    rows = rr.get("rentRollRows") or []
    type_counts = {}
    for g in unit_groups:
        t = (g.get("unitType") or "").strip()
        type_counts[t] = type_counts.get(t, 0) + int(g.get("occupiedCount") or 0)
    for r in rows:
        if (r.get("status") or "").lower() == "occupied":
            t = (r.get("unitType") or "").strip()
            type_counts[t] = type_counts.get(t, 0) + 1

    for unit_type, required_cat in unit_to_cat.items():
        n = type_counts.get(unit_type, 0)
        if n == 0:
            continue
        cat_total = _cat_total(required_cat)
        if cat_total == 0:
            checks.append({
                "item": f"Rent-roll vs income parity: {unit_type}",
                "check": f"Expect non-zero '{required_cat}' income",
                "status": "fail",
                "detail": (f"Rent roll has {n} occupied {unit_type} unit(s) "
                           f"but '{required_cat}' income totals $0. The LLM "
                           f"likely collapsed this revenue into another "
                           f"category (most often Gross Potential Rent)."),
            })
        else:
            checks.append({
                "item": f"Rent-roll vs income parity: {unit_type}",
                "check": f"'{required_cat}' present",
                "status": "ok",
                "detail": f"{n} unit(s) → ${cat_total:,.0f} income"
            })

    # GPR row-count sanity: if rent roll has multiple revenue streams,
    # GPR (lot rent only) should be ONE row, not a sum of all streams.
    gpr_rows = by_cat.get("Gross Potential Rent", [])
    if len(gpr_rows) > 1:
        checks.append({
            "item": "Gross Potential Rent line count",
            "check": "Expect 1 GL line under GPR (lot rent only)",
            "status": "warn",
            "detail": (f"GPR has {len(gpr_rows)} line items. Multiple GL "
                       f"accounts under one GGC category sometimes "
                       f"indicates duplication — confirm none of these are "
                       f"actually RV / Storage / Retail."),
        })

    # NOI sanity — must be positive at the underwritten level.
    inc_sum = sum(float(i.get("ggcUnderwritten") or 0) for i in income)
    exp_sum = sum(float(e.get("ggcUnderwritten") or 0) for e in expenses)
    noi = inc_sum - exp_sum
    if inc_sum > 0:
        if noi <= 0:
            checks.append({
                "item": "Underwritten NOI",
                "check": "NOI > 0",
                "status": "fail",
                "detail": (f"NOI = ${noi:,.0f} (income ${inc_sum:,.0f} − "
                           f"expenses ${exp_sum:,.0f}). Deal does not cover "
                           f"its operating cost."),
            })
        else:
            ratio = exp_sum / inc_sum if inc_sum else 0
            status = ("warn" if (ratio < 0.20 or ratio > 0.65) else "ok")
            checks.append({
                "item": "Expense ratio",
                "check": "25–55% typical for MHC, 35–50% for hybrid",
                "status": status,
                "detail": f"Underwritten expense ratio = {ratio:.1%}",
            })

    # Surface any methodology Pydantic validation issues. These are
    # advisory — a monthly-tie mismatch within ~5-15% usually means the
    # LLM applied a methodology adjustment to an intermediate field, not
    # that the deal is broken. We treat the structural ones (bad
    # ggcCategory enum) as warns, and elevate to fail only when many
    # items share the same root issue (signals the LLM was confused
    # across the board, not noisy on a single line).
    methodology_issues = financials.get("_methodologyValidation") or []
    severity_for_issues = "warn" if len(methodology_issues) < 20 else "fail"
    for msg in methodology_issues:
        checks.append({
            "item": "Methodology schema validation",
            "check": "Per-line Pydantic check",
            "status": severity_for_issues,
            "detail": msg,
        })

    checks.extend(_check_template_wiring(financials))

    # ── EGI − OpEx = NOI identity (§7 T-12 check #3). Catches a stale
    # propertyInfo.noi (or a methodology that under-summed income lines).
    prop = financials.get("propertyInfo") or {}
    egi  = sum(float(i.get("ggcUnderwritten") or 0) for i in income)
    opex = sum(float(e.get("ggcUnderwritten") or 0) for e in expenses)
    implied_noi = egi - opex
    reported_noi = prop.get("noi")
    if isinstance(reported_noi, (int, float)) and reported_noi != 0:
        spread = abs(implied_noi - reported_noi) / max(abs(reported_noi), 1)
        pct = spread * 100
        if pct <= 0.5:
            status, prefix = "ok",   "EGI − OpEx = NOI"
        elif pct <= 2:
            status, prefix = "warn", "EGI − OpEx ≈ NOI"
        else:
            status, prefix = "fail", "EGI − OpEx ≠ NOI"
        checks.append({
            "item": "NOI identity",
            "check": prefix,
            "status": status,
            "detail": (f"Σincome.ggcUnderwritten ${egi:,.0f} − "
                       f"Σexpense.ggcUnderwritten ${opex:,.0f} = "
                       f"${implied_noi:,.0f}; propertyInfo.noi reports "
                       f"${reported_noi:,.0f} ({pct:.1f}% off)"),
        })

    # ── Unit/pad ID uniqueness — if per-row data is present. The rent
    # roll's COUNTIFS in the template assumes unique unit IDs; duplicates
    # would inflate occupancy counts and double-charge GPR.
    rows = (financials.get("rentRoll") or {}).get("rentRollRows") or []
    if isinstance(rows, list) and rows:
        ids = [(r.get("unitId") or "").strip() for r in rows
               if isinstance(r, dict)]
        ids = [i for i in ids if i]
        if ids:
            seen, dupes = set(), []
            for uid in ids:
                if uid in seen:
                    dupes.append(uid)
                seen.add(uid)
            if dupes:
                unique_dupes = sorted(set(dupes))
                preview = ", ".join(unique_dupes[:5])
                more = f" (+{len(unique_dupes) - 5} more)" if len(unique_dupes) > 5 else ""
                checks.append({
                    "item": "Rent-roll unit ID uniqueness",
                    "check": "All unitIds unique",
                    "status": "fail",
                    "detail": (f"Duplicate unit IDs: {preview}{more}. "
                               "COUNTIFS will over-count occupancy."),
                })
            else:
                checks.append({
                    "item": "Rent-roll unit ID uniqueness",
                    "check": "All unitIds unique",
                    "status": "ok",
                    "detail": f"{len(ids)} unique unit IDs.",
                })

    return checks


def _check_template_wiring(financials):
    """§10.4 NOI traceback: simulate what the Underwriting tab's NOI cell
    (I47) will compute to, given the values fill_template is about to
    write. I47 is built from SUMIFS over Data Consolidation column G
    (t12Total), keyed by EXACT ggcCategory strings. Any income/expense
    item whose ggcCategory is not in the canonical lists is silently
    invisible to the SUMIFS — its value vanishes from the workbook NOI
    with no error. This check catches that before write-back so the
    verification gate can block the run.

    Even when every category is canonical, this surfaces the implied
    NOI as a value the reviewer can compare against the workbook's I47
    after opening it. That's the actionable "tie-out" piece §10.4 wants.
    """
    checks = []
    income = financials.get("income") or []
    expenses = financials.get("expenses") or []
    inc_canonical = set(GGC_INCOME_CATEGORIES)
    exp_canonical = set(GGC_EXPENSE_CATEGORIES)

    def _split(items, canonical):
        in_total, out_total = 0.0, 0.0
        bad = []
        for it in items:
            v = float(it.get("t12Total") or it.get("ggcUnderwritten") or 0)
            cat = (it.get("ggcCategory") or "").strip()
            if cat in canonical:
                in_total += v
            else:
                out_total += v
                if cat:
                    bad.append((cat, v))
        return in_total, out_total, bad

    in_inc, out_inc, bad_inc = _split(income, inc_canonical)
    in_exp, out_exp, bad_exp = _split(expenses, exp_canonical)
    expected_egi  = in_inc + out_inc
    expected_opex = in_exp + out_exp
    expected_noi = expected_egi - expected_opex
    implied_noi  = in_inc - in_exp

    if bad_inc or bad_exp:
        bad_strings = sorted(set([c for c, _ in bad_inc + bad_exp]))
        preview = ", ".join(repr(c) for c in bad_strings[:5])
        more = f" (+{len(bad_strings) - 5} more)" if len(bad_strings) > 5 else ""
        loss = expected_noi - implied_noi
        checks.append({
            "item": "Underwriting NOI traceback",
            "check": "All ggcCategory strings match SUMIFS criteria",
            "status": "fail",
            "detail": (f"${(out_inc + out_exp):,.0f} of value lives under "
                       f"non-canonical category string(s) {preview}{more}. "
                       f"These will not match the Underwriting!I47 SUMIFS "
                       f"criteria. Expected NOI ${expected_noi:,.0f}; the "
                       f"workbook will compute ${implied_noi:,.0f} "
                       f"(off by ${loss:,.0f})."),
        })
    else:
        checks.append({
            "item": "Underwriting NOI traceback",
            "check": "Implied Underwriting!I47",
            "status": "ok",
            "detail": (f"All categories canonical. Implied NOI "
                       f"${implied_noi:,.0f} = EGI ${in_inc:,.0f} − "
                       f"OpEx ${in_exp:,.0f}. Compare against I47 "
                       f"after opening the workbook."),
        })

    return checks


# ═══════════════════════════════════════════════════════════════════════════
# N=3 EXTRACTION MAJORITY VOTE (Wang et al. self-consistency)
# Runs extraction in parallel N times and field-merges. Median across runs
# for numerics, mode for strings, union for nested lists. Surfaces a
# "Run-to-run disagreement" check when annualTotals diverge >5% between
# runs — that signal is more useful than the median itself.
# Gated by deep_search so the cost (≈N× extraction tokens) is opt-in.
# ═══════════════════════════════════════════════════════════════════════════

def call_extract_financials_merged(api_key, file_blocks, property_info, n_runs=3):
    print(f"[Claude] Starting {n_runs}× extraction merge...")
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=n_runs) as executor:
        futures = [executor.submit(call_extract_financials, api_key,
                                   file_blocks, property_info)
                   for _ in range(n_runs)]
        results = []
        for i, f in enumerate(as_completed(futures)):
            try:
                results.append(f.result())
                print(f"[Claude] Extraction run {i+1}/{n_runs} complete")
            except Exception as e:
                print(f"[Claude] Extraction run {i+1}/{n_runs} FAILED: {e}")

    if not results:
        raise RuntimeError(f"All {n_runs} extraction runs failed")
    print(f"[Claude] Merged {len(results)} runs in {time.time() - t0:.1f}s")
    return _merge_extraction(results)


def _median_of(vs):
    nums = [v for v in vs if isinstance(v, (int, float))]
    return statistics.median(nums) if nums else None


# Confidence-weighted-median helpers. The LLM emits a confidence enum
# {high, medium, low} per line item but the original merge step
# discarded it — so a high-confidence outlier counted the same as two
# low-confidence inliers (and vice versa). Re-introducing the weight
# pushes the merged value toward whichever runs the model was most
# certain about.
_CONF_WEIGHT = {"high": 3, "medium": 2, "low": 1}

def _weighted_median(values_with_weights):
    """Weighted median: lowest-key value v where cumulative weight
    crosses 50% of total. Falls back to plain median when all weights
    are equal or missing. values_with_weights is an iterable of
    (value, weight) — values that are None are skipped."""
    pairs = [(float(v), max(int(w or 0), 0))
             for v, w in values_with_weights
             if isinstance(v, (int, float))]
    pairs = [(v, w) for v, w in pairs if w > 0]
    if not pairs:
        return None
    pairs.sort(key=lambda p: p[0])
    total = sum(w for _, w in pairs)
    half = total / 2
    running = 0
    for v, w in pairs:
        running += w
        if running >= half:
            return v
    return pairs[-1][0]


def _median_with_confidence(items, value_field, confidence_field="confidence"):
    """Confidence-weighted median across a list of dicts. Each item
    contributes its value_field weighted by _CONF_WEIGHT[confidence_field].
    Items missing both weight and confidence default to medium."""
    pairs = []
    for it in items:
        if not isinstance(it, dict):
            continue
        v = it.get(value_field)
        conf = (it.get(confidence_field) or "medium").lower()
        w = _CONF_WEIGHT.get(conf, _CONF_WEIGHT["medium"])
        pairs.append((v, w))
    return _weighted_median(pairs)


def _mode_with_confidence(items, value_field, confidence_field="confidence"):
    """Mode of a categorical field, weighted by confidence so a
    high-confidence run can outvote two low-confidence runs that agree."""
    weighted = Counter()
    for it in items:
        if not isinstance(it, dict):
            continue
        v = it.get(value_field)
        if v in (None, ""):
            continue
        conf = (it.get(confidence_field) or "medium").lower()
        w = _CONF_WEIGHT.get(conf, _CONF_WEIGHT["medium"])
        weighted[v] += w
    return weighted.most_common(1)[0][0] if weighted else ""


def _mode_of(vs):
    nons = [v for v in vs if v not in (None, "")]
    return Counter(nons).most_common(1)[0][0] if nons else ""


def _merge_extraction(runs):
    """Field-level median for numerics, mode for strings."""
    # ── reportingPeriod: take the modal periodUsed across runs ───────────
    rp_runs = [r.get("reportingPeriod", {}) for r in runs]
    merged_rp = {
        "periodUsed": _mode_of([r.get("periodUsed") for r in rp_runs]),
        "dateRange":  _mode_of([r.get("dateRange")  for r in rp_runs]),
        "monthsCovered": (
            int(_median_of([r.get("monthsCovered") for r in rp_runs
                            if r.get("monthsCovered") is not None]))
            if any(r.get("monthsCovered") is not None for r in rp_runs)
            else None
        ),
        "candidatePeriodsSeen": sorted({
            p for r in rp_runs for p in (r.get("candidatePeriodsSeen") or [])
        }),
        "notes": " | ".join(sorted({
            r.get("notes", "") for r in rp_runs if r.get("notes")
        })),
    }

    # ── income / expenses: match line items by sellerLabel, median fields ──
    def _merge_lines(runs_lists, section_value):
        by_label = {}
        for run_list in runs_lists:
            for ln in run_list or []:
                key = (ln.get("sellerLabel") or "").strip().lower()
                if not key:
                    continue
                by_label.setdefault(key, []).append(ln)

        merged = []
        disagreements = []
        for key, items in by_label.items():
            # Confidence-weighted median: a high-confidence run pulls the
            # merged value toward its number; two low-confidence runs that
            # agree on a wrong figure no longer outvote one high-confidence
            # correct figure.
            annual = _median_with_confidence(items, "annualTotal")
            # detect run-to-run disagreement on this line
            non_null_annuals = [i.get("annualTotal") for i in items
                                if isinstance(i.get("annualTotal"), (int, float))]
            if len(non_null_annuals) >= 2 and max(non_null_annuals) != 0:
                spread = ((max(non_null_annuals) - min(non_null_annuals))
                          / abs(max(non_null_annuals)))
                if spread > 0.05:
                    disagreements.append({
                        "label":  items[0].get("sellerLabel", key),
                        "range":  (min(non_null_annuals), max(non_null_annuals)),
                        "spread": spread,
                    })

            # monthly: confidence-weighted median per index across runs that
            # returned 12 values. Each contributing run carries its own
            # confidence weight at every monthly position.
            monthly_with_weights = [
                (i.get("monthly"),
                 _CONF_WEIGHT.get((i.get("confidence") or "medium").lower(),
                                   _CONF_WEIGHT["medium"]))
                for i in items
                if isinstance(i.get("monthly"), list) and len(i.get("monthly")) == 12
            ]
            if monthly_with_weights:
                monthly = [
                    _weighted_median([(m[k], w) for m, w in monthly_with_weights])
                    for k in range(12)
                ]
            else:
                monthly = None

            # Surface the merged confidence so the methodology stage knows
            # how much to trust each line. Use the minimum: if any
            # contributing run was "low", the merged item is at best "low".
            confidences = [(i.get("confidence") or "").lower() for i in items]
            confidences = [c for c in confidences if c in _CONF_WEIGHT]
            if confidences:
                merged_conf = min(confidences,
                                  key=lambda c: _CONF_WEIGHT.get(c, 0))
            else:
                merged_conf = "medium"

            merged.append({
                "sellerLabel": items[0].get("sellerLabel", key),
                "annualTotal": annual,
                "monthly":     monthly,
                "section":     section_value,
                "isSubtotal":  any(i.get("isSubtotal") for i in items),
                "confidence":  merged_conf,
            })
        return merged, disagreements

    income, inc_disagree = _merge_lines(
        [r.get("income") for r in runs], "income")
    expenses, exp_disagree = _merge_lines(
        [r.get("expenses") for r in runs], "expense")

    # ── rentRoll: median scalars, union of unitTypes by name ─────────────
    rr_runs = [r.get("rentRoll", {}) for r in runs]
    merged_rr = {
        "totalRowsInRentRoll":   int(_median_of([r.get("totalRowsInRentRoll") for r in rr_runs]) or 0) or None,
        "statedTotalRentMonthly": _median_of([r.get("statedTotalRentMonthly") for r in rr_runs]),
        "statedTotalIsMonthly":  bool(rr_runs[0].get("statedTotalIsMonthly", True)) if rr_runs else True,
        "occupiedCount":         int(_median_of([r.get("occupiedCount") for r in rr_runs]) or 0) or None,
        "vacantCount":           int(_median_of([r.get("vacantCount")   for r in rr_runs]) or 0) or None,
    }
    types_by_name = {}
    for r in rr_runs:
        for ut in r.get("unitTypes") or []:
            name = (ut.get("unitType") or "").strip().lower()
            if not name:
                continue
            types_by_name.setdefault(name, []).append(ut)
    merged_types = []
    for name, items in types_by_name.items():
        merged_types.append({
            "unitType":           items[0].get("unitType", name),
            "count":              int(_median_of([i.get("count") for i in items]) or 0),
            "occupiedCount":      int(_median_of([i.get("occupiedCount") for i in items]) or 0),
            "vacantCount":        int(_median_of([i.get("vacantCount") for i in items]) or 0),
            "avgLotRentOccupied": _median_of([i.get("avgLotRentOccupied") for i in items]),
            "hasHomeRentEntries": any(i.get("hasHomeRentEntries") for i in items),
            "avgHomeRent":        _median_of([i.get("avgHomeRent") for i in items]),
        })
    merged_rr["unitTypes"] = merged_types

    # ── documentsSeen + extractionNotes: union / concat ──────────────────
    docs = sorted({d for r in runs for d in (r.get("documentsSeen") or [])})
    notes = " || ".join(sorted({
        r.get("extractionNotes", "") for r in runs if r.get("extractionNotes")
    }))

    # Add a synthetic line so the disagreement is visible downstream
    if inc_disagree or exp_disagree:
        all_disagree = inc_disagree + exp_disagree
        sample = ", ".join(
            f"{d['label']} ${d['range'][0]:,.0f}–${d['range'][1]:,.0f} "
            f"({d['spread']*100:.0f}%)"
            for d in all_disagree[:5]
        )
        notes = (
            (notes + " || " if notes else "")
            + f"Run-to-run disagreement on {len(all_disagree)} line item(s) "
            + f"across {len(runs)} extraction runs. Sample: {sample}"
        )

    return {
        "reportingPeriod": merged_rp,
        "income":          income,
        "expenses":        expenses,
        "rentRoll":        merged_rr,
        "documentsSeen":   docs,
        "extractionNotes": notes,
        "_meta": {
            "merge_runs":   len(runs),
            "disagreements": len(inc_disagree) + len(exp_disagree),
        },
    }


# ═══════════════════════════════════════════════════════════════════════════
# CLAUDE CALL #2 — METHODOLOGY (Fable 5, adaptive thinking, judgment)
# Takes the CLEAN extracted data from Call 1 and applies GGC's categorization
# and underwriting methodology. Because the numbers are already extracted and
# verified, this call focuses purely on analysis instead of fighting raw PDFs.
# ═══════════════════════════════════════════════════════════════════════════
FINANCIAL_PARSE_PROMPT = f"""You are a real estate underwriting analyst at Gary Group Capital (GGC), a private equity firm focused on mobile home parks.

You are given CLEAN, PRE-EXTRACTED financial data that a separate extraction step has already pulled from the seller's documents. The numbers have already been transcribed and the correct reporting period has already been identified for you. Your job is NOT to re-read raw documents — it is to apply GGC's categorization and underwriting methodology to the clean data you are given.

Trust the extracted numbers as your source of truth. The annual totals and monthly arrays in the provided data are what you work from. Map every extracted line item to GGC's standardized categories and apply the methodology below.

CRITICAL OUTPUT RULES:
- Your response MUST start with `{{` (a JSON open brace) — NOTHING before it
- Your response MUST end with `}}` (a JSON close brace) — NOTHING after it
- DO NOT add preamble, explanation, "I'll analyze...", or commentary of any kind
- DO NOT wrap the JSON in markdown code fences
- If the user-provided property info has typos or inconsistencies, use your best interpretation silently — DO NOT explain corrections in the output
- Notes should go in the "notes" field of each line item, NEVER as freestanding text

## GGC Income Categories (use EXACTLY these strings, including punctuation and capitalization — schema validation rejects deviations):
{json.dumps(GGC_INCOME_CATEGORIES)}

## GGC Expense Categories (use EXACTLY these strings, including punctuation. 'Home Rent Expense (MH)' has the parenthetical, 'Cap-Ex Reserve' has the hyphen):
{json.dumps(GGC_EXPENSE_CATEGORIES)}

## Canonical Unit-Type Taxonomy (use EXACTLY these strings — Unit Mix Summary COUNTIFS keys on these exact labels):
{json.dumps(CANONICAL_UNIT_TYPES)}

Map seller's unit-type labels to one of the four canonical buckets:
- Lot Rent / Site Rent / TOH (tenant-owned home) lots / Pad rent → "TOH MH Site"
- POH Rent / Park-owned Home / Rental Home / Infilled Home → "POH-Infilled units"
- RV Lot / RV Site / Annual RV / Long-Term RV → "Long term RV Site"
- Storage / Retail / Commercial Space / Storefront → "Retail/Commercial"
Preserve the raw seller label in `sellerUnitLabel` for audit traceability.

## Income Categorization — GL Account Mapping (CRITICAL: emit ONE ROW per seller GL account)

Sellers' charts of accounts vary, but the GGC bucketing follows account-number prefixes consistently. Use this mapping. EMIT ONE ROW PER SELLER GL ACCOUNT — do NOT aggregate multiple GL accounts that share a GGC category into a single row, AND do NOT emit subtotal/"non-posting" rows from the source P&L (e.g. "4100 Total Rental Income (non-posting)", "5700 Total Personnel"). These are display artifacts in the seller's chart of accounts; emit only the leaf-level GL accounts beneath them, with their actual t12Total values preserved exactly as the extracted data shows.

INCOME:
| Seller GL prefix | Example label | GGC ggcCategory |
|---|---|---|
| 4101 | Lot Rent / Site Rent / Pad Rent | "Gross Potential Rent" |
| 4103 | Long Term RV Lot Rent / RV Site | "RV Site Rental Income" |
| 4108 | Storage Unit Rent / Boat Storage / Parking | "Parking Income" |
| 4110 | Retail Unit Rent / Commercial Space / Storefront | "Retail" |
| 4131 / "Move-in Specials" | Move-in Specials | "Omitt Income" when sellerNotes contains Discontinued / Non-recurring / Seller-Specific; else "Other Income" |
| 6120 / "Bad Debt" | Bad Debt | "Bad Debt" (NEGATIVE) — sign is forced negative by the GGC override pass |
| 4304 | Damages | "Other Income" |
| 4402 negative (water/sewer non-recurring recovery / refund) | Water & Sewer refund | "Omitt Income" (non-operating) |
| 4403 / 4404 | Electric / Garbage tenant pass-through | "Utility Reimbursement" |
| 4905 | Recovered Legal Fees | "Other Income" |
| 4907 | Violation Fines | "Omitt Income" (non-operating fines — gold-standard routing) |
| 4908 | Payment Processing Fee | "Other Income" |
| 4909 | Cable Revenue Sharing | "Other Income" |
| 4910 | Rental Pool Revenue Sharing | "Other Income" |
| 4913 | Application Fees | "Other Income" |
| 4914 | Late Fees | "Other Income" |
| 4915 | NSF Fees | "Other Income" |
| Home Rent / POH Rent / Lease-to-Own income | Home Rent | "Home Rent Income" |

DECISION RULES (income):
- HARD RULE: any line whose sellerNotes field contains "Discontinued", "Non-recurring", "Seller Specific", "Seller-Specific", "One-time", or "Non-operating" → "Omitt Income". This overrides the GL-prefix routing above. The Omitt bucket is the gold-standard exclusion path for items GGC will not underwrite forward.
- "Utility Reimbursement" = tenant pass-through of metered/billed utility consumption (water, sewer, electric, gas, trash). If the line item represents a CONSUMPTION pass-through to a tenant, it's Utility Reimbursement.
- "Other Income" = revenue-sharing arrangements (cable, internet, laundry, vending), application/late/NSF/pet fees, damages, legal recoveries. If the line is a fee, fine, or revenue share rather than a utility pass-through, it's Other Income.
- NEGATIVE income amounts that ARE NOT marked non-recurring: treat as "Other Income" with a `notes` flag explaining the contra. If material negative AND marked non-recurring → "Omitt Income".
- 5407 Tenant Cable TV: in the income block → "Utility Reimbursement"; in the expense block → "G&A".
- HARD RULE — NEVER prefix categories with "Less: ". The strings "Less: Vacancy", "Less: Concessions", "Less: Bad Debt" are template DISPLAY LABELS in the Underwriting tab; they are NOT GGC categories. The SUMIFS in the Underwriting tab look up "Bad Debt" (not "Less: Bad Debt"), "Vacancy" (not "Less: Vacancy"), etc. Emit the BARE category strings only. A "Less: " prefix silently zeros the line.
- Vacancy is computed in the Underwriting tab from the rent roll. Do NOT emit a "Vacancy" line in Data Consolidation unless the seller's P&L has a dedicated vacancy line item with a non-zero T-12 number — and in that case emit the bare "Vacancy" category, not "Less: Vacancy".
- Concessions: route to "Omitt Income" (typical case: discontinued move-in specials) or "Other Income" depending on the sellerNotes. The bare string "Concessions" is acceptable when the seller's P&L has an explicit concessions line.

EXPENSE BUCKETING:
| Seller GL | GGC ggcCategory |
|---|---|
| 5301 Property Tax | "RE Taxes" |
| 5053 Liability Insurance | "Insurance" |
| 5051 Car / Vehicle Insurance | "Omitt Expense" (vehicles are NOT property opex) |
| 5402 Water & Sewer / 5403 Water Testing | "Water and Sewer" |
| 5404 Electric | "Electricity" |
| 5405 Garbage / Trash | "Trash Removal" |
| 5406 Gas / Propane | "Gas/Fuel" |
| 5401 Vehicle Fuel | "Omitt Expense" (vehicle costs are not opex) |
| 5102 Tree / 5104 Grounds / 5103 Pest | "Ground Maintenance" |
| 5107 Septic / 5108 Plumbing / 5109 Misc / 5110 Equipment / 5111 Electrical / 5200 Supplies | "Repair and Maintenance" |
| 5409 Rentals - Coin Laundry (laundry equipment lease) | "Repair and Maintenance" |
| 5000 Management Fees | "Management Fee" (OVERRIDDEN by GGC's % of EGI in a later pass) |
| 5700 series leaf GLs (5701 Wages, 5702 Health Ins, 5703 Casual Labour, 5704 UI, 5705 Payroll Svc, 5706 FUTA, 5708 SS Tax, 5710 Mgr Salary Allocation, 5713 OR WBF, 5716 Workers Comp) | "Payroll" — emit ONE ROW PER GL, NEVER the "5700 Total Personnel" subtotal |
| 5070 Licenses & Permits / 5072 Dues / 5601 Office Supplies / 5602 Internet / 5603 Telephone / 5606 Background Checks / 5650 Bank Fees / 5407 Tenant Cable TV | "G&A" |
| 5605 Postage | "Omitt Expense" when sellerNotes marks it Seller-Specific; else "G&A" |
| 5061 / 5062 / 5066 Professional | "Professional Fees" |
| 5001 Advertising | "Advertising" |
| 5113 Home Repairs / POH Maintenance / "Home" labels | "Home Rent Expense (MH)" |
| 5300 Cap-Ex | "Cap-Ex Reserve" (OVERRIDDEN to $75/site/year) |

DECISION RULES (expense):
- HARD RULE: any expense line whose sellerNotes contains "Discontinued", "Non-recurring", "Seller Specific", "Seller-Specific", "One-time", or "Non-operating" → "Omitt Expense". Same logic as the income side; the Omitt bucket is the gold-standard exclusion path.
- HARD RULE: vehicle-related lines (Car Insurance, Vehicle Fuel, Vehicle Maintenance) → "Omitt Expense" regardless of GL prefix. Vehicles are seller-owned, not property opex.
- HARD RULE: NEVER emit a subtotal row. Lines whose sellerLabel contains "Total", "non-posting", or "Subtotal" are display artifacts. Emit only the leaf GLs beneath them, each with its own row, each with its own t12Total preserved exactly from the source.
  - WORKED EXAMPLE — the seller's P&L shows: `5700 Personnel (non-posting)` (header) then `5701 Wages $48,650`, `5702 Health Ins $1,026`, `5703 Casual Labour $22,505`, `5704 UI $3,118`, `5705 Payroll Svc $1,475`, `5706 FUTA $93`, `5708 SS $4,432`, `5713 OR-WBF $251`, `5716 Workers Comp $1,660`, then `5700 Total Personnel $83,210`. CORRECT methodology output: 9 rows (one per leaf GL, each ggcCategory="Payroll", each with its real t12Total). WRONG output: a single row with sellerName="5700 Total Personnel" and t12Total=$0 — that drops every payroll dollar from the workbook.
- HARD RULE: every emitted row MUST have a numeric t12Total (zero is allowed; null is NOT allowed when the source P&L shows a value). If the extracted data has a value, the methodology row must carry it through. A row whose sellerName contains "Total"/"Subtotal"/"non-posting" AND has t12Total=0 across all value columns will be DROPPED by the write-back as a placeholder — emit the leaf GLs instead.

REJECT any deviation from the exact category strings above. The downstream Excel SUMIFS keys on these exact strings; even a trailing space or different capitalization will silently zero out the line.

## GGC Underwriting Methodology

### COLLECTIONS METHODOLOGY (NRI = Net Rental Income)

Follow this exact 4-step sequence. Do not deviate.

DEFINITIONS:
- NRI (Net Rental Income) = GPR − Vacancy − Concessions − Bad Debt. This is the standard industry term; many sellers and brokers refer to "collections" when they mean NRI.
- EGI (Effective Gross Income) = NRI + Other Income.
- Other Income excludes Home Rent (Home Rent is its own income stream and gets bifurcated, see POH section).

STEP 1 — GPR (Gross Potential Rent):
Pull directly from the rent roll.

- For OCCUPIED units: use contracted lot rent from the rent roll.
- For VACANT units WITH a market rent listed: use that market rent.
- For VACANT units WITHOUT a market rent (i.e. $0, blank, or missing): impute the market rent as the AVERAGE of occupied lot rents within the same unit type. Calculate this per unit type — a vacant Premium Lot uses the average of occupied Premium Lots, not the average of all occupied lots.
- If there is only ONE unit type in the rent roll, use the overall average of occupied rents.
- If a unit type has NO occupied units to average from (all vacant), use the average of all occupied lot rents at the property as the fallback.

- Sum the above to get GPR. This represents the 100%-occupied, market-rate ceiling.
- Note in the income line item "notes" field if you imputed market rents for any vacant units, and how many were imputed.

STEP 2 — Physical Vacancy:
 - Set from the rent roll occupancy. This is the actual physical vacancy at the property today (vacant units × market rent, as a negative line item). Do NOT use historical vacancy from the P&L for this line — physical vacancy ties to the rent roll only.

STEP 3 — Concessions:
Tie directly to T12 historical concessions. No annualization adjustment, no trend factor. If concessions are zero in the T12, set to zero.

STEP 4 — Bad Debt (the "what-if" / goal-seek step):
- This is the plug that makes underwritten NRI tie to the trend in historical collections. The logic:
  (a) Calculate underwritten NRI target: if collections are trending UP, target = T3 annualized. If flat/fluctuating, target = T12. If trending DOWN, target = T3 annualized AND flag heavily (ask why collections are declining).
  (b) Solve for Bad Debt so that: GPR − Physical Vacancy − Concessions − Bad Debt = Target NRI.
  (c) The resulting Bad Debt is the underwritten figure. Sanity check it against historical bad debt — if the underwritten bad debt diverges materially from what the seller's P&L shows, flag it. A large divergence usually means either (i) the seller is hiding something, (ii) physical occupancy changed recently, or (iii) the T3 trend is being driven by a one-time event.

- If there is a distorting one-time item in the T3 window (e.g. a one-month bad debt recovery that inflates a single month), use T6 annualized as the NRI target instead of T3, and note it.
- Output a "notes" field on the Bad Debt line item showing: which NRI target was used (T3/T6/T12), the target $ figure, and any flags.

STEP 4.5 — Income Spike Detection (pre-paid rent / lump sum):

Before annualizing T3 or T6, scan the monthly income data for anomalous spikes — single months where rental income or other income is materially above the surrounding months.

A spike is defined as: any month where a single income line is ≥1.5× the average of the other 11 months in the same line item.

Common causes:
- Tenants pre-paying multiple months at once (residents do this to avoid eviction or to lock in current rent)
- A lump-sum bad debt recovery (someone paid back a large past-due balance)
- A one-time lease-up bonus, settlement, or insurance proceeds miscategorized as rental income

If a spike is detected:
- Do NOT include the spike month in any T3 or T6 annualization calculation. Replace it with the average of the surrounding 11 months when computing T3 or T6.
- Flag the spike in the flags array with severity "medium": "Income spike detected in [month] for [line item]: [amount] vs. [avg] average. Likely pre-paid rent or lump-sum recovery. Excluded from T3/T6 annualization but kept in T12 total. Confirm with seller."
- DO include the spike in the T12 total (because T12 is the trailing reality, not the run-rate forecast).

- This is symmetric with how expense spikes are handled — flag, don't auto-strip the historical, but exclude from annualized forecasts.

### CROSS-CHECK OUTPUT (mandatory)

- In addition to the standard JSON output, include a "sourceMapping" object that shows, for each major line item, which source column/cell the dollar value came from. Format:

"sourceMapping": [
  "gpr": "Rent Roll page 8 'Total Possible Rent' $125,695 × 12 = $1,508,940",
  "reTaxes": "T-12 col 'T-12 Ended 5/23' row '6810·TAXES-PROPERTY' = $44,798",
  "insurance": "T-12 col 'T-12 Ended 5/23' row '6450 INSURANCE' = $50,462",
  ...
]

This lets the reviewer trace every number back to its source.

### BROKER PROFORMA EXTRACTION (REQUIRED — DO NOT EMIT NULL)

If the seller's P&L contains a column titled "PRO FORMA", "Broker Proforma",
"Broker's Proforma", "Y1 Proforma", "Forecast Y1", or any similar header,
you MUST populate `brokerProforma` for EVERY income AND expense line item.
Copy the value directly from that column. Even when the broker's forecast
matches the T12 exactly, write the number — do NOT emit `null`.

Rules:
- If the line item has an explicit value in the Pro Forma column, use it.
- If the Pro Forma cell is empty for a particular line item but the column
  exists in the source, use `t12Total` as the fallback value (NOT null).
- If the P&L genuinely has NO Pro Forma column anywhere, then and only then
  may you emit `null` for `brokerProforma` on every line.
- The downstream Excel template uses column F (Broker Proforma) as a
  side-by-side comparison against T12 and the GGC underwritten figure.
  When you emit null, that column reads as blank and the reviewer loses
  the broker-vs-actuals reconciliation entirely.

### GPR — MULTI-COLUMN SPREADSHEET DISAMBIGUATION (READ CAREFULLY)

Seller financials often contain MULTIPLE potential GPR figures. You must pick
the right one. Common traps:

1. **Multiple T-12 columns in one spreadsheet:**
   The seller may show several annual periods side-by-side (e.g. "T-12 Ended 9/22",
   "T-12 Ended 5/23", "Oct 22 - May 23"). Use the MOST RECENT T-12 ending date.
   The "Oct 22 - May 23"-style header is a PARTIAL period (8 months) — IGNORE it.
   The "T-12 Ended X/YY" column is the annual figure — USE THAT.

2. **Rent roll totals are MONTHLY, not annual:**
   A rent roll's "Total Possible Rent" or "Totals for [property]" row shows the
   MONTHLY rent roll total. To get annual GPR from the rent roll, multiply by 12.

3. **Sanity check (MANDATORY — apply for every deal):**
   - Expected Annual GPR ≈ Total Units × Average Lot Rent × 12
   - For Las Brisas: 295 units × ~$450/mo × 12 = ~$1.59M annual
   - If your parsed GPR is < 50% of this expected figure, you have grabbed the
     wrong column. Re-check the source and pick the column that matches the
     expected magnitude.
   - Output a note explaining which column/source you used and why.

4. **Sub-line items with the property name:**
   A line labeled "RENTAL INCOME [PROPERTY NAME]" or "4015 · RENTAL INCOME LAS
   BRISAS" is STILL rental income. Bucket it under Gross Potential Rent, not
   Other Income. The property name in the label does not change the category.

### DATA QUALITY CROSS-CHECKS (run these first, before any underwriting math)

The user provides two counts in the property form: Total Units and Park-Owned Home (POH) Count. Both must reconcile against the rent roll. If they don't, you must flag it explicitly — these mismatches almost always indicate missing data, not a clean dataset.

CHECK 1 — Total Units vs. Rent Roll Rows:
Compare the user-entered "Total Units" against the number of rent roll rows you find in the uploaded documents.

- If MATCH: proceed normally.
- If RENT ROLL HAS FEWER ROWS than Total Units: this is the most common case. Sellers often omit vacant sites from the rent roll. Assume the missing rows are vacant lots and include them in GPR at market rent (use the average of occupied lot rents in the same unit type as the imputed market rent). Add a "high" severity flag to the flags array with the exact wording: "Rent roll shows [N rows] but property is [Total Units] units — assumed [difference] additional vacant lots at market rent. Confirm with broker: are vacant sites excluded from the rent roll, or is the property actually smaller than stated?"
- If RENT ROLL HAS MORE ROWS than Total Units: flag as "high" severity. This means the user entered the wrong unit count or the rent roll includes something non-unit (model homes, storage spaces, common-area structures). Do not silently override — ask the user to confirm.

CHECK 2 — POH Count vs. Rent Roll Home Rent Entries:
Compare the user-entered "Park-Owned Home Count" against the number of rent roll rows where Home Rent > 0.

- If MATCH (within ±2): proceed normally, no flag.
- If USER SAYS MORE POH than rent roll shows home rent for: the seller is likely hiding home rent income (or comping the home for the on-site manager). Flag as "medium" severity: "User stated [X] POH but only [Y] rent roll entries show home rent. Possible employee allowance (comped lot for manager) or hidden home rent income. Verify with seller."
- If USER SAYS FEWER POH than rent roll shows home rent for: the user count is probably wrong, or some rent roll "home rent" entries are actually other charges miscategorized. Flag as "medium" severity and use the rent roll count as the authoritative POH number.
- If USER STATED 0 POH but rent roll has home rent entries: flag as "high" severity. The user either didn't know or input wrong. Use rent roll count.

POH PERCENTAGE CALCULATION:
After reconciliation, compute POH % = (POH Count / Total Units) × 100. Use this number throughout the rest of the analysis (affects expense ratio benchmarks, NOI bifurcation, and risk flags).

Output a "dataQualityChecks" object in the JSON response showing what you found:
- totalUnitsStated (integer, user input)
- rentRollRowsFound (integer, what you parsed)
- unitCountReconciliation ("match" | "rent_roll_short" | "rent_roll_long")
- pohCountStated (integer, user input)
- pohRowsFound (integer, what you parsed from rent roll)
- pohReconciliation ("match" | "stated_high" | "stated_low" | "stated_zero_but_found")
- finalPohCount (integer, the number to use going forward)
- finalPohPercent (decimal)

### OTHER INCOME
- All other income items (laundry, application fees, pet fees, month-to-month premiums, storage, etc.): use T12 as-is, no annualization adjustment
- Do not apply the T3 annualization logic to other income — only to lot rent collections

### EXPENSES (general rule)
- Most expenses: T12 * 1.03 (trailing 12 months plus 3% inflation factor for year one)
- The 3% reflects standard annual inflation assumption going into year one of ownership
- If an expense line appears materially above or below market on a per-unit basis, flag it and use your judgment — do not blindly apply T12 * 1.03 if the number is clearly distorted

### TAXES (post-acquisition reassessment)

- Underwritten taxes must reflect post-sale reassessment, not the seller's historical number.

- PRIMARY METHOD:
  - Underwritten Taxes = (Purchase Price × 0.65) × Local Tax Rate

- The new assessed value is typically 60-70% of purchase price; use 65% as the default. Apply the local tax rate (look up the millage rate or effective rate for the county/township).

- FALLBACK RULE (when you cannot find a clean tax rate or when the calculation looks suspect):
  - Underwritten Taxes = Historical T12 Taxes × 1.15

- SANITY CHECK (always apply):
  - If the primary method produces a number LOWER than historical T12 taxes, the primary method is wrong. Override it with: Historical T12 Taxes × 1.15. Reassessment after a sale never reduces taxes — if your math says it does, the millage rate or assessed value ratio is off.

NOTES:
- Cook County IL is aggressive: assesses at 10% of market value, then multiplies by a state equalization factor of ~3x, then applies the tax rate. Cook County also chases sales — flag any Cook County property as a high-reassessment-risk diligence item.
- Other counties may only reassess every 10 years.
- If a parcel number is available from the seller's tax bill, use it to look up current assessed value as a cross-check.
- Always show both methods in the notes field so Michael can sanity-check.

### INSURANCE

- Default: T12 × 1.05 (insurance has been inflating faster than general inflation).

- FLOOD ZONE OVERRIDE:
- If the property is in a FEMA flood zone (anything OTHER than Zone X), trend insurance by an additional 15%:
  - Underwritten Insurance = T12 × 1.05 × 1.15 = T12 × 1.2075

- Zone X = no special flood hazard, no override needed. Zones A, AE, AH, AO, AR, A99, V, VE = special flood hazard, apply the override.

- The user will indicate flood zone status via a yes/no input in the form. If "yes," apply the override. If "no," use the base T12 × 1.05. If unknown, default to no override but flag "verify flood zone status via FEMA flood map" as a diligence item.

- NOTE: GGC may obtain better pricing than the seller through their portfolio umbrella policy — if the seller's per-unit insurance figure is materially above market (rough benchmark: $200-300/unit/year for MHC), flag this as a potential expense reduction opportunity.

### MANAGEMENT FEE
Override the seller's management fee entirely.
- Properties under 200 sites: 5% of EGI
- Properties 200 sites or more: 4% of EGI
- EGI = NRI + Other Income (where NRI = GPR − Vacancy − Concessions − Bad Debt)

DO NOT REASSIGN PAYROLL TO MANAGEMENT FEE:
- If the seller's P&L has a "Payroll" or "Wages" line, that's on-site staff (manager, maintenance). It stays in the Payroll category.
- The Management Fee category is GGC's synthetic 4-5% of EGI fee — it's added ON TOP of payroll, not in place of it.
- If the seller already has a "Management Fee" line, override it with GGC's calculated fee. Do NOT add the seller's mgmt fee number to anything else.
- Result: every deal has BOTH a Payroll line (from seller actuals) AND a Management Fee line (synthetic GGC override). They are different categories serving different purposes.

### REPAIR & MAINTENANCE / GROUND MAINTENANCE
- These are discretionary line items — apply judgment, do not blindly use T12 * 1.03
- Look for one-time items: a single month with a spike (e.g. $4,000 vs $1,500 in all other months) likely indicates a one-time project. FLAG these per the One-Time Item Handling rule above — do not silently exclude them.
- ONE-TIME ITEM HANDLING (flag, do not auto-strip):
    - Scan monthly data for anomalous spikes: any month where a single expense line is ≥2× the average of the other 11 months.
    - DO NOT automatically back out these items. Michael (the reviewer) wants to make the inclusion/exclusion decision himself.
    - Instead, flag each spike in the flags array with severity "medium" and these exact fields:
    - item: the line item name
    - issue: "One-time item detected: [month] showed [amount] vs. [avg] average across other months. Variance of [X]×."
    - severity: "medium"
    - recommendation: "Possible one-time project (e.g. tree work, road repair, plumbing). Confirm scope with seller. If genuinely one-time, exclude from T12 × 1.03 underwriting basis. If recurring, leave in."
    - In the line item's "notes" field, also note the spike month and amount, so the reviewer can see it next to the line.
    - Use the T12 total AS-IS for the underwriting basis (T12 × 1.03). The reviewer will manually adjust if they decide a spike was truly one-time.
        - This rule applies to ALL expense lines, not just R&M — Insurance, Professional Fees, Repairs, Ground Maintenance, anything with a discernible monthly pattern. Apply consistently.
- Use per-unit benchmarks as a sanity check — flag if R&M is materially above or below typical range
- R&M covers: road repairs, pothole patching, cement work, common area repairs, occasionally home repairs that get charged back to tenants
- Ground Maintenance covers: landscaping, lawn care, tree trimming, snow removal, etc.

HOME RENT EXPENSE RE-BUCKETING:
Scan all expense line items for labels containing "home," "POH," "park-owned home," "home repairs," "RM Home," or similar. These are NOT general R&M — they belong in the Home Rent Expense bucket (GGC category: "Home Rent Expense (MH)"), not in the general "Repair and Maintenance" line.

Examples to re-bucket:
- "RM Home Repairs" → Home Rent Expense (MH)
- "POH Maintenance" → Home Rent Expense (MH)
- "Home Renovations" → Home Rent Expense (MH)

Items that stay in general R&M:
- "RM Community" → Repair and Maintenance
- "Road Repairs" → Repair and Maintenance
- "Common Area" → Repair and Maintenance

When in doubt, flag the line item and ask in the questions array.

POH percentage cross-check: If POH > 20% and you see no Home Rent Expense in the seller's financials, assume home expenses are embedded inside general R&M and flag it. Expense ratio benchmarks:
- All-TOH (no park-owned homes): 25-35% expense ratio
- Mixed (some POH): 35-40% expense ratio  
- High POH (40%+): 45-50% expense ratio
- If actual expense ratio is materially above these benchmarks for the POH mix, flag "expenses appear inflated by hidden home rent costs."

### PAYROLL
- Typically one on-site manager per community
- Many managers live on-site and receive a free lot as compensation — this is called an Employee Allowance (comped site) and should be broken out as a separate line item, not buried in payroll
- Employee Allowance = lot rent value of the comped site * 12

### PARK-OWNED HOME (POH) BIFURCATION
- This is critical — do not blend lot rent NOI and home rent NOI into a single valuation
- Identify POH percentage from the rent roll: any unit with a home rent column value is a park-owned home
- Separate income into two streams:
  (1) Lot Rent NOI — ground rent only, excludes all home rent income and home rent expense
  (2) Home Rent NOI — home rent income minus home rent expense only
- The home rent business is valued at a much higher (worse) cap rate than the lot rent business because it carries more risk, more expense, and lower rent growth
- Flag POH percentage: under 20% is acceptable, 40%+ signals a fundamentally different and less desirable operating model
- If the seller does not break out home rent expenses separately, flag this and ask for the breakout

### STABILIZED COLUMN
- In addition to the standard underwritten column, create a stabilized NOI column
- Replace current contracted lot rents with market rents (from rent comps) to show what the property generates at full market occupancy and market rents
- This is the "ceiling" valuation — what is this worth once the value-add plan is fully executed
- Stabilized cap rate = Stabilized NOI / Purchase Price — this tells you the ingoing cap rate on a fully stabilized basis

### STABILIZED YIELD ON COST (critical decision metric)

- In addition to Stabilized NOI, compute Stabilized Yield on Cost:
  - Stabilized Yield on Cost = Stabilized NOI / (Purchase Price + CapEx Reserve)

- Why "on cost" not "on price": the CapEx reserve is real money GGC has to raise and deploy to reach the stabilized state. Total cost basis is the honest denominator.

- Also compute Ingoing Cap Rate:
  - Ingoing Cap Rate = Underwritten (year-1) NOI / Purchase Price

- The decision metric is the SPREAD:
  - Spread = Stabilized Yield on Cost − Ingoing Cap Rate

- GGC's rule: the spread must be at least 200 basis points (2.00%) for the deal to clear. If the spread is below 200 bps, flag it as a hard "DOES NOT MEET INVESTMENT CRITERIA" warning at the top of the output. If the spread is above 200 bps, note how much above (e.g. "+250 bps — meets criteria with 50 bps cushion").

- GGC does NOT pay for value it's creating. The recommended purchase price should be based on the INGOING NOI at market cap rate, NOT the stabilized NOI. The stabilized column is a forward-looking check, not a valuation input.

- Return these in propertyInfo:
  - ingoingCapRate (decimal)
  - stabilizedYieldOnCost (decimal)
  - spreadBps (integer, basis points)
  - meetsInvestmentCriteria (boolean)

### CAPEX RESERVE
- Add a CapEx reserve on top of T12 expenses — this is not in the seller's financials
- Standard: $50/unit/year for older or lower-quality assets
- Higher-quality or larger assets: up to $75/unit/year
- This covers road resurfacing, utility infrastructure, common area improvements, and home addition costs

### PER-UNIT BENCHMARKS
- Always calculate and display expense figures on a per-unit basis alongside total figures
- Per-unit figures allow quick sanity checks against GGC's experience across 100+ deals
- If a per-unit figure looks materially off vs. typical ranges, flag it for review — do not just accept the seller's number

### ONE-TIME ITEMS & VARIANCE FLAGS
- Scan all expense line items for month-over-month spikes — any month that is 2x or more the average of other months is likely a one-time item
- Flag these explicitly: identify the month, the amount, and the variance from average
- Michael will decide whether to exclude them from the underwritten run rate
- High professional fees often correlate with high bad debt — may indicate the seller has been filing frequent evictions
- High R&M in a single month often indicates a capital project (water line, road work) that should not recur

### ALTERNATIVE HOUSING (required output)
- Average single-family home sale price within 5-mile and 10-mile radius
- Year-over-year home price appreciation percentage
- Average 2-bedroom apartment rent in the market
- Average 3-bedroom apartment rent in the market
- MHP all-in monthly cost = lot rent + estimated home payment (assume $50,000 home, $3,000 down, $47,000 loan at current rates)
- Compare MHP all-in cost to 2BR and 3BR apartment rents — GGC targets MHP all-in cost at less than comparable rental alternatives
- GGC targets home sale prices at approximately 30% of the average single-family home price in the area — this is their demand threshold

### RECOMMENDED PURCHASE PRICE OUTPUT
- Derive a recommended purchase price using two methods:
  (1) Cap Rate Method: Underwritten NOI / Market Cap Rate (from sale comps)
  (2) Stabilized Cap Rate Method: Stabilized NOI / Market Cap Rate
- Also show: Price per unit at recommended purchase price
- Compare to asking price and flag the gap
- Note: Lot Rent NOI and Home Rent NOI should be capitalized at different rates — do not blend them
  * Lot Rent NOI: capitalize at market cap rate for the asset class and submarket (typically 5-7%)
  * Home Rent NOI: capitalize at a higher cap rate (typically 12-15%) reflecting the lower quality of that income stream

### PROPERTY INFORMATION BLOCK (Underwriting tab right-side metadata)
The Underwriting tab carries a "Property Information" block at columns M:R.
Populate the optional fields below ONLY when the OM, listing, or T12 documents
state them explicitly. Leave as null if not found — do NOT guess.
- yearBuilt: integer year (e.g. 2015). From OM "Year Built" field.
- websiteUrl: broker / property listing URL if present (e.g. flintstoneproperties.com/properties/X)
- acreage: total acres (number). From OM "Acreage" or "Total Acres".
- utilityStructure: short phrase describing water/sewer setup (e.g. "Private Well & Septic")
- electricityNotes: who pays / structure (e.g. "owner paid — sub-metered & billed back")
- trashNotes: who pays / structure (e.g. "covered through county taxes")
- taxAssessorUrl: county tax-assessor lookup URL if mentioned
- taxParcels: array of parcel-level tax records if the OM includes a parcel table
  (each: parcelId, marketValue, taxableValue, taxes, acres). All fields nullable.

## Output (JSON only, no prose, no fences)

{{
  "propertyInfo": {{
    "name": "string", "address": "string", "city": "string", "state": "string",
    "county": "string", "totalUnits": integer, "askingPrice": number,
    "propertyType": "MHC|RV|Hybrid",
    "ingoingCapRate": number,
    "stabilizedYieldOnCost": number,
    "spreadBps": integer,
    "meetsInvestmentCriteria": boolean,
    "yearBuilt": integer or null,
    "websiteUrl": "string" or null,
    "acreage": number or null,
    "utilityStructure": "string" or null,
    "electricityNotes": "string" or null,
    "trashNotes": "string" or null,
    "taxAssessorUrl": "string" or null,
    "taxParcels": [{{"parcelId", "marketValue", "taxableValue", "taxes", "acres"}}] or null
  }},
  "income": [
    {{
      "ggcCategory": "string", "sellerName": "string",
      "fyPrior": number, "fyCurrent": number, "brokerProforma": number,
      "t12Total": number, "monthly": [12 numbers],
      "ggcUnderwritten": number, "confidence": "high|medium|low", "notes": "string"
    }}
  ],
  "expenses": [{{ same shape as income }}],

  "rentRoll": {{
    "totalUnits": integer, "occupiedUnits": integer, "vacantUnits": integer,
    "occupancyRate": number, "avgLotRent": number, "parkOwnedHomes": integer,
    "pohPercent": number,
    "rentRollRows": [
      {{"unitId": "string (lot number, e.g. 'A05', 'B07', 'EL02A')",
        "unitType": "MUST be one of [TOH MH Site, POH-Infilled units, Long term RV Site, Retail/Commercial]",
        "status": "Occupied | Vacant",
        "tenantName": "string (real tenant name from rent roll, or '' for vacant)",
        "lotRent": number,
        "homeRent": number (0 if no POH rent),
        "sellerUnitLabel": "string (raw seller label, e.g. 'WHA Lot', 'Storage') — for audit trail"}}
    ],
    "unitGroups": [
      {{"unitType": "MUST be one of [TOH MH Site, POH-Infilled units, Long term RV Site, Retail/Commercial]",
        "occupiedCount": integer,
        "vacantCount": integer,
        "lotRent": number,
        "pohRent": number (0 if TOH),
        "ltoPremium": number,
        "tenantNamePattern": "string",
        "sellerUnitLabel": "string (raw seller label, for audit)"}}
    ],
    "unitMixSummary": [
      {{"unitType": "one of [TOH MH Site, POH-Infilled units, Long term RV Site, Retail/Commercial]",
        "count": integer, "occupied": integer, "avgRent": number, "marketRent": number}}
    ]
  }},
  "flags": [{{"item", "issue", "severity", "recommendation"}}],
    "dataQualityChecks": {{
    "totalUnitsStated": integer,
    "rentRollRowsFound": integer,
    "unitCountReconciliation": "match|rent_roll_short|rent_roll_long",
    "pohCountStated": integer,
    "pohRowsFound": integer,
    "pohReconciliation": "match|stated_high|stated_low|stated_zero_but_found",
    "finalPohCount": integer,
    "finalPohPercent": number
  }},
  "questions": ["string"],
  "dataQuality": {{"hasT12", "hasT3", "hasRentRoll", "hasMonthlyBreakdown",
    "t12Period": "string", "missingData": ["string"]}}
}}

RENT ROLL OUTPUT GUIDANCE:
- ALWAYS emit `rentRollRows` with one entry per actual rent roll row (preserves real tenant names and Unit IDs). This is required for audit and downstream delinquency analysis.
- ALSO emit `unitGroups` as the aggregated 4-category summary. The backend reads `rentRollRows` first, falling back to `unitGroups` only if missing.
- For `unitGroups`, aggregate ONLY by the 4 canonical categories above — never by raw seller labels like "WHA Lot" or "Storage". A property with 100 WHA Lots + 5 Storage units + 1 Commercial should emit 2-3 `unitGroups`: one "TOH MH Site" with count=100, one "Retail/Commercial" with count=5-6. The COUNTIFS in the Excel template keys on these exact canonical strings."""


def call_parse_financials(api_key, extracted, property_info):
    """
    Call 2 of the financial pipeline: GGC methodology on clean extracted data.
    MODEL_METHODOLOGY (Fable 5, adaptive thinking, effort=THINKING_EFFORT)
    takes the verified extraction output and applies categorization +
    underwriting logic. No raw documents — it works from the clean JSON the
    extraction step produced.
    """
    user_blocks = [{
        "type": "text",
        "text": f"""Property context (user-stated):
- Name: {property_info.get('name', 'N/A')}
- Address: {property_info.get('address', 'N/A')}
- Total Units: {property_info.get('units', 'N/A')}
- Park-Owned Home Count (user-stated): {property_info.get('pohCount', '0')}
- Asking Price: ${property_info.get('askingPrice', 'N/A')}
- Flood Zone Status: {property_info.get('floodZone', 'unknown')}

Below is the CLEAN, PRE-EXTRACTED financial data from the seller's documents. The reporting period has already been identified and the numbers transcribed. Apply GGC's categorization and underwriting methodology to this data and return the structured JSON.

=== EXTRACTED FINANCIAL DATA ===
{json.dumps(extracted, indent=2)[:180000]}
=== END EXTRACTED DATA ===

Apply the GGC methodology and return the structured JSON."""
    }]
    model_id = _model_for_stage(property_info, "methodology")
    print(f"[Claude] Stage 2/2 — METHODOLOGY ({model_id}, "
          f"effort={THINKING_EFFORT})...")
    t0 = time.time()
    response = call_claude(api_key, FINANCIAL_PARSE_PROMPT, user_blocks,
                           use_thinking=True, model=model_id,
                           output_schema=METHODOLOGY_OUTPUT_SCHEMA,
                           max_tokens=MAX_TOKENS_METHODOLOGY)
    elapsed = time.time() - t0
    print(f"[Claude] Methodology returned in {elapsed:.1f}s "
          f"(stop_reason: {response.get('stop_reason', '?')})")
    parsed = extract_json(response)

    # Activate the methodology-line validators. They reject any income
    # row whose ggcCategory isn't in GGC_INCOME_CATEGORIES (likewise for
    # expenses) — the SUMIFS in the Underwriting template will silently
    # zero out anything misspelled, so we want to fail loud here.
    validation_errors = []
    for i, item in enumerate(parsed.get("income") or []):
        try:
            MethodologyIncomeItem.model_validate(item)
        except ValidationError as e:
            validation_errors.append(f"income[{i}]: {e.errors()[0].get('msg', str(e))}")
    for i, item in enumerate(parsed.get("expenses") or []):
        try:
            MethodologyExpenseItem.model_validate(item)
        except ValidationError as e:
            validation_errors.append(f"expenses[{i}]: {e.errors()[0].get('msg', str(e))}")
    if validation_errors:
        # Capture so the Extraction Check tab can render the failures.
        # Don't raise — the rest of the run may still be usable, and the
        # tab is where the reviewer expects to see categorization issues.
        parsed.setdefault("_methodologyValidation", []).extend(validation_errors)
        print(f"[Methodology] {len(validation_errors)} validation warnings "
              f"(see Extraction Check tab)")
    _carry_extraction_through(parsed, extracted)
    _ensure_rent_roll_complete(parsed, property_info)
    apply_ggc_overrides(parsed, property_info)
    return parsed


def _carry_extraction_through(parsed, extracted):
    """Forward per-tenant rent-roll rows AND per-line sellerNotes from the
    extraction output into the methodology output. The methodology schema
    deliberately drops rentRollRows + sellerNotes (to keep its grammar
    compilable for Structured Outputs), so without this copy the data
    vanishes and downstream code can't use it. We match methodology rows to
    extracted rows by token-overlap on sellerName/sellerLabel — looser than
    exact match because the LLM frequently adds prefixes like "5701 Wages,
    Salary" → "Wages" or strips account numbers."""
    if not isinstance(parsed, dict) or not isinstance(extracted, dict):
        return
    parsed_rr = parsed.setdefault("rentRoll", {})
    ext_rr    = extracted.get("rentRoll") or {}
    if isinstance(ext_rr.get("rentRollRows"), list) and not parsed_rr.get("rentRollRows"):
        parsed_rr["rentRollRows"] = ext_rr["rentRollRows"]

    def _norm(s):
        return "".join(c for c in (s or "").lower() if c.isalnum())

    # Match methodology rows to extracted rows so we can copy sellerNotes
    # + proFormaTotal onto each methodology row. The override pass uses
    # these to force Omitt routing on Non-recurring/Discontinued/Seller-
    # Specific lines deterministically (independent of LLM judgment).
    def _attach(meth_list, ext_list):
        if not isinstance(meth_list, list) or not isinstance(ext_list, list):
            return
        ext_by_norm = {}
        for e in ext_list:
            if isinstance(e, dict):
                ext_by_norm.setdefault(_norm(e.get("sellerLabel")), []).append(e)
        for m in meth_list:
            if not isinstance(m, dict):
                continue
            sn = _norm(m.get("sellerName"))
            if not sn:
                continue
            # Try direct match first, then substring (extracted ⊆ methodology
            # OR methodology ⊆ extracted to catch both directions of GL-number
            # stripping).
            matched = ext_by_norm.get(sn)
            if not matched:
                for key, group in ext_by_norm.items():
                    if key and (key in sn or sn in key):
                        matched = group
                        break
            if matched:
                e = matched[0]
                m["_sellerNotes"]   = e.get("sellerNotes") or ""
                m["_proFormaTotal"] = e.get("proFormaTotal")
    _attach(parsed.get("income"),   extracted.get("income"))
    _attach(parsed.get("expenses"), extracted.get("expenses"))


def call_parse_financials_merged(api_key, extracted, property_info, n_runs=3):
    """Self-consistency wrapper around call_parse_financials. Runs N
    methodology calls in parallel against the SAME extracted data and
    merges them: ggcCategory by confidence-weighted mode (the actual fix
    for category drift), numerics by confidence-weighted median. A
    non-unanimous vote on ggcCategory becomes an entry on the Extraction
    Check tab so the reviewer can see the model disagreed with itself.
    Per CLAUDE.md §0 mechanism #2 this is the default, not opt-in.
    """
    if n_runs <= 1:
        return call_parse_financials(api_key, extracted, property_info)

    print(f"[Claude] Starting {n_runs}× methodology merge...")
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=n_runs) as executor:
        futures = [executor.submit(call_parse_financials, api_key,
                                   extracted, property_info)
                   for _ in range(n_runs)]
        results = []
        for i, f in enumerate(as_completed(futures)):
            try:
                results.append(f.result())
                print(f"[Claude] Methodology run {i+1}/{n_runs} complete")
            except Exception as e:
                print(f"[Claude] Methodology run {i+1}/{n_runs} FAILED: {e}")

    if not results:
        raise RuntimeError(f"All {n_runs} methodology runs failed")
    print(f"[Claude] Merged {len(results)} methodology runs in "
          f"{time.time() - t0:.1f}s")
    merged = _merge_methodology(results)
    # Carry per-tenant rent-roll rows + the seller-notes column through to
    # the merged output; _merge_methodology only voted on the methodology
    # fields the LLM emits and would otherwise lose the extraction-stage
    # data needed for the rent-roll write-back.
    _carry_extraction_through(merged, extracted)
    # Re-run rent-roll completeness on the merged output so the deterministic
    # backstop still applies (each run's _ensure_rent_roll_complete output
    # is replaced by the merged unitGroups, which may again be short).
    _ensure_rent_roll_complete(merged, property_info)
    # Force-apply GGC's deterministic methodology rules (§5.4): management
    # fee, taxes, insurance, CapEx, bad-debt sign. These are RULES, not LLM
    # judgment — applying them in Python eliminates run-to-run drift on
    # exactly the lines GGC requires to be exact. Per CLAUDE.md §0
    # mechanism #3 these overrides are not optional.
    apply_ggc_overrides(merged, property_info)
    return merged


def _merge_methodology(runs):
    """Field-level merge for the methodology output. Income / expense rows
    are grouped by sellerName (the input identity that doesn't drift between
    runs); within each group, ggcCategory is voted by confidence-weighted
    mode and every numeric field is reduced by confidence-weighted median.
    A non-unanimous category vote is surfaced as a check so the reviewer
    sees the disagreement.

    propertyInfo and rentRoll scalars merge by median (numerics) or mode
    (strings). unitGroups merge by unitType. Flags union by (item, issue).
    """
    if len(runs) == 1:
        return runs[0]

    category_disagreements = []  # surfaced as checks on the merged output

    def _merge_line_items(field):
        # Group across runs by sellerName so the same input line ends up
        # in the same merge bucket regardless of category drift.
        buckets = {}
        for r_i, run in enumerate(runs):
            for it in (run.get(field) or []):
                if not isinstance(it, dict):
                    continue
                key = (it.get("sellerName") or "").strip()
                buckets.setdefault(key, []).append(it)

        merged_items = []
        for seller_name, items in buckets.items():
            if not items:
                continue
            # Vote on ggcCategory (the high-stakes drift field).
            cats = [(it.get("ggcCategory") or "",
                     _CONF_WEIGHT.get((it.get("confidence") or "medium").lower(),
                                      _CONF_WEIGHT["medium"]))
                    for it in items]
            cat_votes = Counter()
            for c, w in cats:
                if c:
                    cat_votes[c] += w
            ggc_cat = cat_votes.most_common(1)[0][0] if cat_votes else ""
            unique_cats = {c for c, _ in cats if c}
            if len(unique_cats) > 1:
                category_disagreements.append({
                    "field": field,
                    "sellerName": seller_name,
                    "winner": ggc_cat,
                    "candidates": sorted(unique_cats),
                })

            # Confidence-weighted median for numerics.
            t12_total = _median_with_confidence(items, "t12Total") or 0
            ggc_uw    = _median_with_confidence(items, "ggcUnderwritten") or 0
            fy_prior  = _median_with_confidence(items, "fyPrior") or 0
            fy_curr   = _median_with_confidence(items, "fyCurrent") or 0
            broker_pf = _median_with_confidence(items, "brokerProforma") or 0

            # Monthly: median per month-index across the runs that emit it.
            monthly_runs = [it.get("monthly") for it in items
                            if isinstance(it.get("monthly"), list)
                            and len(it["monthly"]) == 12]
            if monthly_runs:
                monthly = [statistics.median(
                    [m[i] for m in monthly_runs
                     if isinstance(m[i], (int, float))])
                    for i in range(12)]
            elif t12_total:
                monthly = [t12_total / 12] * 12
            else:
                monthly = [0] * 12

            # Confidence: mode across runs (high beats medium beats low only
            # if more than half of runs say so).
            conf = _mode_of([it.get("confidence") for it in items]) or "medium"
            notes = " || ".join(sorted({it.get("notes") or "" for it in items
                                         if it.get("notes")}))

            merged_items.append({
                "ggcCategory":     ggc_cat,
                "sellerName":      seller_name,
                "fyPrior":         fy_prior,
                "fyCurrent":       fy_curr,
                "brokerProforma":  broker_pf,
                "t12Total":        t12_total,
                "monthly":         monthly,
                "ggcUnderwritten": ggc_uw,
                "confidence":      conf,
                "notes":           notes,
            })
        return merged_items

    merged = {
        "income":   _merge_line_items("income"),
        "expenses": _merge_line_items("expenses"),
    }

    # propertyInfo: per-field median/mode across runs.
    prop_runs = [r.get("propertyInfo") or {} for r in runs]
    merged_prop = {}
    if prop_runs:
        keys = set().union(*(p.keys() for p in prop_runs))
        for k in keys:
            vs = [p.get(k) for p in prop_runs if p.get(k) is not None]
            if not vs:
                continue
            nums = [v for v in vs if isinstance(v, (int, float))]
            if nums and len(nums) == len(vs):
                merged_prop[k] = statistics.median(nums)
            else:
                merged_prop[k] = _mode_of(vs)
    merged["propertyInfo"] = merged_prop

    # rentRoll: scalars by median, unitGroups merged by unitType.
    rr_runs = [r.get("rentRoll") or {} for r in runs]
    scalar_fields = ("totalUnits", "occupiedUnits", "vacantUnits",
                     "occupancyRate", "avgLotRent", "parkOwnedHomes",
                     "pohPercent")
    merged_rr = {}
    for k in scalar_fields:
        vs = [rr.get(k) for rr in rr_runs if rr.get(k) is not None]
        nums = [v for v in vs if isinstance(v, (int, float))]
        if nums:
            merged_rr[k] = statistics.median(nums)

    group_buckets = {}
    for rr in rr_runs:
        for g in (rr.get("unitGroups") or []):
            if not isinstance(g, dict):
                continue
            t = (g.get("unitType") or "").strip()
            if not t:
                continue
            group_buckets.setdefault(t, []).append(g)
    merged_groups = []
    for unit_type, gs in group_buckets.items():
        merged_groups.append({
            "unitType":      unit_type,
            "occupiedCount": int(statistics.median(
                [g.get("occupiedCount") or 0 for g in gs])),
            "vacantCount":   int(statistics.median(
                [g.get("vacantCount") or 0 for g in gs])),
            "lotRent":       statistics.median(
                [g.get("lotRent") or 0 for g in gs]),
            "pohRent":       statistics.median(
                [g.get("pohRent") or 0 for g in gs]),
            "tenantNamePattern": _mode_of(
                [g.get("tenantNamePattern") for g in gs]),
            "sellerUnitLabel":   _mode_of(
                [g.get("sellerUnitLabel") for g in gs]),
        })
    merged_rr["unitGroups"] = merged_groups
    merged["rentRoll"] = merged_rr

    # Flags: union across runs by (item, issue), keep highest severity seen.
    sev_rank = {"high": 3, "medium": 2, "low": 1}
    flag_by_key = {}
    for r in runs:
        for fl in (r.get("flags") or []):
            if not isinstance(fl, dict):
                continue
            key = ((fl.get("item") or "").strip(),
                   (fl.get("issue") or "").strip()[:160])
            cur = flag_by_key.get(key)
            if cur is None or sev_rank.get(fl.get("severity") or "low", 0) > \
                              sev_rank.get(cur.get("severity") or "low", 0):
                flag_by_key[key] = fl
    merged["flags"] = list(flag_by_key.values())

    # dataQualityChecks: per-field median/mode.
    dqc_runs = [r.get("dataQualityChecks") or {} for r in runs]
    merged_dqc = {}
    if dqc_runs:
        keys = set().union(*(d.keys() for d in dqc_runs))
        for k in keys:
            vs = [d.get(k) for d in dqc_runs if d.get(k) is not None]
            nums = [v for v in vs if isinstance(v, (int, float))]
            if nums and len(nums) == len(vs):
                merged_dqc[k] = statistics.median(nums)
            else:
                merged_dqc[k] = _mode_of(vs)
    merged["dataQualityChecks"] = merged_dqc

    # questions: dedupe-by-text, preserve order of first appearance.
    seen_q = set()
    merged_questions = []
    for r in runs:
        for q in (r.get("questions") or []):
            qn = (q or "").strip()
            if qn and qn not in seen_q:
                seen_q.add(qn)
                merged_questions.append(qn)
    merged["questions"] = merged_questions

    # dataQuality: take the first run's (modal) view; this field is
    # narrative-shaped so a true field-level merge would be lossy.
    merged["dataQuality"] = runs[0].get("dataQuality") or {}

    # Surface the category disagreements as checks (one per disagreed line).
    if category_disagreements:
        check_entries = []
        for d in category_disagreements[:20]:
            check_entries.append({
                "item": f"Methodology vote disagreement: {d['sellerName']!r}",
                "check": "Unanimous ggcCategory across runs",
                "status": "warn",
                "detail": (f"Runs voted {d['candidates']} — picked "
                           f"{d['winner']!r} by weighted mode. Verify the "
                           f"placement is correct."),
            })
        merged.setdefault("_extractionChecks", []).extend(check_entries)
        merged["_methodologyVoteDisagreements"] = len(category_disagreements)
        print(f"[Methodology] {len(category_disagreements)} ggcCategory "
              f"vote disagreement(s) across {len(runs)} runs.")
    merged["_methodologyVoteRuns"] = len(runs)
    return merged


def apply_ggc_overrides(financials, property_info):
    """Force the GGC methodology rules per CLAUDE.md §5.4 onto the merged
    methodology output. These are rules (not LLM judgment), so applying
    them in Python eliminates run-to-run drift on the exact lines GGC
    requires to be exact:

    * Bad debt sign: always negative (t12Total, ggcUnderwritten, monthly).
    * Management fee: 5% of EGI under 200 sites, 4% at 200+ sites,
      EGI-based. Override any seller mgmt-fee line; insert one if missing.
    * Insurance: T12 × 1.05; × 1.15 when the user marked the property
      flood zone.
    * Taxes: never below the historical T12 × 1.15 floor.
    * CapEx reserve: $75/unit/year (gold standard per CorrectOutput I43).

    Mutates `financials` in place. Records every applied override on
    `financials["_ggcOverrides"]` so the Extraction Check tab can list
    exactly what changed.
    """
    income = financials.setdefault("income", [])
    expenses = financials.setdefault("expenses", [])
    rr = financials.get("rentRoll") or {}
    overrides = financials.setdefault("_ggcOverrides", [])

    # Defensive normalization: strip "Less: " prefix and remap variant
    # labels ("General and Administrative" → "G&A") that some methodology
    # runs emit when structured-outputs grammar masking falls back to
    # prompt-only enforcement. Done in place so every downstream string
    # check below sees the canonical enum value.
    normalized_count = 0
    for it in income:
        raw = it.get("ggcCategory")
        norm = _normalize_ggc_category(raw)
        if norm != raw:
            it["ggcCategory"] = norm
            normalized_count += 1
    for it in expenses:
        raw = it.get("ggcCategory")
        norm = _normalize_ggc_category(raw)
        if norm != raw:
            it["ggcCategory"] = norm
            normalized_count += 1
    if normalized_count:
        financials.setdefault("_extractionChecks", []).append({
            "item": "Methodology category normalization",
            "check": "ggcCategory matches canonical enum",
            "status": "warn",
            "detail": (f"{normalized_count} line item(s) carried non-canonical "
                       "category strings (e.g. 'Less: Bad Debt' or 'General "
                       "and Administrative'). Defensively normalized to the "
                       "enum values the Underwriting tab SUMIFS expect. This "
                       "indicates structured-outputs grammar masking was "
                       "unavailable for the methodology call — verify the "
                       "Anthropic beta header / pinned model snapshot."),
        })

    # Drop placeholder rows: section-subtotal labels with zero values
    # across t12 / underwritten / monthly. The methodology prompt forbids
    # these, but prompt-only enforcement lets one through occasionally.
    # Dropping at this layer keeps every downstream consumer (verification,
    # parity checks, write-back) consistent.
    def _is_empty_subtotal_row(it):
        sn = (it.get("sellerName") or "").lower()
        if not any(t in sn for t in ("total", "subtotal", "non-posting")):
            return False
        t12 = it.get("t12Total") or 0
        uw  = it.get("ggcUnderwritten") or 0
        monthly = it.get("monthly") or []
        monthly_sum = sum(m for m in monthly if isinstance(m, (int, float)))
        return t12 == 0 and uw == 0 and monthly_sum == 0
    dropped_subtotals = []
    income_keep, expense_keep = [], []
    for it in income:
        if _is_empty_subtotal_row(it):
            dropped_subtotals.append(it.get("sellerName"))
        else:
            income_keep.append(it)
    for it in expenses:
        if _is_empty_subtotal_row(it):
            dropped_subtotals.append(it.get("sellerName"))
        else:
            expense_keep.append(it)
    income[:]   = income_keep
    expenses[:] = expense_keep
    if dropped_subtotals:
        financials.setdefault("_extractionChecks", []).append({
            "item": "Methodology subtotal collapse",
            "check": "no empty-value 'Total ___' rows",
            "status": "warn",
            "detail": (f"Dropped {len(dropped_subtotals)} placeholder "
                       f"section-subtotal row(s) with zero values: "
                       f"{dropped_subtotals}. The LLM collapsed a section "
                       "(typically '5700 Total Personnel') without bringing "
                       "the leaf GLs across. Underlying dollars likely "
                       "missing from this workbook — verify the source P&L "
                       "leaf GL accounts landed individually."),
        })

    def _record(category, before, after, basis):
        overrides.append({
            "category": category, "before": before, "after": after,
            "basis": basis,
        })

    # ── Force Omitt routing from seller-notes (deterministic) ───────────
    # The methodology prompt instructs the LLM to route Non-recurring /
    # Discontinued / Seller-Specific lines to Omitt, but LLMs miss this
    # rule under prompt-only schema enforcement (Haiku) and on multi-run
    # voting where one run disagrees. Force it in Python so the gold-
    # standard Omitt routing (see CorrectOutput Data Consolidation rows
    # 11, 17, 24, 48, 73, 86) happens every time.
    OMITT_MARKERS = (
        "non-recurring", "non recurring", "nonrecurring",
        "discontinued", "seller specific", "seller-specific",
        "one-time", "one time", "non-operating",
    )
    def _force_omitt(items, target_category):
        for it in items:
            notes = (it.get("_sellerNotes") or "").lower()
            if any(m in notes for m in OMITT_MARKERS):
                current = (it.get("ggcCategory") or "").strip()
                if current and current != target_category:
                    _record(
                        f"{it.get('sellerName', '?')}: {current} → {target_category}",
                        current, target_category,
                        f"sellerNotes={it.get('_sellerNotes')!r}")
                    it["ggcCategory"] = target_category
    _force_omitt(income,   "Omitt Income")
    _force_omitt(expenses, "Omitt Expense")

    # Vehicle-related expenses → Omitt Expense regardless of GL prefix.
    # Catches Car Insurance (5051), Vehicle Fuel (5401), and similar lines
    # the LLM sometimes lands in Insurance or Gas/Fuel.
    VEHICLE_MARKERS = ("car insurance", "vehicle insurance", "vehicle fuel",
                       "fuel for vehicle", "auto insurance", "truck",
                       "vehicle maintenance")
    for it in expenses:
        sn = (it.get("sellerName") or "").lower()
        if any(m in sn for m in VEHICLE_MARKERS):
            current = (it.get("ggcCategory") or "").strip()
            if current and current != "Omitt Expense":
                _record(
                    f"{it.get('sellerName', '?')}: {current} → Omitt Expense",
                    current, "Omitt Expense",
                    "vehicle line — not property opex")
                it["ggcCategory"] = "Omitt Expense"

    # ── Bad debt sign: always negative ──────────────────────────────────
    for it in income:
        if (it.get("ggcCategory") or "").strip() == "Bad Debt":
            for fld in ("t12Total", "ggcUnderwritten", "fyPrior",
                        "fyCurrent", "brokerProforma"):
                v = it.get(fld)
                if isinstance(v, (int, float)) and v > 0:
                    it[fld] = -v
            monthly = it.get("monthly")
            if isinstance(monthly, list):
                it["monthly"] = [
                    -m if isinstance(m, (int, float)) and m > 0 else m
                    for m in monthly
                ]

    # ── Effective Gross Income (post bad-debt sign fix) ────────────────
    # EXCLUDE "Omitt Income" — that's the explicit non-operating bucket;
    # including it would oversize the management-fee percentage (which is
    # the EGI consumer downstream) by the omit amount on every deal.
    egi = 0.0
    for it in income:
        if (it.get("ggcCategory") or "").strip() == "Omitt Income":
            continue
        v = it.get("ggcUnderwritten")
        if isinstance(v, (int, float)):
            egi += v

    # ── Management fee: 5% under 200, 4% at 200+, EGI-based ────────────
    try:
        units_for_mgmt = int(rr.get("totalUnits") or 0) or int(
            str(property_info.get("units", "")).strip() or 0)
    except (ValueError, TypeError):
        units_for_mgmt = 0
    mgmt_pct = 0.04 if units_for_mgmt >= 200 else 0.05
    if egi > 0:
        mgmt_target = round(egi * mgmt_pct, 2)
        mgmt_lines = [e for e in expenses
                      if (e.get("ggcCategory") or "").strip() == "Management Fee"]
        if mgmt_lines:
            primary = mgmt_lines[0]
            before = primary.get("ggcUnderwritten")
            primary["ggcUnderwritten"] = mgmt_target
            primary["monthly"] = [mgmt_target / 12] * 12
            primary["notes"] = (primary.get("notes") or "").strip()
            if primary["notes"]:
                primary["notes"] += " || "
            primary["notes"] += (
                f"GGC override: {mgmt_pct:.0%} of EGI "
                f"(${egi:,.0f}) on {units_for_mgmt} units"
            )
            primary["confidence"] = "high"
            _record("Management Fee", before, mgmt_target,
                    f"{mgmt_pct:.0%} × EGI ${egi:,.0f}")
            # Zero any duplicate mgmt fee rows so they don't double-count.
            for extra in mgmt_lines[1:]:
                _record("Management Fee (duplicate)",
                        extra.get("ggcUnderwritten"), 0,
                        "deduped vs primary mgmt-fee row")
                extra["ggcUnderwritten"] = 0
                extra["monthly"] = [0] * 12
        else:
            expenses.append({
                "ggcCategory":     "Management Fee",
                "sellerName":      "Management Fee (GGC override)",
                "fyPrior":         0, "fyCurrent": 0, "brokerProforma": 0,
                "t12Total":        0,
                "monthly":         [mgmt_target / 12] * 12,
                "ggcUnderwritten": mgmt_target,
                "confidence":      "high",
                "notes": (f"GGC override (no seller line found): "
                          f"{mgmt_pct:.0%} of EGI ${egi:,.0f} "
                          f"on {units_for_mgmt} units"),
            })
            _record("Management Fee (inserted)", None, mgmt_target,
                    f"{mgmt_pct:.0%} × EGI ${egi:,.0f}")

    # ── Insurance: T12 × 1.05 (× 1.15 if flood zone) ───────────────────
    fz_raw = str(property_info.get("floodZone", "")).strip().lower()
    flood = fz_raw in ("yes", "true", "1", "flood", "y")
    ins_mult = 1.15 if flood else 1.05
    for e in expenses:
        if (e.get("ggcCategory") or "").strip() == "Insurance":
            t12 = e.get("t12Total") or e.get("ggcUnderwritten") or 0
            if isinstance(t12, (int, float)) and t12 > 0:
                target = round(t12 * ins_mult, 2)
                before = e.get("ggcUnderwritten")
                e["ggcUnderwritten"] = target
                e["monthly"] = [target / 12] * 12
                e["confidence"] = "high"
                e["notes"] = ((e.get("notes") or "").strip()
                              + (" || " if e.get("notes") else "")
                              + f"GGC override: T12 × {ins_mult:.2f}"
                              + (" (flood zone)" if flood else ""))
                _record("Insurance", before, target,
                        f"T12 ${t12:,.0f} × {ins_mult:.2f}"
                        + (" (flood)" if flood else ""))

    # ── Taxes: never below historical × 1.15 ───────────────────────────
    for e in expenses:
        if (e.get("ggcCategory") or "").strip() == "RE Taxes":
            t12 = e.get("t12Total") or 0
            llm_val = e.get("ggcUnderwritten") or 0
            if isinstance(t12, (int, float)) and t12 > 0:
                floor = round(t12 * 1.15, 2)
                if not isinstance(llm_val, (int, float)) or llm_val < floor:
                    before = e.get("ggcUnderwritten")
                    e["ggcUnderwritten"] = floor
                    e["monthly"] = [floor / 12] * 12
                    e["confidence"] = "high"
                    e["notes"] = ((e.get("notes") or "").strip()
                                  + (" || " if e.get("notes") else "")
                                  + "GGC override: T12 × 1.15 historical floor")
                    _record("RE Taxes", before, floor,
                            f"T12 ${t12:,.0f} × 1.15 (was below floor)")

    # ── CapEx reserve: $75/unit/year (gold standard per CorrectOutput I43) ───────────────────────────────────
    try:
        units = int(rr.get("totalUnits") or 0) or int(
            str(property_info.get("units", "")).strip() or 0)
    except (ValueError, TypeError):
        units = 0
    if units > 0:
        # GGC standard: $75/site/year — matches CorrectOutput's I43 formula
        # (`=J43 × N7` with J43 hardcoded to $75). Earlier CLAUDE.md said
        # $50 but the gold-standard Whaleshead model uses $75.
        capex_per_unit = 75
        capex_target = float(units * capex_per_unit)
        capex_lines = [e for e in expenses
                       if (e.get("ggcCategory") or "").strip() == "Cap-Ex Reserve"]
        if capex_lines:
            primary = capex_lines[0]
            before = primary.get("ggcUnderwritten")
            primary["ggcUnderwritten"] = capex_target
            primary["monthly"] = [capex_target / 12] * 12
            primary["confidence"] = "high"
            primary["notes"] = ((primary.get("notes") or "").strip()
                                + (" || " if primary.get("notes") else "")
                                + f"GGC override: ${capex_per_unit} × {units} units")
            _record("Cap-Ex Reserve", before, capex_target,
                    f"${capex_per_unit} × {units} units")
            for extra in capex_lines[1:]:
                _record("Cap-Ex Reserve (duplicate)",
                        extra.get("ggcUnderwritten"), 0, "deduped")
                extra["ggcUnderwritten"] = 0
                extra["monthly"] = [0] * 12
        else:
            expenses.append({
                "ggcCategory":     "Cap-Ex Reserve",
                "sellerName":      "Cap-Ex Reserve (GGC override)",
                "fyPrior":         0, "fyCurrent": 0, "brokerProforma": 0,
                "t12Total":        0,
                "monthly":         [capex_target / 12] * 12,
                "ggcUnderwritten": capex_target,
                "confidence":      "high",
                "notes": f"GGC override (no seller line): ${capex_per_unit} × {units} units",
            })
            _record("Cap-Ex Reserve (inserted)", None, capex_target,
                    f"${capex_per_unit} × {units} units")

    # Drop the override log if nothing actually changed — keeps the
    # Extraction Check tab focused on real events.
    if not overrides:
        financials.pop("_ggcOverrides", None)
    else:
        # Surface a single summary check so the reviewer sees what changed
        # without scrolling through every individual override. `after` is
        # numeric for value overrides (mgmt fee, capex, ins, taxes) but a
        # string for category-reroute overrides (sellerNotes-driven Omitt
        # forcing) — format accordingly.
        def _fmt(after):
            if isinstance(after, (int, float)):
                return f"${after:,.0f}"
            return str(after) if after else ""
        summary = "; ".join(
            f"{o['category']}: → {_fmt(o['after'])}"
            for o in overrides[:6]
        )
        more = f" (+{len(overrides) - 6} more)" if len(overrides) > 6 else ""
        financials.setdefault("_extractionChecks", []).append({
            "item": "GGC deterministic overrides",
            "check": "§5.4 rules (mgmt fee, ins, taxes, capex, bad-debt sign)",
            "status": "ok",
            "detail": f"{len(overrides)} override(s) applied: {summary}{more}",
        })
        print(f"[GGC Overrides] Applied {len(overrides)} deterministic "
              f"override(s) (mgmt/ins/tax/capex/bad-debt sign).")


def _ensure_rent_roll_complete(financials, property_info):
    """Deterministic backstop for §2.3 / §5.1 vacant-pad imputation.

    The methodology prompt instructs the LLM to impute missing vacant lots
    at per-type market rent when rent-roll rows < stated unit count, but the
    LLM does not always follow through — and when it doesn't, GPR is
    silently undercounted. This runs after Stage 3 returns: if the
    methodology's totalUnits is short of the user-stated unit count, the
    shortfall is added as vacant lots, distributed proportionally across
    existing unit groups (so per-type market rents already on each group
    carry forward). Adds a high-severity flag and updates the rent-roll
    aggregates so downstream consumers (Excel write-back, parity checks,
    KPI tiles) see the corrected totals.
    """
    try:
        stated_units = int(str(property_info.get("units", "")).strip() or 0)
    except (ValueError, TypeError):
        stated_units = 0
    if not stated_units:
        return

    rr = financials.get("rentRoll") or {}
    groups = rr.get("unitGroups") or []
    if not groups:
        return

    sizes = [(g.get("occupiedCount") or 0) + (g.get("vacantCount") or 0)
             for g in groups]
    current_total = sum(sizes)
    if current_total >= stated_units:
        return

    shortfall = stated_units - current_total
    sized = sorted(enumerate(sizes), key=lambda x: -x[1])
    remaining = shortfall
    for idx, sz in sized[:-1]:
        share = int(round(shortfall * sz / current_total)) if current_total else 0
        share = min(share, remaining)
        groups[idx]["vacantCount"] = (groups[idx].get("vacantCount") or 0) + share
        remaining -= share
    if remaining > 0:
        biggest_idx = sized[0][0]
        groups[biggest_idx]["vacantCount"] = (
            groups[biggest_idx].get("vacantCount") or 0) + remaining

    new_occupied = sum(g.get("occupiedCount") or 0 for g in groups)
    new_vacant = sum(g.get("vacantCount") or 0 for g in groups)
    new_total = new_occupied + new_vacant
    rr["unitGroups"] = groups
    rr["totalUnits"] = new_total
    rr["occupiedUnits"] = new_occupied
    rr["vacantUnits"] = new_vacant
    if new_total:
        rr["occupancyRate"] = new_occupied / new_total
    financials["rentRoll"] = rr

    flags = financials.setdefault("flags", [])
    flags.append({
        "item": "Rent roll vs unit count",
        "issue": (f"Rent roll showed {current_total} units but property is "
                  f"{stated_units} units — assumed {shortfall} additional "
                  f"vacant lots at per-type market rent."),
        "severity": "high",
        "recommendation": (
            "Confirm with broker: are vacant sites excluded from the rent "
            "roll, or is the property actually smaller than stated?"
        ),
    })
    print(f"[Methodology] Imputed {shortfall} vacant lots "
          f"({current_total} → {new_total}) to match stated unit count.")


# ═══════════════════════════════════════════════════════════════════════════
# CLAUDE CALL #2 — Market Research
# ═══════════════════════════════════════════════════════════════════════════
MARKET_RESEARCH_PROMPT = """You are a CRE analyst at GGC researching a mobile home park acquisition. Be exhaustive - better data means a better deal decision.

Use web_search aggressively (8 searches available) to find:

1. **Rent comps** - 12-20 neighboring MHPs within ~25 miles. For EACH one capture:
   - Name, full address, city, state
   - Distance from subject (miles)
   - Total units (sites)
   - Lot rent (monthly)
   - Occupancy if disclosed
   - Year built / age class
   - Park-owned home %
   - Amenities list (pool, clubhouse, etc.)
   - Star/quality rating if available (1-5)
   - Source URL

2. **Sale comps** — 8-15 recent MHP transactions in the broader region (last 4 years). For EACH:
   - Property name, location, sale date (MM/YY)
   - # sites
   - Sale price
   - $/unit
   - Cap rate at sale (if disclosed)
   - NOI at sale (if known)
   - Buyer / seller (if disclosed)
   - Source URL

3. **Demographics (rich)** — for both county AND MSA level if different:
   - Total population (current)
   - 1-year, 5-year, and 10-year population growth %
   - Population projection (next 5 years if available)
   - Median household income (current)
   - 5-year household income growth %
   - Unemployment rate (current and 1yr ago for trend)
   - Poverty rate
   - % of households earning under $50K (target MHP demographic)
   - Top 5-10 employers by name with employee count if available
   - Major industries / economic drivers
   - Any large planned developments, factory openings, base expansions, etc.

4. **Alternative housing** —
   - Avg single-family home price + 1yr appreciation %
   - Median home price
   - Avg apartment rent: 1BR, 2BR, 3BR
   - Apartment rent growth (1yr)
   - Vacancy rate of local apartments
   - Construction permits / pipeline for apartments and SFR

5. **MHP affordability calc** — compute MHP all-in monthly cost (lot rent + $50K home @ 8% × 20yr ≈ $418/mo + lot rent) vs 2BR apt rent. Output the savings %.

6. **Landmarks** — distance in miles to:
   - 3 nearest major employers (NAMED, not generic)
   - Walmart / Target / big-box
   - Grocery store
   - Hospital / urgent care
   - Elementary school
   - High school
   - Community college or university
   - Major highway / interstate
   - Downtown (nearest city)
   - Airport

7. **Visual reference URLs** — Google Maps satellite/street view URLs for the subject

CRITICAL OUTPUT RULES:
- Your response MUST start with `{` and end with `}`. NOTHING else.
- DO NOT preface, explain, correct typos out loud, or comment.
- DO NOT use markdown code fences.
- All percentages as decimals (0.045 NOT "4.5%" or "4.5").
- All currency as plain numbers (78000 NOT "$78,000").
- Inside string values, escape any newlines as \\n. NO literal line breaks inside strings.

## Output schema (JSON only):
{
  "rentComps": [{"name", "address", "city", "state", "distance" (string e.g. "4.2 mi"), "units" (int), "lotRent" (number), "occupancy" (decimal 0-1), "yearBuilt" (int or null), "pohPercent" (decimal 0-1 or null), "amenities" (string), "qualityRating" (1-5 or null), "source" (URL)}],
  "saleComps": [{"name", "location", "saleDate" (MM/YY), "units" (int), "salePrice" (number), "pricePerUnit" (number), "capRate" (decimal 0-1 or null), "noi" (number or null), "buyer" (string or null), "seller" (string or null), "source" (URL)}],
  "demographics": {
    "countyName" (string),
    "countyPopulation" (int),
    "populationGrowth1yr" (decimal),
    "populationGrowth5yr" (decimal),
    "populationGrowth10yr" (decimal),
    "populationProjection5yr" (int or null),
    "medianHHIncome" (int),
    "incomeGrowth5yr" (decimal),
    "unemploymentRate" (decimal),
    "unemploymentRate1yrAgo" (decimal or null),
    "povertyRate" (decimal or null),
    "pctHHUnder50k" (decimal or null),
    "majorEmployers" (array of strings, ideally with employee counts),
    "majorIndustries" (array of strings),
    "plannedDevelopments" (string summary or null)
  },
  "altHousing": {
    "avgSFHomePrice" (int),
    "medianSFHomePrice" (int or null),
    "homePriceGrowth1yr" (decimal),
    "avgRent1BR" (int),
    "avgRent2BR" (int),
    "avgRent3BR" (int),
    "apartmentRentGrowth1yr" (decimal or null),
    "apartmentVacancyRate" (decimal or null),
    "mhpAllInCost" (int),
    "mhpVsApartmentSavingsPercent" (number, e.g. 35 for 35%)
  },
  "landmarks": [{"type" (use the exact labels: "Nearest Major Employer #1", "Nearest Major Employer #2", "Nearest Major Employer #3", "Nearest Walmart / Big Box", "Nearest Grocery Store", "Nearest Hospital / Urgent Care", "Nearest Elementary School", "Nearest High School", "Nearest Community College / University", "Nearest Major Highway / Interstate", "Downtown (nearest city)", "Nearest Airport"), "name", "distanceMiles" (number)}],
  "visuals": {"aerialView", "streetViewEntrance", "streetViewInterior", "exampleHome1", "exampleHome2", "directions"},
  "marketRentConclusion": "string (3-4 sentences with specific numbers — where subject sits vs comp range, implied upside)",
  "marketCapRateConclusion": "string (3-4 sentences with specific cap rate ranges from comps)",
  "demandSignal": "STRONG|MODERATE|WEAK",
  "demandRationale": "string (4-6 sentences citing specific demographic, employer, and affordability numbers — escape any newlines as \\n)"
}
"""


def call_market_research(api_key, property_info):
    prompt = f"""Property: {property_info.get('name', '')}
Address: {property_info.get('address', '')}
City/State: {property_info.get('city', '')}, {property_info.get('state', '')}
Units: {property_info.get('units', '')}

Pull comps, demographics, alt housing, landmarks, and visual URLs."""
    # web_search_20260209 adds dynamic filtering on Fable 5 / Opus 4.8+: the
    # model code-filters search results before they enter context, which
    # measurably improves comp accuracy and cuts token burn vs the 20250305
    # version. No beta header needed.
    tools = [{"type": "web_search_20260209", "name": "web_search", "max_uses": 12}]
    model_id = _model_for_stage(property_info, "market")
    print(f"[Claude] Starting market research call ({model_id}, with web_search)...")
    t0 = time.time()
    response = call_claude(api_key, MARKET_RESEARCH_PROMPT,
                            [{"type": "text", "text": prompt}], tools=tools,
                            model=model_id,
                            max_tokens=MAX_TOKENS_MARKET)
    elapsed = time.time() - t0
    print(f"[Claude] Market research returned in {elapsed:.1f}s "
          f"(stop_reason: {response.get('stop_reason', '?')})")
    return extract_json(response)

def call_market_research_merged(api_key, property_info, n_runs=3):
    """
    Run market research multiple times in parallel and merge the results.
    Dedupes rent comps by name, sale comps by name+date, takes the union of
    landmarks and demographics (most common value wins), and concatenates
    the narrative conclusions.

    Cost: ~n_runs × single-run cost. Use for high-stakes deals.
    """
    print(f"[Claude] Starting {n_runs}× market research merge...")
    t0 = time.time()

    # Fire all runs in parallel — they're independent and the bottleneck is
    # network/web-search latency, not local CPU
    with ThreadPoolExecutor(max_workers=n_runs) as executor:
        futures = [executor.submit(call_market_research, api_key, property_info)
                   for _ in range(n_runs)]
        results = []
        for i, f in enumerate(as_completed(futures)):
            try:
                results.append(f.result())
                print(f"[Claude] Market research run {i+1}/{n_runs} complete")
            except Exception as e:
                print(f"[Claude] Market research run {i+1}/{n_runs} FAILED: {e}")
                # Continue with however many succeeded — don't lose the whole job
                # because one of three calls timed out

    if not results:
        raise RuntimeError(f"All {n_runs} market research runs failed")

    elapsed = time.time() - t0
    print(f"[Claude] Merged {len(results)} runs in {elapsed:.1f}s")

    return _merge_market_research(results)


def _merge_market_research(runs):
    """
    Merge logic:
    - rentComps: union by name (case-insensitive), prefer the run with more fields populated
    - saleComps: union by name+saleDate, prefer the run with more fields
    - landmarks: union by type, prefer the closest-distance entry for each type
    - demographics: take the most common non-null value per field (mode), fall back to first
    - altHousing: average the numeric fields, take most common for strings
    - narrative conclusions: concatenate with " | " separator, deduped
    - demandSignal: majority vote across runs
    """
    def _completeness(d):
        """Score how 'full' a comp dict is — used to break ties between duplicate comps."""
        return sum(1 for v in d.values() if v not in (None, "", 0))

    # ── RENT COMPS: dedupe by lowercased name, keep most complete version ──
    rent_by_name = {}
    for run in runs:
        for comp in run.get("rentComps", []) or []:
            key = (comp.get("name") or "").strip().lower()
            if not key:
                continue
            if key not in rent_by_name or _completeness(comp) > _completeness(rent_by_name[key]):
                rent_by_name[key] = comp
    # Parse distance strings like "4.2 mi" or "12 miles" into a float so
    # we sort numerically (lexicographic puts "10.1 mi" before "4.2 mi").
    def _parse_distance_miles(s):
        if isinstance(s, (int, float)):
            return float(s)
        if not isinstance(s, str):
            return float("inf")
        import re
        m = re.search(r"-?\d+(\.\d+)?", s)
        return float(m.group(0)) if m else float("inf")

    merged_rent_comps = sorted(rent_by_name.values(),
                                key=lambda c: _parse_distance_miles(c.get("distance")))

    # ── SALE COMPS: dedupe by name+saleDate, same completeness rule ──
    sale_by_key = {}
    for run in runs:
        for comp in run.get("saleComps", []) or []:
            key = ((comp.get("name") or "").strip().lower(),
                   (comp.get("saleDate") or "").strip())
            if not key[0]:
                continue
            if key not in sale_by_key or _completeness(comp) > _completeness(sale_by_key[key]):
                sale_by_key[key] = comp

    # Parse sale dates ("MM/YY", "MM/YYYY", "YYYY-MM") so we sort
    # newest-first (most recent transactions are the meaningful cap-rate
    # signal; lexicographic puts "01/26" before "12/22").
    def _parse_sale_year(s):
        if not isinstance(s, str):
            return -1
        import re
        # Capture trailing 2- or 4-digit year, optional month prefix.
        m = re.search(r"(\d{1,2})[/\-](\d{2,4})$", s.strip())
        if m:
            yr = int(m.group(2))
            yr = 2000 + yr if yr < 100 else yr
            return yr * 12 + int(m.group(1))
        # ISO-ish fallback "2024-03"
        m = re.match(r"(\d{4})-(\d{1,2})", s.strip())
        if m:
            return int(m.group(1)) * 12 + int(m.group(2))
        return -1

    merged_sale_comps = sorted(sale_by_key.values(),
                                key=lambda c: _parse_sale_year(c.get("saleDate")),
                                reverse=True)

    # ── LANDMARKS: one entry per landmark type, keep closest ──
    landmark_by_type = {}
    for run in runs:
        for lm in run.get("landmarks", []) or []:
            t = lm.get("type", "")
            if not t:
                continue
            dist = lm.get("distanceMiles") or 999
            if t not in landmark_by_type or dist < (landmark_by_type[t].get("distanceMiles") or 999):
                landmark_by_type[t] = lm
    merged_landmarks = list(landmark_by_type.values())

    # ── DEMOGRAPHICS: mode (most common value) per field, falls back to first non-null ──
    from collections import Counter
    def _consensus(field_name, runs_list):
        values = [r.get("demographics", {}).get(field_name) for r in runs_list]
        non_null = [v for v in values if v not in (None, "")]
        if not non_null:
            return None
        # For numeric fields, average; for strings, take mode
        if all(isinstance(v, (int, float)) for v in non_null):
            return sum(non_null) / len(non_null)
        if all(isinstance(v, list) for v in non_null):
            # Lists like majorEmployers — take union, preserve order
            seen, out = set(), []
            for v in non_null:
                for item in v:
                    if item not in seen:
                        seen.add(item)
                        out.append(item)
            return out
        # Strings — return most common
        c = Counter(str(v) for v in non_null)
        return c.most_common(1)[0][0]

    demo_fields = ["countyName", "countyPopulation", "populationGrowth1yr",
                    "populationGrowth5yr", "populationGrowth10yr", "populationProjection5yr",
                    "medianHHIncome", "incomeGrowth5yr", "unemploymentRate",
                    "unemploymentRate1yrAgo", "povertyRate", "pctHHUnder50k",
                    "majorEmployers", "majorIndustries", "plannedDevelopments"]
    merged_demo = {f: _consensus(f, runs) for f in demo_fields}

    # ── ALT HOUSING: average numerics, mode for strings ──
    alt_fields = ["avgSFHomePrice", "medianSFHomePrice", "homePriceGrowth1yr",
                   "avgRent1BR", "avgRent2BR", "avgRent3BR",
                   "apartmentRentGrowth1yr", "apartmentVacancyRate",
                   "mhpAllInCost", "mhpVsApartmentSavingsPercent"]
    merged_alt = {}
    for f in alt_fields:
        values = [r.get("altHousing", {}).get(f) for r in runs]
        non_null = [v for v in values if isinstance(v, (int, float))]
        if non_null:
            merged_alt[f] = sum(non_null) / len(non_null)

    # ── VISUALS: take first non-empty per field ──
    visual_fields = ["aerialView", "streetViewEntrance", "streetViewInterior",
                      "exampleHome1", "exampleHome2", "directions"]
    merged_visuals = {}
    for f in visual_fields:
        for r in runs:
            v = (r.get("visuals", {}) or {}).get(f)
            if v:
                merged_visuals[f] = v
                break

    # ── NARRATIVE CONCLUSIONS: concat, deduped — different runs caught different angles ──
    def _merge_narrative(field):
        seen, parts = set(), []
        for r in runs:
            text = (r.get(field) or "").strip()
            if text and text.lower() not in seen:
                seen.add(text.lower())
                parts.append(text)
        return "  ||  ".join(parts) if parts else ""

    # ── DEMAND SIGNAL: majority vote ──
    signals = [r.get("demandSignal") for r in runs if r.get("demandSignal")]
    if signals:
        signal_vote = Counter(signals).most_common(1)[0][0]
    else:
        signal_vote = "MODERATE"

    return {
        "rentComps": merged_rent_comps,
        "saleComps": merged_sale_comps,
        "landmarks": merged_landmarks,
        "demographics": merged_demo,
        "altHousing": merged_alt,
        "visuals": merged_visuals,
        "marketRentConclusion": _merge_narrative("marketRentConclusion"),
        "marketCapRateConclusion": _merge_narrative("marketCapRateConclusion"),
        "demandSignal": signal_vote,
        "demandRationale": _merge_narrative("demandRationale"),
        "_meta": {
            "merge_runs": len(runs),
            "merged_rent_comps_count": len(merged_rent_comps),
            "merged_sale_comps_count": len(merged_sale_comps),
        }
    }

# ═══════════════════════════════════════════════════════════════════════════
# TEMPLATE FILLING — uses GGC's actual blank template
# ═══════════════════════════════════════════════════════════════════════════
def _protect_formulas(ws):
    """Wrap a worksheet so any value-write via ws.cell(row=, column=, value=)
    is skipped when the target cell already holds a formula. Defense in
    depth per CLAUDE.md §10.2: the template has ~1,651 pre-wired formulas
    (SUM, SUMIFS, IFERROR-based pricing chains, etc.) and a write-back that
    clobbers one of them yields a workbook whose final NOI doesn't trace.
    Returns a mutable counter [int] so the caller can read how many writes
    were blocked. Only `ws.cell()` is patched — the ws["P4"] = v style uses
    `_set_addr()` below.
    """
    blocked = [0]
    orig_cell = ws.cell

    def safe_cell(row, column, value=None):
        c = orig_cell(row=row, column=column)
        if value is not None:
            if isinstance(c.value, str) and c.value.startswith("="):
                blocked[0] += 1
                return c
            c.value = value
        return c

    ws.cell = safe_cell
    ws._formula_blocks = blocked
    return blocked


def _set_addr(ws, addr, value):
    """Set ws[addr].value = value, skipping if a formula is already there."""
    cell = ws[addr]
    if isinstance(cell.value, str) and cell.value.startswith("="):
        blocked = getattr(ws, "_formula_blocks", None)
        if blocked is not None:
            blocked[0] += 1
        return False
    cell.value = value
    return True


def _structural_rows(ws, row_start, row_end, value_cols=(4, 5, 6, 7)):
    """Return the set of row indices inside [row_start, row_end] whose value
    columns (default D, E, F, G — the FY Prior / FY Current / Broker / T12
    cells) hold pre-wired template formulas. These rows hold structural
    subtotals (income SUM at row 23, expense SUM at row 60, NOI at row 64,
    reconciliation IF-checks at rows 25/62, header repeats at row 27) — the
    Data Consolidation write loop must SKIP them or it lands category labels
    in column A on rows whose value columns hold huge SUM formulas, and the
    Underwriting tab's SUMIFS then pulls those formula outputs as line item
    values. (That is the bug that drove 17June's Advertising T-12 to $1.17M
    and R&M to $263k vs the seller's actual $2.5k and $10.4k — see
    Outputs/17June diagnostic for the original symptom.)
    """
    out = set()
    for r in range(row_start, row_end + 1):
        for c in value_cols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("="):
                out.add(r)
                break
    return out


# Common variations the methodology LLM emits when structured-output grammar
# masking is unavailable and prompt-only enforcement is the only line of
# defense (Anthropic structured-outputs is rejected for some pinned model
# snapshots — see _call_anthropic fallback path). The Underwriting tab's
# SUMIFS keys on EXACT template strings, so an unmapped variant silently
# zeros the line. Defense-in-depth: normalize at write-back AND in
# apply_ggc_overrides.
_GGC_CATEGORY_ALIASES = {
    "general and administrative": "G&A",
    "general & administrative":   "G&A",
    "g & a":                      "G&A",
    "g and a":                    "G&A",
    "capex reserve":              "Cap-Ex Reserve",
    "capital expenditures":       "Cap-Ex Reserve",
    "capex":                      "Cap-Ex Reserve",
    "home rent expense":          "Home Rent Expense (MH)",
    "home rent expense (poh)":    "Home Rent Expense (MH)",
    "electrcitiy":                "Electricity",      # gold-typo lands on canonical
    "trash":                      "Trash Removal",
    "water/sewer":                "Water and Sewer",
    "water & sewer":              "Water and Sewer",
}


def _normalize_ggc_category(cat):
    """Strip "Less: " prefixes and map well-known label variants to the
    canonical enum strings the template SUMIFS expect. Called on every
    ggcCategory before it lands in Data Consolidation column A, so the
    workbook is correct even when the methodology LLM emits variant
    strings under prompt-only schema enforcement.
    """
    if not isinstance(cat, str):
        return cat
    s = cat.strip()
    # Drop the "Less: " prefix some runs emit on Bad Debt / Vacancy /
    # Concessions despite the prompt's ban. UW SUMIFS look up the bare
    # category, so the prefixed string silently zeros the line.
    low = s.lower()
    if low.startswith("less:"):
        s = s[5:].strip()
        low = s.lower()
    # Apply alias map.
    if low in _GGC_CATEGORY_ALIASES:
        return _GGC_CATEGORY_ALIASES[low]
    return s


def fill_template(financials, market, output_path):
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"GGC template not found at {TEMPLATE_PATH}. "
            f"Make sure GGC_Blank_Underwriting_Sizer.xlsx is in the same folder as backend.py."
        )

    wb = load_workbook(TEMPLATE_PATH)
    # Per-worksheet counters of writes blocked because a formula already
    # lived in the target cell. Summed and surfaced at the end of fill.
    formula_blocks_total = [0]

    # ── Data Consolidation ────────────────────────────────────────────────
    # Income rows 3-36 (34 slots), Expense rows 43-102 (60 slots) — the
    # patched Underwriting tab's SUMIFS span those ranges.
    # Cols: A=GGC Cat, B=Source Name, D=FY Prior, E=FY Current, F=Broker PF,
    #       G=T12, J-U=monthly (12), H=annualization (formula — don't touch),
    #       V=row total (formula).
    #
    # STRUCTURAL ROWS inside those write ranges hold pre-wired SUM / IF / NOI
    # formulas in the template (rows 22-27 inside the income band; rows 60,
    # 62, 64 inside the expense band). We MUST skip them — writing a category
    # label into column A of one of these rows makes the Underwriting tab's
    # SUMIFS pick up the row's formula output as if it were a line item,
    # which is exactly how 17June ended up with Advertising T-12 = $1.17M
    # and R&M T-12 = $263k (the row-60 expense-subtotal and row-64 NOI
    # formula outputs were summed into the line). Scan once at template
    # load and use the resulting slot lists to lay items out.
    ws = wb["Data Consolidation"]
    income_structural  = _structural_rows(ws, 3, 36)
    expense_structural = _structural_rows(ws, 43, 102)
    income_slots  = [r for r in range(3, 37)  if r not in income_structural]
    expense_slots = [r for r in range(43, 103) if r not in expense_structural]
    _protect_formulas(ws)

    # Strip noise rows before writing. The methodology occasionally emits:
    #   - "Choose Expense Category" — a leftover from the template/prompt
    #     example that should never reach the workbook
    #   - Section-subtotal rows from the seller's chart of accounts
    #     ("4100 Total Rental Income (non-posting)", "5700 Total Personnel")
    #     — display artifacts; only the leaf GLs beneath them carry real
    #     data. The methodology prompt bans emitting them, but defense-in-
    #     depth here in case prompt-only enforcement (structured-outputs
    #     fallback) lets one through.
    _SKIP_SENTINEL_CATEGORIES = {"Choose Income Category",
                                  "Choose Expense Category", ""}
    def _keep(item):
        if _looks_like_subtotal(item.get("sellerName")):
            return False
        # Drop rows with zero T-12 AND zero monthly AND zero underwritten
        # whose sellerName screams "subtotal" — that's the LLM collapsing
        # a section header without bringing the underlying values across.
        # Real $0 line items keep their row.
        t12 = item.get("t12Total") or 0
        uw  = item.get("ggcUnderwritten") or 0
        monthly = item.get("monthly") or []
        monthly_sum = sum(m for m in monthly if isinstance(m, (int, float)))
        if t12 == 0 and uw == 0 and monthly_sum == 0:
            sn = (item.get("sellerName") or "").lower()
            if any(t in sn for t in ("total", "subtotal", "non-posting")):
                return False
        return True

    # Normalize every ggcCategory before write-back. Strip the "Less: "
    # prefix and map LLM variants ("General and Administrative" → "G&A")
    # to the canonical strings the UW SUMIFS expect. Belt-and-suspenders
    # in case structured-outputs grammar masking is silently unavailable
    # (the _call_anthropic fallback prints a warning but still proceeds).
    def _prep(items):
        out = []
        for it in items or []:
            if not _keep(it):
                continue
            cat = _normalize_ggc_category(it.get("ggcCategory"))
            if cat in _SKIP_SENTINEL_CATEGORIES:
                continue
            it = dict(it)  # copy so we don't mutate caller's data
            it["ggcCategory"] = cat
            out.append(it)
        return out

    income_items  = _prep(financials.get("income"))
    expense_items = _prep(financials.get("expenses"))

    # Overflow → hard fail. Silent truncation past the SUMIFS range would
    # understate every UW line and is exactly the kind of invisible
    # accuracy degradation CLAUDE.md §0 forbids.
    if len(income_items) > len(income_slots):
        dropped = [it.get("sellerName") for it in income_items[len(income_slots):]]
        financials.setdefault("_extractionChecks", []).append({
            "item": "Data Consolidation income capacity",
            "check": f"≤ {len(income_slots)} non-structural rows in $A$3:$A$36",
            "status": "fail",
            "detail": (f"Methodology emitted {len(income_items)} income line "
                       f"items; only {len(income_slots)} non-structural rows "
                       f"are available. Dropped: {dropped}"),
        })
    if len(expense_items) > len(expense_slots):
        dropped = [it.get("sellerName") for it in expense_items[len(expense_slots):]]
        financials.setdefault("_extractionChecks", []).append({
            "item": "Data Consolidation expense capacity",
            "check": f"≤ {len(expense_slots)} non-structural rows in $A$43:$A$102",
            "status": "fail",
            "detail": (f"Methodology emitted {len(expense_items)} expense line "
                       f"items; only {len(expense_slots)} non-structural rows "
                       f"are available. Dropped: {dropped}"),
        })

    def _write_item(r, item):
        ws.cell(row=r, column=1, value=item.get("ggcCategory", ""))
        ws.cell(row=r, column=2, value=item.get("sellerName", ""))
        ws.cell(row=r, column=4, value=item.get("fyPrior", 0))
        ws.cell(row=r, column=5, value=item.get("fyCurrent", 0))
        ws.cell(row=r, column=6, value=item.get("brokerProforma", 0))
        ws.cell(row=r, column=7, value=item.get("t12Total", 0))
        monthly = item.get("monthly") or []
        if len(monthly) == 12:
            for m_i, val in enumerate(monthly):
                ws.cell(row=r, column=10 + m_i, value=val)
        elif item.get("t12Total"):
            even = (item["t12Total"] or 0) / 12
            for m_i in range(12):
                ws.cell(row=r, column=10 + m_i, value=even)

    written_income_rows = set()
    for item, r in zip(income_items, income_slots):
        _write_item(r, item)
        written_income_rows.add(r)
    written_expense_rows = set()
    for item, r in zip(expense_items, expense_slots):
        _write_item(r, item)
        written_expense_rows.add(r)

    # Clear any trailing slots the template ships with default text
    # ("Choose Expense Category", "Input Source Data", etc.) so the
    # workbook doesn't display placeholder rows the methodology never
    # populated. Skip structural rows (their formulas must remain) and
    # rows we just wrote.
    for r in range(3, 37):
        if r in income_structural or r in written_income_rows:
            continue
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=None)
    for r in range(43, 103):
        if r in expense_structural or r in written_expense_rows:
            continue
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=None)

    # ── Rent Roll Input ────────────────────────────────────────────────────
    # Restructured column layout (matches Unit Mix Summary COUNTIFS/SUMIFS):
    #   A=Count (formula), B=Unit ID, C=Unit Type (canonical), D=Status,
    #   F=Tenant Name, G=Type detail, H=Type code, I=Lot Rent, J=Home Rent,
    #   K=Combined (formula). Data rows 3-1002.
    #
    # Prefer per-row data from rentRoll.rentRollRows (preserves real tenant
    # names + unit IDs). Fall back to unitGroups expansion when only the
    # aggregated form is available.
    ws = wb["Rent Roll Input"]
    _protect_formulas(ws)
    rr = financials.get("rentRoll") or {}
    per_row = rr.get("rentRollRows") or []
    unit_groups = rr.get("unitGroups") or []

    # Map any non-canonical unit-type string the LLM emits to the closest
    # canonical bucket. Required because we dropped the JSON-schema enum
    # constraint on unitType to keep the structured-outputs grammar
    # compilable. We still want Unit Mix Summary's COUNTIFS to match.
    def _canonicalize_unit_type(raw):
        if not isinstance(raw, str):
            return "TOH MH Site"
        s = raw.strip().lower()
        if not s:
            return "TOH MH Site"
        if "poh" in s or "park owned" in s or "park-owned" in s or "infilled" in s:
            return "POH-Infilled units"
        if "rv" in s or "annual rv" in s or "long term rv" in s or "long-term rv" in s:
            return "Long term RV Site"
        if "retail" in s or "commercial" in s or "storage" in s or "storefront" in s:
            return "Retail/Commercial"
        return "TOH MH Site"

    individual_units = []
    if per_row:
        for row in per_row:
            individual_units.append({
                "unitId":    row.get("unitId", "") or "",
                "unitType":  _canonicalize_unit_type(row.get("unitType")),
                "status":    row.get("status", "Occupied") or "Occupied",
                "tenantName": row.get("tenantName", "") or "",
                "lotRent":   row.get("lotRent", 0) or 0,
                "homeRent":  row.get("homeRent", 0) or 0,
            })
    else:
        for grp in unit_groups:
            ut = _canonicalize_unit_type(grp.get("unitType"))
            lot_rent = grp.get("lotRent", 0) or 0
            home_rent = grp.get("pohRent", 0) or 0
            name_prefix = grp.get("tenantNamePattern", "Tenant")
            occ_count = grp.get("occupiedCount", 0) or 0
            vac_count = grp.get("vacantCount", 0) or 0
            for i in range(occ_count):
                individual_units.append({
                    "unitId": "", "unitType": ut, "status": "Occupied",
                    "tenantName": f"{name_prefix} {len(individual_units) + 1}",
                    "lotRent": lot_rent, "homeRent": home_rent,
                })
            for _ in range(vac_count):
                individual_units.append({
                    "unitId": "", "unitType": ut, "status": "Vacant",
                    "tenantName": "", "lotRent": 0, "homeRent": 0,
                })

    # Template formulas (COUNTIFS / SUMIFS) only scan rows 3:1002, so the
    # rent roll capacity is 1000 rows. Surface a fail check when a deal
    # exceeds that — silent truncation would understate Total Units (N7)
    # and Occupancy (N8), which then poisons every downstream calc.
    if len(individual_units) > 1000:
        financials.setdefault("_extractionChecks", []).append({
            "item": "Rent Roll capacity",
            "check": "≤ 1000 rows",
            "status": "fail",
            "detail": (
                f"Rent roll has {len(individual_units):,} units; template "
                "scans only rows 3:1002 so the last "
                f"{len(individual_units) - 1000:,} were dropped. "
                "Extend the SUMIFS/COUNTIFS ranges in fix_template.py "
                "(search '1002') and the truncate cap below."
            ),
        })

    for i, unit in enumerate(individual_units[:1000]):
        r = 3 + i
        ws.cell(row=r, column=2,  value=unit.get("unitId", ""))      # B
        ws.cell(row=r, column=3,  value=unit.get("unitType", ""))    # C
        ws.cell(row=r, column=4,  value=unit.get("status", ""))      # D
        ws.cell(row=r, column=6,  value=unit.get("tenantName", ""))  # F
        ws.cell(row=r, column=9,  value=unit.get("lotRent", 0))      # I
        ws.cell(row=r, column=10, value=unit.get("homeRent", 0))     # J
        # A (Count) and K (Combined) are formulas seeded by fix_template.py

    # ── Add Miscellaneous tab ──────────────────────────────────────────────
    if "Miscellaneous" in wb.sheetnames:
        del wb["Miscellaneous"]
    add_miscellaneous_tab(wb, financials, market)

    # ── Add Comps Analysis tab ─────────────────────────────────────────────
    if "Comps Analysis" in wb.sheetnames:
        del wb["Comps Analysis"]
    add_comps_analysis_tab(wb, financials, market)

    # ── Subject property cells the template formulas key on ───────────────
    # The patched template puts the Subject pricing block at columns O-P
    # of GGC Underwriting. The P4 (Purchase Price) cell is wired to
    # =IFERROR(IF(ISNUMBER(P9),P9,0),0), so it reads from P9 (Asking
    # Price). Without this write, P4 stays at 0 and the entire Sources
    # and Uses / Loan Scenario / Pro Forma Y0 chain collapses.
    underw = wb["GGC Underwriting"]
    _protect_formulas(underw)
    prop = financials.get("propertyInfo") or {}
    try:
        ask = float(prop.get("askingPrice") or 0)
    except (TypeError, ValueError):
        ask = 0
    if ask > 0:
        # P9 (Asking Price) is the only cell we write — P4 (Purchase Price)
        # is wired to =IFERROR(IF(ISNUMBER(P9),P9,0),0), so it auto-computes
        # from P9 and the reviewer can adjust P9 when negotiating below ask
        # without losing the formula chain. The earlier dual-write clobbered
        # P4's formula with a literal, defeating that override path; the
        # formula-protection guard at _set_addr now blocks it anyway.
        _set_addr(underw, "P9", ask)
    # Property metadata for the M-N subject block. CorrectOutput's row
    # layout is: M4 Name, M5 Address, M6 Type, M7 Units, M8 Occupancy,
    # M9 Acreage, M10 County. Earlier code was off by one (name landed at
    # N5 where the Address label sits) — fixed here to match correct.
    if prop.get("name"):
        _set_addr(underw, "N4", prop.get("name"))
    if prop.get("address"):
        _set_addr(underw, "N5", prop.get("address"))
    if prop.get("propertyType"):
        _set_addr(underw, "N6", prop.get("propertyType"))
    if prop.get("acreage") is not None:
        _set_addr(underw, "N9", prop.get("acreage"))
    if prop.get("county"):
        _set_addr(underw, "N10", prop.get("county"))

    # County tax rate (countyTaxRate) and flood zone (floodZone) are
    # deliberately NOT written to the I22/I23 cells. Partner direction:
    # keep I22 = J22*N7 (per-unit) and I23 = D23*1.05 (T12 × 1.05) — no
    # methodology MAX branches, no flood surcharge. The Flood Zone is
    # still written below to R5 as informational metadata only.
    #
    # Per-site RE Tax rate: CorrectOutput's I22 formula is `=J22 × N7`
    # (a per-site tax assumption × unit count). The template ships J22 as
    # a hardcoded number; if the user provides `tax_per_site` on the form
    # we write it here so a single per-site update flows through I22, the
    # cap rate (P6 = I47/P4), and the Pro Forma chain. Skip when blank so
    # the template's existing J22 value is preserved.
    try:
        tax_per_site = float(prop.get("taxPerSite") or 0)
    except (TypeError, ValueError):
        tax_per_site = 0
    if tax_per_site > 0:
        # Sanity-clamp: real MHC per-site taxes run roughly $100-$2,000/site
        # depending on county. A 10× typo (e.g. 4000 for 400) would silently
        # inflate I22 by an order of magnitude. Flag clearly when outside.
        if tax_per_site < 100 or tax_per_site > 2_000:
            financials.setdefault("_extractionChecks", []).append({
                "item": "Per-site tax assumption",
                "check": "$100-$2,000/site is typical",
                "status": "warn",
                "detail": (f"tax_per_site=${tax_per_site:,.0f} is outside the "
                           f"typical MHC range. Confirm this isn't a typo."),
            })
        _set_addr(underw, "J22", tax_per_site)

    # Underwritten date stamp at N2. Today's date is the default; the
    # reviewer can overwrite in-cell if they want to date-stamp to a
    # different reference period. M2 holds the label (set in
    # fix_template.py section 19d).
    from datetime import datetime as _dt
    if _set_addr(underw, "N2", _dt.now()):
        underw["N2"].number_format = "[$-F800]dddd\\,\\ mmmm\\ dd\\,\\ yyyy"

    # Right-side utility / build metadata block (Q3:R8). Each row is
    # optional — write only when the methodology agent extracted a real
    # value, otherwise leave the template's blank cell alone.
    if prop.get("websiteUrl"):
        _set_addr(underw, "R3", prop.get("websiteUrl"))
    if prop.get("yearBuilt") is not None:
        _set_addr(underw, "R4", prop.get("yearBuilt"))
    # Flood Zone: the user's form value takes precedence (they know the
    # zone from their own due diligence), with methodology extraction as
    # the fallback. Skip the placeholder string "unknown".
    fz = prop.get("floodZone")
    if fz and str(fz).strip().lower() not in ("", "unknown", "none"):
        _set_addr(underw, "R5", fz)
    if prop.get("utilityStructure"):
        _set_addr(underw, "R6", prop.get("utilityStructure"))
    if prop.get("electricityNotes"):
        _set_addr(underw, "R7", prop.get("electricityNotes"))
    if prop.get("trashNotes"):
        _set_addr(underw, "R8", prop.get("trashNotes"))

    # Tax Analysis Section (M19:R33). N19 holds the county assessor URL
    # if the methodology found one; parcel rows go into M26:R32. The
    # summary formulas at N20/N21/N22 are wired in fix_template.py to
    # the parcel table, so populating the table cascades into the
    # Assessed Value / Levy Rate / Estimated Tax cells automatically.
    if prop.get("taxAssessorUrl"):
        _set_addr(underw, "N19", prop.get("taxAssessorUrl"))
    parcels = prop.get("taxParcels") or []
    if isinstance(parcels, list):
        # Write up to 7 parcels into rows 26-32. Excess parcels are
        # dropped (the table is fixed-height in the template); under-7
        # parcel sets leave trailing rows blank.
        for i, parcel in enumerate(parcels[:7]):
            r = 26 + i
            if not isinstance(parcel, dict):
                continue
            if parcel.get("parcelId"):
                _set_addr(underw, f"M{r}", parcel.get("parcelId"))
            if parcel.get("marketValue") is not None:
                _set_addr(underw, f"N{r}", parcel.get("marketValue"))
            if parcel.get("taxableValue") is not None:
                _set_addr(underw, f"O{r}", parcel.get("taxableValue"))
            if parcel.get("taxes") is not None:
                _set_addr(underw, f"P{r}", parcel.get("taxes"))
            if parcel.get("acres") is not None:
                _set_addr(underw, f"R{r}", parcel.get("acres"))

    # Tally formula-protection blocks across the worksheets we wrapped.
    # When a write was deferred to a pre-wired template formula, surface
    # it on the Extraction Check tab so the reviewer can spot any
    # unexpected collisions. A nonzero count is informational, not a
    # failure — the template formula was preserved exactly as intended.
    total_blocks = sum(
        ws._formula_blocks[0]
        for ws in (wb["Data Consolidation"], wb["Rent Roll Input"],
                   wb["GGC Underwriting"])
        if getattr(ws, "_formula_blocks", None) is not None
    )
    if total_blocks:
        financials.setdefault("_extractionChecks", []).append({
            "item": "Template formula protection",
            "check": "Skip writes that would clobber pre-wired formulas",
            "status": "warn",
            "detail": (f"{total_blocks} cell write(s) were deferred because "
                       "the template already held a formula at that "
                       "address. Confirm the cells you expected to "
                       "populate (asking price, property metadata, data "
                       "consolidation) made it through."),
        })
        print(f"[Template] Formula protection blocked {total_blocks} write(s).")

    # ── Add Extraction Check tab (source reconciliation) ───────────────────
    # This is the "do the numbers tie out?" tab Michael asked for in the
    # meeting. Lives at the front so it's the first thing the reviewer
    # sees. Built LAST so it can include the formula-protection tally
    # above and any other checks accumulated during fill_template.
    if "Extraction Check" in wb.sheetnames:
        del wb["Extraction Check"]
    add_extraction_check_tab(wb, financials)

    # Force Excel to recalculate every formula when the user opens the
    # output. Without these flags, openpyxl-written formulas show as
    # blank cells in Excel until F9 is pressed manually — which the user
    # observed as "all my Pro Forma and Loan Scenario cells are empty".
    # Re-asserting here (in addition to fix_template.py) protects against
    # openpyxl silently resetting calculation properties on save.
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcCompleted = False
    wb.calculation.calcOnSave = True

    # Atomic save: write to a sibling temp file, then rename. A mid-save
    # crash would otherwise leave the destination half-written, and
    # /api/download would happily ship the corrupt bytes.
    output_path = Path(output_path)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, output_path)
    except Exception:
        # Don't leave a stale .tmp behind on failure.
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise
    return output_path


def add_extraction_check_tab(wb, financials):
    """
    Source-reconciliation tab. Shows what the extraction step pulled from the
    documents and whether it ties out — the verification checkpoint that lets
    the reviewer trust the numbers before reading the underwriting. Placed at
    index 0 so it's the first tab in the workbook.
    """
    ws = wb.create_sheet("Extraction Check", 0)
    ws.sheet_view.showGridLines = False

    NAVY = "1F3864"
    MID_BLUE = "2E5090"
    GREEN = "16A34A"
    AMBER = "D97706"
    RED = "DC2626"
    GRAY = "6B7280"
    LIGHT_YEL = "FFF2CC"
    WHITE = "FFFFFF"
    GREEN_BG = "C6EFCE"
    AMBER_BG = "FFEB9C"
    RED_BG = "FFC7CE"

    def style(cell, value=None, bold=False, color="000000", size=10,
              bg=None, align="left", v_align="center", wrap=False, italic=False):
        if value is not None:
            cell.value = value
        cell.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
        cell.alignment = Alignment(horizontal=align, vertical=v_align, wrap_text=wrap)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    for col, w in enumerate([2, 38, 30, 12, 52], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells("B2:F2")
    style(ws["B2"], "EXTRACTION CHECK — SOURCE RECONCILIATION", bold=True,
          color=WHITE, size=16, bg=NAVY, align="center")
    ws.row_dimensions[2].height = 30
    ws.merge_cells("B3:F3")
    style(ws["B3"], "What the model pulled from the documents, and whether it ties to the source. Review before trusting the underwriting.",
          color=GRAY, size=10, align="center")
    ws.row_dimensions[3].height = 18

    extraction = financials.get("_extraction", {}) or {}
    checks = financials.get("_extractionChecks", []) or []

    # ── Reporting period summary ─────────────────────────────────────────
    rp = extraction.get("reportingPeriod", {}) or {}
    ws.merge_cells("B5:F5")
    style(ws["B5"], "  REPORTING PERIOD USED", bold=True, color=WHITE, size=12, bg=NAVY)
    ws.row_dimensions[5].height = 22

    period_rows = [
        ("Period used", rp.get("periodUsed", "—")),
        ("Date range", rp.get("dateRange", "—")),
        ("Months covered", rp.get("monthsCovered", "—")),
        ("Other periods seen in doc", ", ".join(str(c) for c in (rp.get("candidatePeriodsSeen") or [])) or "—"),
        ("Notes", rp.get("notes", "") or "—"),
    ]
    for i, (label, val) in enumerate(period_rows):
        r = 6 + i
        style(ws.cell(row=r, column=2), label, bold=True, size=10)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        style(ws.cell(row=r, column=3), str(val), bg=LIGHT_YEL, color="0000FF", wrap=True)
        ws.row_dimensions[r].height = 18

    # ── Checks table ─────────────────────────────────────────────────────
    checks_start = 6 + len(period_rows) + 1
    ws.merge_cells(start_row=checks_start, start_column=2, end_row=checks_start, end_column=6)
    style(ws.cell(row=checks_start, column=2), "  RECONCILIATION CHECKS", bold=True,
          color=WHITE, size=12, bg=NAVY)
    ws.row_dimensions[checks_start].height = 22

    # Summary counts
    n_ok = sum(1 for c in checks if c.get("status") == "ok")
    n_warn = sum(1 for c in checks if c.get("status") == "warn")
    n_fail = sum(1 for c in checks if c.get("status") == "fail")
    summ_r = checks_start + 1
    ws.merge_cells(start_row=summ_r, start_column=2, end_row=summ_r, end_column=6)
    if n_fail:
        summ_text = f"⚠ {n_fail} FAILED, {n_warn} warnings, {n_ok} passed — review the failures before using these numbers."
        summ_bg = RED
    elif n_warn:
        summ_text = f"{n_warn} warnings, {n_ok} passed — numbers mostly tie out; check the warnings."
        summ_bg = AMBER
    else:
        summ_text = f"✓ All {n_ok} checks passed — extracted numbers tie to the source."
        summ_bg = GREEN
    style(ws.cell(row=summ_r, column=2), summ_text, bold=True, color=WHITE, size=11,
          bg=summ_bg, align="center")
    ws.row_dimensions[summ_r].height = 24

    # Column headers
    hdr_r = summ_r + 1
    for col, label in [(2, "Item"), (3, "Check"), (4, "Status"), (5, "Detail")]:
        style(ws.cell(row=hdr_r, column=col), label, bold=True, color=WHITE,
              size=9, bg=MID_BLUE, align="center")
    ws.row_dimensions[hdr_r].height = 18

    status_map = {
        "ok":   ("✓ OK",   GREEN_BG, GREEN),
        "warn": ("⚠ WARN", AMBER_BG, AMBER),
        "fail": ("✗ FAIL", RED_BG,   RED),
    }
    # Sort so failures float to the top, then warnings, then OK
    order = {"fail": 0, "warn": 1, "ok": 2}
    sorted_checks = sorted(checks, key=lambda c: order.get(c.get("status"), 3))
    for i, c in enumerate(sorted_checks):
        r = hdr_r + 1 + i
        label, bg, fg = status_map.get(c.get("status"), ("?", "FFFFFF", "000000"))
        style(ws.cell(row=r, column=2), c.get("item", ""), size=9, wrap=True)
        style(ws.cell(row=r, column=3), c.get("check", ""), size=9, wrap=True)
        style(ws.cell(row=r, column=4), label, bold=True, color=fg, bg=bg, align="center", size=9)
        style(ws.cell(row=r, column=5), c.get("detail", ""), size=9, wrap=True)
        ws.row_dimensions[r].height = 24

    if not checks:
        r = hdr_r + 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        style(ws.cell(row=r, column=2), "No checks were generated for this run.",
              italic=True, color=GRAY, align="center")

    ws.sheet_properties.tabColor = "DC2626"
    return ws


def add_comps_analysis_tab(wb, financials, market):
    """Dedicated comps analysis tab — rent comps, sale comps, statistics,
    subject vs market comparison, and demographics expansion."""
    ws = wb.create_sheet("Comps Analysis", 1)  # Place right after Miscellaneous
    ws.sheet_view.showGridLines = False

    NAVY = "1F3864"
    MID_BLUE = "2E5090"
    GRAY_HDR = "404040"
    LIGHT_YEL = "FFF2CC"
    LIGHT_GRAY = "F2F2F2"
    GREEN_BG = "C6EFCE"
    RED_BG = "FFC7CE"
    WHITE = "FFFFFF"

    def style(cell, value=None, bold=False, color="000000", size=10,
              bg=None, align="left", v_align="center", wrap=False, italic=False, fmt=None):
        if value is not None:
            cell.value = value
        cell.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
        cell.alignment = Alignment(horizontal=align, vertical=v_align, wrap_text=wrap)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if fmt:
            cell.number_format = fmt

    def section_header(row, end_col, title):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end_col)
        style(ws.cell(row=row, column=2), title, bold=True, color=WHITE, size=12, bg=NAVY)
        ws.row_dimensions[row].height = 22

    def col_header(row, col, label, bg=GRAY_HDR):
        style(ws.cell(row=row, column=col), label, bold=True, color=WHITE,
              size=9, bg=bg, align="center", wrap=True)

    # Column widths — wide enough for all comp data
    widths = [2, 26, 22, 9, 9, 11, 11, 11, 9, 22, 18, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells("B2:L2")
    style(ws["B2"], "COMPS ANALYSIS — DETAILED COMPETITIVE SET", bold=True,
          color=WHITE, size=16, bg=NAVY, align="center")
    ws.row_dimensions[2].height = 32
    ws.merge_cells("B3:L3")
    style(ws["B3"], "Rent comps, sale comps, statistics, and subject-vs-market positioning",
          color="6B7280", size=10, align="center")
    ws.row_dimensions[3].height = 18

    # Data-source disclaimer. Comps in this tab are generated by an LLM
    # (Claude Opus) via web_search, not pulled from a structured comp DB
    # like CoStar or MHVillage. The data and the source URL come from
    # different snippets of the same search, so the linked page often does
    # NOT contain the lot rent / occupancy / cap rate shown in the row.
    # This banner exists so partners reading the workbook understand the
    # reliability limit before they cite a number in IC.
    ws.merge_cells("B5:L5")
    style(ws["B5"], "  ⚠  DATA SOURCE — READ BEFORE USING", bold=True,
          color=WHITE, size=11, bg="C65911", align="left")
    ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:L8")
    _disclaimer = (
        "Comps below are sourced via AI web research, not from a structured "
        "comp database (CoStar / MHVillage / county records). The Source URL "
        "in each row is the page the AI cited, but it may NOT contain the "
        "specific lot rent, occupancy, or cap rate populated here — those "
        "fields are inferred from multiple web snippets. Treat all comp data "
        "as a directional starting point and verify every number against the "
        "primary source before using it in IC materials."
    )
    style(ws["B6"], _disclaimer, size=10, color="7B3F00", bg="FFF2CC",
          align="left", v_align="top", wrap=True)
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 18

    # Investment Criteria Check was removed: GGC reviewers consistently
    # flagged the ingoing cap rate / stabilized YOC / spread numbers as
    # wrong (the LLM's spreadBps drifts off the rates it emits, the
    # ingoing cap depends on a purchase price that may be a placeholder,
    # and the verdict was being shown with high visual prominence on
    # values nobody trusted). The 200-bps rule still lives in the prompt
    # and in the methodology audit trail — but it no longer prints a
    # verdict box on the Comps tab that contradicts what the underwriter
    # sees when they actually read the Pro Forma.

    rent_comps = market.get("rentComps", []) or []
    sale_comps = market.get("saleComps", []) or []

    # ── SUBJECT vs MARKET POSITIONING ────────────────────────────────────────
    subject_rent = financials.get("rentRoll", {}).get("avgLotRent", 0) or 0
    subject_units = financials.get("propertyInfo", {}).get("totalUnits", 0) or 0
    subject_occ = financials.get("rentRoll", {}).get("occupancyRate", 0) or 0
    asking = financials.get("propertyInfo", {}).get("askingPrice", 0) or 0
    ppu_ask = (asking / subject_units) if subject_units else 0

    # Comp set statistics
    rents = [c.get("lotRent") for c in rent_comps if isinstance(c.get("lotRent"), (int, float))]
    units_list = [c.get("units") for c in rent_comps if isinstance(c.get("units"), (int, float))]
    # Normalize occupancy to decimal (LLM may emit 95 to mean 95%, or 0.95).
    # Same heuristic as to_decimal_pct: values >= 1.5 are presumed percent
    # form and divided by 100. Otherwise % cells show 9500%.
    occ_list = [to_decimal_pct(c.get("occupancy")) for c in rent_comps]
    occ_list = [v for v in occ_list if isinstance(v, (int, float))]

    sale_caps = [to_decimal_pct(c.get("capRate")) for c in sale_comps]
    sale_caps = [v for v in sale_caps if isinstance(v, (int, float))]
    sale_ppu = [c.get("pricePerUnit") for c in sale_comps if isinstance(c.get("pricePerUnit"), (int, float))]

    def safe_min(lst): return min(lst) if lst else None
    def safe_max(lst): return max(lst) if lst else None
    def safe_avg(lst): return (sum(lst) / len(lst)) if lst else None
    def safe_median(lst):
        if not lst: return None
        s = sorted(lst)
        n = len(s)
        return s[n//2] if n % 2 else (s[n//2 - 1] + s[n//2]) / 2

    section_header(11, 12, "  SUBJECT vs MARKET — KEY METRICS")
    headers = ["Metric", "Subject", "Comp Min", "Comp Max", "Comp Avg", "Comp Median", "Subject Position"]
    for i, h in enumerate(headers):
        col_header(12, 2 + i, h)
    ws.row_dimensions[12].height = 28

    def positioning(subj, comp_avg):
        if subj is None or comp_avg is None or comp_avg == 0:
            return ("N/A", "F2F2F2")
        delta = (subj - comp_avg) / comp_avg
        if delta < -0.05:
            return (f"{abs(delta)*100:.1f}% below avg ↑ upside", GREEN_BG)
        elif delta > 0.05:
            return (f"{delta*100:.1f}% above avg", RED_BG)
        else:
            return (f"In line ({delta*100:+.1f}%)", LIGHT_GRAY)

    pos_text, pos_bg = positioning(subject_rent, safe_avg(rents))
    metrics = [
        ("Lot Rent ($/mo)", subject_rent, safe_min(rents), safe_max(rents),
         safe_avg(rents), safe_median(rents), pos_text, pos_bg, "$#,##0"),
        ("# Sites", subject_units, safe_min(units_list), safe_max(units_list),
         safe_avg(units_list), safe_median(units_list), "", LIGHT_GRAY, "#,##0"),
        ("Occupancy", subject_occ, safe_min(occ_list), safe_max(occ_list),
         safe_avg(occ_list), safe_median(occ_list), "", LIGHT_GRAY, "0.0%"),
    ]
    for i, m in enumerate(metrics):
        r = 13 + i
        style(ws.cell(row=r, column=2), m[0], bold=True, size=10)
        for j in range(5):
            val = m[1 + j]
            style(ws.cell(row=r, column=3 + j), val, bg=LIGHT_YEL,
                  color="0000FF", align="right", fmt=m[8])
        style(ws.cell(row=r, column=8), m[6], bold=True, size=10, bg=m[7], align="center")
        ws.row_dimensions[r].height = 18

    # ── RENT COMPS TABLE ─────────────────────────────────────────────────────
    rc_start = 18
    section_header(rc_start, 12, f"  RENT COMPS  ({len(rent_comps)} found)")
    rc_headers = ["Property", "Location", "Distance", "# Sites", "Lot Rent",
                  "Occupancy", "Year Built", "POH %", "Amenities", "Quality", "Source"]
    for i, h in enumerate(rc_headers):
        col_header(rc_start + 1, 2 + i, h)
    ws.row_dimensions[rc_start + 1].height = 28

    for i, c in enumerate(rent_comps[:30]):
        r = rc_start + 2 + i
        loc = f"{c.get('city', '')}, {c.get('state', '')}".strip(", ")
        rating = c.get("qualityRating")
        rating_str = ("★" * int(rating)) if isinstance(rating, (int, float)) and rating else ""
        cells = [
            (c.get("name", ""),         "left",   None),
            (loc,                       "left",   None),
            (c.get("distance", ""),     "center", None),
            (c.get("units", 0),         "center", "#,##0"),
            (c.get("lotRent", 0),       "right",  "$#,##0"),
            (to_decimal_pct(c.get("occupancy")),  "center", "0.0%"),
            (c.get("yearBuilt", ""),    "center", None),
            (to_decimal_pct(c.get("pohPercent")), "center", "0.0%"),
            ((c.get("amenities") or "")[:80], "left", None),
            (rating_str,                "center", None),
            (c.get("source", ""),       "left",   None),
        ]
        for j, (val, align, fmt) in enumerate(cells):
            bg = LIGHT_GRAY if i % 2 == 0 else WHITE
            style(ws.cell(row=r, column=2 + j), val, size=9, bg=bg,
                  align=align, fmt=fmt, wrap=(j == 8))
        ws.row_dimensions[r].height = 30

    # Stats row at the bottom of the rent comps table
    if rent_comps:
        stats_r = rc_start + 2 + len(rent_comps[:30])
        style(ws.cell(row=stats_r, column=2), "STATISTICS", bold=True,
              color=WHITE, bg=MID_BLUE)
        style(ws.cell(row=stats_r, column=3), "", bg=MID_BLUE)
        style(ws.cell(row=stats_r, column=4), "Range:", bold=True, bg=MID_BLUE,
              color=WHITE, align="right")
        if rents:
            style(ws.cell(row=stats_r, column=5),
                  f"{safe_min(units_list) or 0:,.0f}-{safe_max(units_list) or 0:,.0f}",
                  bold=True, color=WHITE, bg=MID_BLUE, align="center")
            style(ws.cell(row=stats_r, column=6),
                  f"${safe_min(rents):,.0f}-${safe_max(rents):,.0f}",
                  bold=True, color=WHITE, bg=MID_BLUE, align="center")
        ws.row_dimensions[stats_r].height = 20

        avg_r = stats_r + 1
        style(ws.cell(row=avg_r, column=4), "Average:", bold=True, color=WHITE,
              bg=MID_BLUE, align="right")
        if units_list:
            style(ws.cell(row=avg_r, column=5),
                  f"{safe_avg(units_list):,.0f}", bold=True, color=WHITE,
                  bg=MID_BLUE, align="center")
        if rents:
            style(ws.cell(row=avg_r, column=6),
                  f"${safe_avg(rents):,.0f}", bold=True, color=WHITE,
                  bg=MID_BLUE, align="center")
        if occ_list:
            style(ws.cell(row=avg_r, column=7),
                  f"{safe_avg(occ_list)*100:.1f}%", bold=True, color=WHITE,
                  bg=MID_BLUE, align="center")

    # ── MARKET RENT CONCLUSION ──────────────────────────────────────────────
    rc_concl_r = rc_start + 4 + len(rent_comps[:30])
    section_header(rc_concl_r, 12, "  MARKET RENT CONCLUSION")
    style(ws.cell(row=rc_concl_r + 1, column=2), market.get("marketRentConclusion", ""),
          size=11, wrap=True, bg=LIGHT_YEL, color="0000FF", v_align="top")
    ws.merge_cells(start_row=rc_concl_r + 1, start_column=2, end_row=rc_concl_r + 1, end_column=12)
    ws.row_dimensions[rc_concl_r + 1].height = 60

    # ── SALE COMPS TABLE ─────────────────────────────────────────────────────
    sc_start = rc_concl_r + 3
    section_header(sc_start, 12, f"  SALE COMPS  ({len(sale_comps)} found)")
    sc_headers = ["Property", "Location", "Sale Date", "# Sites", "Sale Price",
                  "$ / Unit", "Cap Rate", "NOI", "Buyer", "Seller", "Source"]
    for i, h in enumerate(sc_headers):
        col_header(sc_start + 1, 2 + i, h)
    ws.row_dimensions[sc_start + 1].height = 28

    for i, c in enumerate(sale_comps[:30]):
        r = sc_start + 2 + i
        cells = [
            (c.get("name", ""),         "left",   None),
            (c.get("location", ""),     "left",   None),
            (c.get("saleDate", ""),     "center", None),
            (c.get("units", 0),         "center", "#,##0"),
            (c.get("salePrice", 0),     "right",  "$#,##0"),
            (c.get("pricePerUnit", 0),  "right",  "$#,##0"),
            (c.get("capRate"),          "center", "0.00%"),
            (c.get("noi"),              "right",  "$#,##0"),
            (c.get("buyer", ""),        "left",   None),
            (c.get("seller", ""),       "left",   None),
            (c.get("source", ""),       "left",   None),
        ]
        for j, (val, align, fmt) in enumerate(cells):
            bg = LIGHT_GRAY if i % 2 == 0 else WHITE
            style(ws.cell(row=r, column=2 + j), val, size=9, bg=bg, align=align, fmt=fmt)
        ws.row_dimensions[r].height = 18

    # Sale comp statistics — only $/Unit averaged. Cap-rate aggregates
    # (Avg Cap, trimmed mean, median, std dev, Implied Valuation, Market
    # Cap Rate Conclusion) were removed: GGC reviewers consistently said
    # the per-comp cap rates the LLM scrapes are unreliable (sources
    # rarely disclose cap rates honestly, and what's printed is often a
    # broker's marketing number rather than the actual deal economics).
    # Averaging unreliable inputs into a single "market cap rate" amplifies
    # the noise. Per-comp cap rates remain visible in the table as raw
    # data — but the tool stops computing aggregates from them.
    if sale_comps:
        stats_r = sc_start + 2 + len(sale_comps[:30])
        style(ws.cell(row=stats_r, column=2), "STATISTICS", bold=True,
              color=WHITE, bg=MID_BLUE)
        if sale_ppu:
            style(ws.cell(row=stats_r, column=7),
                  f"${safe_avg(sale_ppu):,.0f}", bold=True, color=WHITE,
                  bg=MID_BLUE, align="center")
            style(ws.cell(row=stats_r, column=6), "Avg $/Unit:", bold=True,
                  color=WHITE, bg=MID_BLUE, align="right")
        ws.row_dimensions[stats_r].height = 20

    # cap_concl_r is kept defined to anchor downstream sections that
    # were calculated relative to it (Demographics block, etc.).
    cap_concl_r = sc_start + 4 + len(sale_comps[:30])

    # ── DEMOGRAPHICS DEEP DIVE ───────────────────────────────────────────────
    demo = market.get("demographics", {}) or {}
    dem_start = cap_concl_r + 3
    section_header(dem_start, 12, "  DEMOGRAPHICS — DEEP DIVE")

    demo_rows = [
        ("County", demo.get("countyName", "")),
        ("Total Population", demo.get("countyPopulation")),
        ("1-Year Pop Growth", safe_pct(demo.get("populationGrowth1yr"))),
        ("5-Year Pop Growth", safe_pct(demo.get("populationGrowth5yr"))),
        ("10-Year Pop Growth", safe_pct(demo.get("populationGrowth10yr"))),
        ("5-Yr Population Projection", demo.get("populationProjection5yr")),
        ("Median HH Income", safe_money(demo.get("medianHHIncome"))),
        ("5-Year Income Growth", safe_pct(demo.get("incomeGrowth5yr"))),
        ("Unemployment (Current)", safe_pct(demo.get("unemploymentRate"))),
        ("Unemployment (1yr Ago)", safe_pct(demo.get("unemploymentRate1yrAgo"))),
        ("Poverty Rate", safe_pct(demo.get("povertyRate"))),
        ("% HHs Earning < $50K", safe_pct(demo.get("pctHHUnder50k"))),
        ("Major Industries", ", ".join(demo.get("majorIndustries", []) or [])),
        ("Planned Developments", demo.get("plannedDevelopments", "")),
    ]
    for i, (label, val) in enumerate(demo_rows):
        r = dem_start + 1 + i
        style(ws.cell(row=r, column=2), label, bold=True, size=10)
        style(ws.cell(row=r, column=3), val, bg=LIGHT_YEL, color="0000FF")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
        ws.row_dimensions[r].height = 18

    # Top employers as a separate sub-section
    emp_start = dem_start + 1 + len(demo_rows) + 1
    style(ws.cell(row=emp_start, column=2), "TOP EMPLOYERS",
          bold=True, color=WHITE, bg=MID_BLUE, size=11)
    ws.merge_cells(start_row=emp_start, start_column=2, end_row=emp_start, end_column=12)
    ws.row_dimensions[emp_start].height = 22

    employers = demo.get("majorEmployers", []) or []
    for i, emp in enumerate(employers[:15]):
        r = emp_start + 1 + i
        style(ws.cell(row=r, column=2), f"#{i+1}", bold=True, size=10, align="center")
        style(ws.cell(row=r, column=3), str(emp), bg=LIGHT_YEL, color="0000FF")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
        ws.row_dimensions[r].height = 18

    ws.sheet_properties.tabColor = "70AD47"  # green tab


def add_miscellaneous_tab(wb, financials, market):
    """Add a Miscellaneous tab at the front with property links, demographics,
    images, landmarks, and amenities — value-add data on top of GGC's template."""
    ws = wb.create_sheet("Miscellaneous", 0)
    ws.sheet_view.showGridLines = False

    NAVY = "1F3864"
    MID_BLUE = "2E5090"
    GRAY_HDR = "404040"
    LIGHT_YEL = "FFF2CC"
    WHITE = "FFFFFF"

    def style_cell(cell, value=None, bold=False, color="000000", size=10,
                   bg=None, align="left", v_align="center", wrap=False, italic=False):
        if value is not None:
            cell.value = value
        cell.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
        cell.alignment = Alignment(horizontal=align, vertical=v_align, wrap_text=wrap)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    def section_header(row, end_col_letter, title):
        ws.merge_cells(f"B{row}:{end_col_letter}{row}")
        style_cell(ws[f"B{row}"], title, bold=True, color=WHITE, size=12, bg=NAVY)
        ws.row_dimensions[row].height = 22

    for col, w in enumerate([2, 28, 38, 38, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("B2:E2")
    style_cell(ws["B2"], "PROPERTY OVERVIEW", bold=True, color=WHITE, size=16, bg=NAVY, align="center")
    ws.row_dimensions[2].height = 32
    ws.merge_cells("B3:E3")
    style_cell(ws["B3"], "Location context, visuals, demographics, and qualitative assessment",
               color="6B7280", size=10, align="center")
    ws.row_dimensions[3].height = 18

    prop = financials.get("propertyInfo", {})
    full_addr = f"{prop.get('address','')}, {prop.get('city','')}, {prop.get('state','')}".strip(", ")
    addr_encoded = quote_plus(full_addr) if full_addr else ""

    # PROPERTY LINKS
    section_header(5, "E", "  PROPERTY LINKS")
    links = [
        ("Property Name", prop.get("name", "")),
        ("Address", full_addr),
        ("Property Type", prop.get("propertyType", "")),
        ("Total Units", prop.get("totalUnits", "")),
        ("County", prop.get("county", "")),
        ("Asking Price", prop.get("askingPrice", "")),
        ("CoStar Link", ""),
        ("Property Website", ""),
        ("Broker / OM", ""),
        ("County Assessor", ""),
        ("FEMA Flood Map", ""),
    ]
    for i, (label, val) in enumerate(links):
        r = 6 + i
        style_cell(ws.cell(row=r, column=2), label, bold=True, size=10)
        style_cell(ws.cell(row=r, column=3), val, color="0000FF", bg=LIGHT_YEL)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 18

    # DEMOGRAPHICS
    demo_start = 6 + len(links) + 1
    section_header(demo_start, "E", "  DEMOGRAPHICS")
    demo = market.get("demographics", {}) or {}
    demo_rows = [
        ("County Population", demo.get("countyPopulation", "")),
        ("5-Year Population Growth", safe_pct(demo.get("populationGrowth5yr"))),
        ("Median Household Income", safe_money(demo.get("medianHHIncome"))),
        ("Unemployment Rate", safe_pct(demo.get("unemploymentRate"))),
        ("Top Employers", ", ".join(demo.get("majorEmployers", []) or [])),
    ]
    for i, (label, val) in enumerate(demo_rows):
        r = demo_start + 1 + i
        style_cell(ws.cell(row=r, column=2), label, bold=True, size=10)
        style_cell(ws.cell(row=r, column=3), val, color="0000FF", bg=LIGHT_YEL)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 18

    # ALTERNATIVE HOUSING
    alt_start = demo_start + 1 + len(demo_rows) + 1
    section_header(alt_start, "E", "  ALTERNATIVE HOUSING (affordability)")
    alt = market.get("altHousing", {}) or {}
    alt_rows = [
        ("Avg Single-Family Home Price", safe_money(alt.get("avgSFHomePrice"))),
        ("Avg 1BR Apartment Rent",       safe_money(alt.get("avgRent1BR"), "/mo")),
        ("Avg 2BR Apartment Rent",       safe_money(alt.get("avgRent2BR"), "/mo")),
        ("Avg 3BR Apartment Rent",       safe_money(alt.get("avgRent3BR"), "/mo")),
        ("MHP All-In Monthly Cost",      safe_money(alt.get("mhpAllInCost"), "/mo")),
        ("MHP Savings vs 2BR Apt",       safe_pct(alt.get("mhpVsApartmentSavingsPercent"))),
    ]
    for i, (label, val) in enumerate(alt_rows):
        r = alt_start + 1 + i
        style_cell(ws.cell(row=r, column=2), label, bold=True, size=10)
        style_cell(ws.cell(row=r, column=3), val, color="0000FF", bg=LIGHT_YEL)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 18

    # AERIAL & ROADMAP
    aerial_start = alt_start + 1 + len(alt_rows) + 1
    section_header(aerial_start, "E", "  AERIAL & LOCATION MAP")
    style_cell(ws.cell(row=aerial_start+1, column=2), "Satellite / Aerial",
               bold=True, color=WHITE, size=10, bg=MID_BLUE, align="center")
    ws.merge_cells(start_row=aerial_start+1, start_column=2, end_row=aerial_start+1, end_column=3)
    style_cell(ws.cell(row=aerial_start+1, column=4), "Roadmap (with landmarks)",
               bold=True, color=WHITE, size=10, bg=MID_BLUE, align="center")
    ws.merge_cells(start_row=aerial_start+1, start_column=4, end_row=aerial_start+1, end_column=5)

    IMG_HEIGHT_ROWS = 15
    for r in range(aerial_start + 2, aerial_start + 2 + IMG_HEIGHT_ROWS):
        ws.row_dimensions[r].height = 20

    if full_addr:
        aerial_path = fetch_google_static_map(full_addr, "satellite", 17, "600x400")
        if aerial_path:
            embed_image_in_cell(ws, aerial_path, f"B{aerial_start+2}")
        else:
            style_cell(ws.cell(row=aerial_start+2, column=2),
                       "(set GOOGLE_MAPS_API_KEY to embed images)",
                       italic=True, color="9CA3AF", size=9, align="center")
            ws.merge_cells(start_row=aerial_start+2, start_column=2,
                           end_row=aerial_start+2+IMG_HEIGHT_ROWS-1, end_column=3)

        roadmap_path = fetch_google_static_map(full_addr, "roadmap", 14, "600x400")
        if roadmap_path:
            embed_image_in_cell(ws, roadmap_path, f"D{aerial_start+2}")
        else:
            style_cell(ws.cell(row=aerial_start+2, column=4),
                       "(set GOOGLE_MAPS_API_KEY to embed images)",
                       italic=True, color="9CA3AF", size=9, align="center")
            ws.merge_cells(start_row=aerial_start+2, start_column=4,
                           end_row=aerial_start+2+IMG_HEIGHT_ROWS-1, end_column=5)

    # Street View intentionally removed. Google's Street View cars only
    # cover public roadways, so for MH/RV parks (private roads inside the
    # property) the nearest panorama is almost always 100m+ away on the
    # public road outside. The "Main Entrance" and "Interior Road" images
    # were both rotations of that same external panorama — misleading,
    # not informative. Aerial/satellite is the right tool for property
    # context; reviewers can pop the Street View URL link below if they
    # want to spot-check approach quality.

    # CLICKABLE URLs
    url_start = aerial_start + 2 + IMG_HEIGHT_ROWS + 1
    section_header(url_start, "E", "  CLICKABLE URLs")
    visuals = market.get("visuals", {}) or {}
    fb_aerial = f"https://www.google.com/maps/search/?api=1&query={addr_encoded}&t=k" if addr_encoded else ""
    fb_street = f"https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={addr_encoded}" if addr_encoded else ""
    fb_dir = f"https://www.google.com/maps/dir/?api=1&destination={addr_encoded}" if addr_encoded else ""
    url_rows = [
        ("Aerial / Satellite", visuals.get("aerialView") or fb_aerial),
        ("Roadmap", fb_aerial.replace("&t=k", "") if fb_aerial else ""),
        ("Street View (Google Maps)", fb_street),
        ("Driving Directions", visuals.get("directions") or fb_dir),
    ]
    for i, (label, url) in enumerate(url_rows):
        r = url_start + 1 + i
        style_cell(ws.cell(row=r, column=2), label, bold=True, size=10)
        style_cell(ws.cell(row=r, column=3), url, color="0000FF", bg=LIGHT_YEL)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 18

    # LANDMARK DISTANCES
    lm_start = url_start + 1 + len(url_rows) + 1
    section_header(lm_start, "E", "  PROXIMITY TO KEY LANDMARKS")
    style_cell(ws.cell(row=lm_start+1, column=2), "Landmark Type",
               bold=True, color=WHITE, size=9, bg=GRAY_HDR, align="center")
    style_cell(ws.cell(row=lm_start+1, column=3), "Name / Address",
               bold=True, color=WHITE, size=9, bg=GRAY_HDR, align="center")
    ws.merge_cells(start_row=lm_start+1, start_column=3, end_row=lm_start+1, end_column=4)
    style_cell(ws.cell(row=lm_start+1, column=5), "Distance (mi)",
               bold=True, color=WHITE, size=9, bg=GRAY_HDR, align="center")

    # Fuzzy match: Claude often returns "Walmart" instead of "Nearest Walmart / Big Box"
    # Match by keyword presence rather than exact string equality
    raw_landmarks = market.get("landmarks") or []
    def find_landmark(target_label):
        target_lower = target_label.lower()
        # Extract key tokens from the target (skip "nearest", articles, etc.)
        target_keywords = [w for w in re.findall(r"\w+", target_lower)
                          if w not in {"nearest", "the", "a", "of", "or", "and", "/"}]
        # Score each AI-returned landmark by keyword overlap
        best_match = None
        best_score = 0
        for lm in raw_landmarks:
            lm_text = (lm.get("type", "") + " " + lm.get("name", "")).lower()
            score = sum(1 for kw in target_keywords if kw in lm_text)
            # Boost score for exact "type" match
            if lm.get("type", "").lower() == target_lower:
                score += 100
            if score > best_score:
                best_score = score
                best_match = lm
        return best_match if best_score > 0 else {}

    landmark_types = [
        "Nearest Major Employer #1", "Nearest Major Employer #2", "Nearest Major Employer #3",
        "Nearest Walmart / Big Box", "Nearest Grocery Store", "Nearest Hospital / Urgent Care",
        "Nearest Elementary School", "Nearest High School",
        "Nearest Community College / University", "Nearest Major Highway / Interstate",
        "Downtown (nearest city)", "Nearest Airport",
    ]
    used_landmarks = set()
    for i, lm_type in enumerate(landmark_types):
        r = lm_start + 2 + i
        lm = find_landmark(lm_type)
        # Avoid using the same landmark twice
        lm_id = id(lm) if lm else None
        if lm_id in used_landmarks:
            lm = {}
        else:
            if lm_id:
                used_landmarks.add(lm_id)
        style_cell(ws.cell(row=r, column=2), lm_type, bold=True, size=10)
        style_cell(ws.cell(row=r, column=3), lm.get("name", ""), color="0000FF", bg=LIGHT_YEL)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        style_cell(ws.cell(row=r, column=5), lm.get("distanceMiles"),
                   color="0000FF", bg=LIGHT_YEL, align="right")
        ws.row_dimensions[r].height = 18

    # DILIGENCE FLAGS
    flags_start = lm_start + 2 + len(landmark_types) + 1
    section_header(flags_start, "E", "  DILIGENCE FLAGS")
    flags = financials.get("flags", []) or []
    for i, f in enumerate(flags[:10]):
        r = flags_start + 1 + i
        sev = (f.get("severity", "") or "").lower()
        sev_color = {"high": "DC2626", "medium": "D97706", "low": "16A34A"}.get(sev, "6B7280")
        style_cell(ws.cell(row=r, column=2), f.get("severity", "").upper(),
                   bold=True, color=WHITE, size=9, bg=sev_color, align="center")
        msg = f"{f.get('item', '')}: {f.get('issue', '')} — Recommendation: {f.get('recommendation', '')}"
        style_cell(ws.cell(row=r, column=3), msg, size=10, wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 36

    # QUESTIONS FOR SELLER
    q_start = flags_start + 1 + len(flags[:10]) + 1
    section_header(q_start, "E", "  QUESTIONS FOR SELLER / BROKER")
    questions = financials.get("questions", []) or []
    for i, q in enumerate(questions[:10]):
        r = q_start + 1 + i
        style_cell(ws.cell(row=r, column=2), f"#{i+1}", bold=True, size=10, align="center")
        style_cell(ws.cell(row=r, column=3), q, size=10, wrap=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 30

    ws.sheet_properties.tabColor = "F59E0B"


# ═══════════════════════════════════════════════════════════════════════════
# JOB ORCHESTRATION
# ═══════════════════════════════════════════════════════════════════════════
def _set_job(job_id, **fields):
    """Atomically update a job's mutable fields under JOBS_LOCK."""
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(fields)
    # Hosted mode: mirror the light fields to Firestore so the web app's run
    # history survives instance restarts. The full result payload stays out —
    # Firestore docs cap at 1MB and the dashboard only needs status/progress;
    # /api/status keeps serving the full result from memory.
    _fb_run_upsert(job_id, {k: v for k, v in fields.items()
                            if k in ("status", "progress", "error")})


def run_analysis_job(job_id, api_key, file_blocks, property_info):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if job is None:
        return
    try:
        # Begin per-job token accounting. Every call_claude in this thread
        # appends its usage to the thread-local accumulator. Snapshot at the
        # end of the job so we can attach total $ cost to the result.
        reset_usage_tracking()
        _set_job(job_id, status="running", progress="Starting analysis...")

        deep_search = property_info.get("deepSearch", "off") == "on"
        skip_market = bool(property_info.get("_skipMarket"))
        market_fn = (lambda *a, **k: call_market_research_merged(*a, **k, n_runs=3)) if deep_search else call_market_research

        # The financial side is now a 4-step sequence with two opt-in stages:
        #   0. CACHE LOOKUP                   — fingerprint hit returns instantly
        #   1. EXTRACT (Fable 5, no thinking) — N=1 default, N=3 when deep_search
        #      → field-level median across N runs
        #   2. VERIFY  (pure Python)          — tie-outs, 2σ rents, POH, cross-doc
        #   3. METHODOLOGY (Fable 5, adaptive thinking) — GGC categorization +
        #      underwriting
        #   4. CACHE WRITE                    — if no hard-fail in verification
        # Market research is independent, so we run it in parallel with the
        # whole financial sequence.
        n_extract_runs     = (FINANCIAL_PARSE_RUNS_DEEP if deep_search
                              else FINANCIAL_PARSE_RUNS)
        n_methodology_runs = (METHODOLOGY_RUNS_DEEP     if deep_search
                              else METHODOLOGY_RUNS)
        # Form override beats the default when present (1, 3, or 5 only).
        n_runs_override = int(property_info.get("_nRunsOverride") or 0)
        if n_runs_override in (1, 3, 5):
            n_extract_runs = n_runs_override
            n_methodology_runs = n_runs_override
            print(f"[CostMode] User override: n_runs={n_runs_override} on both stages")

        def financial_pipeline():
            cache_key = extraction_cache_key(
                file_blocks, property_info, n_extract_runs,
                EXTRACTION_PROMPT, FINANCIAL_PARSE_PROMPT,
                n_methodology_runs=n_methodology_runs)
            cached = extraction_cache_get(cache_key)
            if cached is not None:
                print(f"[Cache] HIT key={cache_key[:8]}... — returning memoized "
                      "financials (no Claude calls this run)")
                cached.setdefault("_cache", {})
                cached["_cache"].update({"hit": True, "key": cache_key[:8]})
                return cached
            print(f"[Cache] MISS key={cache_key[:8]} "
                  f"(deep_search={deep_search}, "
                  f"n_extract={n_extract_runs}, n_method={n_methodology_runs})")

            if n_extract_runs > 1:
                extracted = call_extract_financials_merged(
                    api_key, file_blocks, property_info, n_runs=n_extract_runs)
            else:
                extracted = call_extract_financials(
                    api_key, file_blocks, property_info)
            # If the user left the unit count blank on the form, derive it
            # from the rent roll the extraction step transcribed. This makes
            # the field truly optional (the form's job is to capture domain
            # knowledge the docs don't carry, not to duplicate what they do)
            # and skips the rent-roll-rows-vs-units cross-check in
            # verify_extraction by leaving stated_units==0. The derived
            # value flows into the methodology prompt + apply_ggc_overrides
            # so management-fee tiering, etc. still work.
            try:
                stated_units_form = int(
                    str(property_info.get("units", "")).strip() or 0)
            except (ValueError, TypeError):
                stated_units_form = 0
            if stated_units_form == 0:
                derived = ((extracted.get("rentRoll") or {})
                           .get("totalRowsInRentRoll") or 0)
                if isinstance(derived, (int, float)) and derived > 0:
                    property_info["units"] = str(int(derived))
                    extracted["extractionNotes"] = (
                        (extracted.get("extractionNotes") or "")
                        + (" || " if extracted.get("extractionNotes") else "")
                        + f"Total Units auto-derived from rent roll "
                          f"({int(derived)} rows) — user left the field blank."
                    )
                    print(f"[Auto-derive] units = {int(derived)} "
                          f"(from rent-roll row count, form was blank)")
            checks = verify_extraction(extracted, property_info)
            print(f"[Verify/Extract] {len(checks)} checks: "
                  f"{sum(1 for c in checks if c['status'] == 'fail')} fail, "
                  f"{sum(1 for c in checks if c['status'] == 'warn')} warn")
            financials = call_parse_financials_merged(
                api_key, extracted, property_info, n_runs=n_methodology_runs)
            # Carry the user-provided county tax rate through into
            # financials.propertyInfo so fill_template can stamp it into
            # the Underwriting tab (P12) — the RE Taxes override formula
            # uses it as the preferred reassessment method.
            if property_info.get("countyTaxRate"):
                financials.setdefault("propertyInfo", {})["countyTaxRate"] = \
                    property_info.get("countyTaxRate")
            # Same pattern for the user's flood-zone input: methodology
            # may extract one from the OM but the user's form value is
            # more authoritative (they've consulted FEMA maps). Falls
            # back to the methodology value when no form input was given.
            if property_info.get("floodZone") and \
               str(property_info.get("floodZone")).strip().lower() not in ("", "unknown"):
                financials.setdefault("propertyInfo", {})["floodZone"] = \
                    property_info.get("floodZone")
            # Methodology-side checks run AFTER categorization because they
            # need both the income.ggcCategory tags and the rent roll's
            # canonical unit types. This is where the lot-rent / RV-rent
            # collapse bug would surface.
            checks.extend(verify_methodology(financials))
            # Recompute fail/warn counts now that methodology-side checks are
            # in. Earlier versions only counted extraction fails, so a clean
            # extraction with a broken methodology categorization slipped
            # through both the cache gate and the write-back gate.
            n_fail = sum(1 for c in checks if c["status"] == "fail")
            n_warn = sum(1 for c in checks if c["status"] == "warn")
            print(f"[Verify/Total] {len(checks)} checks: "
                  f"{n_fail} fail, {n_warn} warn")
            # Carry the raw extraction + checks through so they can be rendered
            # on the Extraction Check tab.
            financials["_extraction"] = extracted
            financials["_extractionChecks"] = checks
            financials["_verification"] = {
                "hardFails": n_fail,
                "warnings": n_warn,
                "failedCheckNames": [c["item"] for c in checks
                                     if c["status"] == "fail"][:20],
            }
            financials["_cache"] = {"hit": False, "key": cache_key[:8],
                                     "n_extract_runs": n_extract_runs}

            # Only memoize a clean result. If hard fails remain we want the
            # next run to re-try, not return the broken cached output.
            if n_fail == 0:
                extraction_cache_put(cache_key, financials)
            else:
                print(f"[Cache] Skipping write — {n_fail} hard failures remain")
            return financials

        # Skip market research when the user toggles it off — saves one Claude
        # call (the most expensive single call thanks to web_search + thinking).
        # The Excel template gracefully handles an empty market dict: comp
        # tables stay empty, the Comps Analysis tab notes "no data", and
        # Underwriting fields that key off market rents fall back to the
        # contracted lot rents on the rent roll.
        empty_market = {
            "rentComps": [], "saleComps": [], "demographics": {},
            "landmarks": [], "demandSignal": "skipped",
            "_skipped": "Market research skipped (cost-mode override).",
        }
        with ThreadPoolExecutor(max_workers=2) as executor:
            future_financials = executor.submit(financial_pipeline)
            if skip_market:
                future_market = None
            else:
                future_market = executor.submit(market_fn, api_key, property_info)

            results = {}
            futures = [future_financials] + ([future_market] if future_market else [])
            for future in as_completed(futures):
                if future is future_financials:
                    results["financials"] = future.result()
                    _set_job(job_id, progress="✓ Financials extracted, verified, and underwritten.")
                else:
                    results["market"] = future.result()
                    _set_job(job_id, progress="✓ Market research complete.")
            if skip_market:
                results["market"] = empty_market
                _set_job(job_id, progress="✓ Market research skipped per cost-mode.")

        # Verification surfacing (formerly a hard gate). We ALWAYS produce
        # the workbook now — the Extraction Check tab at sheet 0 carries
        # every failure with OK/WARN/FAIL coloring, and the API result also
        # carries the count + failed-check names so the result panel can
        # show a clear "review this before trusting" banner. Blocking the
        # workbook entirely was too aggressive: a reviewer would rather see
        # a flagged draft than nothing at all.
        verification = results["financials"].get("_verification") or {}
        n_fail = int(verification.get("hardFails") or 0)
        failed_names = verification.get("failedCheckNames") or []
        usage_summary = get_usage_summary()
        if n_fail > 0:
            preview = ", ".join(failed_names[:3])
            more = f" (+{len(failed_names) - 3} more)" if len(failed_names) > 3 else ""
            print(f"[{job_id}] Verification: {n_fail} hard fail(s) — "
                  f"workbook still produced. Failed checks: {preview}{more}")

        _set_job(job_id, progress="Filling GGC template...")
        output_path = JOBS_DIR / f"{job_id}.xlsx"
        fill_template(results["financials"], results["market"], output_path)
        # Hosted mode: persist the finished model to Firebase Storage so it
        # survives instance restarts and shows up in the web app's history.
        _fb_store_output(job_id, output_path)

        _set_job(job_id,
                 status="complete",
                 progress="Done.",
                 result={
                     "financials": results["financials"],
                     "market": results["market"],
                     "download_url": f"/api/download/{job_id}",
                     "verification": verification,
                     "usage": usage_summary,
                 })
        print(f"[{job_id}] Job cost: ${usage_summary['totals']['cost_usd']:.2f} "
              f"across {usage_summary['calls']} Claude calls "
              f"(deep_search={deep_search})")
    except Exception as e:
        # Log full traceback server-side; the /api/status response is
        # sanitized so the client never sees internal error text.
        traceback.print_exc()
        _set_job(job_id, status="error",
                 error=str(e)[:200],
                 progress="Error: analysis failed — check server logs.")


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════════════════════
@app.route("/")
def root():
    return send_from_directory(".", "index.html")


@app.route("/api/config")
def config():
    """Frontend pulls this on page load to learn what optional integrations
    the server has configured. The server's Anthropic key is NEVER returned
    over the wire — the frontend either provides its own key per request or
    falls back to the server key only inside `run_analysis_job`. Returning
    a key here would let any unauthenticated visitor read GGC's billing
    credential."""
    return jsonify({
        "default_api_key_present": bool(DEFAULT_ANTHROPIC_KEY),
        "google_maps_enabled": bool(GOOGLE_MAPS_API_KEY),
    })


@app.route("/api/analyze", methods=["POST"])
@require_auth
def analyze():
    api_key = (request.form.get("api_key") or "").strip() or DEFAULT_ANTHROPIC_KEY
    if not api_key:
        return jsonify({"error": "API key required"}), 400

    # Cost-mode controls. "economy" overrides extraction+methodology to the
    # cheap model, n_runs lets the user dial self-consistency down to 1, and
    # skip_market eliminates the market-research call entirely. These let
    # the user trade accuracy for spend during demos/testing without losing
    # the option to flip back to max-accuracy mode for real deals.
    cost_mode = (request.form.get("cost_mode", "max") or "max").lower()
    if cost_mode not in {"economy", "balanced", "max"}:
        cost_mode = "max"
    try:
        n_runs_override = int(request.form.get("n_runs", "0") or 0)
    except (TypeError, ValueError):
        n_runs_override = 0
    skip_market = (request.form.get("skip_market", "0") or "0") in ("1", "true", "on")

    property_info = {
        "name":        request.form.get("property_name", ""),
        "address":     request.form.get("address", ""),
        "city":        request.form.get("city", ""),
        "county":      request.form.get("county", ""),
        "countyTaxRate": request.form.get("county_tax_rate", ""),
        # Per-site RE tax assumption ($/unit/year). When provided, the
        # write-back overrides the template's J22 default. CorrectOutput
        # used $400/site for Whaleshead (148 sites × $400 = $59,200 I22).
        "taxPerSite":    request.form.get("tax_per_site", ""),
        "pohCount":    request.form.get("poh_count", "0"),
        "state":       request.form.get("state", ""),
        "units":       request.form.get("units", ""),
        "askingPrice": request.form.get("asking_price", ""),
        "floodZone":   request.form.get("flood_zone", "unknown"),
        "deepSearch":  request.form.get("deep_search", "off"),
        "_costMode":   cost_mode,
        "_nRunsOverride": n_runs_override,
        "_skipMarket": skip_market,
    }

    if not property_info["city"] or not property_info["state"]:
        return jsonify({"error": "City and State are required for market research"}), 400

    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "At least one file required"}), 400

    # Whitelist file extensions; reject anything else before reading bytes.
    for f in files:
        ext = (Path(f.filename or "").suffix or "").lower()
        if ext not in ALLOWED_UPLOAD_EXTS:
            return jsonify({"error": f"Unsupported file type: {ext or '<none>'}. "
                                     f"Allowed: {sorted(ALLOWED_UPLOAD_EXTS)}"}), 400

    file_blocks = [encode_file_for_claude(f) for f in files]
    job_id = _new_job_id()
    with JOBS_LOCK:
        JOBS[job_id] = {"status": "queued", "progress": "Queued", "result": None,
                        # Ownership for hosted mode: only the creator's uid may
                        # poll or download this job (None when auth is off).
                        "uid": getattr(g, "user_uid", None),
                        "email": getattr(g, "user_email", None)}
        _evict_old_jobs()
    _fb_run_create(job_id, property_info)

    Thread(target=run_analysis_job, args=(job_id, api_key, file_blocks, property_info),
           daemon=True).start()

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
@require_auth
def status(job_id):
    # Validate the path component before any dict / filesystem lookup so
    # /api/status/../../etc/passwd can't even probe state.
    if not _valid_job_id(job_id):
        return jsonify({"error": "Invalid job id"}), 400
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return jsonify({"error": "Job not found"}), 404
        # Hosted mode: a signed-in user may only see their own jobs. 404 (not
        # 403) so other users' job ids aren't confirmed to exist.
        if REQUIRE_AUTH and job.get("uid") != g.user_uid:
            return jsonify({"error": "Job not found"}), 404
        # Return progress fields + the full analysis result (financials,
        # market, download_url) — the frontend's showResults() reads
        # result.financials / result.market to populate the KPI cards.
        # Internal error strings are still scrubbed: a server-side
        # exception text may include prompt snippets or API-response
        # excerpts and isn't useful to the client anyway.
        public = {
            "status":   job.get("status"),
            "progress": job.get("progress"),
            "result":   job.get("result") if isinstance(job.get("result"), dict) else None,
        }
        if job.get("status") == "error":
            public["error"] = "Analysis failed — check server logs."
    return jsonify(public)


@app.route("/api/download/<job_id>")
@require_auth
def download(job_id):
    if not _valid_job_id(job_id):
        return jsonify({"error": "Invalid job id"}), 400
    if REQUIRE_AUTH:
        # Same ownership rule as /api/status. If the job aged out of memory,
        # the durable copy lives in Firebase Storage (runs/{uid}/{job_id}.xlsx)
        # and the web app downloads it there with per-user rules instead.
        with JOBS_LOCK:
            owner = (JOBS.get(job_id) or {}).get("uid")
        if owner != g.user_uid:
            return jsonify({"error": "Job not found"}), 404
    file_path = JOBS_DIR / f"{job_id}.xlsx"
    # Defense in depth: even with the regex guard, resolve and verify the
    # final path stays inside JOBS_DIR before serving.
    try:
        resolved = file_path.resolve(strict=True)
        if JOBS_DIR.resolve() not in resolved.parents:
            return jsonify({"error": "Invalid job id"}), 400
    except FileNotFoundError:
        return jsonify({"error": "File not ready"}), 404
    with JOBS_LOCK:
        name = (JOBS.get(job_id, {}).get("result", {}) or {}) \
            .get("financials", {}).get("propertyInfo", {}).get("name", "Property")
    safe = re.sub(r"[^A-Za-z0-9_-]", "_", name)[:40] or "Property"
    return send_file(file_path, as_attachment=True, download_name=f"GGC_UW_{safe}.xlsx")


if __name__ == "__main__":
    # Cloud Run injects PORT; localhost keeps the historical 5001. (In the
    # container this file is served by gunicorn instead — see Dockerfile.)
    _port = int(os.environ.get("PORT", "5001"))
    print(" ╔═════════════════════════════════════════════════════════════════════════════════════╗")
    print(" ║  GGC Deal Engine — Backend Server v6 (playbook hardening)                           ║")
    print(f"║  Extraction:   {MODEL_EXTRACTION:<48s}                                 ║")
    print(f"║  Methodology:  {MODEL_METHODOLOGY:<48s}                                 ║")
    print(f"║  Market:       {MODEL_MARKET:<48s}                                 ║")
    print(f"║  Pipeline: parse → extract (N={FINANCIAL_PARSE_RUNS_DEEP} deep) → verify → methodology → cache → fill ║")
    print(f"║  Parser:       {PARSER_BACKEND:<48s}                                 ║")
    print(f"║  Structured outputs (beta):  {'ENABLED' if USE_STRUCTURED_OUTPUTS else 'DISABLED':<48s}             ║")
    print(f"║  Extraction cache:           {'ENABLED' if EXTRACTION_CACHE_ENABLED else 'DISABLED':<48s}             ║")
    print(f"║  Validator retries:          {MAX_PARSE_RETRIES}                                                      ║")
    print(f"║  Template: GGC_Blank_Underwriting_Sizer_Extended (1000 rows)                        ║")
    print(f"║  Google Maps: {'ENABLED' if GOOGLE_MAPS_API_KEY else 'DISABLED (no key set)':<43s}  ║")
    print(f"║  Document AI: {'ENABLED' if DOC_AI_ENABLED else 'DISABLED (no GCP config)':<43s} ║")
    print(f"║  Auth: {'REQUIRED (Firebase)' if REQUIRE_AUTH else 'off (local mode)':<43s}                                  ║")
    print(f"║  Open: http://localhost:{_port:<5d}                                                       ║")
    print(" ╚═════════════════════════════════════════════════════════════════════════════════════╝")
    app.run(host="0.0.0.0", port=_port, debug=False, threaded=True)
