"""Pull every GGC category string used on the Data Consolidation sheet
plus the seller's source label and rough placement (income vs expense)."""
from openpyxl import load_workbook
PATH = "/Users/nicholasrevencomacbookair/Desktop/GGC/V5/ggc-deal-engine/Outputs/CorrectOutput copy.xlsx"
wb = load_workbook(PATH, data_only=True)
ws = wb["Data Consolidation"]

# rows 3..36 are income, rows 43..102 are expenses (per UW SUMIFS ranges)
def section(start, end, label):
    print(f"\n===== {label} (rows {start}-{end}) =====")
    print(f"{'ROW':<5} {'GGC CATEGORY (col A)':<40} {'SOURCE LABEL (col B)':<60} {'T12 (G)':>12}")
    print("-" * 130)
    for r in range(start, end + 1):
        cat = ws.cell(row=r, column=1).value
        src = ws.cell(row=r, column=2).value
        t12 = ws.cell(row=r, column=7).value
        if cat or src or (t12 not in (None, 0)):
            cat_s = str(cat or "")[:38]
            src_s = str(src or "")[:58].strip()
            t12_s = f"{t12:,.2f}" if isinstance(t12, (int, float)) else str(t12 or "")
            print(f"{r:<5} {cat_s:<40} {src_s:<60} {t12_s:>12}")

section(3, 36, "INCOME")
section(43, 102, "EXPENSES")

# Print all unique GGC category strings (col A)
print("\n\n===== UNIQUE GGC CATEGORY STRINGS (col A across whole sheet) =====")
seen = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is None:
        continue
    s = str(v).strip()
    if not s or s in seen:
        continue
    seen.append(s)
for s in seen:
    print(repr(s))
